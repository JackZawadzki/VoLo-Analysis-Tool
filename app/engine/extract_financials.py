"""
extract_financials.py — single-source-of-truth financial extraction.

Built to fix a specific class of bug in the legacy extractor: pulling values
from multiple sheets and concatenating them along the time axis. The legacy
extractor would silently switch sources at the year one sheet ran out of
columns, producing sequences like $923, $923, $923, $5.4B, $5.4B, ...

Guarantees provided by this extractor:
  1. Every row in the output comes from exactly ONE sheet.
  2. If the chosen sheet doesn't have year Y, the output value is null.
     Never fall back to a different sheet for "missing" years.
  3. Scope is named explicitly: output.scope.sheet + scope_description.
  4. Candidates are discovered, ranked, and one is picked with a recorded
     selection_rationale.
  5. Self-verification runs after extraction; broken selections are
     detected and the agent either re-selects or surfaces the ambiguity.

Design is workbook-agnostic. No hardcoded sheet names, row indices, or
column positions. Discovery is driven by label heuristics and year-header
detection.

Usage:
    from app.engine.extract_financials import extract
    result = extract("/path/to/workbook.xlsx")
    # result is a dict matching the schema documented in README below.

    # Command line:
    python -m app.engine.extract_financials path/to/workbook.xlsx
"""
from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

# ─────────────────────────────────────────────────────────────────────
# Configuration — canonical metrics and their synonyms
# ─────────────────────────────────────────────────────────────────────
# Each canonical metric has a primary label and synonyms. When multiple
# matches are found in a sheet, we prefer the "Total X" variant and drop
# header-only (value-less) rows. Sign conventions are used in verification
# to flag rows where numbers contradict expectations.

CANONICAL_METRICS: Dict[str, Dict[str, Any]] = {
    "revenue": {
        "synonyms": [
            "total revenue", "total facility revenue", "product revenue",
            "net revenue", "total sales", "net sales", "sales revenue",
            "revenue", "sales",
        ],
        "expected_sign": "positive",
        "prefer": ["total"],  # prefer synonym containing this word
    },
    "cogs": {
        "synonyms": [
            "total cogs", "cost of goods sold", "cost of sales",
            "total cost of revenue", "cogs",
        ],
        "expected_sign": "any",  # some models show COGS positive, some negative
        "prefer": ["total"],
    },
    "gross_profit": {
        "synonyms": [
            "gross profit", "total gross profit", "gross margin $",
            "gross profit/(loss)",
        ],
        "expected_sign": "any",
        "prefer": ["total"],
    },
    "operating_expenses": {
        "synonyms": [
            "total operating expenses", "total opex", "operating expenses",
            "total operating costs", "sg&a", "total sg&a",
        ],
        "expected_sign": "any",
        "prefer": ["total"],
    },
    "ebitda": {
        "synonyms": [
            "adj. ebitda", "adjusted ebitda", "ebitda",
        ],
        "expected_sign": "any",
        "prefer": ["adj", "total"],
    },
    "operating_income": {
        "synonyms": [
            "operating income", "operating profit", "ebit",
            "operating income / (loss)",
        ],
        "expected_sign": "any",
        "prefer": [],
    },
    "net_income": {
        "synonyms": [
            "net income", "net profit", "net earnings", "net income / (loss)",
        ],
        "expected_sign": "any",
        "prefer": [],
    },
    "capex": {
        "synonyms": [
            "total capex", "capital expenditures", "capital expenditure",
            "(-) capex", "capex", "capex spend", "gross capex",
        ],
        "expected_sign": "any",
        "prefer": ["total", "gross"],
    },
}

# Tokens that make a sheet name LOOK like a canonical annual P&L.
SCOPE_BONUS_TOKENS = {
    "combined": 5, "unified": 5, "consolidated": 5, "consolidation": 5,
    "group": 3, "total": 3, "summary": 2, "overview": 2,
    "annual": 2,  # "FS - Annual" is P&L-ish but single-scope
    "fs": 2, "financials": 2, "p&l": 3, "pnl": 3,
    "income statement": 4, "statements": 2,
}
# Tokens that suggest a single-facility / single-line view — penalized.
SCOPE_PENALTY_TOKENS = {
    "l1": 4, "l2": 4, "line 1": 4, "line 2": 4, "line1": 4, "line2": 4,
    "facility": 3, "plant": 3, "standalone": 4,
    # Specific-place suffixes (Smyrna, Muskegon, Bowling Green, Kentucky
    # in Mitra; Plano, Austin in other workbooks). We match known tokens
    # plus a general rule at runtime that any sheet name ending in a
    # proper noun but lacking scope-bonus tokens gets a small penalty.
}
# Purely non-P&L sheets — heavy penalty so they never bubble up.
NON_PL_TOKENS = {
    "changelog": 20, "convention": 20, "scenario": 10,
    "assumption": 8, "input": 8,
    "waterfall": 6, "distribution": 6,
    "capital": 6,  # capital structure, not P&L
    "opco": 6,  # operator agreements
    "balance sheet": 4, "bs ": 4, "cfs ": 4, "cash flow": 4,
    "dcf": 4,  # DCF is a valuation, not source P&L
    "ira": 4, "financing": 4, "working capital": 4,
    "raw material": 4, "expenses": 4,  # these are supporting schedules
    "revenue": 0,  # supporting schedule may still be useful
}

# Year range we consider plausible.
MIN_YEAR, MAX_YEAR = 2015, 2055

# Minimum years on a year-row for it to count as a year axis.
MIN_YEAR_COUNT = 4

# Max rows / cols we'll scan per sheet (performance cap).
MAX_SCAN_ROWS = 200
MAX_SCAN_COLS = 60

# First N columns to search for metric labels.
LABEL_COL_SEARCH_WIDTH = 10


# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────

def _norm(s: Any) -> str:
    """Normalize a label string for matching: lowercase, collapse whitespace,
    strip leading/trailing punctuation."""
    if s is None:
        return ""
    t = str(s).lower().strip()
    # Remove common decorative prefixes: "(-) ", "+ ", "- ", bullets
    t = re.sub(r"^[\-\+\(\)\s\u2022\u25aa\u25ba]+", "", t)
    # Collapse whitespace
    t = re.sub(r"\s+", " ", t)
    # Strip trailing parenthetical notes like "(net)"
    t = re.sub(r"\s*\([^)]*\)\s*$", "", t).strip()
    return t


def _is_year(v: Any) -> bool:
    if isinstance(v, bool):  # in Python, bool is an int subclass — exclude it
        return False
    if isinstance(v, int) and MIN_YEAR <= v <= MAX_YEAR:
        return True
    if isinstance(v, float) and v.is_integer() and MIN_YEAR <= int(v) <= MAX_YEAR:
        return True
    if isinstance(v, str):
        m = re.fullmatch(r"\s*(20\d{2})\s*", v)
        if m:
            y = int(m.group(1))
            if MIN_YEAR <= y <= MAX_YEAR:
                return True
    return False


def _to_year(v: Any) -> Optional[int]:
    if isinstance(v, bool):
        return None
    if isinstance(v, int) and MIN_YEAR <= v <= MAX_YEAR:
        return int(v)
    if isinstance(v, float) and v.is_integer() and MIN_YEAR <= int(v) <= MAX_YEAR:
        return int(v)
    if isinstance(v, str):
        m = re.fullmatch(r"\s*(20\d{2})\s*", v)
        if m:
            y = int(m.group(1))
            if MIN_YEAR <= y <= MAX_YEAR:
                return y
    return None


def _is_numeric(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _detect_unit(label: str, nearby_cells: List[Any]) -> Optional[str]:
    """Look for [$mn], ($M), in millions, etc. in the label itself and in a few
    cells immediately after the label column."""
    candidates = [label] + [str(c) for c in nearby_cells if c is not None]
    for c in candidates:
        s = str(c)
        m = re.search(r"\[\s*\$\s*(mm|mn|m|k|bn|b)\s*\]", s, re.IGNORECASE)
        if m:
            tok = m.group(1).lower()
            return {"mm": "USD_M", "mn": "USD_M", "m": "USD_M",
                    "k": "USD_K", "bn": "USD_B", "b": "USD_B"}[tok]
        if re.search(r"\$\s*(in\s+)?millions", s, re.IGNORECASE):
            return "USD_M"
        if re.search(r"\$\s*(in\s+)?thousands", s, re.IGNORECASE):
            return "USD_K"
        if re.search(r"\$\s*(in\s+)?billions", s, re.IGNORECASE):
            return "USD_B"
        if re.search(r"\b(gwh|kwh|mwh|mw|kw)\b", s, re.IGNORECASE):
            return "ENERGY"
        if re.search(r"\btons?\b|\btonnes?\b", s, re.IGNORECASE):
            return "MASS"
    return None


# ─────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────

@dataclass
class YearAxis:
    """A row in a sheet that serves as a year header."""
    row_idx: int            # 0-indexed
    col_to_year: Dict[int, int]  # {col_idx: year}
    span: int               # count of year columns
    year_min: int
    year_max: int


@dataclass
class MetricHit:
    """A row that matched a canonical metric's synonym."""
    canonical: str          # "revenue", "cogs", etc.
    matched_label: str      # the raw label we saw
    row_idx: int
    label_col: int
    synonym_matched: str
    has_total_prefix: bool  # label contains a preferred token (e.g. "total")
    numeric_cell_count: int # how many year columns had numeric values


@dataclass
class SheetSignal:
    """Everything we learned about a sheet during discovery."""
    name: str
    year_axes: List[YearAxis]
    metric_hits: Dict[str, List[MetricHit]]  # canonical -> list of hits
    rows: List[List[Any]]    # raw grid (bounded)
    label_col_used: Optional[int]  # most common label column across hits
    score: float = 0.0
    score_breakdown: Dict[str, float] = field(default_factory=dict)
    disqualified: bool = False
    disqualification_reason: Optional[str] = None


# ─────────────────────────────────────────────────────────────────────
# Discovery — find candidate P&L sheets and what they contain
# ─────────────────────────────────────────────────────────────────────

def _load_grid(ws) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row or 0, MAX_SCAN_ROWS),
        max_col=min(ws.max_column or 0, MAX_SCAN_COLS),
        values_only=True,
    ):
        rows.append(list(row))
    return rows


def _find_year_axes(rows: List[List[Any]]) -> List[YearAxis]:
    """Find every row that has ≥ MIN_YEAR_COUNT integer-year cells, taking
    the LONGEST CONTIGUOUS RUN of years in that row as the axis. A year
    row only counts when adjacent years are within 1 of each other (we
    accept gaps of up to 1 to tolerate blank separator columns)."""
    axes: List[YearAxis] = []
    for r_idx, row in enumerate(rows):
        year_cells: List[Tuple[int, int]] = []
        for c_idx, v in enumerate(row):
            y = _to_year(v)
            if y is not None:
                year_cells.append((c_idx, y))
        if len(year_cells) < MIN_YEAR_COUNT:
            continue
        # Take longest run where adjacent years differ by 1
        best: List[Tuple[int, int]] = []
        cur: List[Tuple[int, int]] = []
        for pair in year_cells:
            if not cur or pair[1] == cur[-1][1] + 1:
                cur.append(pair)
            elif pair[1] == cur[-1][1]:
                # duplicate year — skip
                continue
            else:
                if len(cur) > len(best):
                    best = cur
                cur = [pair]
        if len(cur) > len(best):
            best = cur
        if len(best) >= MIN_YEAR_COUNT:
            col_to_year = {c: y for c, y in best}
            axes.append(YearAxis(
                row_idx=r_idx,
                col_to_year=col_to_year,
                span=len(best),
                year_min=best[0][1],
                year_max=best[-1][1],
            ))
    return axes


def _find_metric_hits(rows: List[List[Any]],
                      year_axes: List[YearAxis]) -> Dict[str, List[MetricHit]]:
    """For each canonical metric, find every row whose label matches a
    synonym. A hit is qualified only if the row has at least one numeric
    value in a year column of some year axis (rules out pure section
    headers).
    """
    year_cols_all = set()
    for ax in year_axes:
        year_cols_all.update(ax.col_to_year.keys())

    hits: Dict[str, List[MetricHit]] = {k: [] for k in CANONICAL_METRICS}

    for r_idx, row in enumerate(rows):
        # Find the label cell (leftmost string in the first N cols)
        label_col: Optional[int] = None
        label_str: Optional[str] = None
        for c_idx in range(min(LABEL_COL_SEARCH_WIDTH, len(row))):
            v = row[c_idx]
            if isinstance(v, str) and v.strip():
                label_col = c_idx
                label_str = v
                break
        if label_str is None:
            continue
        norm = _norm(label_str)
        if not norm:
            continue

        # Count numeric cells in year columns
        numeric_in_year_cols = sum(
            1 for c in year_cols_all if c < len(row) and _is_numeric(row[c])
        )

        # Match against each canonical metric
        for canonical, cfg in CANONICAL_METRICS.items():
            best_synonym = None
            for syn in cfg["synonyms"]:
                if syn == norm or syn in norm.split(" / ") or syn in norm:
                    # Require word-boundary-ish match for short synonyms
                    if len(syn.split()) == 1 and len(syn) <= 5:
                        # e.g. "ebit" or "sales" — require word boundary
                        if re.search(rf"\b{re.escape(syn)}\b", norm):
                            best_synonym = syn
                            break
                    else:
                        best_synonym = syn
                        break
            if best_synonym is None:
                continue
            # Skip if no numeric values at all (header-only row)
            if numeric_in_year_cols == 0:
                continue
            has_pref = any(tok in norm for tok in cfg.get("prefer", []))
            hits[canonical].append(MetricHit(
                canonical=canonical,
                matched_label=label_str,
                row_idx=r_idx,
                label_col=label_col,
                synonym_matched=best_synonym,
                has_total_prefix=has_pref,
                numeric_cell_count=numeric_in_year_cols,
            ))
            break  # row counted for at most one canonical metric

    # Sort each metric's hits so the preferred row comes first
    for canonical, lst in hits.items():
        lst.sort(key=lambda h: (
            not h.has_total_prefix,       # prefer "total X" variants first
            -h.numeric_cell_count,        # then rows with more data
            len(h.matched_label),         # then shorter labels
        ))
    return hits


def discover_sheets(wb) -> List[SheetSignal]:
    signals: List[SheetSignal] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = _load_grid(ws)
        if not rows:
            continue
        year_axes = _find_year_axes(rows)
        metric_hits = _find_metric_hits(rows, year_axes)
        # Dominant label column across all hits (helps later when we pick
        # the "primary" one).
        from collections import Counter
        cols_used = Counter()
        for hits in metric_hits.values():
            for h in hits:
                cols_used[h.label_col] += 1
        label_col_used = cols_used.most_common(1)[0][0] if cols_used else None
        signals.append(SheetSignal(
            name=name,
            year_axes=year_axes,
            metric_hits=metric_hits,
            rows=rows,
            label_col_used=label_col_used,
        ))
    return signals


# ─────────────────────────────────────────────────────────────────────
# Ranking — score each sheet for "is this the canonical annual P&L?"
# ─────────────────────────────────────────────────────────────────────

def _score_sheet(s: SheetSignal) -> None:
    """Populate s.score, s.score_breakdown, and s.disqualified in place."""
    bd: Dict[str, float] = {}
    low_name = s.name.lower()

    # --- Disqualifiers ---
    # No year axis → cannot be a P&L.
    if not s.year_axes:
        s.disqualified = True
        s.disqualification_reason = "no year axis detected"
        return
    # No matched P&L metrics → not a P&L.
    total_hits = sum(len(v) for v in s.metric_hits.values())
    if total_hits == 0:
        s.disqualified = True
        s.disqualification_reason = "no P&L metric labels matched"
        return
    # Strong non-P&L keyword match → penalize heavily but not DQ outright.
    non_pl_penalty = 0.0
    for tok, weight in NON_PL_TOKENS.items():
        if tok in low_name:
            non_pl_penalty += weight
    bd["non_pl_keyword_penalty"] = -non_pl_penalty

    # --- Metric completeness ---
    # How many canonical metrics did we find?
    found_metrics = [k for k, v in s.metric_hits.items() if v]
    bd["metric_completeness"] = 3.0 * len(found_metrics)
    # Bonus for having the core P&L triad
    core = {"revenue", "cogs", "gross_profit"}
    if core.issubset(set(found_metrics)):
        bd["core_pl_triad_bonus"] = 6.0
    # Bonus for having EBITDA / Net Income
    if "ebitda" in found_metrics:
        bd["ebitda_bonus"] = 2.0
    if "net_income" in found_metrics:
        bd["net_income_bonus"] = 2.0
    if "capex" in found_metrics:
        bd["capex_bonus"] = 2.0

    # --- Year span ---
    best_axis = max(s.year_axes, key=lambda a: a.span)
    bd["year_span"] = 0.5 * best_axis.span  # 20 years = 10 points

    # --- Scope breadth (sheet-name keywords) ---
    scope_bonus = 0.0
    for tok, weight in SCOPE_BONUS_TOKENS.items():
        if tok in low_name:
            scope_bonus += weight
    bd["scope_bonus"] = scope_bonus

    scope_penalty = 0.0
    for tok, weight in SCOPE_PENALTY_TOKENS.items():
        if tok in low_name:
            scope_penalty += weight
    # Also: any sheet name that looks like a plant/city name (ends in a
    # capitalized proper noun not in the bonus list) gets a small penalty.
    # We detect this by the presence of an underscore/space followed by
    # a capitalized word that isn't a known scope token.
    words = re.split(r"[\s_\-]+", s.name)
    place_like = any(
        w and w[0].isupper() and w.lower() not in SCOPE_BONUS_TOKENS
        and w.lower() not in ("fs", "pnl", "p&l", "income")
        for w in words[1:]  # skip first word (often just "FS")
    )
    if place_like and scope_bonus == 0:
        scope_penalty += 1
    bd["scope_penalty"] = -scope_penalty

    # --- Data density on revenue row (if found) ---
    rev_hits = s.metric_hits.get("revenue", [])
    if rev_hits:
        top = rev_hits[0]
        # Ratio of filled year cols
        density = top.numeric_cell_count / max(best_axis.span, 1)
        bd["revenue_density"] = 4.0 * min(density, 1.0)

    s.score = sum(bd.values())
    s.score_breakdown = bd


def rank_sheets(signals: List[SheetSignal]) -> List[SheetSignal]:
    for s in signals:
        _score_sheet(s)
    candidates = [s for s in signals if not s.disqualified]
    candidates.sort(key=lambda s: s.score, reverse=True)
    return candidates


# ─────────────────────────────────────────────────────────────────────
# Extraction — pull one metric row per canonical metric, strictly from
#              the chosen sheet, mapped against the chosen year axis.
# ─────────────────────────────────────────────────────────────────────

def _pick_year_axis_for_metrics(s: SheetSignal) -> Optional[YearAxis]:
    """Choose the year axis nearest ABOVE the majority of metric rows.
    If multiple axes exist (e.g. FS_Combined has rows 6 and 7), prefer
    the one that covers the MOST metric row columns with its own columns.
    """
    if not s.year_axes:
        return None
    # Columns used by metric rows (across all canonical metrics)
    all_metric_cols: set = set()
    for hits in s.metric_hits.values():
        for h in hits:
            row = s.rows[h.row_idx]
            for c_idx, v in enumerate(row):
                if _is_numeric(v) and c_idx > h.label_col:
                    all_metric_cols.add(c_idx)
    if not all_metric_cols:
        # Just pick the widest axis above any metric row
        best = max(s.year_axes, key=lambda a: a.span)
        return best
    best_axis = None
    best_overlap = -1
    # Only consider axes ABOVE the first metric hit row
    first_metric_row = min(
        (h.row_idx for hits in s.metric_hits.values() for h in hits),
        default=9999,
    )
    for ax in s.year_axes:
        if ax.row_idx >= first_metric_row:
            continue
        overlap = len(set(ax.col_to_year.keys()) & all_metric_cols)
        # Tie-break: wider axis wins
        key = (overlap, ax.span)
        if best_axis is None or key > (best_overlap, best_axis.span):
            best_axis = ax
            best_overlap = overlap
    if best_axis is None:
        best_axis = max(s.year_axes, key=lambda a: a.span)
    return best_axis


def extract_from_sheet(s: SheetSignal) -> Dict[str, Any]:
    """Pull canonical metrics from `s` using a single year axis. Every
    output value is from `s` or is null. Never reach into another sheet."""
    axis = _pick_year_axis_for_metrics(s)
    if axis is None:
        return {"error": "no year axis", "sheet": s.name}

    year_cols = axis.col_to_year                       # col_idx -> year
    all_years = sorted(year_cols.values())

    metrics_out: Dict[str, Any] = {}
    for canonical, cfg in CANONICAL_METRICS.items():
        hits = s.metric_hits.get(canonical, [])
        if not hits:
            metrics_out[canonical] = None
            continue
        top = hits[0]
        row = s.rows[top.row_idx]
        # Detect unit from label or nearby cells (immediately after label col)
        nearby = [row[top.label_col + i]
                  for i in range(1, 5)
                  if top.label_col + i < len(row)]
        unit = _detect_unit(top.matched_label, nearby)
        # Extract values year-by-year
        values: Dict[str, Optional[float]] = {}
        for col_idx, year in sorted(year_cols.items()):
            if col_idx >= len(row):
                values[str(year)] = None
                continue
            v = row[col_idx]
            if _is_numeric(v):
                values[str(year)] = float(v)
            else:
                values[str(year)] = None
        metrics_out[canonical] = {
            "label": top.matched_label,
            "synonym_matched": top.synonym_matched,
            "source_row_idx_0based": top.row_idx,
            "source_row_excel_addr": openpyxl.utils.get_column_letter(top.label_col + 1) + str(top.row_idx + 1),
            "unit": unit,
            "values": values,
            "alternates_not_picked": [
                {"label": h.matched_label, "row_idx": h.row_idx}
                for h in hits[1:4]
            ],
        }

    return {
        "sheet": s.name,
        "year_axis": {
            "row_idx_0based": axis.row_idx,
            "year_min": axis.year_min,
            "year_max": axis.year_max,
            "years_covered": all_years,
        },
        "metrics": metrics_out,
    }


# ─────────────────────────────────────────────────────────────────────
# Verification — sanity-check the extracted values. Catches the class
#                of bug the legacy extractor produced.
# ─────────────────────────────────────────────────────────────────────

def _extract_numeric_series(metric_block: Dict[str, Any]) -> Dict[int, float]:
    """Return {year:int -> value:float} for cells that have a numeric value."""
    if not metric_block or "values" not in metric_block:
        return {}
    out: Dict[int, float] = {}
    for y_str, v in metric_block["values"].items():
        if v is None:
            continue
        try:
            out[int(y_str)] = float(v)
        except Exception:
            pass
    return out


def verify_extraction(extraction: Dict[str, Any]) -> Dict[str, Any]:
    """Run sanity checks on a single-sheet extraction and return a
    verification block with warnings/errors."""
    checks_passed: List[str] = []
    warnings: List[str] = []
    errors: List[str] = []

    metrics = extraction.get("metrics", {})
    rev = _extract_numeric_series(metrics.get("revenue"))
    cogs = _extract_numeric_series(metrics.get("cogs"))
    gp = _extract_numeric_series(metrics.get("gross_profit"))
    ebitda = _extract_numeric_series(metrics.get("ebitda"))

    # 1. Gross Profit ≈ Revenue - COGS (within 1%)
    if rev and cogs and gp:
        discrepancies: List[str] = []
        for y in sorted(set(rev) & set(cogs) & set(gp)):
            r, c, g = rev[y], cogs[y], gp[y]
            # COGS sign convention varies — test both.
            expected_a = r - abs(c)
            expected_b = r - c  # if model stores COGS negative, subtract as-is
            # Pick the closer of the two expected values.
            expected = min((expected_a, expected_b),
                           key=lambda x: abs(x - g))
            denom = max(abs(r), 1.0)
            if abs(g - expected) / denom > 0.01:
                discrepancies.append(
                    f"{y}: R={r:.1f} C={c:.1f} GP={g:.1f} (expected ~{expected:.1f})"
                )
        if discrepancies:
            warnings.append(
                "gross_profit != revenue - cogs beyond 1% for "
                f"{len(discrepancies)} year(s); first: {discrepancies[0]}"
            )
        else:
            checks_passed.append("gross_profit ≈ revenue - cogs for all overlapping years")

    # 2. Monotonic/reasonable revenue growth — flag any >5x jump year-to-year.
    #    Elevate to ERROR when the jump is surrounded by plateaus on both
    #    sides: that signature (flat → jump → flat) is diagnostic of a
    #    concatenation/source-switch bug. A legitimate commercial ramp shows
    #    multi-year acceleration, not flat-then-jump-then-flat.
    if rev:
        years_sorted = sorted(rev)
        suspicious: List[Tuple[int, int, float, float]] = []
        for prev_y, cur_y in zip(years_sorted, years_sorted[1:]):
            pv, cv = rev[prev_y], rev[cur_y]
            if pv != 0 and abs(cv / pv) >= 5.0:
                suspicious.append((prev_y, cur_y, pv, cv))

        def _is_plateau(years_window: List[int]) -> bool:
            """True if all non-null values in window are within 1% of each
            other (3+ flat consecutive years)."""
            vals = [rev[y] for y in years_window if y in rev]
            if len(vals) < 3:
                return False
            mx, mn = max(vals), min(vals)
            return mx > 0 and (mx - mn) / mx < 0.01

        for prev_y, cur_y, pv, cv in suspicious:
            pre_window = [y for y in years_sorted if prev_y - 3 <= y <= prev_y]
            post_window = [y for y in years_sorted if cur_y <= y <= cur_y + 3]
            plateau_sandwich = _is_plateau(pre_window) and _is_plateau(post_window)
            msg = f"{prev_y}->{cur_y}: {pv:.1f} -> {cv:.1f} ({cv/pv:.1f}x)"
            if plateau_sandwich:
                errors.append(
                    f"revenue shows plateau->jump->plateau pattern at {msg}. "
                    "This is diagnostic of a source-switch/concatenation bug "
                    "(values were likely pulled from two different sheets). "
                    "The extractor should re-select its source sheet."
                )
            else:
                warnings.append(
                    f"revenue has a >=5x year-over-year jump: {msg} "
                    "(could be legitimate commercial ramp; verify manually)"
                )
        if not suspicious:
            checks_passed.append("no >=5x revenue jumps detected")

    # 3. Sign consistency within each metric
    for canonical, series in (("revenue", rev), ("cogs", cogs),
                              ("ebitda", ebitda), ("gross_profit", gp)):
        if not series:
            continue
        positives = sum(1 for v in series.values() if v > 0)
        negatives = sum(1 for v in series.values() if v < 0)
        if positives and negatives:
            # Some metrics legitimately cross zero (EBITDA, Net Income).
            # But Revenue should be strictly non-negative.
            if canonical == "revenue":
                errors.append(
                    f"{canonical} has both positive and negative values "
                    f"(pos={positives}, neg={negatives})"
                )
            else:
                # Informational — not a hard error.
                checks_passed.append(
                    f"{canonical} crosses zero (pos={positives}, neg={negatives}); "
                    "consistent with J-curve / ramp"
                )

    # 4. Unit consistency — check that at least one metric has a unit
    unit_info: Dict[str, Optional[str]] = {}
    for name, m in metrics.items():
        if m:
            unit_info[name] = m.get("unit")
    declared_units = {u for u in unit_info.values() if u}
    if len(declared_units) > 1:
        warnings.append(
            f"mixed unit markers across metrics: {unit_info}. "
            "This may be OK (ENERGY + USD_M is legitimate for battery models) "
            "but verify revenue is USD, not GWh."
        )
    elif declared_units:
        checks_passed.append(f"units consistent across metrics ({declared_units.pop()})")
    else:
        warnings.append(
            "no explicit unit markers found on any metric row; scale "
            "inference will rely on magnitude heuristics downstream"
        )

    # 5. Plateau detection — every metric plateauing at identical value for
    #    many years can indicate stale formulas. Informational only.
    if rev and len(rev) >= 8:
        sorted_vals = [rev[y] for y in sorted(rev)]
        tail = sorted_vals[-8:]
        if len(set(round(v, 2) for v in tail)) == 1 and tail[0] > 0:
            checks_passed.append(
                f"revenue plateaus at {tail[0]:.1f} for last {len(tail)} years "
                "(common when terminal year is held flat)"
            )

    # 6. Data density — flag if too few years are populated
    if rev:
        populated = sum(1 for v in rev.values() if v > 0)
        total = len(extraction.get("year_axis", {}).get("years_covered", []))
        if total and populated / total < 0.3:
            warnings.append(
                f"revenue populated for only {populated}/{total} years "
                "of declared coverage; extraction may be shallow"
            )

    return {
        "checks_passed": checks_passed,
        "warnings": warnings,
        "errors": errors,
        "ok": len(errors) == 0,
    }


# ─────────────────────────────────────────────────────────────────────
# Selection rationale + clarifying questions
# ─────────────────────────────────────────────────────────────────────

def _format_candidate_summary(s: SheetSignal) -> Dict[str, Any]:
    best = max(s.year_axes, key=lambda a: a.span) if s.year_axes else None
    return {
        "sheet": s.name,
        "score": round(s.score, 2),
        "score_breakdown": {k: round(v, 2) for k, v in s.score_breakdown.items()},
        "year_span": f"{best.year_min}..{best.year_max}" if best else None,
        "metrics_found": [k for k, v in s.metric_hits.items() if v],
    }


def _should_ask_user(ranked: List[SheetSignal]) -> Optional[str]:
    """Return a clarifying-question string if the top two candidates are
    too close to call, or None if the top candidate is clearly best."""
    if len(ranked) < 2:
        return None
    top, runner = ranked[0], ranked[1]
    gap = top.score - runner.score
    # Only ask if the gap is small AND both have the scope-bonus signal,
    # meaning both claim to be "combined/consolidated". If the runner-up
    # is clearly a plant-level sheet we shouldn't bother asking.
    top_scope = top.score_breakdown.get("scope_bonus", 0)
    runner_scope = runner.score_breakdown.get("scope_bonus", 0)
    if gap < 3.0 and top_scope > 0 and runner_scope > 0:
        top_years = max(top.year_axes, key=lambda a: a.span) if top.year_axes else None
        runner_years = max(runner.year_axes, key=lambda a: a.span) if runner.year_axes else None
        return (
            f"Two candidate canonical P&L sheets scored similarly: "
            f"'{top.name}' ({top_years.year_min if top_years else '?'}–"
            f"{top_years.year_max if top_years else '?'}) and "
            f"'{runner.name}' ({runner_years.year_min if runner_years else '?'}–"
            f"{runner_years.year_max if runner_years else '?'}). "
            f"Which should be used?"
        )
    return None


# ─────────────────────────────────────────────────────────────────────
# Top-level: run the full pipeline
# ─────────────────────────────────────────────────────────────────────

def extract(xlsx_path: str, user_selected_sheet: Optional[str] = None) -> Dict[str, Any]:
    """Full extraction pipeline.

    Args:
        xlsx_path: path to the workbook.
        user_selected_sheet: if the caller has already resolved a
            clarifying question, pass the chosen sheet name here and we
            extract from that sheet directly (skipping ranking).

    Returns a dict matching the documented output schema.
    """
    path = Path(xlsx_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    signals = discover_sheets(wb)
    ranked = rank_sheets(signals)

    # Summary of what we saw (useful for debugging / UI display)
    candidate_summary = [_format_candidate_summary(s) for s in ranked[:6]]

    if not ranked:
        return {
            "error": "no P&L-like sheets found in workbook",
            "file": str(path),
            "sheets_scanned": [s.name for s in signals],
        }

    # Selection
    if user_selected_sheet:
        chosen = next((s for s in ranked if s.name == user_selected_sheet), None)
        if chosen is None:
            return {
                "error": f"user-selected sheet '{user_selected_sheet}' "
                         "is not a valid P&L candidate",
                "available": [s.name for s in ranked],
            }
        rationale = f"User explicitly chose '{chosen.name}'."
        clarifying_question = None
    else:
        question = _should_ask_user(ranked)
        if question:
            # Return early, let caller ask the user, re-invoke with
            # user_selected_sheet set.
            return {
                "file": str(path),
                "status": "clarifying_question",
                "clarifying_question": question,
                "candidates": candidate_summary,
            }
        chosen = ranked[0]
        top_reasons = sorted(chosen.score_breakdown.items(),
                             key=lambda kv: -kv[1])[:4]
        rationale = (
            f"Picked '{chosen.name}' (score {chosen.score:.1f}). "
            f"Top contributions: "
            + ", ".join(f"{k}={v:+.1f}" for k, v in top_reasons)
            + "."
        )
        if len(ranked) > 1:
            runner_up = ranked[1]
            gap = chosen.score - runner_up.score
            rationale += (
                f" Runner-up was '{runner_up.name}' (gap {gap:.1f}); "
                "not chosen because "
            )
            if runner_up.score_breakdown.get("scope_penalty", 0) < 0:
                rationale += "of single-scope/plant-level sheet-name tokens."
            elif runner_up.score_breakdown.get("scope_bonus", 0) < chosen.score_breakdown.get("scope_bonus", 0):
                rationale += "the chosen sheet's name advertises broader scope (combined/consolidated/unified)."
            elif runner_up.score_breakdown.get("year_span", 0) < chosen.score_breakdown.get("year_span", 0):
                rationale += "the chosen sheet covers more years."
            else:
                rationale += "it ranked lower on metric completeness."
        clarifying_question = None

    # Extract + verify
    extraction = extract_from_sheet(chosen)
    verification = verify_extraction(extraction)

    # If verification produced hard errors AND an alternate exists,
    # retry once on the runner-up and note both outcomes.
    fallback_attempted = None
    if not verification["ok"] and not user_selected_sheet and len(ranked) > 1:
        runner_up = ranked[1]
        alt_extraction = extract_from_sheet(runner_up)
        alt_verification = verify_extraction(alt_extraction)
        fallback_attempted = {
            "sheet": runner_up.name,
            "verification": alt_verification,
        }
        # If alternate is clean and original is dirty, switch
        if alt_verification["ok"] and not verification["ok"]:
            rationale = (
                f"Initially chose '{chosen.name}' but verification failed: "
                f"{verification['errors']}. Switched to '{runner_up.name}', "
                "which passed verification."
            )
            chosen = runner_up
            extraction = alt_extraction
            verification = alt_verification

    # Build output
    year_axis = extraction.get("year_axis", {})
    scope_description = _compose_scope_description(chosen, year_axis)

    return {
        "file": str(path),
        "status": "ok" if verification["ok"] else "ok_with_warnings",
        "scope": {
            "sheet": chosen.name,
            "scope_description": scope_description,
            "years_covered": year_axis.get("years_covered", []),
            "year_axis_row_idx_0based": year_axis.get("row_idx_0based"),
        },
        "selection_rationale": rationale,
        "clarifying_question": clarifying_question,
        "metrics": extraction.get("metrics", {}),
        "verification": verification,
        "fallback_attempted": fallback_attempted,
        "candidates_considered": candidate_summary,
        "extractor_version": "0.1",
    }


def _compose_scope_description(s: SheetSignal, year_axis: Dict[str, Any]) -> str:
    """Human-readable description of what the chosen sheet covers."""
    parts: List[str] = []
    low = s.name.lower()

    # Scope breadth
    if any(t in low for t in ("combined", "unified", "consolidated", "group")):
        parts.append("multi-entity combined/consolidated view")
    elif any(t in low for t in ("summary", "overview")):
        parts.append("summary roll-up")
    elif any(t in low for t in ("l1", "l2", "line 1", "line 2", "standalone", "facility", "plant")):
        parts.append("single-facility / single-line view")
    else:
        parts.append("scope not explicitly advertised in sheet name")

    # Year span
    ymin = year_axis.get("year_min") or (min(year_axis.get("years_covered", [0])) if year_axis.get("years_covered") else None)
    ymax = year_axis.get("year_max") or (max(year_axis.get("years_covered", [0])) if year_axis.get("years_covered") else None)
    if ymin and ymax:
        parts.append(f"{ymin}–{ymax}")

    # Periodicity hint
    if "annual" in low:
        parts.append("annual periodicity")
    elif "quarter" in low:
        parts.append("quarterly periodicity")

    return "; ".join(parts)


# ─────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────

def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python -m app.engine.extract_financials <path/to/workbook.xlsx> [--sheet SHEET]")
        sys.exit(2)
    xlsx_path = sys.argv[1]
    user_sheet = None
    if "--sheet" in sys.argv:
        i = sys.argv.index("--sheet")
        if i + 1 < len(sys.argv):
            user_sheet = sys.argv[i + 1]
    result = extract(xlsx_path, user_selected_sheet=user_sheet)
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()

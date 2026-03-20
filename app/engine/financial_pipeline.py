"""
Deterministic Financial Model Extraction Pipeline

Extracts annual financial metrics from Excel files, builds formula DAGs,
and generates evidence images.
Zero-hallucination: if anything cannot be proven, FAIL loudly.
"""

import os
import re
import json
import hashlib
import logging
import subprocess
import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import Image, ImageDraw, ImageFont

logger = logging.getLogger("financial_pipeline")

METRIC_ALLOWLIST = [
    "units_sold", "units_deployed", "units_operational",
    "revenue", "ebitda", "arr",
    "capital_raised", "cash", "total_assets",
    "total_liabilities", "total_debt",
    "net_income", "gross_profit", "gross_profit_high", "gross_profit_low",
    "revenue_high", "revenue_low", "cogs", "opex", "token_revenue",
    # Extended SaaS / subscription metrics
    "mrr", "bookings", "ndr", "customer_count", "churn_rate",
    # Burn / runway metrics
    "burn_rate", "runway_months",
    # Energy-specific
    "lcoe", "capacity_factor", "generation_gwh",
    # Biotech / clinical
    "patients_enrolled", "trial_count",
    # Fintech
    "aum", "tpv",
]

FLOW_METRICS = {
    "units_sold", "revenue", "ebitda", "capital_raised",
    "net_income", "gross_profit", "gross_profit_high", "gross_profit_low",
    "revenue_high", "revenue_low", "cogs", "opex", "token_revenue",
    "bookings", "burn_rate", "generation_gwh",
    "patients_enrolled", "tpv",
}
STOCK_METRICS = {
    "arr", "cash", "total_assets", "total_liabilities", "total_debt",
    "mrr", "ndr", "customer_count", "churn_rate",
    "runway_months", "lcoe", "capacity_factor", "trial_count", "aum",
    "units_deployed", "units_operational",
}

# Metrics that represent physical quantities (not dollars) — scale factors do NOT apply
UNIT_METRICS = {
    "units_sold", "units_deployed", "units_operational",
    "customer_count", "patients_enrolled", "trial_count",
    "lcoe", "capacity_factor", "generation_gwh",
    "churn_rate", "ndr", "runway_months",
}

# Sheet name relevance scoring for multi-sheet workbooks
SHEET_PRIORITY_KEYWORDS = {
    "high": ["summary", "p&l", "income statement", "income_statement",
             "financial model", "financial_model", "financials",
             "consolidated", "pro forma", "proforma",
             "annual", "yearly"],
    "medium": ["revenue", "operating", "balance sheet", "cash flow",
               "cashflow", "metrics", "kpi", "dashboard", "output",
               "forecast", "quarterly", "monthly"],
    "low": ["notes", "assumptions", "inputs", "cover", "contents",
            "glossary", "instructions", "readme", "changelog", "version",
            "production", "product table", "change log", "forecast period"],
}

# Keywords that suggest a segment/entity/geography sheet rather than consolidated
_SEGMENT_KEYWORDS = [
    "north america", "europe", "asia", "apac", "emea", "latam",
    "norway", "usa", "uk", "germany", "japan", "china", "india",
    "canada", "australia", "france", "brazil", "mexico",
    "segment", "division", "region", "subsidiary", "entity",
    "product a", "product b", "product c",
]

def _score_sheet_relevance(sheet_name: str) -> int:
    """Score a sheet's relevance for financial extraction (higher = more relevant).

    Consolidated/summary sheets score highest (100-150).
    Segment/geography sheets are penalised to prevent them overriding consolidated data.
    """
    name_lower = sheet_name.lower().strip()

    # Check for segment/entity sheets — these should NEVER override consolidated
    is_segment = any(kw in name_lower for kw in _SEGMENT_KEYWORDS)

    # Consolidated sheets get a bonus on top of other scores
    is_consolidated = "consolidated" in name_lower or "consol" in name_lower

    base_score = 25  # default
    for kw in SHEET_PRIORITY_KEYWORDS["high"]:
        if kw in name_lower:
            base_score = max(base_score, 100)
            break
    else:
        for kw in SHEET_PRIORITY_KEYWORDS["medium"]:
            if kw in name_lower:
                base_score = max(base_score, 50)
                break
        else:
            for kw in SHEET_PRIORITY_KEYWORDS["low"]:
                if kw in name_lower:
                    base_score = 5
                    break

    # Boost consolidated sheets
    if is_consolidated:
        base_score = max(base_score, 120)

    # Penalise segment sheets — they should only fill in gaps, not override
    if is_segment:
        base_score = min(base_score, 15)

    return base_score

LABEL_SYNONYMS = {
    "units_sold": [
        r"(?i)^units?\s*sold$", r"(?i)^total\s*units?\s*sold$",
        r"(?i)^units?\s*shipped$", r"(?i)^volume$", r"(?i)^units?$",
    ],
    "units_deployed": [
        r"(?i)^units?\s*deployed$", r"(?i)^total\s*units?\s*deployed$",
        r"(?i)^cumulative\s*units?\s*deployed$",
        r"(?i)^installed\s*capacity$", r"(?i)^capacity\s*deployed$",
        r"(?i)^mw\s*deployed$", r"(?i)^systems?\s*deployed$",
        r"(?i)^mw\s*deployed\s*by\b", r"(?i)^gw\s*deployed$",
    ],
    "units_operational": [
        r"(?i)^units?\s*operational$", r"(?i)^total\s*units?\s*operational$",
        r"(?i)^cumulative\s*units?\s*operational$",
        r"(?i)^operational\s*capacity$", r"(?i)^mw\s*operational$",
        r"(?i)^systems?\s*operational$", r"(?i)^units?\s*in\s*operation$",
        r"(?i)^gw\s*operational$",
    ],
    "revenue": [
        r"(?i)^total\s*revenue$", r"(?i)^revenue$", r"(?i)^net\s*revenue$",
        r"(?i)^total\s*sales$", r"(?i)^sales$", r"(?i)^net\s*sales$",
        r"(?i)^gross\s*revenue$",
        r"(?i)^revenues$", r"(?i)^total\s*revenues$",
        r"(?i)^net\s*revenues$", r"(?i)^gross\s*revenues$",
        r"(?i)^total\s*monthly\s*revenue$", r"(?i)^total\s*annual\s*revenue$",
        r"(?i)^api\s*token\s*net\s*revenue$",
        r"(?i)^annual\s*revenue\s*@\s*[\d\.]+%.*$",
        r"(?i)^revenue\s*(low|high)\s*\(trillions?\s*tokens?\)$",
        r"(?i)^total\s*revenue\s*\(\$[MKB]?\)$",
    ],
    "ebitda": [
        r"(?i)^ebitda$", r"(?i)^adj\.?\s*ebitda$", r"(?i)^adjusted\s*ebitda$",
        r"(?i)^total\s*ebitda$", r"(?i)^group\s*ebitda$",
        r"(?i)^ebitda\s*(excl|incl|ex|inc)\.?\s*projects?$",
        r"(?i)^non.?project\s*ebitda$",
        r"(?i)^operating\s*profit$", r"(?i)^operating\s*income$",
        r"(?i)^ebit$",
        r"(?i)^operating\s*income\s*\(loss\)$",
        r"(?i)^operating\s*income\s*/?\s*\(loss\)$",
        r"(?i)^ebitda\s*\(?loss\)?$",
        r"(?i)^ebitda\s*/\s*\(?loss\)?$",
        r"(?i)^ebitda\s*\(operating\)$",
        r"(?i)^ebitda\s*margin$",
        r"(?i)^total\s*operating\s*profit$",
        r"(?i)^total\s*operating\s*income$",
        r"(?i)^operating\s*profit\s*/?\s*\(?loss\)?$",
        r"(?i)^operating\s*result$",
        r"(?i)^result\s*before\s*tax$",
        r"(?i)^profit\s*before\s*tax$",
    ],
    "net_income": [
        r"(?i)^net\s*profit\s*/?\s*loss$", r"(?i)^net\s*profit$",
        r"(?i)^net\s*income$", r"(?i)^net\s*loss$",
        r"(?i)^total\s*annual\s*net\s*profit\s*/?\s*loss$",
        r"(?i)^profit\s*/?\s*loss$",
        r"(?i)^net\s*income\s*/?\s*\(?loss\)?$",
        r"(?i)^net\s*profit\s*\(?loss\)?$",
        r"(?i)^total\s*net\s*income$",
        r"(?i)^total\s*net\s*profit$",
        r"(?i)^profit\s*after\s*tax$",
        r"(?i)^profit\s*for\s*the\s*(year|period)$",
        r"(?i)^net\s*result$",
    ],
    "gross_profit": [
        r"(?i)^gross\s*profit$", r"(?i)^gross\s*margin$",
        r"(?i)^gross\s*profit\s+annual\s*\(?(low|high)?\)?",
        r"(?i)^gross\s*profit\s*\(net\s*mrr\)\s*(low|high)$",
        r"(?i)^gross\s*margins?\s*[-–]\s*\$?\s*$",
        r"(?i)^gross\s*margins?\s*\$?\s*$",
    ],
    "gross_profit_high": [
        r"(?i)^gross\s*profit\s+annual\s*\(?high\)?",
        r"(?i)^gross\s*profit\s*\(net\s*mrr\)\s*high$",
    ],
    "gross_profit_low": [
        r"(?i)^gross\s*profit\s+annual\s*\(?low\s*\)?",
        r"(?i)^gross\s*profit\s*\(net\s*mrr\)\s*low$",
    ],
    "revenue_high": [
        r"(?i)^revenue\s+high\b.*",
        r"(?i)^revenue\s*\(high\)$",
    ],
    "revenue_low": [
        r"(?i)^revenue\s+low\b.*",
        r"(?i)^revenue\s*\(low\)$",
    ],
    "cogs": [
        r"(?i)^cogs$", r"(?i)^cost\s*of\s*goods\s*sold$",
        r"(?i)^cost\s*of\s*revenue$", r"(?i)^cost\s*of\s*sales$",
        r"(?i)^cogs\s*(low|high)\s*\([\d]+%\)$",
        r"(?i)^cost\s*of\s*goods\s*\(product\)$",
        r"(?i)^cost\s*of\s*goods\s*\(maintenance\)$",
        r"(?i)^total\s*cost\s*of\s*goods\s*sold$",
        r"(?i)^total\s*cogs$",
    ],
    "opex": [
        r"(?i)^opex$", r"(?i)^operating\s*expenses?$",
        r"(?i)^total\s*opex$", r"(?i)^opex\s*cost.*$",
        r"(?i)^total\s*cost$",
        r"(?i)^total\s*operating\s*expenses?$",
        r"(?i)^operating\s*expenses?\s*[+&]\s*r\s*&?\s*d$",
        r"(?i)^total\s*operating\s*expenses?\s*[+&]\s*r\s*&?\s*d$",
    ],
    "token_revenue": [
        r"(?i)^api\s*(model\s*)?revenue\s*(forecast)?$",
        r"(?i)^token\s*revenue$",
    ],
    "arr": [
        r"(?i)^arr$", r"(?i)^annual\s*recurring\s*revenue$",
        r"(?i)^annualized\s*recurring\s*revenue$",
    ],
    "capital_raised": [
        r"(?i)^capital\s*raised$", r"(?i)^total\s*capital\s*raised$",
        r"(?i)^funding$", r"(?i)^total\s*funding$",
        r"(?i)^proceeds\s*from\s*financing$",
    ],
    "cash": [
        r"(?i)^cash$", r"(?i)^cash\s*and\s*cash\s*equivalents$",
        r"(?i)^cash\s*&\s*cash\s*equivalents$", r"(?i)^total\s*cash$",
    ],
    "total_assets": [
        r"(?i)^total\s*assets$",
    ],
    "total_liabilities": [
        r"(?i)^total\s*liabilities$",
    ],
    "total_debt": [
        r"(?i)^total\s*debt$", r"(?i)^long[\s-]*term\s*debt$",
    ],
    # ── Physical unit metrics ─────────────────────────────────────────────────
    # Placed LAST so financial metric labels take priority in any ambiguous match.
    #
    
    # ── SaaS / Subscription metrics ──────────────────────────────────────────
    "mrr": [
        r"(?i)^mrr$", r"(?i)^monthly\s*recurring\s*revenue$",
        r"(?i)^net\s*mrr$", r"(?i)^total\s*mrr$",
    ],
    "bookings": [
        r"(?i)^bookings$", r"(?i)^total\s*bookings$",
        r"(?i)^new\s*bookings$", r"(?i)^gross\s*bookings$",
        r"(?i)^net\s*new\s*bookings$",
    ],
    "ndr": [
        r"(?i)^ndr$", r"(?i)^net\s*dollar\s*retention$",
        r"(?i)^net\s*revenue\s*retention$", r"(?i)^nrr$",
    ],
    "customer_count": [
        r"(?i)^customer\s*count$", r"(?i)^total\s*customers?$",
        r"(?i)^number\s*of\s*customers?$", r"(?i)^active\s*customers?$",
        r"(?i)^paying\s*customers?$", r"(?i)^(total\s+)?clients?$",
    ],
    "churn_rate": [
        r"(?i)^churn\s*rate$", r"(?i)^(monthly|annual)\s*churn$",
        r"(?i)^logo\s*churn$", r"(?i)^customer\s*churn$",
        r"(?i)^gross\s*churn$", r"(?i)^revenue\s*churn$",
    ],
    # ── Burn / runway ────────────────────────────────────────────────────────
    "burn_rate": [
        r"(?i)^(monthly\s*)?burn\s*rate$", r"(?i)^net\s*burn$",
        r"(?i)^cash\s*burn$", r"(?i)^monthly\s*burn$",
    ],
    "runway_months": [
        r"(?i)^runway$", r"(?i)^cash\s*runway$",
        r"(?i)^runway\s*\(months?\)$", r"(?i)^months?\s*of\s*runway$",
    ],
    # ── Energy-specific ──────────────────────────────────────────────────────
    "lcoe": [
        r"(?i)^lcoe$", r"(?i)^levelized\s*cost\s*of\s*energy$",
        r"(?i)^lcoe\s*\(\$\/mwh\)$", r"(?i)^levelized\s*cost$",
    ],
    "capacity_factor": [
        r"(?i)^capacity\s*factor$", r"(?i)^net\s*capacity\s*factor$",
        r"(?i)^cf\s*\(%\)$",
    ],
    "generation_gwh": [
        r"(?i)^(total\s+)?generation\s*\(gwh\)$",
        r"(?i)^(annual\s+)?energy\s*generation$",
        r"(?i)^(total\s+)?gwh\s*(generated|produced)$",
    ],
    # ── Biotech / clinical ───────────────────────────────────────────────────
    "patients_enrolled": [
        r"(?i)^patients?\s*enrolled$", r"(?i)^total\s*patients?$",
        r"(?i)^enrollment$", r"(?i)^(total\s+)?subjects?$",
    ],
    "trial_count": [
        r"(?i)^(active\s+)?trials?$", r"(?i)^clinical\s*trials?$",
        r"(?i)^studies?\s*active$",
    ],
    # ── Fintech ──────────────────────────────────────────────────────────────
    "aum": [
        r"(?i)^aum$", r"(?i)^assets?\s*under\s*management$",
        r"(?i)^total\s*aum$",
    ],
    "tpv": [
        r"(?i)^tpv$", r"(?i)^total\s*payment\s*volume$",
        r"(?i)^gross\s*payment\s*volume$", r"(?i)^gpv$",
    ],
    # ── Additional common financial metrics ──────────────────────────────────
    "free_cash_flow": [
        r"(?i)^free\s*cash\s*flow$", r"(?i)^fcf$",
        r"(?i)^unlevered\s*free\s*cash\s*flow$",
        r"(?i)^levered\s*free\s*cash\s*flow$",
    ],
    "capex": [
        r"(?i)^cap(?:ital)?\s*ex(?:penditures?)?$",
        r"(?i)^capital\s*expenditures?$", r"(?i)^total\s*capex$",
        r"(?i)^purchase\s*of\s*(?:property|pp&?e|fixed\s*assets)$",
    ],
    "depreciation": [
        r"(?i)^depreciation$",
        r"(?i)^depreciation\s*(?:&|and)\s*amortization$",
        r"(?i)^d&a$", r"(?i)^total\s*depreciation$",
    ],
    "operating_income": [
        r"(?i)^operating\s*income$", r"(?i)^operating\s*profit$",
        r"(?i)^ebit$", r"(?i)^income\s*from\s*operations$",
    ],
    "headcount": [
        r"(?i)^headcount$", r"(?i)^fte$",
        r"(?i)^full[\s\-]*time\s*equivalents?$",
        r"(?i)^employees?$", r"(?i)^total\s*headcount$",
    ],
}

SCALE_PATTERNS = [
    # Factor converts cell value → actual USD.
    # "in thousands" cell 500 means $500,000 → factor 1_000
    (r"(?i)\bin\s*thousands\b", 1_000),
    (r"(?i)\$\s*in\s*thousands\b", 1_000),
    (r"(?i)\(\s*\$?\s*000s?\s*\)", 1_000),
    (r"(?i)\(\s*thousands\s*\)", 1_000),
    (r"(?i)\ball\s+values?\s+in\s+\$?\s*000", 1_000),
    (r"(?i)^\$\s*000s?$", 1_000),
    (r"(?i)\bin\s*millions\b", 1_000_000),
    (r"(?i)\$\s*in\s*millions\b", 1_000_000),
    (r"(?i)\(\s*\$?\s*MM?\s*\)", 1_000_000),
    (r"(?i)\(\s*millions\s*\)", 1_000_000),
    (r"(?i)\bin\s*billions\b", 1_000_000_000),
    (r"(?i)\$\s*in\s*billions\b", 1_000_000_000),
    (r"(?i)\(\s*\$?\s*B\s*\)", 1_000_000_000),
    (r"(?i)\(\s*billions\s*\)", 1_000_000_000),
    (r"(?i)\(\s*\$K\s*\)", 1_000),
    (r"(?i)\(\s*\$M\s*\)", 1_000_000),
    (r"(?i)\(\s*\$B\s*\)", 1_000_000_000),
]

PERIOD_PATTERNS = {
    "yearly": re.compile(
        r"(?i)(?:FY|CY)?\s*'?(\d{4})$"
    ),
    "yearly_short": re.compile(
        r"(?i)(?:FY|CY)\s*'?(\d{2})$"
    ),
    "quarter": re.compile(
        r"(?i)Q([1-4])\s*'?(\d{2,4})"
    ),
    "quarter_alt": re.compile(
        r"(?i)(\d{4})\s*Q([1-4])"
    ),
    "month_year": re.compile(
        r"(?i)(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s*'?(\d{2,4})"
    ),
    "date_iso": re.compile(
        r"(\d{4})-(\d{2})-(\d{2})"
    ),
    "quarter_hyphen": re.compile(
        r"(?i)(\d{4})\s*[-/]\s*Q([1-4])"
    ),
    "yearly_suffix": re.compile(
        r"(?i)(?:FY|CY)?\s*'?(\d{4})\s*[AEFPB]$"
    ),
    "ltm_ntm": re.compile(
        r"(?i)(LTM|NTM|TTM)\s*'?(\d{2,4})?"
    ),
}

ACTUAL_FORECAST_RE = re.compile(r"(?i)\b(actual|budget|forecast|projected|estimate)\b")

# ── Scenario detection ────────────────────────────────────────────────────────
# Canonical scenario names and the keywords / regex patterns that map to them.
# "base" is the default when no scenario indicator is found.

SCENARIO_CANONICAL = {
    "bear": re.compile(
        r"(?i)\b(bear|bearish|downside|pessimistic|worst[\s\-]?case|low[\s\-]?case"
        r"|conservative|stress|floor)\b"
    ),
    "base": re.compile(
        r"(?i)\b(base|baseline|mid|middle|expected|management|central"
        r"|most[\s\-]?likely|reference|target)\b"
    ),
    "bull": re.compile(
        r"(?i)\b(bull|bullish|upside|optimistic|best[\s\-]?case|high[\s\-]?case"
        r"|aggressive|stretch|ceiling|blue[\s\-]?sky)\b"
    ),
}

# Section header patterns: rows that contain ONLY a scenario label (not a metric)
_SCENARIO_SECTION_RE = re.compile(
    r"(?i)^\s*(bear|bull|base|upside|downside|optimistic|pessimistic|aggressive"
    r"|conservative|worst|best|low|mid|high|management|stress|stretch|blue[\s\-]?sky)"
    r"[\s\-]*(case|scenario|forecast|projection|estimate)?\s*$"
)


def detect_scenario_from_sheet_name(sheet_name: str) -> str:
    """Return canonical scenario ('bear', 'base', 'bull') from a sheet name, or 'base'."""
    s = (sheet_name or "").strip()
    for scenario, pattern in SCENARIO_CANONICAL.items():
        if pattern.search(s):
            return scenario
    return "base"


_SCENARIO_RESIDUAL_RE = re.compile(r"(?i)\s*(case|scenario|forecast|projection|estimate)\s*")


def detect_scenario_from_label(label: str) -> tuple:
    """
    Check if a metric label embeds a scenario qualifier.
    Returns (cleaned_label, scenario) where scenario is 'bear'/'base'/'bull'
    or None if no scenario indicator found.

    Examples:
        "Revenue - Bear Case"   -> ("Revenue", "bear")
        "EBITDA (Upside)"       -> ("EBITDA", "bull")
        "Revenue"               -> ("Revenue", None)
    """
    if not label:
        return (label, None)
    s = str(label).strip()
    for scenario, pattern in SCENARIO_CANONICAL.items():
        m = pattern.search(s)
        if m:
            cleaned = pattern.sub("", s)
            cleaned = _SCENARIO_RESIDUAL_RE.sub(" ", cleaned)
            cleaned = re.sub(r"[\s\-\u2013\u2014]+$", "", cleaned).strip()
            cleaned = re.sub(r"^[\s\-\u2013\u2014]+", "", cleaned).strip()
            cleaned = re.sub(r"\s*[\(\)]\s*", " ", cleaned).strip()
            cleaned = re.sub(r"\s*[-\u2013\u2014]\s*$", "", cleaned).strip()
            if cleaned:
                return (cleaned, scenario)
    return (s, None)


def detect_scenario_sections(ws, max_rows: int = 500) -> dict:
    """
    Scan a worksheet for section-header rows that indicate scenario blocks.
    Returns {row_number: scenario_name} for each detected section header.

    A section header is a row where column A (or B) contains text matching
    a scenario keyword but NOT a metric label — it's a divider row.
    """
    sections = {}
    for row in range(1, min(max_rows + 1, ws.max_row + 1)):
        for col in range(1, min(4, ws.max_column + 1)):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            s = str(val).strip()
            if not s or len(s) > 60:
                continue
            if match_metric_label(s):
                continue
            if _SCENARIO_SECTION_RE.match(s):
                for scenario, pattern in SCENARIO_CANONICAL.items():
                    if pattern.search(s):
                        sections[row] = scenario
                        break
    return sections


def resolve_scenario_for_row(row: int, scenario_sections: dict, sheet_scenario: str) -> str:
    """
    Determine the scenario for a given row based on:
    1. Section headers above the row (nearest one wins)
    2. Sheet-level scenario (fallback)
    3. 'base' as ultimate default
    """
    if scenario_sections:
        nearest_section_row = None
        for sec_row in scenario_sections:
            if sec_row < row:
                if nearest_section_row is None or sec_row > nearest_section_row:
                    nearest_section_row = sec_row
        if nearest_section_row is not None:
            return scenario_sections[nearest_section_row]
    return sheet_scenario

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def convert_xls_to_xlsx(input_path: str, out_dir: str) -> str:
    result = subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", out_dir, input_path],
        capture_output=True, text=True, timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice conversion failed: {result.stderr}")
    base = Path(input_path).stem + ".xlsx"
    converted = os.path.join(out_dir, base)
    if not os.path.exists(converted):
        raise RuntimeError(f"Converted file not found: {converted}")
    return converted


def _quarter_end(year: int, quarter: int) -> str:
    ends = {1: (3, 31), 2: (6, 30), 3: (9, 30), 4: (12, 31)}
    m, d = ends[quarter]
    return f"{year}-{m:02d}-{d:02d}"


def _month_end(year: int, month: int) -> str:
    import calendar
    _, last_day = calendar.monthrange(year, month)
    return f"{year}-{month:02d}-{last_day:02d}"


def _fy_from_period_end(period_end: str, fy_end_month: int) -> int:
    try:
        dt = datetime.date.fromisoformat(period_end)
    except (ValueError, TypeError):
        # Malformed date (e.g. year 0) — fall back to current year
        logger.warning(f"Invalid period_end '{period_end}', defaulting to current year")
        return datetime.date.today().year
    if dt.year < 1900 or dt.year > 2100:
        logger.warning(f"Suspicious year {dt.year} in period_end '{period_end}', skipping")
        return datetime.date.today().year
    if fy_end_month == 12:
        return dt.year
    if dt.month <= fy_end_month:
        return dt.year
    return dt.year + 1


def parse_period_header(text) -> Optional[dict]:
    if text is None:
        return None
    import datetime as _dt
    if isinstance(text, (_dt.datetime, _dt.date)):
        if text.year < 2015 or text.year > 2045:
            return None
        return {
            "type": "month",
            "month": text.month,
            "year": text.year,
            "period_end": _month_end(text.year, text.month),
            "raw": str(text),
        }
    # Handle raw numeric years from Excel cells (int 2024 or float 2024.0)
    if isinstance(text, (int, float)):
        yr = int(text)
        if 2015 <= yr <= 2050 and (isinstance(text, int) or text == float(yr)):
            return {
                "type": "year",
                "year": yr,
                "period_end": f"{yr}-12-31",
                "raw": str(yr),
            }
    s = str(text).strip()
    if not s:
        return None
    # Normalize float years: "2024.0" → "2024"
    if re.match(r"^\d{4}\.0$", s):
        s = s.split(".")[0]

    m = PERIOD_PATTERNS["date_iso"].search(s)
    if m:
        return {"type": "date", "period_end": m.group(0), "raw": s}

    m = PERIOD_PATTERNS["quarter"].search(s)
    if m:
        q = int(m.group(1))
        yr = int(m.group(2))
        if yr < 100:
            yr += 2000
        return {"type": "quarter", "quarter": q, "year": yr,
                "period_end": _quarter_end(yr, q), "raw": s}

    m = PERIOD_PATTERNS["quarter_alt"].search(s)
    if m:
        yr = int(m.group(1))
        q = int(m.group(2))
        return {"type": "quarter", "quarter": q, "year": yr,
                "period_end": _quarter_end(yr, q), "raw": s}

    m = PERIOD_PATTERNS["quarter_hyphen"].search(s)
    if m:
        yr = int(m.group(1))
        q = int(m.group(2))
        return {"type": "quarter", "quarter": q, "year": yr,
                "period_end": _quarter_end(yr, q), "raw": s}

    m = PERIOD_PATTERNS["month_year"].search(s)
    if m:
        mo = MONTH_MAP.get(m.group(1).lower()[:3])
        yr = int(m.group(2))
        if yr < 100:
            yr += 2000
        if mo:
            return {"type": "month", "month": mo, "year": yr,
                    "period_end": _month_end(yr, mo), "raw": s}

    m = PERIOD_PATTERNS["yearly"].search(s)
    if m:
        yr = int(m.group(1))
        if yr < 1900 or yr > 2100:
            return None
        return {"type": "year", "year": yr,
                "period_end": f"{yr}-12-31", "raw": s}

    m = PERIOD_PATTERNS["yearly_suffix"].search(s)
    if m:
        yr = int(m.group(1))
        if yr < 1900 or yr > 2100:
            return None
        suffix = s[-1].upper()
        data_class = {"A": "actual", "E": "estimate", "F": "forecast", "P": "projected", "B": "budget"}.get(suffix)
        result = {"type": "year", "year": yr, "period_end": f"{yr}-12-31", "raw": s}
        if data_class:
            result["data_class"] = data_class
        return result

    m = PERIOD_PATTERNS["yearly_short"].search(s)
    if m:
        yr = int(m.group(1))
        yr = yr + 2000 if yr < 100 else yr
        if yr < 1900 or yr > 2100:
            return None
        forecast_flag = "forecast" if ACTUAL_FORECAST_RE.search(s) else None
        result = {"type": "year", "year": yr, "period_end": f"{yr}-12-31", "raw": s}
        if forecast_flag:
            result["data_class"] = forecast_flag
        return result

    m = PERIOD_PATTERNS["ltm_ntm"].search(s)
    if m:
        label = m.group(1).upper()
        yr_part = m.group(2)
        yr = int(yr_part) if yr_part else datetime.date.today().year
        if yr < 100:
            yr += 2000
        return {"type": "year", "year": yr, "period_end": f"{yr}-12-31",
                "raw": s, "data_class": "ltm" if label in ("LTM", "TTM") else "ntm"}

    return None


def match_metric_label(text: str) -> Optional[str]:
    if not text:
        return None
    s = str(text).strip()
    for metric, patterns in LABEL_SYNONYMS.items():
        for pat in patterns:
            if re.match(pat, s):
                return metric
    # Try after stripping embedded scale annotations: "Revenue ($M)" -> "Revenue"
    stripped = re.sub(r'\s*\(\$?[MKBmkb]\w*\)\s*$', '', s).strip()
    if stripped != s:
        for metric, patterns in LABEL_SYNONYMS.items():
            for pat in patterns:
                if re.match(pat, stripped):
                    return metric
    # Try after stripping leading account codes: "6050 Revenue" -> "Revenue"
    code_stripped = re.sub(r'^\d{3,5}\s+', '', s)
    if code_stripped != s:
        for metric, patterns in LABEL_SYNONYMS.items():
            for pat in patterns:
                if re.match(pat, code_stripped):
                    return metric
    # Try after stripping scenario qualifiers: "Revenue - Bear Case" -> "Revenue"
    scenario_cleaned, detected_scenario = detect_scenario_from_label(s)
    if detected_scenario and scenario_cleaned != s:
        result = match_metric_label(scenario_cleaned)
        if result:
            return result
    # Try after stripping common prefixes/suffixes that financial models add:
    # "Total EBITDA" -> "EBITDA", "Group Revenue" -> "Revenue",
    # "Blaze EBITDA" -> "EBITDA", "EBITDA - Consolidated" -> "EBITDA"
    _PREFIX_STRIP = [
        r'^(?:total|group|consolidated|consol|combined|company|net)\s+',
        r'^[A-Z][a-zA-Z]+\s+(?=(?:ebitda|revenue|net\s|gross\s|operating|cogs|opex|arr|mrr|cash|ebit)\b)',  # company name prefix
    ]
    _SUFFIX_STRIP = [
        r'\s*[-–—]\s*(?:total|consolidated|consol|combined|group|net)\s*$',
        r'\s*[-–—]\s*(?:annual|yearly|fy)\s*$',
    ]
    for pat in _PREFIX_STRIP:
        cleaned = re.sub(pat, '', s, flags=re.IGNORECASE).strip()
        if cleaned and cleaned != s:
            for metric, patterns in LABEL_SYNONYMS.items():
                for p in patterns:
                    if re.match(p, cleaned):
                        return metric
    for pat in _SUFFIX_STRIP:
        cleaned = re.sub(pat, '', s, flags=re.IGNORECASE).strip()
        if cleaned and cleaned != s:
            for metric, patterns in LABEL_SYNONYMS.items():
                for p in patterns:
                    if re.match(p, cleaned):
                        return metric
    return None


def fuzzy_match_label(text: str, threshold: float = 0.88) -> Optional[tuple]:
    if not text:
        return None
    s = str(text).strip().lower()
    # Also try with common prefixes stripped for fuzzy
    candidates = [s]
    stripped = re.sub(r'^(?:total|group|consolidated|combined|net)\s+', '', s, flags=re.IGNORECASE).strip()
    if stripped != s:
        candidates.append(stripped)
    best_metric = None
    best_score = 0
    for candidate in candidates:
        for metric, patterns in LABEL_SYNONYMS.items():
            for pat in patterns:
                clean_pat = pat.replace(r"(?i)", "").replace("^", "").replace("$", "")
                clean_pat = re.sub(r"\\s[*+]", " ", clean_pat).replace(r"\s", " ")
                clean_pat = clean_pat.replace(r"\.?", "").replace(r"[\s-]*", " ").lower()
                if not clean_pat:
                    continue
                len_ratio = min(len(candidate), len(clean_pat)) / max(len(candidate), len(clean_pat)) if max(len(candidate), len(clean_pat)) > 0 else 0
                if len_ratio < 0.6:
                    continue
                common = sum(1 for a, b in zip(candidate, clean_pat) if a == b)
                score = common / max(len(candidate), len(clean_pat))
                if score > best_score:
                    best_score = score
                    best_metric = metric
    if best_score >= threshold and best_metric:
        return (best_metric, "FUZZY", best_score)
    return None


def detect_scale_evidence(ws, max_rows: int = 60) -> dict:
    """
    Scan for explicit scale annotations (in thousands, $MM, etc.).
    Returns a dict whose ``_multiply_by`` converts cell values → actual USD.

    If none found, default to RAW (multiply by 1) — startup models typically
    report exact dollar figures without scale headers.
    """
    for row in range(1, min(max_rows + 1, ws.max_row + 1)):
        for col in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is None:
                continue
            s = str(val).strip()
            for pat, factor in SCALE_PATTERNS:
                if re.search(pat, s):
                    return {
                        "normalized_to": "USD",
                        "evidence_cell": f"{get_column_letter(col)}{row}",
                        "evidence_text": s,
                        "_multiply_by": factor,
                    }
    return {
        "normalized_to": "RAW_USD",
        "evidence_cell": None,
        "evidence_text": "No scale header found; cell values assumed to be raw USD.",
        "_multiply_by": 1,
        "_scale_assumed": True,
    }


def detect_label_scale(label: str) -> Optional[int]:
    """Detect scale annotations embedded in metric labels like 'Revenue ($M)'.
    Returns the factor to convert cell value → actual USD."""
    if re.search(r'\(\$?M\)|\(\$?\s*millions?\)', label, re.IGNORECASE):
        return 1_000_000
    if re.search(r'\(\$?K\)|\(\$?\s*thousands?\)', label, re.IGNORECASE):
        return 1_000
    if re.search(r'\(\$?B\)|\(\$?\s*billions?\)', label, re.IGNORECASE):
        return 1_000_000_000
    return None


def _parse_month_number_header(val, col: int, row: int) -> Optional[dict]:
    """Recognise plain sequential integers (1-120) as ordinal month headers."""
    if not isinstance(val, (int, float)):
        return None
    n = int(val)
    if n < 1 or n > 120:
        return None
    # Anchor to the current calendar year so annualisation works for present-day models
    anchor_year = datetime.date.today().year
    year = anchor_year + (n - 1) // 12
    month = ((n - 1) % 12) + 1
    return {
        "type": "month",
        "month": month,
        "year": year,
        "period_end": _month_end(year, month),
        "raw": str(n),
        "col": col,
        "cell": f"{get_column_letter(col)}{row}",
        "_ordinal_month": n,
    }


def _parse_year_label_header(val, col: int, row: int) -> Optional[dict]:
    """Recognise 'Year 1', 'Year 2', '18 Months', 'Year N' labels as period markers."""
    if not isinstance(val, str):
        return None
    s = val.strip()
    m = re.match(r"(?i)^year\s*(\d+)$", s)
    if m:
        n = int(m.group(1))
        # Anchor to current year so Year 1 = this year
        yr = datetime.date.today().year + n - 1
        return {
            "type": "year",
            "year": yr,
            "period_end": f"{yr}-12-31",
            "raw": s,
            "col": col,
            "cell": f"{get_column_letter(col)}{row}",
        }
    m = re.match(r"(?i)^(\d+)\s*months?$", s)
    if m:
        total_mo = int(m.group(1))
        yr = datetime.date.today().year + (total_mo - 1) // 12
        mo = ((total_mo - 1) % 12) + 1
        return {
            "type": "month",
            "month": mo,
            "year": yr,
            "period_end": _month_end(yr, mo),
            "raw": s,
            "col": col,
            "cell": f"{get_column_letter(col)}{row}",
        }
    return None


def _cluster_periods(periods: list, gap_threshold: int = 4) -> list:
    """Split period headers into contiguous column clusters.
    Two columns belong to different zones if:
    - The gap between them exceeds gap_threshold, OR
    - The period type changes (e.g., quarter -> year) indicating an
      annual summary block adjacent to quarterly detail.
    """
    if len(periods) <= 1:
        return [periods] if periods else []
    sorted_p = sorted(periods, key=lambda p: p["col"])
    clusters = [[sorted_p[0]]]
    for p in sorted_p[1:]:
        prev = clusters[-1][-1]
        col_gap = p["col"] - prev["col"] > gap_threshold
        type_change = p.get("type") != prev.get("type")
        if col_gap or type_change:
            clusters.append([p])
        else:
            clusters[-1].append(p)
    return [c for c in clusters if len(c) >= 2]


def _get_cell_value_with_merge(ws, row, col):
    """Get cell value, resolving merged cells by looking up the merge range.
    Only returns the value for the FIRST column in a merged range to avoid duplicates.
    """
    val = ws.cell(row=row, column=col).value
    if val is not None:
        return val
    # Check if this cell is part of a merged range
    for merge_range in ws.merged_cells.ranges:
        if (merge_range.min_row <= row <= merge_range.max_row and
            merge_range.min_col <= col <= merge_range.max_col):
            # Only resolve for the first column of the merge to avoid duplicates
            if col == merge_range.min_col:
                return ws.cell(row=merge_range.min_row, column=merge_range.min_col).value
            return None  # Non-first columns in merge return None
    return None


def find_all_period_header_rows(ws, max_rows: int = 100) -> list:
    """
    Return ALL period header rows found within a sheet (for multi-section sheets).
    Each entry is a dict with 'row' and 'periods' keys.
    If a single row has multiple column clusters separated by a gap, each cluster
    is returned as a separate entry so multi-zone sheets are handled correctly.

    Ordinal-number headers (1,2,3...) are only used if no calendar-based
    headers exist on the same sheet, to avoid treating "PeriodNumber" index
    rows as time periods.

    Handles merged cells by resolving the merge range to the top-left cell value.
    """
    calendar_found = []
    ordinal_found = []
    seen_keys = set()

    for row in range(1, min(max_rows + 1, ws.max_row + 1)):
        calendar_periods = []
        ordinal_periods = []
        for col in range(1, ws.max_column + 1):
            val = _get_cell_value_with_merge(ws, row, col)
            if val is None:
                continue
            parsed = parse_period_header(val)
            if parsed:
                parsed["col"] = col
                parsed["cell"] = f"{get_column_letter(col)}{row}"
                calendar_periods.append(parsed)
                continue
            parsed_ord = _parse_month_number_header(val, col, row)
            if parsed_ord:
                ordinal_periods.append(parsed_ord)
                continue
            parsed_yr = _parse_year_label_header(val, col, row)
            if parsed_yr:
                parsed_yr["col"] = col
                parsed_yr["cell"] = f"{get_column_letter(col)}{row}"
                calendar_periods.append(parsed_yr)

        if len(calendar_periods) >= 2:
            clusters = _cluster_periods(calendar_periods)
            for cluster in clusters:
                col_range = (cluster[0]["col"], cluster[-1]["col"])
                key = (row, col_range)
                if key not in seen_keys:
                    seen_keys.add(key)
                    calendar_found.append({"row": row, "periods": cluster, "_col_range": col_range})

        if len(ordinal_periods) >= 2:
            clusters = _cluster_periods(ordinal_periods)
            for cluster in clusters:
                col_range = (cluster[0]["col"], cluster[-1]["col"])
                key = (row, col_range)
                if key not in seen_keys:
                    seen_keys.add(key)
                    ordinal_found.append({"row": row, "periods": cluster, "_col_range": col_range})

    if calendar_found:
        return calendar_found
    return ordinal_found


# Regex for recognising physical unit-of-measure strings in adjacent cells
_UNIT_TYPE_RE = re.compile(
    r"^(MW|GW|kW|MWh|GWh|kWh|TWh|Barrels?|Bbl|BOE|BOEPD|"
    r"Tonnes?|MT|kg|Liters?|Gallons?|"
    r"Subscribers?|Customers?|Seats?|Licenses?|Users?|Sites?|Projects?|Wells?|"
    r"Units?|Installations?|Assets?)$",
    re.IGNORECASE,
)


def _detect_unit_type(ws, row: int, label_col: int, max_look: int = 5) -> Optional[str]:
    """
    Look in columns immediately to the right of the label cell for a
    unit-of-measure string (e.g. "MW", "GWh", "barrels", "subscribers").
    Returns the unit string, or None if not found.
    Stops scanning as soon as it hits a numeric data cell.
    """
    for offset in range(1, max_look + 1):
        col = label_col + offset
        if col > ws.max_column:
            break
        v = ws.cell(row=row, column=col).value
        if v is None:
            continue
        if isinstance(v, str):
            s = v.strip()
            if _UNIT_TYPE_RE.match(s):
                return s
            # Stop once we hit a non-unit string (section label, etc.)
            break
        elif isinstance(v, (int, float)):
            # Reached numeric data — unit column was not found
            break
    return None


def find_metric_rows(ws, max_rows: int = 500) -> list:
    """
    Scan for rows whose label matches a known metric.
    Scans up to max_rows rows and up to 8 label columns to handle
    multi-section sheets where labels appear in columns B-D (e.g. Refiant).
    Also sniffs the unit type for UNIT_METRICS rows (MW, GWh, barrels, etc.)
    from the cell(s) adjacent to the label.
    Deduplicates: if the same metric appears multiple times in the same
    column range, all occurrences are returned so multi-scenario sheets
    (Low/High) are captured.
    Now also detects scenario qualifiers embedded in labels (e.g.
    "Revenue - Bear Case") and stores them in the result dict.
    """
    results = []
    seen = set()
    for row in range(1, min(max_rows + 1, ws.max_row + 1)):
        for col in range(1, min(9, ws.max_column + 1)):
            val = ws.cell(row=row, column=col).value
            # Also try resolving merged cells for labels
            if val is None:
                val = _get_cell_value_with_merge(ws, row, col)
            if val is None:
                continue
            s = str(val).strip()
            if not s:
                continue
            metric = match_metric_label(s)
            match_type = "EXACT"
            if not metric:
                fuzzy = fuzzy_match_label(s)
                if fuzzy:
                    metric, match_type, _ = fuzzy
            if metric:
                key = (metric, row, col)
                if key not in seen:
                    seen.add(key)
                    unit_type = None
                    if metric in UNIT_METRICS:
                        unit_type = _detect_unit_type(ws, row, col)
                    _, label_scenario = detect_scenario_from_label(s)
                    results.append({
                        "metric": metric,
                        "row": row,
                        "label_col": col,
                        "label_cell": f"{get_column_letter(col)}{row}",
                        "matched_label": s,
                        "match_type": match_type,
                        "unit_type": unit_type,
                        "label_scenario": label_scenario,
                    })
    return results


def _looks_like_year(v) -> bool:
    """Return True if a numeric value looks like a calendar year, not financial data."""
    if isinstance(v, float) and v != int(v):
        return False
    return 2015 <= int(v) <= 2060


def extract_metric_values(ws, metric_info: dict, period_header: dict) -> dict:
    row = metric_info["row"]
    is_unit = metric_info["metric"] in UNIT_METRICS
    values = []
    value_cells = []
    raw_values = []
    header_cells = []
    parsed_period_ends = []

    for p in period_header["periods"]:
        col = p["col"]
        cell_val = ws.cell(row=row, column=col).value
        cell_ref = f"{get_column_letter(col)}{row}"

        if cell_val is not None and isinstance(cell_val, (int, float)):
            if not is_unit and _looks_like_year(cell_val):
                continue
            values.append({"value": cell_val, "period": p, "cell": cell_ref})
            value_cells.append(cell_ref)
            raw_values.append(cell_val)
            header_cells.append(p["cell"])
            parsed_period_ends.append(p["period_end"])
        elif cell_val is not None:
            try:
                num = float(str(cell_val).replace(",", "").replace("$", "").strip())
                values.append({"value": num, "period": p, "cell": cell_ref})
                value_cells.append(cell_ref)
                raw_values.append(num)
                header_cells.append(p["cell"])
                parsed_period_ends.append(p["period_end"])
            except (ValueError, TypeError):
                pass

    return {
        "values": values,
        "value_cells": value_cells,
        "raw_values": raw_values,
        "header_cells": header_cells,
        "parsed_period_ends": parsed_period_ends,
    }


def annualize_metric(metric: str, values: list, fy_end_month: int) -> list:
    if not values:
        return []

    period_type = values[0]["period"].get("type", "year")
    by_fy = {}

    for v in values:
        pe = v["period"]["period_end"]
        fy = _fy_from_period_end(pe, fy_end_month)
        by_fy.setdefault(fy, []).append(v)

    results = []
    for fy, fy_vals in sorted(by_fy.items()):
        if period_type == "year":
            if len(fy_vals) == 1:
                results.append({
                    "fiscal_year": fy,
                    "period_end": fy_vals[0]["period"]["period_end"],
                    "value": fy_vals[0]["value"],
                    "method": "YEAR_REPORTED",
                    "coverage_count": 1,
                    "coverage_required": 1,
                    "source_periods": "YEAR",
                    "status": "PASS",
                    "cells_used": [fy_vals[0]["cell"]],
                    "raw_used": [fy_vals[0]["value"]],
                })
            continue

        if metric in FLOW_METRICS:
            if period_type == "quarter":
                required = 4
                method = "SUM_4Q"
                source = "QUARTER"
                min_for_extrapolation = 3  # 3 of 4 quarters → extrapolate
            elif period_type == "month":
                required = 12
                method = "SUM_12M"
                source = "MONTH"
                min_for_extrapolation = 9  # 9 of 12 months → extrapolate
            else:
                required = 1
                method = "YEAR_REPORTED"
                source = "YEAR"
                min_for_extrapolation = 1

            n = len(fy_vals)
            if n >= required:
                # Full coverage — sum directly
                total = sum(v["value"] for v in fy_vals[:required])
                last_pe = max(v["period"]["period_end"] for v in fy_vals[:required])
                results.append({
                    "fiscal_year": fy,
                    "period_end": last_pe,
                    "value": total,
                    "method": method,
                    "coverage_count": n,
                    "coverage_required": required,
                    "source_periods": source,
                    "status": "PASS",
                    "cells_used": [v["cell"] for v in fy_vals[:required]],
                    "raw_used": [v["value"] for v in fy_vals[:required]],
                })
            elif n >= min_for_extrapolation:
                # Partial coverage — extrapolate (annualise from available periods)
                partial_sum = sum(v["value"] for v in fy_vals)
                extrapolated = partial_sum * (required / n)
                last_pe = max(v["period"]["period_end"] for v in fy_vals)
                results.append({
                    "fiscal_year": fy,
                    "period_end": last_pe,
                    "value": extrapolated,
                    "method": f"{method}_EXTRAPOLATED",
                    "coverage_count": n,
                    "coverage_required": required,
                    "source_periods": source,
                    "status": "PASS",
                    "details": f"Extrapolated from {n}/{required} {source.lower()}s.",
                    "cells_used": [v["cell"] for v in fy_vals],
                    "raw_used": [v["value"] for v in fy_vals],
                })
            else:
                results.append({
                    "fiscal_year": fy,
                    "period_end": f"{fy}-{fy_end_month:02d}-28",
                    "value": None,
                    "method": method,
                    "coverage_count": n,
                    "coverage_required": required,
                    "source_periods": source,
                    "status": "FAIL",
                    "fail_code": "INSUFFICIENT_COVERAGE_NO_EXTRAPOLATION",
                    "details": f"Only {n} {source.lower()}s found for FY{fy}; require {min_for_extrapolation}+ to extrapolate, {required} for exact.",
                    "cells_used": [v["cell"] for v in fy_vals],
                    "raw_used": [v["value"] for v in fy_vals],
                })

        elif metric in STOCK_METRICS:
            last_val = max(fy_vals, key=lambda v: v["period"]["period_end"])
            results.append({
                "fiscal_year": fy,
                "period_end": last_val["period"]["period_end"],
                "value": last_val["value"],
                "method": "PERIOD_END",
                "coverage_count": len(fy_vals),
                "coverage_required": 1,
                "source_periods": period_type.upper(),
                "status": "PASS",
                "cells_used": [last_val["cell"]],
                "raw_used": [last_val["value"]],
            })

    return results


def render_evidence_image(
    ws, label_row: int, label_col: int,
    value_cells: list, header_cells: list, header_row: int,
    sheet_name: str, caption: str, out_path: str,
    scale_info: dict = None, annualization_method: str = "",
):
    rows_above = 2
    rows_below = 2
    start_row = max(1, min(label_row, header_row) - rows_above)
    end_row = min(ws.max_row, max(label_row, header_row) + rows_below)
    start_col = max(1, label_col - 1)

    value_cols = set()
    for vc in value_cells:
        m = re.match(r"([A-Z]+)(\d+)", vc)
        if m:
            value_cols.add(column_index_from_string(m.group(1)))
    header_cols = set()
    for hc in header_cells:
        m = re.match(r"([A-Z]+)(\d+)", hc)
        if m:
            header_cols.add(column_index_from_string(m.group(1)))

    all_cols = sorted(set([label_col]) | value_cols | header_cols)
    if all_cols:
        end_col = max(all_cols)
        start_col = min(start_col, min(all_cols))
    else:
        end_col = min(ws.max_column, label_col + 5)

    cell_w = 120
    cell_h = 28
    pad = 20
    caption_h = 80
    header_h = 30

    num_cols = end_col - start_col + 1
    num_rows = end_row - start_row + 1

    img_w = pad * 2 + num_cols * cell_w
    img_h = pad + header_h + num_rows * cell_h + caption_h + pad

    img = Image.new("RGB", (img_w, img_h), "#FFFFFF")
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", 11)
        font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono-Bold.ttf", 11)
        font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", 9)
    except (OSError, IOError):
        font = ImageFont.load_default()
        font_bold = font
        font_small = font

    draw.text((pad, pad // 2), f"Sheet: {sheet_name}", fill="#333333", font=font_bold)

    highlight_cells = set()
    for vc in value_cells:
        m = re.match(r"([A-Z]+)(\d+)", vc)
        if m:
            highlight_cells.add((int(m.group(2)), column_index_from_string(m.group(1)), "#E8F5E9"))
    for hc in header_cells:
        m = re.match(r"([A-Z]+)(\d+)", hc)
        if m:
            highlight_cells.add((int(m.group(2)), column_index_from_string(m.group(1)), "#E3F2FD"))
    highlight_cells.add((label_row, label_col, "#FFF9C4"))

    for ri, row in enumerate(range(start_row, end_row + 1)):
        for ci, col in enumerate(range(start_col, end_col + 1)):
            x = pad + ci * cell_w
            y = pad + header_h + ri * cell_h

            bg_color = "#FFFFFF"
            for (hr, hc_val, color) in highlight_cells:
                if hr == row and hc_val == col:
                    bg_color = color
                    break

            draw.rectangle([x, y, x + cell_w - 1, y + cell_h - 1], fill=bg_color, outline="#CCCCCC")

            cell_val = ws.cell(row=row, column=col).value
            text = ""
            if cell_val is not None:
                if isinstance(cell_val, float):
                    text = f"{cell_val:,.2f}"
                elif isinstance(cell_val, int):
                    text = f"{cell_val:,}"
                else:
                    text = str(cell_val)[:18]

            cell_ref = f"{get_column_letter(col)}{row}"
            cell_label = f"{cell_ref}: "
            draw.text((x + 3, y + 3), cell_label, fill="#999999", font=font_small)
            draw.text((x + 3, y + 14), text, fill="#000000", font=font)

    caption_y = pad + header_h + num_rows * cell_h + 8
    lines = caption.split("\n")
    for i, line in enumerate(lines[:4]):
        draw.text((pad, caption_y + i * 14), line[:120], fill="#555555", font=font_small)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    img.save(out_path)
    return out_path


def build_formula_dag(wb) -> dict:
    dag = {"nodes": {}, "edges": [], "leaf_inputs": set()}
    formula_re = re.compile(r"([A-Z]+[a-z]*[\w\s]*!)?(\$?[A-Z]{1,3}\$?\d+)")

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell_ref = f"'{ws_name}'!{get_column_letter(col)}{row}"

                if cell.data_type == 'f' and cell.value and str(cell.value).startswith('='):
                    formula = str(cell.value)
                    dag["nodes"][cell_ref] = {
                        "type": "formula",
                        "formula": formula,
                        "sheet": ws_name,
                        "row": row,
                        "col": col,
                    }
                    refs = formula_re.findall(formula)
                    for sheet_part, cell_part in refs:
                        clean_ref = cell_part.replace("$", "")
                        if sheet_part:
                            ref_sheet = sheet_part.rstrip("!").strip("'")
                        else:
                            ref_sheet = ws_name
                        dep_ref = f"'{ref_sheet}'!{clean_ref}"
                        dag["edges"].append((dep_ref, cell_ref))
                elif cell.value is not None and isinstance(cell.value, (int, float)):
                    if cell_ref not in dag["nodes"]:
                        dag["nodes"][cell_ref] = {
                            "type": "input",
                            "value": cell.value,
                            "sheet": ws_name,
                            "row": row,
                            "col": col,
                        }

    formula_nodes = {ref for ref, info in dag["nodes"].items() if info["type"] == "formula"}
    targets = {src for src, dst in dag["edges"]}
    for ref, info in dag["nodes"].items():
        if info["type"] == "input" and ref in targets:
            dag["leaf_inputs"].add(ref)

    dag["leaf_inputs"] = list(dag["leaf_inputs"])
    return dag


def get_upstream_inputs(dag: dict, output_cell: str, max_depth: int = 50) -> list:
    reverse_edges = {}
    for src, dst in dag["edges"]:
        reverse_edges.setdefault(dst, []).append(src)

    visited = set()
    inputs = []

    def dfs(node, depth):
        if depth > max_depth or node in visited:
            return
        visited.add(node)
        if node in dag["nodes"] and dag["nodes"][node]["type"] == "input":
            inputs.append(node)
            return
        for parent in reverse_edges.get(node, []):
            dfs(parent, depth + 1)

    dfs(output_cell, 0)
    return inputs


def detect_revenue_type(dag: dict, revenue_cell: str) -> str:
    if revenue_cell not in dag["nodes"]:
        return "UNKNOWN"
    node = dag["nodes"][revenue_cell]
    if node["type"] != "formula":
        return "UNKNOWN"
    formula = node.get("formula", "")

    if re.search(r"(?i)SUM\s*\(", formula):
        refs = re.findall(r"[A-Z]+\d+", formula.replace("$", ""))
        if len(refs) >= 2:
            return "SEGMENT_SUM"

    if "*" in formula:
        return "UNITS_X_PRICE"

    if re.search(r"(?i)(ARR|MRR|churn)", formula):
        return "SUBSCRIPTION_ARR"

    return "UNKNOWN"


def compute_extraction_confidence(record: dict) -> float:
    """
    Score 0.0-1.0 confidence for a single extracted record based on:
    - Label match quality (EXACT=1.0, FUZZY=0.6)
    - Scale evidence (explicit=1.0, assumed=0.5)
    - Coverage completeness
    - Cross-sheet consistency (added post-hoc)
    """
    score = 0.0
    weights = {"label": 0.35, "scale": 0.25, "coverage": 0.25, "status": 0.15}

    quality = record.get("quality", {})
    prov = record.get("provenance", {})
    ann = record.get("annualization", {})

    label_grade = quality.get("evidence_grade", "B")
    score += weights["label"] * (1.0 if label_grade == "A" else 0.6)

    units = prov.get("units", {})
    if units.get("normalized_to") == "RAW_ASSUMED":
        score += weights["scale"] * 0.5
    elif units.get("evidence_cell"):
        score += weights["scale"] * 1.0
    else:
        score += weights["scale"] * 0.3

    required = ann.get("coverage_required", 1)
    actual = ann.get("coverage_count", 0)
    if required > 0:
        score += weights["coverage"] * min(1.0, actual / required)

    if quality.get("status") == "PASS":
        score += weights["status"] * 1.0

    return round(score, 3)


def cross_validate_records(records: list) -> list:
    """
    Detect inconsistencies across sheets for the same metric×fiscal_year.
    Returns warnings list.
    """
    warnings = []
    by_key: dict = {}
    for r in records:
        key = (r["metric"], r["fiscal_year"])
        by_key.setdefault(key, []).append(r)

    for (metric, fy), group in by_key.items():
        if len(group) < 2:
            continue
        values = [r["value_usd"] for r in group if r["value_usd"] is not None]
        if len(values) < 2:
            continue
        vals_set = set(values)
        if len(vals_set) == 1:
            continue
        mean_val = sum(values) / len(values)
        max_dev = max(abs(v - mean_val) for v in values)
        if mean_val != 0 and max_dev / abs(mean_val) > 0.05:
            sheets = [r["provenance"]["source_sheet"] for r in group]
            warnings.append({
                "type": "CROSS_SHEET_MISMATCH",
                "metric": metric,
                "fiscal_year": fy,
                "values": values,
                "sheets": sheets,
                "max_deviation_pct": round(max_dev / abs(mean_val) * 100, 1),
            })

    return warnings


class FinancialPipeline:
    def __init__(self, input_path: str, company_id: str,
                 fy_end_month: int = 12, out_dir: str = "./out"):
        self.input_path = input_path
        self.company_id = company_id
        self.fy_end_month = fy_end_month
        self.out_dir = out_dir
        self.records = []
        self.failures = []
        self.evidence_manifest = []
        self.model_summary = {}
        self.file_sha256 = ""
        self.file_name = ""
        self.wb = None
        self.wb_data = None

        os.makedirs(out_dir, exist_ok=True)
        os.makedirs(os.path.join(out_dir, "evidence_images"), exist_ok=True)

    def run(self) -> dict:
        logger.info(f"Starting pipeline for {self.company_id} with file {self.input_path}")
        # Set up file-based debug logging so diagnostics persist after request
        _debug_log_path = os.path.join(self.out_dir, "extraction_debug.log")
        _file_handler = logging.FileHandler(_debug_log_path, mode='w')
        _file_handler.setLevel(logging.DEBUG)
        _file_handler.setFormatter(logging.Formatter('%(levelname)s %(message)s'))
        logger.addHandler(_file_handler)
        logger.setLevel(logging.DEBUG)

        self.file_name = os.path.basename(self.input_path)
        self.file_sha256 = sha256_file(self.input_path)

        working_path = self.input_path
        ext = Path(self.input_path).suffix.lower()
        if ext == ".xls":
            logger.info("Converting .xls to .xlsx via LibreOffice")
            working_path = convert_xls_to_xlsx(self.input_path, self.out_dir)
        elif ext == ".csv":
            self._convert_csv_to_xlsx(working_path)
            working_path = os.path.join(self.out_dir, Path(self.input_path).stem + ".xlsx")

        self.wb = openpyxl.load_workbook(working_path)
        self.wb_data = openpyxl.load_workbook(working_path, data_only=True)
        self._all_sheet_names = list(self.wb_data.sheetnames)
        logger.info(f"Workbook sheets: {self._all_sheet_names}")
        for sn in self._all_sheet_names:
            logger.info(f"  Sheet '{sn}': score={_score_sheet_relevance(sn)}")

        self._extract_metrics()
        self._derive_margin()
        self._build_model_summary()
        self._write_outputs()

        return self._build_result()

    def _convert_csv_to_xlsx(self, csv_path: str):
        import pandas as pd
        df = pd.read_csv(csv_path)
        xlsx_path = os.path.join(self.out_dir, Path(csv_path).stem + ".xlsx")
        df.to_excel(xlsx_path, index=False)

    # ── Extraction helpers ────────────────────────────────────────────────────

    def _extract_metrics(self):
        """
        Top-level orchestrator: iterates over all sheets (sorted by relevance
        score), finds period headers and metric rows, then delegates per-row
        processing to helpers.

        Scenario awareness:
        1. Sheet-level: sheet name contains scenario keyword (e.g. "Bull Case")
        2. Section-level: section header row within a sheet (e.g. a row reading
           "Bear Case" that divides blocks of metrics)
        3. Label-level: metric label itself embeds a scenario
           (e.g. "Revenue - Downside")
        Priority: label > section > sheet > "base" (default)
        """
        sheet_order = sorted(
            self.wb_data.sheetnames,
            key=lambda s: _score_sheet_relevance(s),
            reverse=True,
        )
        logger.info("Sheet extraction order: %s", sheet_order)

        # Track which (scenario, metric, fy) combos were already extracted
        # from higher-priority sheets to avoid overwriting with lower-quality
        # per-product or per-segment breakdowns.
        self._extracted_keys = set()
        self._extracted_scores = {}  # dedup_key -> sheet_score of winning sheet

        # Extraction diagnostics for debugging
        self._sheet_diagnostics = []

        for ws_name in sheet_order:
            ws = self.wb_data[ws_name]
            sheet_score = _score_sheet_relevance(ws_name)
            diag = {"sheet": ws_name, "score": sheet_score, "status": "processing",
                    "period_headers": 0, "metric_rows": 0, "metrics_found": [],
                    "skipped_reason": None, "data_only_fallback": False}

            sheet_scenario = detect_scenario_from_sheet_name(ws_name)
            scenario_sections = detect_scenario_sections(ws)
            if scenario_sections:
                logger.info(
                    "Detected scenario sections in '%s': %s",
                    ws_name, scenario_sections
                )

            scale = detect_scale_evidence(ws)
            all_headers = find_all_period_header_rows(ws)

            # FALLBACK: if data_only worksheet has no headers, try the formula worksheet
            # (which has text labels but formulas as strings instead of calculated values).
            # This handles the case where the file was saved without cached formula results.
            if not all_headers and ws_name in self.wb.sheetnames:
                ws_formula = self.wb[ws_name]
                all_headers_formula = find_all_period_header_rows(ws_formula)
                if all_headers_formula:
                    logger.info(
                        f"Sheet '{ws_name}' (score={sheet_score}): data_only had no headers, "
                        f"but formula sheet has {len(all_headers_formula)} — using formula sheet for headers"
                    )
                    all_headers = all_headers_formula
                    diag["data_only_fallback"] = True

            if not all_headers:
                diag["status"] = "skipped"
                diag["skipped_reason"] = "no_period_headers"
                logger.info(f"Sheet '{ws_name}' (score={sheet_score}): NO period headers found — skipping")
                # Log what the first few rows actually contain for debugging
                sample_cells = []
                for r in range(1, min(6, ws.max_row + 1)):
                    row_vals = []
                    for c in range(1, min(15, ws.max_column + 1)):
                        v = ws.cell(row=r, column=c).value
                        if v is not None:
                            row_vals.append(f"({c})={repr(v)[:40]}")
                    if row_vals:
                        sample_cells.append(f"  Row {r}: {', '.join(row_vals)}")
                if sample_cells:
                    logger.info(f"  First rows of '{ws_name}':\n" + "\n".join(sample_cells))
                self._sheet_diagnostics.append(diag)
                continue

            diag["period_headers"] = len(all_headers)
            # Log what periods were found
            for h in all_headers[:3]:
                periods_desc = [f"{p.get('raw','?')}@col{p.get('col','?')}" for p in h.get("periods", [])[:5]]
                logger.info(f"Sheet '{ws_name}' (score={sheet_score}): header row {h['row']}: {periods_desc}")

            metric_rows = find_metric_rows(ws)
            if not metric_rows:
                diag["status"] = "skipped"
                diag["skipped_reason"] = "no_metric_labels"
                logger.info(f"Sheet '{ws_name}' (score={sheet_score}): NO metric labels found — skipping")
                # Log labels in first column for debugging
                sample_labels = []
                for r in range(1, min(50, ws.max_row + 1)):
                    for c in range(1, min(5, ws.max_column + 1)):
                        v = ws.cell(row=r, column=c).value
                        if isinstance(v, str) and v.strip():
                            sample_labels.append(f"  ({c},{r}): {v.strip()[:50]}")
                if sample_labels:
                    logger.info(f"  Labels in '{ws_name}':\n" + "\n".join(sample_labels[:20]))
                self._sheet_diagnostics.append(diag)
                continue

            diag["metric_rows"] = len(metric_rows)
            diag["metrics_found"] = [f"{m['metric']}@{m['label_cell']}({m['matched_label'][:30]})" for m in metric_rows]
            logger.info(
                f"Sheet '{ws_name}' (score={sheet_score}): found {len(metric_rows)} metric rows: "
                f"{[m['metric'] for m in metric_rows]}"
            )

            for mi in metric_rows:
                label_scenario = mi.get("label_scenario")
                row_scenario = resolve_scenario_for_row(
                    mi["row"], scenario_sections, sheet_scenario
                )
                mi["_resolved_scenario"] = label_scenario or row_scenario or "base"
                mi["_sheet_score"] = sheet_score

                period_header = self._nearest_header(all_headers, mi["row"], mi["label_col"])
                if period_header:
                    self._process_metric_row(ws, ws_name, mi, period_header, scale)

            diag["status"] = "processed"
            self._sheet_diagnostics.append(diag)

    def _nearest_header(self, all_headers: list, metric_row: int, label_col: int) -> Optional[dict]:
        """Return the header zone whose row is <= metric_row and whose column range
        is compatible with the metric label's column position.

        Rejects headers whose leftmost period column is too far right of the
        metric label (prevents cross-region pairing, e.g. a P&L label at col 3
        pairing with a scenario table header starting at col 20).

        When multiple candidates exist at the same row, prefer annual-type headers
        over quarterly/monthly since they contain pre-annualized values.
        """
        MAX_COL_GAP = 15

        candidates = []
        for h in all_headers:
            if h["row"] > metric_row:
                continue
            periods = h.get("periods", [])
            if not periods:
                continue
            first_col = periods[0].get("col", 1)
            last_col = periods[-1].get("col", first_col)
            if label_col > last_col + 4:
                continue
            if first_col - label_col > MAX_COL_GAP:
                continue
            candidates.append(h)
        if not candidates:
            return all_headers[0] if all_headers else None

        def _header_score(h):
            """Higher score = preferred header.
            Prefer: closest row, annual type over sub-annual, more periods."""
            row_proximity = h["row"]
            periods = h.get("periods", [])
            ptype = periods[0].get("type", "") if periods else ""
            type_bonus = 100 if ptype == "year" else 0
            return (row_proximity, type_bonus, len(periods))

        return max(candidates, key=_header_score)

    def _process_metric_row(self, ws, ws_name: str, mi: dict, period_header: dict, scale: dict):
        """
        Extract and annualize values for one metric row, build records/failures,
        and render evidence images.
        """
        label_scale = detect_label_scale(mi.get("matched_label", ""))
        if label_scale is not None:
            scale = {
                "normalized_to": "USD",
                "evidence_cell": mi["label_cell"],
                "evidence_text": f"Scale inferred from label: {mi['matched_label']}",
                "_multiply_by": label_scale,
            }
        extracted = extract_metric_values(ws, mi, period_header)
        if not extracted["values"]:
            # Check if the formula workbook has formulas in these cells (data_only=None issue)
            has_formulas = False
            if ws_name in self.wb.sheetnames:
                ws_f = self.wb[ws_name]
                sample_formula_cells = []
                for p in period_header.get("periods", [])[:5]:
                    fc = ws_f.cell(row=mi["row"], column=p["col"]).value
                    if fc is not None:
                        sample_formula_cells.append(f"col{p['col']}={repr(fc)[:60]}")
                        if isinstance(fc, str) and fc.startswith("="):
                            has_formulas = True
                if sample_formula_cells:
                    logger.info(
                        f"  NO_NUMERIC_VALUES for '{mi['matched_label']}' on '{ws_name}' "
                        f"— formula sheet cells: {sample_formula_cells}"
                    )

            fail_detail = f"No numeric values found for '{mi['matched_label']}' on sheet '{ws_name}'"
            if has_formulas:
                fail_detail += " (cells contain formulas — file may need to be opened in Excel and re-saved to cache values)"

            self.failures.append({
                "metric": mi["metric"],
                "fiscal_year": None,
                "quality": {
                    "status": "FAIL",
                    "fail_code": "NO_NUMERIC_VALUES",
                    "details": fail_detail,
                    "_has_formulas": has_formulas,
                },
                "provenance": {
                    "source_file": {"file_name": self.file_name, "file_sha256": self.file_sha256},
                    "source_sheet": ws_name,
                },
            })
            return

        annualized = annualize_metric(mi["metric"], extracted["values"], self.fy_end_month)
        scenario = mi.get("_resolved_scenario", "base")
        sheet_score = mi.get("_sheet_score", 25)

        for ann in annualized:
            fy = ann.get("fiscal_year")
            dedup_key = (scenario, mi["metric"], fy)

            # If this (scenario, metric, fy) was already extracted from a
            # higher-priority sheet, skip it.  We also store the score of the
            # winning sheet so we can compare properly.
            if dedup_key in self._extracted_keys:
                prev_score = self._extracted_scores.get(dedup_key, 0)
                if sheet_score <= prev_score:
                    logger.debug(
                        "Skipping duplicate %s FY%s from sheet '%s' (score %d <= %d)",
                        mi["metric"], fy, ws_name, sheet_score, prev_score
                    )
                    continue
                else:
                    # Higher-priority sheet found later — replace the old record
                    logger.info(
                        "Replacing %s FY%s with higher-priority sheet '%s' (score %d > %d)",
                        mi["metric"], fy, ws_name, sheet_score, prev_score
                    )
                    self.records = [
                        r for r in self.records
                        if not (r.get("scenario") == scenario
                                and r.get("metric") == mi["metric"]
                                and r.get("fiscal_year") == fy)
                    ]

            record = self._build_record_for(ws, ws_name, mi, ann, period_header, scale)
            if ann["status"] == "FAIL":
                self.failures.append(record)
            else:
                self.records.append(record)
                self._extracted_keys.add(dedup_key)
                self._extracted_scores[dedup_key] = sheet_score

            evidence_path = record["provenance"]["evidence_image"]["path"]
            img_sha = record["provenance"]["evidence_image"]["sha256"]
            if evidence_path:
                self.evidence_manifest.append({
                    "path": evidence_path,
                    "sha256": img_sha,
                    "proves": f"{mi['metric']} FY{ann['fiscal_year']} extraction from {ws_name}",
                    "metric": mi["metric"],
                    "fiscal_year": ann["fiscal_year"],
                })

    def _compute_value(
        self, ann: dict, scale: dict, is_unit_metric: bool
    ) -> tuple:
        """
        Return (value_usd, fail_code_or_None).

        For unit metrics: raw count is stored directly (scale does not apply).
        For financial metrics: cell value × _multiply_by → actual USD.
        """
        if is_unit_metric:
            return ann["value"], None

        if scale.get("_scale_failed"):
            return None, "NO_SCALE_EVIDENCE"

        multiply = scale.get("_multiply_by")
        if multiply is None:
            return None, "NO_SCALE_EVIDENCE"

        val = ann["value"]
        if val is None:
            return None, None
        return val * multiply, None

    def _render_evidence(
        self, ws, ws_name: str, mi: dict, ann: dict,
        period_header: dict, scale: dict
    ) -> tuple:
        """Render evidence image and return (path, sha256). Returns ('','') on failure."""
        evidence_path = os.path.join(
            self.out_dir, "evidence_images",
            f"{mi['metric']}_FY{ann['fiscal_year']}_{ws_name}.png"
        )
        caption = (
            f"Metric: {mi['metric']} | FY{ann['fiscal_year']} | Sheet: {ws_name}\n"
            f"Label: {mi['label_cell']} ({mi['matched_label']}) | Match: {mi['match_type']}\n"
            f"Values: {', '.join(ann['cells_used'])} | Scale: {scale.get('evidence_text', 'N/A')}\n"
            f"Method: {ann['method']} | Coverage: {ann['coverage_count']}/{ann['coverage_required']}"
        )
        try:
            render_evidence_image(
                ws, mi["row"], mi["label_col"],
                ann["cells_used"],
                [p["cell"] for p in period_header["periods"]],
                period_header["row"], ws_name, caption, evidence_path,
                scale, ann["method"],
            )
            return evidence_path, sha256_file(evidence_path)
        except Exception as e:
            logger.warning(f"Failed to render evidence image: {e}")
            return "", ""

    def _build_record_for(
        self, ws, ws_name: str, mi: dict, ann: dict,
        period_header: dict, scale: dict
    ) -> dict:
        """Assemble the full record dict for one metric x fiscal_year observation."""
        is_unit = mi["metric"] in UNIT_METRICS
        unit_type = mi.get("unit_type") if is_unit else None
        currency_field = (unit_type or "UNITS") if is_unit else "USD"

        value_usd, fail_code = self._compute_value(ann, scale, is_unit)
        if fail_code and ann["status"] != "FAIL":
            ann["status"] = "FAIL"
            ann["fail_code"] = fail_code
            ann["details"] = "Cannot normalize to thousands: no scale evidence found in sheet"

        evidence_path, img_sha = self._render_evidence(ws, ws_name, mi, ann, period_header, scale)

        quality = {
            "status": ann["status"],
            "evidence_grade": "A" if mi["match_type"] == "EXACT" else "B",
        }
        if ann["status"] == "FAIL":
            quality["fail_code"] = ann.get("fail_code", "")
            quality["details"] = ann.get("details", "")

        return {
            "metric": mi["metric"],
            "fiscal_year": ann["fiscal_year"],
            "period_end": ann["period_end"],
            "scenario": mi.get("_resolved_scenario", "base"),
            "currency": currency_field,
            "value_usd": round(value_usd, 4) if value_usd is not None else None,
            "unit_type": unit_type,
            "annualization": {
                "method": ann["method"],
                "coverage_required": ann["coverage_required"],
                "coverage_count": ann["coverage_count"],
                "source_periods": ann["source_periods"],
            },
            "provenance": {
                "source_file": {"file_name": self.file_name, "file_sha256": self.file_sha256},
                "source_sheet": ws_name,
                "label": {
                    "matched_label": mi["matched_label"],
                    "match_type": mi["match_type"],
                    "label_cell": mi["label_cell"],
                },
                "period_headers": {
                    "header_cells": [p["cell"] for p in period_header["periods"]],
                    "parsed_period_ends": [p["period_end"] for p in period_header["periods"]],
                },
                "values": {
                    "value_cells": ann["cells_used"],
                    "raw_values": ann["raw_used"],
                },
                "units": {
                    "scale_multiply_by": None if is_unit else scale.get("_multiply_by"),
                    "normalized_to": (unit_type or "RAW_UNITS") if is_unit else "USD",
                    "evidence_cell": None if is_unit else scale.get("evidence_cell"),
                    "evidence_text": (
                        f"Physical units ({unit_type or 'UNITS'}) -- no dollar scale applied"
                        if is_unit else scale.get("evidence_text")
                    ),
                },
                "evidence_image": {
                    "path": evidence_path,
                    "sha256": img_sha,
                    "highlights": {
                        "label_cell": mi["label_cell"],
                        "header_cells": [p["cell"] for p in period_header["periods"]],
                        "value_cells": ann["cells_used"],
                    },
                },
            },
            "quality": quality,
        }

    # ── Post-extraction steps ─────────────────────────────────────────────────

    def _derive_margin(self):
        rev_by_fy = {}
        ebitda_by_fy = {}
        for r in self.records:
            if r["metric"] == "revenue" and r["value_usd"]:
                rev_by_fy[r["fiscal_year"]] = r["value_usd"]
            if r["metric"] == "ebitda" and r["value_usd"]:
                ebitda_by_fy[r["fiscal_year"]] = r["value_usd"]

        for fy in rev_by_fy:
            if fy in ebitda_by_fy and rev_by_fy[fy] != 0:
                margin = ebitda_by_fy[fy] / rev_by_fy[fy]
                self.records.append({
                    "metric": "margin_pct_annual",
                    "fiscal_year": fy,
                    "period_end": f"{fy}-12-31",
                    "currency": "USD",
                    "value_usd": round(margin, 6),
                    "annualization": {
                        "method": "DERIVED",
                        "coverage_required": 1,
                        "coverage_count": 1,
                        "source_periods": "DERIVED",
                    },
                    "provenance": {
                        "source_file": {"file_name": self.file_name, "file_sha256": self.file_sha256},
                        "source_sheet": "DERIVED",
                        "label": {
                            "matched_label": "margin_pct_annual = ebitda / revenue",
                            "match_type": "DERIVED",
                            "label_cell": "N/A",
                        },
                    },
                    "quality": {"status": "PASS", "evidence_grade": "A"},
                })

    def _build_model_summary(self):
        try:
            dag = build_formula_dag(self.wb)
        except Exception as e:
            logger.warning(f"DAG construction failed: {e}")
            dag = {"nodes": {}, "edges": [], "leaf_inputs": []}

        outputs_info = {}
        output_cells = {}

        for metric in ["revenue", "ebitda", "arr"]:
            metric_records = [r for r in self.records if r["metric"] == metric]
            if not metric_records:
                continue
            best = max(metric_records, key=lambda r: r["fiscal_year"])
            sheet = best["provenance"]["source_sheet"]
            label_cell = best["provenance"]["label"]["label_cell"]
            value_cells = best["provenance"]["values"]["value_cells"]

            out_cell_ref = value_cells[-1] if value_cells else label_cell
            full_ref = f"'{sheet}'!{out_cell_ref}"

            upstream = get_upstream_inputs(dag, full_ref)

            top_inputs = []
            for inp_ref in upstream[:20]:
                node = dag["nodes"].get(inp_ref, {})
                inp_parts = inp_ref.split("!")
                if len(inp_parts) == 2:
                    inp_sheet = inp_parts[0].strip("'")
                    inp_cell = inp_parts[1]
                    try:
                        ws_d = self.wb_data[inp_sheet]
                        m = re.match(r"([A-Z]+)(\d+)", inp_cell)
                        if m:
                            c = column_index_from_string(m.group(1))
                            r = int(m.group(2))
                            label_val = ws_d.cell(row=r, column=max(1, c - 1)).value
                            top_inputs.append({
                                "cell": inp_ref,
                                "label": str(label_val or "")[:60],
                                f"elasticity_{metric}": 0,
                            })
                    except Exception:
                        pass

            rev_type = "UNKNOWN"
            if metric == "revenue":
                rev_type = detect_revenue_type(dag, full_ref)

            outputs_info[metric] = {
                "output_cell": f"{sheet}!{out_cell_ref}",
                "fiscal_year": best["fiscal_year"],
                "revenue_type": rev_type if metric == "revenue" else None,
                "top_inputs": top_inputs,
            }
            m = re.match(r"([A-Z]+)(\d+)", out_cell_ref)
            if m:
                output_cells[metric] = {"sheet": sheet, "cell_ref": out_cell_ref}

        verified = []
        for metric, info in outputs_info.items():
            verified.append(
                f"{metric.title()} output cell is {info['output_cell']} (FY{info['fiscal_year']})."
            )
            if info.get("top_inputs"):
                top = info["top_inputs"][0]
                verified.append(
                    f"Top {metric} upstream input is {top['cell']} (label '{top['label']}')."
                )

        self.model_summary = {
            "company_id": self.company_id,
            "outputs": outputs_info,
            "verified_statements": verified,
            "evidence": {
                "graph_stats": {
                    "nodes": len(dag["nodes"]),
                    "edges": len(dag["edges"]),
                },
                "leaf_inputs_count": len(dag["leaf_inputs"]),
                "notes": [],
            },
        }

    # ── Output writers ────────────────────────────────────────────────────────

    def _write_outputs(self):
        metrics_out = {
            "company_id": self.company_id,
            "fy_end_month": self.fy_end_month,
            "records": self.records,
            "failures": self.failures,
        }
        with open(os.path.join(self.out_dir, "annual_metrics.json"), "w") as f:
            json.dump(metrics_out, f, indent=2, default=str)

        with open(os.path.join(self.out_dir, "model_summary.json"), "w") as f:
            json.dump(self.model_summary, f, indent=2, default=str)

        with open(os.path.join(self.out_dir, "evidence_manifest.json"), "w") as f:
            json.dump(self.evidence_manifest, f, indent=2, default=str)

        self._write_human_summary()

    def _write_human_summary(self):
        lines = [
            f"# Financial Model Summary — {self.company_id}",
            "",
            f"**Source File:** {self.file_name}",
            f"**Fiscal Year End Month:** {self.fy_end_month}",
            "",
            "## Annual Metrics",
            "",
            "| Metric | FY | Value (thousands) | Method | Status |",
            "|--------|----|--------------------|--------|--------|",
        ]

        for r in sorted(self.records, key=lambda x: (x["fiscal_year"], x["metric"])):
            val = f"{r['value_usd']:,.2f}" if r["value_usd"] is not None else "N/A"
            lines.append(
                f"| {r['metric']} | {r['fiscal_year']} | {val} | "
                f"{r['annualization']['method']} | {r['quality']['status']} |"
            )

        if self.failures:
            lines.extend([
                "",
                "## Failures",
                "",
                "| Metric | FY | Fail Code | Details |",
                "|--------|----|-----------|---------|",
            ])
            for f in self.failures:
                fy = f.get("fiscal_year", "N/A")
                q = f.get("quality", {})
                lines.append(
                    f"| {f['metric']} | {fy} | {q.get('fail_code', '')} | {q.get('details', '')} |"
                )

        outputs = self.model_summary.get("outputs", {})
        if outputs:
            lines.extend(["", "## Model Structure", ""])
            for metric, info in outputs.items():
                lines.append(f"**{metric.title()}:** Output cell `{info.get('output_cell', 'N/A')}`")
                if metric == "revenue" and info.get("revenue_type"):
                    lines.append(f"  Revenue type: {info['revenue_type']}")
                top = info.get("top_inputs", [])[:5]
                if top:
                    lines.append(f"  Top drivers:")
                    for t in top:
                        lines.append(f"  - `{t['cell']}` ({t['label']})")
                lines.append("")

        if self.model_summary.get("verified_statements"):
            lines.extend(["## Verified Statements", ""])
            for stmt in self.model_summary["verified_statements"]:
                lines.append(f"- {stmt}")

        lines.extend(["", f"*Generated: {datetime.datetime.now().isoformat()}*", ""])

        md_path = os.path.join(self.out_dir, "human_summary.md")
        with open(md_path, "w") as f:
            f.write("\n".join(lines))

    def _build_result(self) -> dict:
        """
        Assemble the final result dict with confidence scores and cross-validation.

        Pivot tables included for downstream consumers:
          result["financials"]   -> {metric: {fiscal_year: value_usd}}  (base scenario)
          result["units"]        -> {metric: {fiscal_year: {...}}}           (base scenario)
          result["fiscal_years"] -> sorted list of all fiscal years in records
          result["scenarios"]    -> {scenario_name: {"financials": {...}, "units": {...}}}
        """
        current_year = datetime.date.today().year
        min_fy = current_year - 10
        max_fy = current_year + 15
        self.records = [r for r in self.records if min_fy <= r.get("fiscal_year", 0) <= max_fy]

        # Build per-scenario pivots
        scenarios_financials: dict = {}
        scenarios_units: dict = {}
        detected_scenarios = set()

        for r in self.records:
            r["confidence"] = compute_extraction_confidence(r)
            scenario = r.get("scenario", "base")
            detected_scenarios.add(scenario)
            metric = r["metric"]
            fy = r["fiscal_year"]
            if metric in UNIT_METRICS:
                scenarios_units.setdefault(scenario, {}).setdefault(metric, {})[fy] = {
                    "value": r["value_usd"],
                    "unit_type": r.get("unit_type"),
                    "confidence": r["confidence"],
                }
            elif r["value_usd"] is not None:
                scenarios_financials.setdefault(scenario, {}).setdefault(metric, {})[fy] = r["value_usd"]

        # Primary financials/units use "base" scenario (backward compatible)
        financials = scenarios_financials.get("base", {})
        units = scenarios_units.get("base", {})

        # If no "base" but other scenarios exist, pick the one with most records
        if not financials and scenarios_financials:
            best_scenario = max(scenarios_financials, key=lambda s: sum(
                len(v) for v in scenarios_financials[s].values()
            ))
            financials = scenarios_financials[best_scenario]
            units = scenarios_units.get(best_scenario, {})
            logger.info(
                "No 'base' scenario found; using '%s' as primary (%d metrics)",
                best_scenario, len(financials)
            )

        # Build structured scenarios dict for downstream consumers
        scenarios_output = {}
        for sc in sorted(detected_scenarios):
            scenarios_output[sc] = {
                "financials": scenarios_financials.get(sc, {}),
                "units": scenarios_units.get(sc, {}),
            }

        xval_warnings = cross_validate_records(self.records)
        if xval_warnings:
            logger.warning("Cross-validation found %d issue(s)", len(xval_warnings))

        all_fy = sorted({r["fiscal_year"] for r in self.records if r.get("fiscal_year")})
        evidence_paths = [e["path"] for e in self.evidence_manifest]

        avg_conf = (
            sum(r.get("confidence", 0) for r in self.records) / len(self.records)
            if self.records else 0
        )

        return {
            "company_id": self.company_id,
            "status": "completed",
            "records_count": len(self.records),
            "failures_count": len(self.failures),
            "records": self.records,
            "failures": self.failures,
            "financials": financials,
            "units": units,
            "fiscal_years": all_fy,
            "scenarios": scenarios_output,
            "detected_scenarios": sorted(detected_scenarios),
            "model_summary": self.model_summary,
            "evidence_paths": evidence_paths,
            "output_dir": self.out_dir,
            "cross_validation_warnings": xval_warnings,
            "extraction_confidence": {
                "average": round(avg_conf, 3),
                "total_records": len(self.records),
                "high_confidence": sum(1 for r in self.records if r.get("confidence", 0) >= 0.7),
                "low_confidence": sum(1 for r in self.records if r.get("confidence", 0) < 0.4),
            },
            "files": {
                "annual_metrics": os.path.join(self.out_dir, "annual_metrics.json"),
                "model_summary": os.path.join(self.out_dir, "model_summary.json"),
                "human_summary": os.path.join(self.out_dir, "human_summary.md"),
                "evidence_manifest": os.path.join(self.out_dir, "evidence_manifest.json"),
            },
            "_sheet_diagnostics": getattr(self, '_sheet_diagnostics', []),
            "_all_sheets": [{"name": s, "score": _score_sheet_relevance(s)} for s in getattr(self, '_all_sheet_names', [])],
        }


# ── Public API ────────────────────────────────────────────────────────────────

def run_pipeline(input_path: str, company_id: str,
                 fy_end_month: int = 12, out_dir: str = "./out") -> dict:
    """
    Run the financial extraction pipeline on an Excel (or CSV) model file.

    Args:
        input_path:    Path to the .xlsx, .xls, or .csv file.
        company_id:    Identifier string for the company (used in outputs).
        fy_end_month:  Fiscal year end month (1–12). Defaults to 12 (Dec).
        out_dir:       Directory where output files are written.

    Returns:
        Result dict with keys: company_id, status, records, failures,
        financials, units, fiscal_years, model_summary, evidence_paths, files.

    Raises:
        FileNotFoundError: if input_path does not exist.
        ValueError:        if file type is unsupported or arguments are invalid.
    """
    p = Path(input_path)
    if not p.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if p.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
        raise ValueError(
            f"Unsupported file type '{p.suffix}'; expected .xlsx, .xls, or .csv"
        )
    if not company_id or not str(company_id).strip():
        raise ValueError("company_id must be a non-empty string")
    if not isinstance(fy_end_month, int) or not (1 <= fy_end_month <= 12):
        raise ValueError(
            f"fy_end_month must be an integer between 1 and 12, got {fy_end_month!r}"
        )

    pipeline = FinancialPipeline(input_path, company_id, fy_end_month, out_dir)
    return pipeline.run()


def get_metric_series(result: dict, metric: str) -> dict:
    """
    Return {fiscal_year: value_usd} for a financial metric.

    Example:
        rev = get_metric_series(result, "revenue")
        # {2027: 67234.5, 2028: 145000.0, ...}
    """
    return result.get("financials", {}).get(metric, {})


def get_units_series(result: dict, metric: str = "units_sold") -> dict:
    """
    Return {fiscal_year: {"value": float, "unit_type": str}} for a physical-unit metric.

    Example:
        mw = get_units_series(result, "units_sold")
        # {2027: {"value": 150.0, "unit_type": "MW"}, ...}
    """
    return result.get("units", {}).get(metric, {})


def pivot_by_year(result: dict, metrics: Optional[list] = None) -> dict:
    """
    Return {fiscal_year: {metric: value_usd}} for financial metrics.

    Args:
        result:  Pipeline result dict from run_pipeline().
        metrics: Optional list of metric names to include. Defaults to all.

    Example:
        table = pivot_by_year(result, metrics=["revenue", "ebitda"])
        # {2027: {"revenue": 67234.5, "ebitda": 40340.7}, ...}
    """
    fin = result.get("financials", {})
    if metrics:
        fin = {m: v for m, v in fin.items() if m in metrics}
    all_fy = sorted({fy for series in fin.values() for fy in series})
    return {
        fy: {m: fin[m].get(fy) for m in fin}
        for fy in all_fy
    }


def to_dataframe(result: dict, metrics: Optional[list] = None,
                 include_units: bool = False):
    """
    Return a pandas DataFrame indexed by fiscal_year.

    Columns are financial metric names (and optionally unit metric names).
    Requires pandas to be installed.

    Args:
        result:        Pipeline result dict from run_pipeline().
        metrics:       Optional list of metric names to include.
        include_units: If True, also include unit metrics (raw counts).

    Example:
        df = to_dataframe(result, metrics=["revenue", "ebitda", "units_sold"],
                          include_units=True)
    """
    import pandas as pd

    fin = result.get("financials", {})
    if metrics:
        fin = {m: v for m, v in fin.items() if m in metrics}

    all_fy = sorted({fy for series in fin.values() for fy in series})
    data = {m: {fy: fin[m].get(fy) for fy in all_fy} for m in fin}

    if include_units:
        for metric, series in result.get("units", {}).items():
            if not metrics or metric in metrics:
                data[metric] = {fy: info.get("value") for fy, info in series.items()}
                # Fill any fiscal years not present in the unit series
                for fy in all_fy:
                    data[metric].setdefault(fy, None)

    df = pd.DataFrame(data, index=all_fy)
    df.index.name = "fiscal_year"
    return df

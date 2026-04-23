"""Deterministic Excel reader — no domain knowledge, no label normalization.

Opens a workbook twice: once with formulas visible (`data_only=False`),
once with cached values (`data_only=True`). Every method returns plain
JSON-serializable data so it can be fed to an LLM tool-use loop.

The agent's job is to decide WHAT to read. This module's job is to
be unwrong about WHAT'S THERE.
"""
from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


class ExcelWorkbook:
    """Wraps an .xlsx file for agent-driven inspection."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")

        # data_only=True returns cached values; data_only=False returns formulas.
        # We keep both so we can expose each cell's value AND formula on demand.
        self._wb_values = load_workbook(self.path, data_only=True, read_only=False)
        self._wb_formulas = load_workbook(self.path, data_only=False, read_only=False)

    # ── Discovery ────────────────────────────────────────────────────────

    def list_sheets(self) -> list[dict[str, Any]]:
        """Return metadata about each sheet in the workbook."""
        out = []
        for name in self._wb_values.sheetnames:
            ws = self._wb_values[name]
            out.append({
                "name": name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "max_col_letter": get_column_letter(ws.max_column),
                "hidden": ws.sheet_state != "visible",
            })
        return out

    def get_named_ranges(self) -> list[dict[str, Any]]:
        """Return any workbook-level defined names (named ranges)."""
        out = []
        for name in self._wb_values.defined_names:
            dn = self._wb_values.defined_names[name]
            out.append({
                "name": name,
                "value": str(dn.value) if dn.value else None,
            })
        return out

    # ── Reading ──────────────────────────────────────────────────────────

    # Default cap on preview width. Wider sheets (Mitra's plant-FS tabs are 45+
    # cols) caused individual preview_sheet calls to return 80–120KB of JSON,
    # which blew the Claude context window after ~6 previews. If the agent
    # legitimately needs more columns, it can pass an explicit `cols` value.
    _PREVIEW_COLS_DEFAULT = 20

    def preview_sheet(self, sheet: str, rows: int = 30, cols: Optional[int] = None) -> dict[str, Any]:
        """Return a compact preview of a sheet.

        For each row, we emit only the non-null cells as [address, value,
        (formula?)] tuples. Empty cells are silently skipped. This cuts a
        typical wide-sheet preview from ~100KB to ~10–15KB while keeping the
        information the agent actually uses — where labels/values live.

        The agent can always fall back to `read_range` or `read_cell` when it
        needs formulas or the full grid of a specific region.
        """
        self._require_sheet(sheet)
        ws = self._wb_values[sheet]
        end_row = min(rows, ws.max_row) if ws.max_row else 0
        effective_cols = cols if cols else self._PREVIEW_COLS_DEFAULT
        end_col = min(effective_cols, ws.max_column) if ws.max_column else 0

        rows_out: list[dict[str, Any]] = []
        total_cells_scanned = 0
        total_cells_emitted = 0
        truncated = False

        # Soft byte budget so one preview call can't dominate the context
        # even if the sheet is unusually dense with long string cells.
        BYTE_BUDGET = 12_000
        approx_bytes = 0

        for r in range(1, end_row + 1):
            row_cells: list[list[Any]] = []
            for c in range(1, end_col + 1):
                cell = ws.cell(r, c)
                total_cells_scanned += 1
                v = self._serialize(cell.value)
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    continue
                addr = cell.coordinate
                # Compact tuple: [address, value] — formula appended only if present.
                entry: list[Any] = [addr, v]
                try:
                    fc = self._wb_formulas[sheet][addr]
                    if isinstance(fc.value, str) and fc.value.startswith("="):
                        entry.append(fc.value)
                except KeyError:
                    pass
                row_cells.append(entry)
                total_cells_emitted += 1
                approx_bytes += len(str(entry))
                if approx_bytes > BYTE_BUDGET:
                    truncated = True
                    break
            if row_cells:
                rows_out.append({"r": r, "cells": row_cells})
            if truncated:
                break

        return {
            "sheet": sheet,
            "rows_returned": end_row,
            "cols_returned": end_col,
            "sheet_max_row": ws.max_row,
            "sheet_max_col": ws.max_column,
            "cells_emitted": total_cells_emitted,
            "cells_scanned": total_cells_scanned,
            "truncated": truncated,
            # Compact sparse format: one entry per non-empty row, each with
            # a list of [address, value, formula?] tuples. Empty cells are
            # omitted — the agent should infer "missing" from absence.
            "rows": rows_out,
            "format_note": "Sparse: each row dict has 'r' (row number) and "
                           "'cells' (list of [address, value, optional formula]). "
                           "Empty cells omitted. Use read_range for full grid.",
        }

    def read_range(self, sheet: str, range_str: str) -> dict[str, Any]:
        """Read a cell range (e.g. 'C12:I12' or 'A1:H40').

        For a horizontal range (single row), returns a flat list of cells.
        For a rectangular range, returns a 2D grid.
        """
        self._require_sheet(sheet)
        ws = self._wb_values[sheet]

        # openpyxl always returns a tuple of row-tuples for range strings,
        # even for a single-row range like "B2:D2" — so we detect that here
        # and flatten to 1d.
        selection = ws[range_str]
        if not isinstance(selection, tuple):
            # Single cell (openpyxl returns a bare cell for "A1")
            return {
                "sheet": sheet,
                "range": range_str,
                "shape": "single",
                "cells": [self._cell_to_dict(selection, sheet)],
            }

        # Normalize to a 2D grid of cells.
        if selection and not isinstance(selection[0], tuple):
            # Shouldn't happen for range strings, but handle defensively.
            grid = [list(selection)]
        else:
            grid = [list(row) for row in selection]

        n_rows = len(grid)
        n_cols = len(grid[0]) if grid else 0

        # Flatten single-row or single-column selections to 1d.
        if n_rows == 1:
            return {
                "sheet": sheet,
                "range": range_str,
                "shape": "1d",
                "cells": [self._cell_to_dict(c, sheet) for c in grid[0]],
            }
        if n_cols == 1:
            return {
                "sheet": sheet,
                "range": range_str,
                "shape": "1d",
                "cells": [self._cell_to_dict(row[0], sheet) for row in grid],
            }

        return {
            "sheet": sheet,
            "range": range_str,
            "shape": "2d",
            "rows": n_rows,
            "cols": n_cols,
            "grid": [[self._cell_to_dict(c, sheet) for c in row] for row in grid],
        }

    def read_cell(self, sheet: str, address: str) -> dict[str, Any]:
        """Read one cell. Returns value, formula (if any), and data type."""
        self._require_sheet(sheet)
        ws_v = self._wb_values[sheet]
        ws_f = self._wb_formulas[sheet]
        cv = ws_v[address]
        cf = ws_f[address]
        return self._cell_to_dict(cv, sheet, formula_cell=cf)

    # ── Search ───────────────────────────────────────────────────────────

    def find_label(self, pattern: str, case_insensitive: bool = True, max_hits: int = 200) -> list[dict[str, Any]]:
        """Find cells whose text content matches a regex.

        Use this to locate candidate labels across sheets without having to
        preview every sheet first.
        """
        flags = re.IGNORECASE if case_insensitive else 0
        try:
            pat = re.compile(pattern, flags)
        except re.error as exc:
            return [{"error": f"invalid regex: {exc}"}]

        hits = []
        for sheet_name in self._wb_values.sheetnames:
            ws = self._wb_values[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and pat.search(cell.value):
                        hits.append({
                            "sheet": sheet_name,
                            "address": cell.coordinate,
                            "text": cell.value,
                        })
                        if len(hits) >= max_hits:
                            return hits
        return hits

    # ── Internal ─────────────────────────────────────────────────────────

    def _require_sheet(self, sheet: str) -> None:
        if sheet not in self._wb_values.sheetnames:
            raise ValueError(
                f"Sheet '{sheet}' not found. Available: {self._wb_values.sheetnames}"
            )

    def _cell_to_dict(self, cell, sheet: str, formula_cell=None) -> dict[str, Any]:
        """Normalize a cell into a JSON-safe dict.

        Pulls the formula from the parallel formula-workbook if available.
        """
        value = self._serialize(cell.value)

        # Figure out the formula, if any
        formula = None
        if formula_cell is None:
            try:
                fc = self._wb_formulas[sheet][cell.coordinate]
                if isinstance(fc.value, str) and fc.value.startswith("="):
                    formula = fc.value
            except KeyError:
                pass
        else:
            if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                formula = formula_cell.value

        return {
            "sheet": sheet,
            "address": cell.coordinate,
            "row": cell.row,
            "col": cell.column,
            "col_letter": get_column_letter(cell.column),
            "value": value,
            "formula": formula,
            "data_type": cell.data_type,
        }

    @staticmethod
    def _serialize(v: Any) -> Any:
        """Convert cell values to JSON-safe forms."""
        if v is None:
            return None
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float, str)):
            return v
        if isinstance(v, (datetime, date, time)):
            return v.isoformat()
        return str(v)


# ── Helpers for tool-use loop integration ──────────────────────────────

def address_range(start_addr: str, end_addr: str) -> str:
    """Compose a range string from two cell addresses (validates them)."""
    coordinate_from_string(start_addr)
    coordinate_from_string(end_addr)
    return f"{start_addr}:{end_addr}"


def row_range(start_addr: str, end_col_letter: str) -> str:
    """Build a horizontal range on a single row from a start address to an end column.

    Example: row_range("C12", "I") -> "C12:I12"
    """
    col, row = coordinate_from_string(start_addr)
    column_index_from_string(end_col_letter)  # validates
    return f"{start_addr}:{end_col_letter}{row}"

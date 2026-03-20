"""
Data loader for Carta financing profiles, NREL ATB cost/deployment projections,
and Lazard LCOE/LCOS benchmarks. Produces queryable parameter store dicts.
"""

import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "sources"


def _val(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "n/a", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return s


def load_carta_rounds() -> dict:
    """
    Returns {sector: {stage: {metric: value}}} from Carta Feb26 rounds sheet.
    Metrics include percentile distributions for round size, pre-money, post-money,
    graduation rates, months to graduation, ESOP, and sample sizes.
    """
    path = DATA_DIR / "Carta Insights_Fund Forecasting Profiles.xlsx"
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet = wb["Feb26 rounds"]
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    header = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = {}

    for row in rows[1:]:
        vals = {header[i]: _val(row[i]) for i in range(len(header)) if i < len(row)}
        sector = vals.get("SECTOR")
        stage = vals.get("SERIES")
        if not sector or not stage:
            continue
        result.setdefault(sector, {})[stage] = {
            "round_size": {
                "p10": vals.get("P10_ROUND_SIZE"),
                "p25": vals.get("P25_ROUND_SIZE"),
                "p50": vals.get("MEDIAN_ROUND_SIZE"),
                "p75": vals.get("P75_ROUND_SIZE"),
                "p90": vals.get("P90_ROUND_SIZE"),
            },
            "pre_money": {
                "p10": vals.get("P10_PREMONEY_VAL"),
                "p25": vals.get("P25_PREMONEY_VAL"),
                "p50": vals.get("MEDIAN_PREMONEY_VAL"),
                "p75": vals.get("P75_PREMONEY_VAL"),
                "p90": vals.get("P90_PREMONEY_VAL"),
            },
            "post_money": {
                "p10": vals.get("P10_POSTMONEY_VAL"),
                "p25": vals.get("P25_POSTMONEY_VAL"),
                "p50": vals.get("MEDIAN_POSTMONEY_VAL"),
                "p75": vals.get("P75_POSTMONEY_VAL"),
                "p90": vals.get("P90_POSTMONEY_VAL"),
            },
            "sample_size_rounds": vals.get("SAMPLE_SIZE_A"),
            "median_esop": vals.get("MEDIAN_ESOP_SIZE"),
            "sample_size_esop": vals.get("SAMPLE_SIZE_B"),
            "graduation_rate": vals.get("GRAD_PERCENT"),
            "median_months_to_grad": vals.get("MEDIAN_MONTHS_TO_GRAD"),
            "sample_size_grad": vals.get("SAMPLE_SIZE_C"),
        }

    return result


def load_carta_funds() -> list:
    """
    Returns list of fund performance records with IRR/TVPI percentiles
    by vintage year, AUM cohort, and time since vintage.
    """
    path = DATA_DIR / "Carta Insights_Fund Forecasting Profiles.xlsx"
    if not path.exists():
        return []

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet = wb["Feb26 funds"]
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    header = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []

    for row in rows[1:]:
        vals = {header[i]: _val(row[i]) for i in range(min(len(header), len(row)))}
        if vals.get("FUND_VINTAGE_YEAR") is None:
            continue
        result.append({
            "vintage": int(vals["FUND_VINTAGE_YEAR"]),
            "aum_cohort": vals.get("FUND_AUM_COHORT"),
            "years_since_vintage": vals.get("YEARS_SINCE_VINTAGE"),
            "irr": {
                "p10": vals.get("P10_IRR"),
                "p25": vals.get("P25_IRR"),
                "p50": vals.get("MEDIAN_IRR"),
                "p75": vals.get("P75_IRR"),
                "p90": vals.get("P90_IRR"),
            },
            "tvpi": {
                "p10": vals.get("P10_TVPI"),
                "p25": vals.get("P25_TVPI"),
                "p50": vals.get("MEDIAN_TVPI"),
                "p75": vals.get("P75_TVPI"),
                "p90": vals.get("P90_TVPI"),
            },
            "sample_size": vals.get("SAMPLE_SIZE"),
        })

    return result


def load_atb_lcoe_summary() -> dict:
    """
    Returns {technology: {cost_case: {year: lcoe_value}}} from ATB Summary_LCOE.
    Only loads a representative class per technology for the default view.
    """
    path = DATA_DIR / "Annual Tech Baseline 2024_v3_Workbook.xlsx"
    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet = wb["Summary_LCOE"]
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    years = [int(float(str(h))) for h in header[7:] if h is not None]
    result = {}

    representative_classes = {
        "UtilityPV": "Class5",
        "CommPV": "Class5",
        "ResPV": "Class5",
        "LandbasedWind": "Class4",
        "OffShoreWind": "Class3",
        "Geothermal": "Hydro",
        "Hydropower": "NPD1",
        "Nuclear": None,
        "Biopower": "Dedicated",
        "CSP": "Class2",
        "Utility-Scale PV-Plus-Battery": "Class5",
        "DistributedWind": "Midsize",
    }

    for row in rows[1:]:
        if row[4] is None:
            continue
        tech = str(row[4]).strip()
        detail = str(row[5]).strip() if row[5] else ""
        cost_case = str(row[2]).strip() if row[2] else ""
        display_name = str(row[6]).strip() if row[6] else ""

        if tech not in representative_classes:
            continue

        rep_class = representative_classes[tech]
        if rep_class and rep_class not in detail:
            continue

        year_values = {}
        for i, yr in enumerate(years):
            idx = 7 + i
            if idx < len(row) and row[idx] is not None:
                try:
                    year_values[yr] = float(row[idx])
                except (ValueError, TypeError):
                    pass

        if not year_values:
            continue

        result.setdefault(tech, {}).setdefault(cost_case, {
            "display_name": display_name,
            "detail": detail,
            "values": {},
        })["values"].update(year_values)

    return result


def load_lazard_lcoe() -> dict:
    """
    Returns current Lazard LCOE ranges by technology (from the v18.0 headline numbers).
    Hand-coded from the PDF extraction since it's a static snapshot.
    """
    return {
        "Solar PV - Utility": {"low": 38, "high": 78, "unit": "$/MWh"},
        "Solar PV - Community & C&I": {"low": 81, "high": 217, "unit": "$/MWh"},
        "Solar PV + Storage - Utility": {"low": 50, "high": 131, "unit": "$/MWh"},
        "Geothermal": {"low": 66, "high": 109, "unit": "$/MWh"},
        "Wind - Onshore": {"low": 37, "high": 86, "unit": "$/MWh"},
        "Wind + Storage - Onshore": {"low": 44, "high": 123, "unit": "$/MWh"},
        "Wind - Offshore": {"low": 70, "high": 157, "unit": "$/MWh"},
        "Gas Peaking": {"low": 149, "high": 251, "unit": "$/MWh"},
        "Nuclear": {"low": 141, "high": 220, "unit": "$/MWh"},
        "Coal": {"low": 71, "high": 173, "unit": "$/MWh"},
        "Gas Combined Cycle": {"low": 48, "high": 109, "unit": "$/MWh"},
    }


def load_lazard_lcos() -> dict:
    """Lazard LCOS v10.0 ranges."""
    return {
        "Utility-Scale Standalone (2hr)": {"low": 129, "high": 277, "unit": "$/MWh"},
        "Utility-Scale Standalone (4hr)": {"low": 115, "high": 254, "unit": "$/MWh"},
        "C&I Standalone (2hr)": {"low": 319, "high": 506, "unit": "$/MWh"},
        "Residential Standalone (4hr)": {"low": 547, "high": 860, "unit": "$/MWh"},
    }


TECHNOLOGY_ARCHETYPE_MAP = {
    "utility_solar": {
        "atb_tech": "UtilityPV",
        "lazard_key": "Solar PV - Utility",
        "description": "Utility-scale solar PV",
        "typical_capacity_factor": (0.20, 0.30),
        "typical_capex_kw": (1150, 1600),
    },
    "commercial_solar": {
        "atb_tech": "CommPV",
        "lazard_key": "Solar PV - Community & C&I",
        "description": "Commercial / C&I distributed solar",
        "typical_capacity_factor": (0.15, 0.20),
        "typical_capex_kw": (1600, 3300),
    },
    "residential_solar": {
        "atb_tech": "ResPV",
        "lazard_key": "Solar PV - Community & C&I",
        "description": "Residential rooftop solar",
        "typical_capacity_factor": (0.14, 0.20),
        "typical_capex_kw": (2600, 4200),
    },
    "onshore_wind": {
        "atb_tech": "LandbasedWind",
        "lazard_key": "Wind - Onshore",
        "description": "Land-based onshore wind",
        "typical_capacity_factor": (0.30, 0.55),
        "typical_capex_kw": (1900, 2300),
    },
    "offshore_wind": {
        "atb_tech": "OffShoreWind",
        "lazard_key": "Wind - Offshore",
        "description": "Fixed-bottom offshore wind",
        "typical_capacity_factor": (0.45, 0.55),
        "typical_capex_kw": (3450, 6550),
    },
    "geothermal": {
        "atb_tech": "Geothermal",
        "lazard_key": "Geothermal",
        "description": "Geothermal (hydrothermal + EGS)",
        "typical_capacity_factor": (0.80, 0.90),
        "typical_capex_kw": (5000, 6460),
    },
    "battery_storage_utility": {
        "atb_tech": "Utility-Scale PV-Plus-Battery",
        "lazard_key": "Solar PV + Storage - Utility",
        "description": "Utility-scale battery storage / solar+storage",
        "typical_capacity_factor": None,
        "typical_capex_kw": None,
    },
    "nuclear_smr": {
        "atb_tech": "Nuclear",
        "lazard_key": "Nuclear",
        "description": "Nuclear / SMR",
        "typical_capacity_factor": (0.89, 0.92),
        "typical_capex_kw": (9020, 14820),
    },
    "ev_electrification": {
        "atb_tech": None,
        "lazard_key": None,
        "description": "EV / fleet electrification / charging infrastructure",
        "typical_capacity_factor": None,
        "typical_capex_kw": None,
    },
    "climate_software": {
        "atb_tech": None,
        "lazard_key": None,
        "description": "Climate SaaS / carbon accounting / grid software",
        "typical_capacity_factor": None,
        "typical_capex_kw": None,
    },
    "industrial_decarb": {
        "atb_tech": None,
        "lazard_key": None,
        "description": "Industrial decarbonization / process electrification",
        "typical_capacity_factor": None,
        "typical_capex_kw": None,
    },
}


def load_all() -> dict:
    """Load all data sources into a unified parameter store."""
    return {
        "carta_rounds": load_carta_rounds(),
        "carta_funds": load_carta_funds(),
        "atb_lcoe": load_atb_lcoe_summary(),
        "lazard_lcoe": load_lazard_lcoe(),
        "lazard_lcos": load_lazard_lcos(),
        "archetypes": TECHNOLOGY_ARCHETYPE_MAP,
    }

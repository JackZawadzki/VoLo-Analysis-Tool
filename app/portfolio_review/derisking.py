"""
Derisking Quartile scoring — ports the logic from the Fund I 2025 / Fund II
2025 tabs of the "Fund I Rankings (Derisking Quadrants)" workbook.

Scoring rubric (per the original sheet):
  - Each company is scored on 7 derisking dimensions, each +1 or -1
    (or 0 for partial / not yet known)
      1. Rapid innovation and adoption
      2. Business model
      3. Technology
      4. Incentive management
      5. Team
      6. Product and growth
      7. IP and Data
  - Plus an exit indicator: 0 if exited (zeros out the total), else 1
  - TOTAL = SUM of the 7 dimension scores (range: -7 to +7)
  - Quartile cutoffs:
        TOTAL >=  5  → Q4 (most derisked)
        TOTAL >=  3  → Q3
        TOTAL >=  1  → Q2
        TOTAL <   1  → Q1 (highest residual risk)

The 7 dimensions also map to slide 2 of the All-Team PortCo Updates deck
(Company Health Index — qualitative risk dimensions), so a single scoring
record powers both the deck-style report and the quartile ranking.
"""

from __future__ import annotations

import logging
import re
import sqlite3
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ── Dimension metadata ────────────────────────────────────────────────────────
DIMENSIONS = [
    ("rapid_innovation_adopt", "Rapid innovation and adoption"),
    ("business_model",         "Business model"),
    ("technology",             "Technology"),
    ("incentive_management",   "Incentive management"),
    ("team",                   "Team"),
    ("product_growth",         "Product and growth"),
    ("ip_and_data",            "IP and Data"),
]
DIMENSION_KEYS = [k for k, _ in DIMENSIONS]


# ── Quartile calculator ───────────────────────────────────────────────────────
def compute_quartile(total: float) -> int:
    """Map a derisking total score (-7..+7) to a quartile (1..4).
    Matches the IFS thresholds from the source workbook."""
    if total >= 5:
        return 4
    if total >= 3:
        return 3
    if total >= 1:
        return 2
    return 1


def compute_total(scores: dict[str, Any], is_exited: bool = False) -> float:
    """Sum the 7 dimension scores (treating None as 0). If exited, the
    sheet zeroes out the total via the 'Q4 0 if Exit' multiplier — we mirror
    that by returning 0 when is_exited."""
    if is_exited:
        return 0.0
    total = 0.0
    for k in DIMENSION_KEYS:
        v = scores.get(k)
        if v is None or v == "":
            continue
        try:
            total += float(v)
        except (TypeError, ValueError):
            continue
    return total


def score_company(scores: dict[str, Any], is_exited: bool = False) -> dict:
    """Compute total + quartile from a dict of dimension → +1/0/-1 values."""
    total = compute_total(scores, is_exited)
    q = compute_quartile(total) if not is_exited else 0
    return {"total_score": total, "quartile": q, "is_exited": int(is_exited)}


# ── Excel import ──────────────────────────────────────────────────────────────
# Column indices in the source sheet (1-indexed, headers on row 4, data row 5+).
SHEET_COLS = {
    "name":                   1,
    "rapid_innovation_adopt": 2,
    "business_model":         3,
    "technology":             4,
    "incentive_management":   5,
    "team":                   6,
    "product_growth":         7,
    "ip_and_data":            8,
    "exit_indicator":         9,   # 0 if exited, 1 otherwise
}


def _norm_score(v: Any) -> Optional[float]:
    """Coerce a cell value to -1/0/+1 or None."""
    if v is None or v == "":
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    # Clamp to -1, 0, 1 (the source uses only these)
    if f >= 1:
        return 1.0
    if f <= -1:
        return -1.0
    return 0.0


def _normalize_name(name: str) -> str:
    s = (name or "").lower().strip()
    s = re.sub(r"\b(inc|incorporated|corp|corporation|ltd|llc|co|company|technologies|public benefit corporation|pbc)\.?\b", "", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _match_company(conn: sqlite3.Connection, deck_name: str) -> Optional[int]:
    norm = _normalize_name(deck_name)
    rows = conn.execute("SELECT id, name FROM pr_companies").fetchall()
    by_norm = {_normalize_name(r["name"]): r["id"] for r in rows}
    if norm in by_norm:
        return by_norm[norm]
    for cn, cid in by_norm.items():
        if not cn or not norm:
            continue
        if cn.startswith(norm) or norm.startswith(cn) or norm in cn or cn in norm:
            if min(len(cn), len(norm)) >= 3:
                return cid
    return None


def import_derisking_sheet(workbook_path: Path, conn: sqlite3.Connection,
                            sheet_name: str = "Fund I 2025",
                            period: Optional[str] = None,
                            fund: Optional[str] = None) -> dict:
    """Read a Fund X YYYY tab from the Derisking Quadrants workbook and
    persist one pr_derisking_scores row per company."""
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Default period from sheet name (e.g., 'Fund I 2025' → 'FY2025')
    if period is None:
        m = re.search(r"(\d{4})", sheet_name)
        period = f"FY{m.group(1)}" if m else "current"
    if fund is None:
        m = re.search(r"(Fund\s*I+|Fund\s*II)", sheet_name, re.I)
        fund = (m.group(1).replace(" ", " ").title() if m else "Fund I")

    counts = {"matched": 0, "unmatched": 0, "scored": 0}
    unmatched: list[str] = []

    # Data rows start at row 5; stop when name is empty for several rows.
    blank_streak = 0
    for r in range(5, ws.max_row + 1):
        name_cell = ws.cell(row=r, column=SHEET_COLS["name"]).value
        name = (str(name_cell).strip() if name_cell else "")
        if not name:
            blank_streak += 1
            if blank_streak > 3:
                break
            continue
        # Stop on threshold/legend rows
        if name.lower().startswith(("quartile", "threshold", "column", "key", "score")):
            break
        blank_streak = 0

        # Score the row
        scores: dict[str, Any] = {}
        for key in DIMENSION_KEYS:
            scores[key] = _norm_score(ws.cell(row=r, column=SHEET_COLS[key]).value)
        exit_val = _norm_score(ws.cell(row=r, column=SHEET_COLS["exit_indicator"]).value)
        is_exited = (exit_val == 0)

        result = score_company(scores, is_exited=is_exited)

        cid = _match_company(conn, name)
        if not cid:
            unmatched.append(name)
            counts["unmatched"] += 1
            continue
        counts["matched"] += 1

        # UPSERT on (company_id, period)
        conn.execute(
            """INSERT INTO pr_derisking_scores
            (company_id, period, fund, rapid_innovation_adopt, business_model,
             technology, incentive_management, team, product_growth, ip_and_data,
             is_exited, total_score, quartile)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(company_id, period) DO UPDATE SET
              fund=excluded.fund,
              rapid_innovation_adopt=excluded.rapid_innovation_adopt,
              business_model=excluded.business_model,
              technology=excluded.technology,
              incentive_management=excluded.incentive_management,
              team=excluded.team,
              product_growth=excluded.product_growth,
              ip_and_data=excluded.ip_and_data,
              is_exited=excluded.is_exited,
              total_score=excluded.total_score,
              quartile=excluded.quartile,
              scored_at=datetime('now')""",
            (
                cid, period, fund,
                scores["rapid_innovation_adopt"],
                scores["business_model"],
                scores["technology"],
                scores["incentive_management"],
                scores["team"],
                scores["product_growth"],
                scores["ip_and_data"],
                result["is_exited"],
                result["total_score"],
                result["quartile"],
            ),
        )
        counts["scored"] += 1
    conn.commit()
    counts["unmatched_names"] = unmatched
    counts["sheet_name"] = sheet_name
    counts["period"] = period
    counts["fund"] = fund
    return counts


def import_full_workbook(workbook_path: Path, conn: sqlite3.Connection) -> dict:
    """Run the importer for both Fund I and Fund II tabs (most recent year)."""
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, data_only=True)

    results = []
    for sn in wb.sheetnames:
        # Match "Fund I 2025", "Fund II 2025" etc.
        if re.match(r"Fund\s*I+\s*\d{4}$", sn):
            try:
                results.append(import_derisking_sheet(workbook_path, conn, sn))
            except Exception as e:
                logger.exception(f"Failed to import {sn}")
                results.append({"sheet_name": sn, "error": str(e)})
    return {"results": results, "sheets_processed": len(results)}

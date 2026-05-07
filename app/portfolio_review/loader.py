"""
Excel → DB loader for portfolio_review.

Pins to specific column indices for each sheet (much more robust than
fuzzy header matching against truncated/parenthesized header text).

Currently implemented:
  - Companies + investments + per-company metadata (PortCo Data - Fund I)
  - Fund II investments (Fund II MSOI)
  - Returns / IRR snapshots (IRR sheet)
  - Board seats (derived from PortCo Data Fund I)
  - Contact info (Contact Info sheet)
  - Per-year revenue financials (Financials Basic)

Run from CLI:
    python -m app.portfolio_review.loader path/to/Portfolio\\ Review.xlsx
"""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────────────────
def _norm_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _norm_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "").replace("x", "")
    if s in ("", "-", "N/A", "DNP", "N/A - DNP"):
        return None
    try:
        if s.endswith("%"):
            return float(s[:-1]) / 100.0
        return float(s)
    except ValueError:
        return None


def _norm_date(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if not s or s.upper() in ("N/A", "TBD"):
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _norm_bool(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("yes", "y", "true", "1", "x")


def _cell(ws, row: int, col: int):
    """Read a single cell value at 1-indexed (row, col)."""
    return ws.cell(row=row, column=col).value


# ── Company UPSERT ────────────────────────────────────────────────────────────
def _upsert_company(conn: sqlite3.Connection, name: str, **fields) -> int:
    """Insert or update by company name; UPDATE only fills empty fields so we
    don't clobber data set by a prior sheet's import."""
    name = name.strip()
    if not name:
        raise ValueError("Company name cannot be empty")
    cur = conn.execute("SELECT * FROM pr_companies WHERE name = ?", (name,))
    row = cur.fetchone()
    updatable = [
        "fund", "brief_description", "sector", "submarket", "business_model",
        "hw_sw", "commercial_status", "ceo_name", "ceo_email", "cfo_name",
        "cfo_email", "address", "website", "fume_date", "first_year_revenue",
        "hyperscale", "notable_partners", "next_round_expect",
    ]
    if row:
        company_id = row["id"]
        # Fill only fields that are currently empty (so subsequent sheets can
        # add missing data without overwriting earlier data).
        sets, vals = [], []
        for k, v in fields.items():
            if k not in updatable or v in (None, "", 0) and k != "hyperscale":
                continue
            existing_value = row[k] if k in row.keys() else None
            if existing_value not in (None, "", 0):
                continue
            sets.append(f"{k} = ?")
            vals.append(int(v) if k == "hyperscale" else v)
        if sets:
            sets.append("updated_at = datetime('now')")
            vals.append(company_id)
            conn.execute(f"UPDATE pr_companies SET {', '.join(sets)} WHERE id = ?", vals)
        return company_id
    # Insert
    cols = ["name"] + list(fields.keys())
    vals: list[Any] = [name] + [
        int(v) if k == "hyperscale" else v for k, v in fields.items()
    ]
    placeholders = ", ".join("?" * len(cols))
    cur = conn.execute(
        f"INSERT INTO pr_companies ({', '.join(cols)}) VALUES ({placeholders})",
        vals,
    )
    return cur.lastrowid


# ── PortCo Data - Fund I ──────────────────────────────────────────────────────
# Header row is row 3. Columns we care about (1-indexed):
PORTCO_FUND_I_COLS = {
    "participated":     1,    # Y/N
    "oc":               2,    # O / C / F (Original / Conversion / Follow-on)
    "name":             3,
    "investment_date":  5,
    "investment_amount":6,
    "notes":            7,
    "no_part_reason":   8,
    "board_seat":       9,
    "board_member":     10,
    "investment_type":  11,
    "pre_money":        17,
    "round_size":       18,
    "latest_share_px":  19,
    "valuation":        20,
    "mark_carta":       22,
    "irr_carta":        23,
    "deal_source":      24,
    "deal_lead":        26,
    "co_investors":     27,
    "sector":           28,
    "submarket":        29,
    "business_model":   30,
    "location":         31,
    "brief_desc":       32,
    "address":          35,
    "cfo_email":        38,
    "ceo_email":        39,
    "stage":            40,
}


def import_portco_fund_i(conn: sqlite3.Connection, ws) -> dict:
    counts = {"companies_inserted": 0, "investments_inserted": 0,
              "company_meta_updated": 0, "rows_skipped": 0}
    company_metadata: dict[str, dict] = {}

    # Pass 1: collect company-level metadata (first non-empty value across rows)
    for r in range(4, ws.max_row + 1):
        name = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["name"]))
        if not name or name.lower().startswith(("do not delete", "color (1=on)")):
            continue
        meta = company_metadata.setdefault(name, {})
        for field, col in PORTCO_FUND_I_COLS.items():
            if field in ("name", "investment_date", "investment_amount", "oc",
                         "notes", "participated", "no_part_reason"):
                continue  # these are per-investment, not per-company
            v = _cell(ws, r, col)
            if v in (None, "") or field in meta:
                continue
            meta[field] = v

    # Pass 2: upsert companies + insert investments
    for r in range(4, ws.max_row + 1):
        name = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["name"]))
        if not name or name.lower().startswith(("do not delete", "color (1=on)")):
            counts["rows_skipped"] += 1
            continue

        meta = company_metadata.get(name, {})
        company_existed = conn.execute(
            "SELECT 1 FROM pr_companies WHERE name=?", (name,)
        ).fetchone() is not None

        company_id = _upsert_company(
            conn, name,
            fund="Fund I",
            brief_description=_norm_str(meta.get("brief_desc")),
            sector=_norm_str(meta.get("sector")),
            submarket=_norm_str(meta.get("submarket")),
            business_model=_norm_str(meta.get("business_model")),
            address=_norm_str(meta.get("address")),
            ceo_email=_norm_str(meta.get("ceo_email")),
            cfo_email=_norm_str(meta.get("cfo_email")),
        )
        if not company_existed:
            counts["companies_inserted"] += 1
        else:
            counts["company_meta_updated"] += 1

        # Now the investment row itself
        inv_date = _norm_date(_cell(ws, r, PORTCO_FUND_I_COLS["investment_date"]))
        amount = _norm_float(_cell(ws, r, PORTCO_FUND_I_COLS["investment_amount"]))
        oc = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["oc"]))[:1].upper() or "O"
        notes = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["notes"]))
        no_part_reason = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["no_part_reason"]))
        participated = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["participated"])).upper().startswith("Y")
        round_label = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["investment_type"]))
        pre_money = _norm_float(_cell(ws, r, PORTCO_FUND_I_COLS["pre_money"]))
        round_size = _norm_float(_cell(ws, r, PORTCO_FUND_I_COLS["round_size"]))
        valuation = _norm_float(_cell(ws, r, PORTCO_FUND_I_COLS["valuation"]))
        deal_lead = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["deal_lead"]))
        board_seat = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["board_seat"])) or "No"
        board_member = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["board_member"]))

        # Skip rows that have no real investment payload (just a name with everything empty)
        if inv_date is None and amount is None and not round_label:
            continue

        # Dedupe — same (company, date, amount) is the same investment
        existing = conn.execute(
            """SELECT id FROM pr_investments
               WHERE company_id=? AND COALESCE(investment_date,'')=COALESCE(?,'')
                 AND COALESCE(investment_amount,-1)=COALESCE(?,-1)""",
            (company_id, inv_date, amount),
        ).fetchone()
        if existing:
            continue

        full_notes = notes
        if no_part_reason:
            full_notes = (full_notes + " — " + no_part_reason).strip(" —")

        conn.execute(
            """INSERT INTO pr_investments
            (company_id, investment_date, original_or_conversion, investment_amount,
             round_label, round_size, pre_money, post_money, deal_lead,
             board_seat, board_member, notes, participated)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (company_id, inv_date, oc, amount, round_label, round_size,
             pre_money, valuation, deal_lead, board_seat, board_member,
             full_notes, 1 if participated else 0),
        )
        counts["investments_inserted"] += 1

    return counts


# ── Fund II MSOI ──────────────────────────────────────────────────────────────
# Header on row 1. Columns:
FUND_II_COLS = {
    "name":             1,
    "ceo_email":        2,
    "cfo_email":        3,
    "investment_date":  4,
    "investment_amount":5,
    "notes":            6,
    "board_seat":       7,
    "board_member":     8,
}


def import_portco_fund_ii(conn: sqlite3.Connection, ws) -> dict:
    counts = {"companies_inserted": 0, "investments_inserted": 0}
    for r in range(2, ws.max_row + 1):
        name = _norm_str(_cell(ws, r, FUND_II_COLS["name"]))
        if not name:
            continue
        existed = conn.execute("SELECT 1 FROM pr_companies WHERE name=?", (name,)).fetchone()
        company_id = _upsert_company(
            conn, name, fund="Fund II",
            ceo_email=_norm_str(_cell(ws, r, FUND_II_COLS["ceo_email"])),
            cfo_email=_norm_str(_cell(ws, r, FUND_II_COLS["cfo_email"])),
        )
        if not existed:
            counts["companies_inserted"] += 1

        inv_date = _norm_date(_cell(ws, r, FUND_II_COLS["investment_date"]))
        amount = _norm_float(_cell(ws, r, FUND_II_COLS["investment_amount"]))
        if inv_date is None and amount is None:
            continue
        notes = _norm_str(_cell(ws, r, FUND_II_COLS["notes"]))
        board_seat = _norm_str(_cell(ws, r, FUND_II_COLS["board_seat"])) or "No"
        board_member = _norm_str(_cell(ws, r, FUND_II_COLS["board_member"]))

        existing = conn.execute(
            """SELECT id FROM pr_investments WHERE company_id=?
                 AND COALESCE(investment_date,'')=COALESCE(?,'')
                 AND COALESCE(investment_amount,-1)=COALESCE(?,-1)""",
            (company_id, inv_date, amount),
        ).fetchone()
        if existing:
            continue
        conn.execute(
            """INSERT INTO pr_investments
            (company_id, investment_date, investment_amount, board_seat, board_member, notes)
            VALUES (?, ?, ?, ?, ?, ?)""",
            (company_id, inv_date, amount, board_seat, board_member, notes),
        )
        counts["investments_inserted"] += 1
    return counts


# ── IRR sheet ─────────────────────────────────────────────────────────────────
# Header row 1. Columns:
IRR_COLS = {
    "company": 1, "cost": 2, "proceeds": 3, "interest": 4,
    "fmv": 5, "total_value": 6, "gain_loss": 7, "multiple": 8, "irr": 9,
}


def import_irr_sheet(conn: sqlite3.Connection, ws, as_of_date: str) -> int:
    added = 0
    for r in range(2, ws.max_row + 1):
        name = _norm_str(_cell(ws, r, IRR_COLS["company"]))
        if not name or name.lower() in ("total", "totals", "fund total"):
            continue
        company_id = _upsert_company(conn, name)
        cost = _norm_float(_cell(ws, r, IRR_COLS["cost"]))
        proceeds = _norm_float(_cell(ws, r, IRR_COLS["proceeds"]))
        interest = _norm_float(_cell(ws, r, IRR_COLS["interest"]))
        fmv = _norm_float(_cell(ws, r, IRR_COLS["fmv"]))
        total_value = _norm_float(_cell(ws, r, IRR_COLS["total_value"]))
        gain_loss = _norm_float(_cell(ws, r, IRR_COLS["gain_loss"]))
        multiple = _norm_float(_cell(ws, r, IRR_COLS["multiple"]))
        irr = _norm_float(_cell(ws, r, IRR_COLS["irr"]))
        if all(x is None for x in (cost, fmv, total_value, multiple)):
            continue
        conn.execute(
            """INSERT INTO pr_returns
            (company_id, as_of_date, cost, proceeds, interest, fmv, total_value, gain_loss, multiple, irr)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(company_id, as_of_date) DO UPDATE SET
              cost=excluded.cost, proceeds=excluded.proceeds, interest=excluded.interest,
              fmv=excluded.fmv, total_value=excluded.total_value, gain_loss=excluded.gain_loss,
              multiple=excluded.multiple, irr=excluded.irr""",
            (company_id, as_of_date, cost, proceeds, interest, fmv, total_value, gain_loss, multiple, irr),
        )
        added += 1
    return added


# ── Board seats from PortCo Data Fund I ───────────────────────────────────────
def import_board_seats(conn: sqlite3.Connection, ws) -> int:
    seen: set[tuple[int, str, str]] = set()
    added = 0
    for r in range(4, ws.max_row + 1):
        name = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["name"]))
        if not name or name.lower().startswith(("do not delete",)):
            continue
        seat = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["board_seat"]))
        member = _norm_str(_cell(ws, r, PORTCO_FUND_I_COLS["board_member"]))
        if seat.lower() in ("", "no", "n/a") or not member or member.upper() == "N/A":
            continue
        seat_type = "Observer" if "observer" in seat.lower() else "Director"
        company_id = _upsert_company(conn, name)
        key = (company_id, seat_type, member)
        if key in seen:
            continue
        seen.add(key)
        existing = conn.execute(
            "SELECT 1 FROM pr_board_seats WHERE company_id=? AND seat_type=? AND board_member=?",
            (company_id, seat_type, member),
        ).fetchone()
        if existing:
            continue
        conn.execute(
            "INSERT INTO pr_board_seats (company_id, seat_type, board_member, active) VALUES (?, ?, ?, 1)",
            (company_id, seat_type, member),
        )
        added += 1
    return added


# ── Contact Info ──────────────────────────────────────────────────────────────
# Header row 6. Columns:
CONTACT_COLS = {
    "name":         2,
    "ceo":          3,
    "ceo_email":    4,
    "looped_in":    6,
    "address":      8,
}


def import_contact_info(conn: sqlite3.Connection, ws) -> int:
    updated = 0
    for r in range(7, ws.max_row + 1):
        name = _norm_str(_cell(ws, r, CONTACT_COLS["name"]))
        if not name or name.lower() in ("color (1=on)", "do not delete"):
            continue
        ceo_name = _norm_str(_cell(ws, r, CONTACT_COLS["ceo"]))
        ceo_email = _norm_str(_cell(ws, r, CONTACT_COLS["ceo_email"]))
        address = _norm_str(_cell(ws, r, CONTACT_COLS["address"]))
        if not (ceo_name or ceo_email or address):
            continue
        # _upsert_company only fills empty fields
        _upsert_company(
            conn, name,
            ceo_name=ceo_name,
            ceo_email=ceo_email,
            address=address,
        )
        updated += 1
    return updated


# ── PortCo Desc by Industry — backfills sector/submarket/business_model ───────
PORTCO_DESC_COLS = {
    "sector":         2,
    "name":           4,
    "brief_desc":     5,
    "submarket":      6,
    "business_model": 7,
    "board_seat":     9,
}


def import_portco_desc(conn: sqlite3.Connection, ws) -> int:
    updated = 0
    for r in range(7, ws.max_row + 1):
        name = _norm_str(_cell(ws, r, PORTCO_DESC_COLS["name"]))
        if not name or name.lower() in ("company name", "do not delete"):
            continue
        _upsert_company(
            conn, name,
            sector=_norm_str(_cell(ws, r, PORTCO_DESC_COLS["sector"])),
            brief_description=_norm_str(_cell(ws, r, PORTCO_DESC_COLS["brief_desc"])),
            submarket=_norm_str(_cell(ws, r, PORTCO_DESC_COLS["submarket"])),
            business_model=_norm_str(_cell(ws, r, PORTCO_DESC_COLS["business_model"])),
        )
        updated += 1
    return updated


# ── Financials Basic — per-year revenue ───────────────────────────────────────
# Header row 6. Columns:
#   col 3 = company name
#   cols 12..16 = FY2020..FY2024 revenue
FIN_BASIC_NAME_COL = 3
FIN_BASIC_YEAR_COLS = {"FY2020": 12, "FY2021": 13, "FY2022": 14, "FY2023": 15, "FY2024": 16}


def import_financials_basic(conn: sqlite3.Connection, ws) -> int:
    added = 0
    for r in range(7, ws.max_row + 1):
        name = _norm_str(_cell(ws, r, FIN_BASIC_NAME_COL))
        if not name or name.lower() in ("portfolio company name", "do not delete", "color"):
            continue
        company_existed = conn.execute(
            "SELECT id FROM pr_companies WHERE name=?", (name,)
        ).fetchone()
        if not company_existed:
            continue
        company_id = company_existed[0]
        for period, col in FIN_BASIC_YEAR_COLS.items():
            rev = _norm_float(_cell(ws, r, col))
            if rev is None:
                continue
            conn.execute(
                """INSERT INTO pr_financials (company_id, period, revenue)
                VALUES (?, ?, ?)
                ON CONFLICT(company_id, period) DO UPDATE SET revenue = excluded.revenue""",
                (company_id, period, rev),
            )
            added += 1
    return added


# ── Public entry point ───────────────────────────────────────────────────────
def run_import(workbook_path: str | Path, conn: sqlite3.Connection,
               user_id: Optional[int] = None,
               as_of_date: Optional[str] = None) -> dict:
    import warnings
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    from openpyxl import load_workbook

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    as_of_date = as_of_date or datetime.utcnow().date().isoformat()

    cur = conn.execute(
        "INSERT INTO pr_imports (user_id, source_file, status) VALUES (?, ?, 'running')",
        (user_id, str(workbook_path)),
    )
    import_id = cur.lastrowid

    counts = {
        "companies": 0, "investments": 0, "returns": 0, "board_seats": 0,
        "financials": 0, "contact_updates": 0, "metadata_updates": 0,
    }
    errors: list[str] = []

    try:
        wb = load_workbook(workbook_path, data_only=True)

        if "PortCo Data - Fund I" in wb.sheetnames:
            try:
                r = import_portco_fund_i(conn, wb["PortCo Data - Fund I"])
                counts["companies"] += r["companies_inserted"]
                counts["investments"] += r["investments_inserted"]
                counts["board_seats"] += import_board_seats(conn, wb["PortCo Data - Fund I"])
            except Exception as e:
                errors.append(f"Fund I: {e}")

        if "Fund II (MSOI)" in wb.sheetnames:
            try:
                r = import_portco_fund_ii(conn, wb["Fund II (MSOI)"])
                counts["companies"] += r["companies_inserted"]
                counts["investments"] += r["investments_inserted"]
            except Exception as e:
                errors.append(f"Fund II: {e}")

        if "PortCo Desc by Industry" in wb.sheetnames:
            try:
                counts["metadata_updates"] += import_portco_desc(conn, wb["PortCo Desc by Industry"])
            except Exception as e:
                errors.append(f"PortCo Desc: {e}")

        if "Contact Info" in wb.sheetnames:
            try:
                counts["contact_updates"] += import_contact_info(conn, wb["Contact Info"])
            except Exception as e:
                errors.append(f"Contact Info: {e}")

        if "IRR" in wb.sheetnames:
            try:
                counts["returns"] += import_irr_sheet(conn, wb["IRR"], as_of_date)
            except Exception as e:
                errors.append(f"IRR: {e}")

        if "Financials Basic" in wb.sheetnames:
            try:
                counts["financials"] += import_financials_basic(conn, wb["Financials Basic"])
            except Exception as e:
                errors.append(f"Financials Basic: {e}")

        conn.commit()
        status = "partial" if errors else "success"
    except Exception as e:
        conn.rollback()
        errors.append(f"FATAL: {e}")
        status = "failed"

    total_rows = sum(counts.values())
    conn.execute(
        """UPDATE pr_imports SET status=?, rows_imported=?, error_summary=?, finished_at=datetime('now')
           WHERE id=?""",
        (status, total_rows, " | ".join(errors), import_id),
    )
    conn.commit()

    return {
        "import_id": import_id,
        "status": status,
        "counts": counts,
        "errors": errors,
        "as_of_date": as_of_date,
    }


# ── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from app.database import get_db, init_db
    from app.portfolio_review.schema import apply_schema

    parser = argparse.ArgumentParser(description="Import a Portfolio Review workbook into the app DB.")
    parser.add_argument("workbook", help="Path to the .xlsx workbook")
    parser.add_argument("--as-of", default=None, help="As-of date for the snapshot (YYYY-MM-DD)")
    args = parser.parse_args()

    init_db()
    with get_db() as conn:
        apply_schema(conn)
        result = run_import(args.workbook, conn, as_of_date=args.as_of)
    import json
    print(json.dumps(result, indent=2))

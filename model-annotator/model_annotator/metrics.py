"""Conditional derived-metric library (pipeline Phase 4).

Metrics are selected by applicability — the tool computes what the model's
structure supports, not a fixed checklist. Every result carries its input
cells, a human-readable computation, and 2-3 checkpoint values for spot
verification. All arithmetic happens here, in Python, on numbers read from
the workbook; nothing is estimated.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from typing import Callable, Optional

from .ingest import WorkbookData, a1, is_number
from .mapping import LineItem, MappingResult
from .schema import DerivedMetric, Granularity, Orientation, SheetRole, format_ref
from .structure import StructureResult

log = logging.getLogger(__name__)

EPS = 1e-9
SMALL_BASE_FRACTION = 0.01     # growth on a base < 1% of terminal scale is pre-product noise


@dataclass
class MetricsResult:
    metrics: list[DerivedMetric] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def get(self, metric_id: str) -> Optional[DerivedMetric]:
        for m in self.metrics:
            if m.metric_id == metric_id:
                return m
        return None

    def series(self, metric_id: str) -> dict[str, Optional[float]]:
        m = self.get(metric_id)
        return dict(m.series) if m else {}

    def scalar(self, metric_id: str) -> Optional[float]:
        m = self.get(metric_id)
        return m.scalar if m else None


@dataclass
class Ctx:
    wbd: WorkbookData
    structure: StructureResult
    mapping: MappingResult
    benchmarks: dict
    sheet: str
    periods: list[str]
    months_per_period: float

    def item(self, cid: str) -> Optional[LineItem]:
        return self.mapping.get(cid, self.sheet)

    def items(self, cid: str) -> list[LineItem]:
        return [it for it in self.mapping.get_all(cid) if it.sheet == self.sheet] or \
               self.mapping.get_all(cid)

    def series(self, cid: str) -> Optional[dict[str, Optional[float]]]:
        it = self.item(cid)
        if it is None:
            return None
        return {p: it.value_for(p) for p in self.periods}

    def sum_series(self, cid: str) -> tuple[list[LineItem], Optional[dict[str, Optional[float]]]]:
        items = self.items(cid)
        if not items:
            return [], None
        items, _parents = resolve_hierarchy(items, self.periods)  # never sum a subtotal with its parts
        out: dict[str, Optional[float]] = {}
        for p in self.periods:
            vals = [it.value_for(p) for it in items]
            vals = [v for v in vals if v is not None]
            out[p] = sum(vals) if vals else None
        return items, out

    def refs(self, *cids: str) -> list[str]:
        out: list[str] = []
        for cid in cids:
            it = self.item(cid)
            if it:
                out.extend(r for r in it.refs if r)
        return out[:80]

    def units_label(self) -> str:
        u = self.structure.units.get(self.sheet)
        return u.label if u else "model units"


# ---------------------------------------------------------------------------
# Series arithmetic (None-safe throughout)
# ---------------------------------------------------------------------------

def growth(s: dict[str, Optional[float]], periods: list[str],
           small_base: Optional[float] = None,
           positive_only: bool = False) -> dict[str, Optional[float]]:
    out: dict[str, Optional[float]] = {}
    prev: Optional[float] = None
    for p in periods:
        v = s.get(p)
        if v is None or prev is None or abs(prev) < EPS:
            out[p] = None
        elif positive_only and (prev <= 0 or v <= 0):
            # a % growth off a non-positive base is meaningless (and the |prev|
            # denominator makes sign-flips read as spurious "growth"); suppress
            out[p] = None
        elif small_base is not None and abs(prev) < small_base:
            out[p] = None     # suppressed: base too small to mean anything
        else:
            out[p] = (v - prev) / abs(prev)
        if v is not None:
            prev = v
    return out


def ratio(a: dict[str, Optional[float]], b: dict[str, Optional[float]],
          periods: list[str]) -> dict[str, Optional[float]]:
    out: dict[str, Optional[float]] = {}
    for p in periods:
        x, y = a.get(p), b.get(p)
        out[p] = (x / y) if (x is not None and y is not None and abs(y) > EPS) else None
    return out


def ex_ratio(num: dict[str, Optional[float]], base: dict[str, Optional[float]],
             periods: list[str]) -> tuple[dict[str, Optional[float]], list[str]]:
    """Ratio whose denominator is a SUBTRACTION-derived base (an 'ex-X' / net
    figure, e.g. revenue minus grants). A period is UNDEFINED when that base is
    <= EPS: the excluded item meets or exceeds the whole, so the base is zero or
    negative and a ratio on it is meaningless. This also blocks the neg/neg
    sign-flip — (negative numerator) / (negative base) would otherwise render as
    a healthy POSITIVE margin. Returns (series, undefined_periods)."""
    out: dict[str, Optional[float]] = {}
    undef: list[str] = []
    for p in periods:
        n, d = num.get(p), base.get(p)
        if n is None or d is None:
            out[p] = None
        elif d <= EPS:                 # zero or NEGATIVE base -> undefined
            out[p] = None
            undef.append(p)
        else:
            out[p] = n / d
    return out, undef


def combine(periods: list[str], fn: Callable[..., Optional[float]],
            *series: dict[str, Optional[float]]) -> dict[str, Optional[float]]:
    out: dict[str, Optional[float]] = {}
    for p in periods:
        vals = [s.get(p) for s in series]
        out[p] = None if any(v is None for v in vals) else fn(*vals)
    return out


def cumulative(s: dict[str, Optional[float]], periods: list[str]) -> dict[str, Optional[float]]:
    out: dict[str, Optional[float]] = {}
    total = 0.0
    seen = False
    for p in periods:
        v = s.get(p)
        if v is not None:
            total += v
            seen = True
        out[p] = total if seen else None
    return out


def terminal_scale(s: dict[str, Optional[float]], periods: list[str]) -> float:
    vals = [abs(s[p]) for p in periods if s.get(p) is not None]
    return max(vals) if vals else 0.0


def resolve_hierarchy(items: list[LineItem], periods: list[str], tol: float = 0.02):
    """Detect subtotal rows and return (leaves, parents_dropped).

    A row is a subtotal if, in every period where it and >=2 of the other rows
    are populated, it equals the sum of the others (within ``tol``). Aggregating
    a category total must use ONE level only — never a subtotal together with its
    own components (which double-counts). Illustrative failure: a payroll calc that
    summed 'Total salaries' AND its three component salary rows = 2x payroll.
    """
    if len(items) < 2:
        return list(items), []
    arr = {id(it): [it.value_for(p) for p in periods] for it in items}
    parents: set[int] = set()
    for i, it in enumerate(items):
        others = [o for j, o in enumerate(items) if j != i]
        checked = matched = 0
        for k in range(len(periods)):
            mine = arr[id(it)][k]
            comp = [arr[id(o)][k] for o in others]
            comp = [c for c in comp if c is not None]
            if mine is None or len(comp) < 2:
                continue
            checked += 1
            s = sum(comp)
            if abs(mine - s) <= tol * max(abs(mine), abs(s), 1.0):
                matched += 1
        if checked >= 1 and matched == checked:
            parents.add(id(it))
    leaves = [it for it in items if id(it) not in parents]
    dropped = [it for it in items if id(it) in parents]
    return (leaves or list(items)), dropped


def coverage(series: dict[str, Optional[float]], periods: list[str]) -> tuple[int, int, int]:
    """(periods populated, periods materially non-zero, total periods)."""
    n_pop = sum(1 for p in periods if series.get(p) is not None)
    n_nz = sum(1 for p in periods if series.get(p) is not None and abs(series[p]) > EPS)
    return n_pop, n_nz, len(periods)


def is_supported(series: dict[str, Optional[float]], periods: list[str]) -> bool:
    """A ratio/insight row earns its place only if its source is materially
    non-zero in at least one period — an all-zero / all-blank source is a
    conclusion-shaped row with no data behind it (drop it, don't interpret it)."""
    _, n_nz, _ = coverage(series, periods)
    return n_nz >= 1


def coverage_note(series: dict[str, Optional[float]], periods: list[str]) -> Optional[str]:
    n_pop, _, n = coverage(series, periods)
    return None if n_pop == n else f"{n_pop} of {n} periods populated — partial coverage"


def prescale_periods(rev: dict[str, Optional[float]], periods: list[str],
                     frac: float = 0.05) -> set[str]:
    """Periods where revenue is below ``frac`` of peak revenue — a near-zero
    denominator that turns a real margin into an artifact (e.g. -1573%). Such
    periods are 'n/m (pre-scale)', not operating ratios."""
    peak = terminal_scale(rev, periods)
    if peak <= 0:
        return set()
    return {p for p in periods if rev.get(p) is not None and abs(rev[p]) < frac * peak}


def mask(series: dict[str, Optional[float]], drop: set[str]) -> dict[str, Optional[float]]:
    return {p: (None if p in drop else v) for p, v in series.items()}


def pre_horizon_nonzero(wbd, item) -> Optional[float]:
    """If the cells immediately LEFT of a row's first forecast cell (the
    historical/actuals columns) carry a non-zero value, return it — used to catch
    a working-capital line that was real, then switched off in the plan."""
    if item is None or not item.coords:
        return None
    r, c0 = item.coords[0]
    sd = wbd.sheets.get(item.sheet)
    if sd is None:
        return None
    for c in range(c0 - 1, max(0, c0 - 4), -1):
        rec = sd.cell(r, c)
        v = rec.value if rec else None
        if isinstance(v, (int, float)) and abs(v) > EPS:
            return float(v)
    return None


def _rising(series: dict[str, Optional[float]], periods: list[str], min_rel: float = 0.15) -> bool:
    """True if the series ends materially higher than it starts — so an
    interpretive caption ('rising intensity…') only appears when the data rises."""
    vals = [series[p] for p in periods if series.get(p) is not None]
    if len(vals) < 2:
        return False
    base = max(abs(vals[0]), abs(vals[-1]), EPS)
    return (vals[-1] - vals[0]) / base > min_rel


def _years_between(a: str, b: str) -> Optional[float]:
    ma, mb = re.match(r"(\d{4})", a), re.match(r"(\d{4})", b)
    return float(int(mb.group(1)) - int(ma.group(1))) if (ma and mb) else None


def _checkpoints(series: dict[str, Optional[float]]) -> dict[str, Optional[float]]:
    keys = [p for p, v in series.items() if v is not None]
    if not keys:
        return {}
    picks = sorted({keys[0], keys[len(keys) // 2], keys[-1]}, key=keys.index)
    return {p: series[p] for p in picks}


def mk(metric_id: str, label: str, *, applicability: str, inputs: list[str],
       computation: str, series: Optional[dict[str, Optional[float]]] = None,
       scalar: Optional[float] = None, units: Optional[str] = None,
       notes: Optional[str] = None, external: bool = False) -> DerivedMetric:
    return DerivedMetric(
        metric_id=metric_id, label=label, applicability=applicability,
        series=series or {}, scalar=scalar, inputs=inputs, computation=computation,
        checkpoints=_checkpoints(series) if series else {}, units=units,
        notes=notes, external_benchmark=external,
    )


# ---------------------------------------------------------------------------
# Scalar assumption scan (valuation params, tax rate, cost per head)
# ---------------------------------------------------------------------------

def find_scalar_assumptions(
    wbd: WorkbookData, structure: StructureResult, pattern: str,
    value_range: Optional[tuple[float, float]] = None,
) -> list[tuple[str, float, str]]:
    """``(qualified ref, value, label)`` for typed scalars whose row label
    matches ``pattern`` on assumptions/valuation/documentation-free sheets."""
    rx = re.compile(pattern, re.I)
    out: list[tuple[str, float, str]] = []
    for name, sd in wbd.sheets.items():
        role, _ = structure.roles.get(name, (SheetRole.unknown, 0.0))
        if role in (SheetRole.documentation, SheetRole.presentation):
            continue
        for (r, c), rec in sd.cells.items():
            if not isinstance(rec.value, str) or not rx.search(rec.value) or len(rec.value) > 80:
                continue
            for cc in range(c + 1, min(c + 8, sd.max_col + 1)):
                vrec = sd.cell(r, cc)
                if vrec is not None and is_number(vrec.value):
                    v = float(vrec.value)
                    if value_range is None or value_range[0] <= v <= value_range[1]:
                        out.append((format_ref(name, a1(r, cc)), v, rec.value.strip()))
                    break
    return out


# ---------------------------------------------------------------------------
# Metric builders
# ---------------------------------------------------------------------------

def _core_growth_and_margins(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    pairs = [("revenue_total", "revenue_growth", "Revenue growth"),
             ("cogs_total", "cogs_growth", "COGS growth"),
             ("opex_total", "opex_growth", "Opex growth"),
             ("ebitda_or_ebit", "ebitda_growth", "EBITDA growth")]
    for cid, mid, label in pairs:
        s = ctx.series(cid)
        if s is None:
            continue
        out.metrics.append(mk(
            mid, label, applicability=f"{cid} mapped on {ctx.sheet!r}",
            inputs=ctx.refs(cid), computation=f"({cid}[t] - {cid}[t-1]) / |{cid}[t-1]|",
            series=growth(s, P), units="x (period growth)",
            notes="small-base periods are acquitted at the findings layer, not hidden here"))

    rev = ctx.series("revenue_total")
    cogs = ctx.series("cogs_total")
    gp = ctx.series("gross_profit")
    if rev and gp:
        # prefer the model's OWN gross-profit row (respects adjustments) when it
        # ties to revenue - |COGS|; the renderer then shows GP as the basis row
        out.metrics.append(mk(
            "gross_margin", "Gross margin (from the model's gross-profit row)",
            applicability="revenue_total and gross_profit mapped",
            inputs=ctx.refs("gross_profit", "revenue_total"),
            computation="gross_profit[t] / revenue_total[t]",
            series=ratio(gp, rev, P), units="fraction"))
    elif rev and cogs:
        # no GP row: derive from revenue and COGS, orienting COGS PER PERIOD
        # (a single credit period must not flip the whole series)
        gm = combine(P, lambda r, c: (r - abs(c)) / r if abs(r) > EPS else None, rev, cogs)
        out.metrics.append(mk(
            "gross_margin", "Gross margin (computed from revenue and COGS)",
            applicability="revenue_total and cogs_total mapped",
            inputs=ctx.refs("revenue_total", "cogs_total"),
            computation="(revenue_total[t] - |cogs_total[t]|) / revenue_total[t]",
            series=gm, units="fraction"))
    # near-zero-revenue years turn real margins into artifacts (e.g. -1573%);
    # mask them as n/m so the revenue-denominated ratios stay readable
    pre = prescale_periods(rev, P) if rev else set()
    ebitda = ctx.series("ebitda_or_ebit")
    eb_item = ctx.item("ebitda_or_ebit")
    if rev and ebitda:
        # label by what the operand actually is: an EBIT row is NOT EBITDA (they
        # diverge by D&A), and calling it EBITDA is a real misstatement
        is_ebitda = bool(eb_item) and "ebitda" in (eb_item.label or "").lower()
        m_label = "EBITDA margin" if is_ebitda else "EBIT margin"
        m_note = None if is_ebitda else "row is EBIT, not EBITDA (D&A not added back); labeled accordingly"
        out.metrics.append(mk(
            "ebitda_margin", m_label,
            applicability="revenue_total and ebitda_or_ebit mapped",
            inputs=ctx.refs("revenue_total", "ebitda_or_ebit"),
            computation=("ebitda[t] / revenue_total[t]" if is_ebitda else "ebit[t] / revenue_total[t]"),
            series=mask(ratio(ebitda, rev, P), pre), units="fraction", notes=m_note))
    opex = ctx.series("opex_total")
    if rev and opex:
        out.metrics.append(mk(
            "opex_over_revenue", "Opex / revenue",
            applicability="revenue_total and opex_total mapped",
            inputs=ctx.refs("revenue_total", "opex_total"),
            computation="opex_total[t] / revenue_total[t]",
            series=mask(ratio(opex, rev, P), pre), units="fraction"))

    # Full-horizon CAGR (analysts always state the multi-year compounding rate)
    for cid, mid, label in (("revenue_total", "revenue_cagr", "Revenue CAGR (full horizon)"),
                            ("gross_profit", "gross_profit_cagr", "Gross profit CAGR (full horizon)")):
        s = ctx.series(cid)
        if not s:
            continue
        pts = [(p, s[p]) for p in P if s.get(p) is not None and s[p] > 0]
        if len(pts) >= 3:
            (p0, v0), (pn, vn) = pts[0], pts[-1]
            yrs = _years_between(p0, pn)
            if yrs and yrs > 0:
                cagr = (vn / v0) ** (1.0 / yrs) - 1.0
                out.metrics.append(mk(
                    mid, label, applicability=f"{cid} mapped, >=3 positive periods",
                    inputs=ctx.refs(cid),
                    computation=f"({cid}[{pn}]/{cid}[{p0}])^(1/{yrs:g}) - 1",
                    scalar=cagr, units="fraction",
                    notes=f"compounded {p0}->{pn}; smooths the volatile year-over-year series"))

    # Operating leverage: incremental EBITDA margin (ΔEBITDA / ΔRevenue)
    if rev and ebitda:
        oplev: dict[str, Optional[float]] = {}
        prev_r = prev_e = None
        for p in P:
            r, ev = rev.get(p), ebitda.get(p)
            if r is not None and ev is not None and prev_r is not None and abs(r - prev_r) > EPS:
                oplev[p] = (ev - prev_e) / (r - prev_r)
            else:
                oplev[p] = None
            if r is not None:
                prev_r, prev_e = r, ev
        out.metrics.append(mk(
            "operating_leverage", "Operating leverage (ΔEBITDA / ΔRevenue)",
            applicability="revenue and EBITDA mapped",
            inputs=ctx.refs("revenue_total", "ebitda_or_ebit"),
            computation="(ebitda[t]-ebitda[t-1]) / (revenue[t]-revenue[t-1])",
            series=oplev, units="incremental fraction",
            notes="how much of each new revenue dollar drops to EBITDA"))

    # Rule of 40: a SCALED recurring-revenue heuristic. Withhold it unless the
    # company is actually in that regime — on a pre-/hyper-scale business (growth
    # in the hundreds of %, deeply negative margins) it is mathematically valid
    # but analytically meaningless (e.g. -540%), and dilutes the real signal.
    if rev and ebitda:
        rg = growth(rev, P)
        em = ratio(ebitda, rev, P)
        applicable = [p for p in P if rg.get(p) is not None and em.get(p) is not None
                      and -0.10 <= rg[p] <= 1.0 and em[p] > -0.5]
        if len(applicable) >= 2:
            r40 = combine(P, lambda g, m: g + m, rg, em)
            out.metrics.append(mk(
                "rule_of_40", "Rule of 40 (revenue growth % + EBITDA margin %)",
                applicability=f"scaled-regime periods present ({len(applicable)})",
                inputs=ctx.refs("revenue_total", "ebitda_or_ebit"),
                computation="revenue_growth[t] + ebitda_margin[t]",
                series=mask(r40, set(P) - set(applicable)), units="fraction",
                notes="growth-stage health check; >=0.40 is the conventional bar"))
        else:
            out.notes.append("Rule of 40 withheld: company is pre-/hyper-scale (growth or margins "
                             "far outside the scaled recurring-revenue regime the metric assumes).")


def _capital_efficiency(ctx: Ctx, out: MetricsResult) -> None:
    """Burn multiple, capex intensity, and the EBITDA-relative ratios an analyst
    adds by hand (Lithios blue rows: Net Interest/EBITDA, Taxes/EBITDA,
    Capex/Revenue, Cumulative Capex)."""
    P = ctx.periods
    rev = ctx.series("revenue_total")
    ebitda = ctx.series("ebitda_or_ebit")
    capex = ctx.series("capex")
    interest = ctx.series("interest_net")
    taxes = ctx.series("taxes")
    cf_op = ctx.series("cf_operating")
    cf_inv = ctx.series("cf_investing")

    # Burn multiple = net cash burn / net new revenue (the headline capital-efficiency number).
    # Prefer the CF statement; fall back to net_change_cash ex-financing.
    burn_series: Optional[dict[str, Optional[float]]] = None
    burn_inputs: list[str] = []
    burn_expr = ""
    if cf_op is not None and cf_inv is not None:
        burn_series = combine(P, lambda o, i: -(o + i), cf_op, cf_inv)
        burn_inputs = ctx.refs("cf_operating", "cf_investing")
        burn_expr = "-(cf_op[t]+cf_inv[t])"
    else:
        chg = ctx.series("net_change_cash")
        cf_fin = ctx.series("cf_financing")
        if chg is not None and cf_fin is not None:
            burn_series = combine(P, lambda n, f: -(n - f), chg, cf_fin)
            burn_inputs = ctx.refs("net_change_cash", "cf_financing")
            burn_expr = "-(net_change_cash[t] - cf_financing[t])"
    if rev is not None and burn_series is not None:
        bm: dict[str, Optional[float]] = {}
        prev_r = None
        for p in P:
            r, burn = rev.get(p), burn_series.get(p)
            new_rev = (r - prev_r) if (r is not None and prev_r is not None) else None
            if burn is not None and new_rev is not None and new_rev > EPS and burn > 0:
                bm[p] = burn / new_rev
            else:
                bm[p] = None
            if r is not None:
                prev_r = r
        if any(v is not None for v in bm.values()):
            out.metrics.append(mk(
                "burn_multiple", "Burn multiple (net burn / net new revenue)",
                applicability="cash burn and revenue mapped",
                inputs=burn_inputs + ctx.refs("revenue_total"),
                computation=f"{burn_expr} / (revenue[t]-revenue[t-1]); burn-positive, growth-positive periods",
                series=bm, units="x",
                notes="dollars burned per incremental dollar of revenue; <1 is efficient, >2 is heavy"))

    if rev and capex and is_supported(capex, P):
        pre = prescale_periods(rev, P)
        cov = coverage_note(capex, P)
        out.metrics.append(mk(
            "capex_over_revenue", "Capex / revenue",
            applicability="capex and revenue mapped",
            inputs=ctx.refs("capex", "revenue_total"),
            computation="|capex[t]| / revenue_total[t]",
            series=mask(combine(P, lambda c, r: abs(c) / r if abs(r) > EPS else None, capex, rev), pre),
            units="fraction", notes="capital intensity of the build" + ("; " + cov if cov else "")))
        out.metrics.append(mk(
            "cumulative_capex", "Cumulative capex",
            applicability="capex mapped", inputs=ctx.refs("capex"),
            computation="cumsum(|capex|)",
            series=cumulative({p: (abs(capex[p]) if capex[p] is not None else None) for p in P}, P),
            units=ctx.units_label(), notes="total capital sunk into the asset base over the horizon"))

    if ebitda and interest:
        out.metrics.append(mk(
            "net_interest_over_ebitda", "Net interest / EBITDA",
            applicability="interest_net and EBITDA mapped",
            inputs=ctx.refs("interest_net", "ebitda_or_ebit"),
            computation="interest_net[t] / ebitda[t] (EBITDA-positive periods)",
            series=combine(P, lambda i, e: i / e if abs(e) > EPS and e > 0 else None, interest, ebitda),
            units="fraction", notes="debt-service load relative to operating earnings"))
    if ebitda and taxes:
        out.metrics.append(mk(
            "taxes_over_ebitda", "Taxes / EBITDA",
            applicability="taxes and EBITDA mapped",
            inputs=ctx.refs("taxes", "ebitda_or_ebit"),
            computation="|taxes[t]| / ebitda[t] (EBITDA-positive periods)",
            series=combine(P, lambda t, e: abs(t) / e if abs(e) > EPS and e > 0 else None, taxes, ebitda),
            units="fraction"))


_CASH_BAL_RX = re.compile(
    r"ending cash|closing cash|cash (?:and|&) cash equiv|cash balance|cash position|"
    r"unrestricted cash|net cash position|cash on hand", re.I)
# a cash BALANCE row, never a cash FLOW / change / movement line
_CASH_FLOW_VETO_RX = re.compile(
    r"change in|net change|increase|decrease|inflow|outflow|cash flow|flow from|"
    r"from operating|from investing|from financing|movement|used in|provided by", re.I)


def _finest_cash_series(ctx: Ctx):
    """P1-1: the cash-balance row with the FINEST period axis anywhere in the model
    (monthly/quarterly), scaled to the primary sheet's units. Survival troughs hide
    between annual year-ends, so we look for a sub-annual cash series — matched by
    label, since the fine series is often on a monthly statement sheet that the
    primary (annual) mapping never reads. Returns (nper, label, sheet, {period:val})
    or None. Generic: no cell addresses, no-op when no finer series exists."""
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    prim = ctx.structure.units.get(ctx.sheet)
    prim_scale = (prim.scale if prim else 1.0) or 1.0
    best = None
    for sheet, axes in ctx.structure.axes.items():
        ax = axes[0] if axes else None
        if ax is None or ax.orientation != Orientation.periods_in_columns:
            continue
        if len(ax.periods) <= len(ctx.periods):
            continue                                     # not finer than the plan's own axis
        sd = ctx.wbd.sheets.get(sheet)
        if sd is None:
            continue
        su = ctx.structure.units.get(sheet)
        conv = (su.scale if su else 1.0) / prim_scale
        cols = []
        for hc in ax.header_cells:
            try:
                cl, _r = coordinate_from_string(hc.split("!")[-1])
                cols.append(column_index_from_string(cl))
            except Exception:
                cols.append(None)
        first_col = min((c for c in cols if c), default=4)
        for r in range(1, sd.max_row + 1):
            lab = next((sd.cell(r, c).value.strip() for c in range(1, first_col)
                        if sd.cell(r, c) and isinstance(sd.cell(r, c).value, str)
                        and sd.cell(r, c).value.strip()), None)
            if not lab or not _CASH_BAL_RX.search(lab) or _CASH_FLOW_VETO_RX.search(lab):
                continue
            series = {}
            for p, c in zip(ax.periods, cols):
                if c is None:
                    continue
                rec = sd.cell(r, c)
                v = rec.value if rec else None
                series[p] = v * conv if isinstance(v, (int, float)) and not isinstance(v, bool) else None
            nz = [v for v in series.values() if v is not None and abs(v) > 1e-9]
            if len(nz) < 2:
                continue
            low = lab.lower()
            # prefer the END/closing balance over a beginning-of-period line
            pref = 0 if ("beginning" in low or "opening" in low) else \
                (2 if re.search(r"ending|closing|equiv", low) else 1)
            key = (len(ax.periods), pref)
            if best is None or key > (best[0], best[1]):
                best = (len(ax.periods), pref, lab[:40], sheet, series)
    return (best[0], best[2], best[3], best[4]) if best else None


def _runway(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    end = ctx.series("ending_cash")
    cash_src = "ending_cash"
    if end is None:
        # fall back to the cash-on-hand at period start: real models often carry
        # only a 'beginning cash' line, but survival is the first question and the
        # trajectory is the same answer.
        end = ctx.series("beginning_cash")
        cash_src = "beginning_cash"
    opex = ctx.series("opex_total")
    mpp = ctx.months_per_period

    # Survival: the lowest cash balance across the plan and WHEN it happens — the
    # first question for any pre-profitability company. A negative trough means
    # the plan needs a raise before that year or it runs out of money.
    if end:
        pts = [(p, end[p]) for p in P if end.get(p) is not None]
        if pts:
            mp, mv = min(pts, key=lambda kv: kv[1])
            neg = mv < 0
            src_note = "" if cash_src == "ending_cash" else " (from beginning-of-period cash; no ending-cash row mapped)"
            out.metrics.append(mk(
                "minimum_cash", f"Minimum cash across the plan ({mp})",
                applicability=f"{cash_src} mapped", inputs=ctx.refs(cash_src),
                computation=f"min over periods of {cash_src}[t]",
                scalar=mv, units=ctx.units_label(),
                notes=(f"trough in {mp}; cash goes NEGATIVE — the model assumes a raise lands before then"
                       if neg else f"trough in {mp}; stays positive across the plan") + src_note))
        # P1-1: the annual year-end snapshot hides the intra-period trough where
        # survival risk actually lives. If a finer (monthly/quarterly) cash series
        # exists, report its true minimum and the trough period.
        fine = _finest_cash_series(ctx)
        if fine is not None:
            _n, flab, fsheet, fseries = fine
            fpts = [(p, v) for p, v in fseries.items() if v is not None and abs(v) > 1e-9]
            annual_low = min((v for _p, v in pts), default=None)
            if fpts:
                tp2, tv2 = min(fpts, key=lambda kv: kv[1])
                # only surface if the finer trough is materially below the annual snapshot
                if annual_low is None or annual_low <= 0 or tv2 < 0.9 * annual_low:
                    out.metrics.append(mk(
                        "minimum_cash_intraperiod", f"Minimum cash — intra-period trough ({tp2})",
                        applicability=f"finest cash series: {fsheet} {flab!r} ({_n} periods)",
                        inputs=ctx.refs(cash_src),
                        computation=f"min over the finest cash-balance series ({fsheet}, {_n} periods), scaled to plan units",
                        scalar=tv2, units=ctx.units_label(),
                        notes=(f"the annual year-end low is {annual_low:.3g}, but the true intra-period "
                               f"trough is {tv2:.3g} in {tp2} — annual snapshots hide where the cash "
                               f"actually bottoms out" if annual_low is not None
                               else f"intra-period cash trough is {tv2:.3g} in {tp2}")))
    if end and opex:
        same = combine(P, lambda e, o: e / (abs(o) / mpp) if abs(o) > EPS else None, end, opex)
        out.metrics.append(mk(
            "runway_same_period_months", "Months of opex in the bank (same period)",
            applicability="ending_cash and opex_total mapped",
            inputs=ctx.refs("ending_cash", "opex_total"),
            computation=f"ending_cash[t] / (opex_total[t] / {mpp:g})",
            series=same, units="months"))
        fwd: dict[str, Optional[float]] = {}
        for i, p in enumerate(P):
            e = end.get(p)
            o_next = opex.get(P[i + 1]) if i + 1 < len(P) else None
            fwd[p] = e / (abs(o_next) / mpp) if (e is not None and o_next is not None and abs(o_next) > EPS) else None
        out.metrics.append(mk(
            "runway_forward_months", "Months of NEXT period's opex in the bank (forward-looking)",
            applicability="ending_cash and opex_total mapped",
            inputs=ctx.refs("ending_cash", "opex_total"),
            computation=f"ending_cash[t] / (opex_total[t+1] / {mpp:g})",
            series=fwd, units="months",
            notes="the honest runway when budgets step up year over year"))
        div = combine(P, lambda a, b: a - b, same, fwd)
        out.metrics.append(mk(
            "runway_divergence_months", "Same-period minus forward runway",
            applicability="both runway variants computed",
            inputs=ctx.refs("ending_cash", "opex_total"),
            computation="runway_same_period_months[t] - runway_forward_months[t]",
            series=div, units="months",
            notes="large positive divergence means the same-year number flatters"))

    # burn-based forward runway when the CF statement is mapped
    cf_op = ctx.series("cf_operating")
    cf_inv = ctx.series("cf_investing")
    if end and cf_op and cf_inv:
        fwd_burn: dict[str, Optional[float]] = {}
        for i, p in enumerate(P):
            e = end.get(p)
            if i + 1 < len(P):
                o, v = cf_op.get(P[i + 1]), cf_inv.get(P[i + 1])
            else:
                o = v = None
            if e is None or o is None or v is None:
                fwd_burn[p] = None
            else:
                burn = -(o + v)
                fwd_burn[p] = e / (burn / mpp) if burn > EPS else None
        out.metrics.append(mk(
            "runway_forward_burn_months", "Forward runway on next period's pre-financing burn",
            applicability="ending_cash, cf_operating and cf_investing mapped",
            inputs=ctx.refs("ending_cash", "cf_operating", "cf_investing"),
            computation=f"ending_cash[t] / (-(cf_op[t+1]+cf_inv[t+1]) / {mpp:g}); None when next period is cash-positive",
            series=fwd_burn, units="months"))

    chg = ctx.series("net_change_cash")
    if chg:
        out.metrics.append(mk(
            "avg_monthly_cash_change", "Average monthly net cash change (incl. financing)",
            applicability="net_change_cash mapped",
            inputs=ctx.refs("net_change_cash"),
            computation=f"net_change_cash[t] / {mpp:g}",
            series=combine(P, lambda n: n / mpp, chg), units=ctx.units_label() + " per month",
            notes="includes financing inflows — a raise year reads positive; see operating burn for the spend rate"))
    # operating burn ex-financing (the real spend rate)
    cf_op2 = ctx.series("cf_operating")
    cf_inv2 = ctx.series("cf_investing")
    if cf_op2 and cf_inv2:
        out.metrics.append(mk(
            "avg_monthly_operating_burn", "Average monthly operating burn (ex-financing)",
            applicability="cf_operating and cf_investing mapped",
            inputs=ctx.refs("cf_operating", "cf_investing"),
            computation=f"-(cf_op[t]+cf_inv[t]) / {mpp:g}",
            series=combine(P, lambda o, i: -(o + i) / mpp, cf_op2, cf_inv2),
            units=ctx.units_label() + " per month",
            notes="financing stripped out — positive = burning cash, negative = self-funding"))


def _cumulative_capital(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    cf_op = ctx.series("cf_operating")
    cf_inv = ctx.series("cf_investing")
    chg = ctx.series("net_change_cash")
    cf_fin = ctx.series("cf_financing")

    if cf_op:
        cum_op = cumulative(cf_op, P)
        out.metrics.append(mk(
            "cumulative_operating_cf", "Cumulative operating cash flow",
            applicability="cf_operating mapped", inputs=ctx.refs("cf_operating"),
            computation="cumsum(cf_operating)", series=cum_op, units=ctx.units_label()))

    fcf: Optional[dict[str, Optional[float]]] = None
    fcf_inputs: list[str] = []
    fcf_comp = ""
    if cf_op and cf_inv:
        fcf = combine(P, lambda o, v: o + v, cf_op, cf_inv)
        fcf_inputs = ctx.refs("cf_operating", "cf_investing")
        fcf_comp = "cf_operating[t] + cf_investing[t]"
    elif chg and cf_fin:
        fcf = combine(P, lambda n, f: n - f, chg, cf_fin)
        fcf_inputs = ctx.refs("net_change_cash", "cf_financing")
        fcf_comp = "net_change_cash[t] - cf_financing[t]"
    if fcf is not None:
        cum_fcf = cumulative(fcf, P)
        out.metrics.append(mk(
            "cumulative_fcf", "Cumulative free cash flow (pre-financing)",
            applicability="cash-flow statement mapped", inputs=fcf_inputs,
            computation=f"cumsum({fcf_comp})", series=cum_fcf, units=ctx.units_label()))
        valid = [(p, v) for p, v in cum_fcf.items() if v is not None]
        if valid:
            peak_p, peak_v = min(valid, key=lambda t: t[1])
            out.metrics.append(mk(
                "peak_cumulative_consumption", "Peak cumulative cash consumption (capital at risk)",
                applicability="cumulative_fcf computed", inputs=fcf_inputs,
                computation=f"min(cumsum({fcf_comp})) — reached in {peak_p}",
                scalar=peak_v, units=ctx.units_label(),
                notes=f"trough reached in {peak_p}"))

    eq = ctx.series("equity_proceeds")
    debt = ctx.series("debt_proceeds")
    cum_ext: Optional[dict[str, Optional[float]]] = None
    if eq:
        out.metrics.append(mk(
            "cumulative_equity", "Cumulative equity raised",
            applicability="equity_proceeds mapped", inputs=ctx.refs("equity_proceeds"),
            computation="cumsum(equity_proceeds)", series=cumulative(eq, P), units=ctx.units_label()))
    if debt:
        out.metrics.append(mk(
            "cumulative_debt", "Cumulative debt drawn",
            applicability="debt_proceeds mapped", inputs=ctx.refs("debt_proceeds"),
            computation="cumsum(debt_proceeds)", series=cumulative(debt, P), units=ctx.units_label()))
    if eq or debt:
        zero = {p: 0.0 for p in P}
        cum_ext = combine(P, lambda a, b: a + b,
                          cumulative(eq or zero, P), cumulative(debt or zero, P))
        out.metrics.append(mk(
            "cumulative_external_capital", "Cumulative external capital (equity + debt)",
            applicability="financing rows mapped",
            inputs=ctx.refs("equity_proceeds", "debt_proceeds"),
            computation="cumsum(equity_proceeds) + cumsum(debt_proceeds)",
            series=cum_ext, units=ctx.units_label()))

    peak = out.get("peak_cumulative_consumption")
    if peak is not None and peak.scalar is not None and peak.scalar < 0 and cum_ext is not None:
        peak_p = (peak.notes or "").replace("trough reached in ", "") or None
        ext_at_peak = cum_ext.get(peak_p) if peak_p else None
        if ext_at_peak is not None:
            out.metrics.append(mk(
                "cushion_ratio", "External capital raised vs peak consumption (cushion)",
                applicability="peak consumption and external capital computed",
                inputs=peak.inputs + ctx.refs("equity_proceeds", "debt_proceeds"),
                computation=f"cumulative_external_capital[{peak_p}] / |peak_cumulative_consumption|",
                scalar=ext_at_peak / abs(peak.scalar) if abs(peak.scalar) > EPS else None,
                units="x", notes="below ~1.1x means the plan runs with almost no slack"))

    ebitda = ctx.series("ebitda_or_ebit")
    if fcf is not None and ebitda:
        conv = combine(P, lambda f, e: f / e if e > EPS else None, fcf, ebitda)
        out.metrics.append(mk(
            "cash_conversion", "FCF conversion (FCF / EBITDA, EBITDA-positive periods)",
            applicability="FCF and EBITDA available", inputs=fcf_inputs + ctx.refs("ebitda_or_ebit"),
            computation=f"({fcf_comp}) / ebitda[t] where ebitda > 0",
            series=conv, units="fraction"))
        # cumulative: over the whole plan, does profitability become cash, or is
        # it absorbed by working capital and capex?
        conv_cum = combine(P, lambda f, e: f / e if e > EPS else None,
                           cumulative(fcf, P), cumulative(ebitda, P))
        out.metrics.append(mk(
            "cash_conversion_cumulative", "Cumulative FCF conversion (Σ FCF / Σ EBITDA)",
            applicability="FCF and EBITDA available", inputs=fcf_inputs + ctx.refs("ebitda_or_ebit"),
            computation="cumsum(FCF)[t] / cumsum(EBITDA)[t] where cumulative EBITDA > 0",
            series=conv_cum, units="fraction",
            notes="whether reported profit turns into cash over the plan, or is eaten by working capital and capex"))


def _grants(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    items = ctx.items("grant_or_subsidy_revenue")
    rev = ctx.series("revenue_total")
    if not items or rev is None:
        return
    # Statements repeat grant rows per section (revenue / COGS / gross-profit
    # echo at 100% margin). Keep revenue-section rows; failing that, dedupe
    # identical value series so echoes are not double-counted.
    rev_section = [it for it in items
                   if "revenue" in it.section and not any(k in it.section for k in ("gross", "cogs"))]
    pool = rev_section or [it for it in items
                           if not any(k in it.section for k in ("gross profit", "gross margin", "cogs"))] or items
    uniq: list[LineItem] = []
    seen_series: list[tuple] = []
    for it in pool:
        key = tuple(round(v, 6) if v is not None else None for v in it.values)
        if all(v in (None, 0.0) for v in key):
            sig = None
        else:
            sig = key
        if sig is not None and sig in seen_series:
            continue
        seen_series.append(sig)
        uniq.append(it)
    items = uniq
    grants: dict[str, Optional[float]] = {}
    for p in P:
        vals = [it.value_for(p) for it in items]
        vals = [v for v in vals if v is not None]
        grants[p] = sum(vals) if vals else None
    # P2-2 category-presence gating: the grant category is only ACTIVE if it is
    # populated with non-zero values. When grants are absent OR zero every period,
    # every grant-dependent row (Grants/revenue = 0%, Revenue ex-grants, ex-grants
    # growth/margin) would just duplicate its base metric — suppress the whole
    # block rather than emit empty-calorie twins. (A model with real grant revenue
    # keeps them and they get flagged if material.)
    if all(v is None or abs(v) <= EPS for v in grants.values()):
        return
    grant_refs = [r for it in items for r in it.refs][:60]
    share = ratio(grants, rev, P)
    out.metrics.append(mk(
        "grant_share_of_revenue", "Grants / total revenue",
        applicability=f"{len(items)} grant/subsidy row(s) mapped",
        inputs=grant_refs + ctx.refs("revenue_total"),
        computation="sum(grant rows)[t] / revenue_total[t]",
        series=share, units="fraction"))
    ex = combine(P, lambda r, g: r - g, rev, grants)
    out.metrics.append(mk(
        "revenue_ex_grants", "Revenue excluding grants",
        applicability="grants mapped", inputs=grant_refs + ctx.refs("revenue_total"),
        computation="revenue_total[t] - sum(grant rows)[t]",
        series=ex, units=ctx.units_label()))
    out.metrics.append(mk(
        "revenue_ex_grants_growth", "Revenue ex-grants growth",
        applicability="grants mapped", inputs=grant_refs + ctx.refs("revenue_total"),
        computation="(rev_ex_grants[t] - rev_ex_grants[t-1]) / rev_ex_grants[t-1], shown only where the prior-year commercial base is positive",
        series=growth(ex, P, positive_only=True), units="x (period growth)",
        notes="the commercial growth rate once 100%-margin grant revenue is removed; blank in years where ex-grant revenue is zero or negative (grants exceed total revenue, so there is no commercial base to grow from)"))
    gp = ctx.series("gross_profit")
    if gp:
        gm_num = combine(P, lambda g, gr: g - gr, gp, grants)
        gm_ex, gm_undef = ex_ratio(gm_num, ex, P)         # ex = revenue - grants (the commercial base)
        note = None
        if gm_undef:
            note = (f"undefined in {', '.join(gm_undef)}: grants exceed revenue, so the commercial base "
                    "(revenue ex-grants) is zero or negative — a margin on it is meaningless. The company "
                    "has effectively no commercial revenue those years; see Grants / revenue.")
        out.metrics.append(mk(
            "gross_margin_ex_grants", "Gross margin ex-grants (grants treated as 100% margin)",
            applicability="grants and gross_profit mapped",
            inputs=grant_refs + ctx.refs("gross_profit", "revenue_total"),
            computation="(gross_profit[t] - grants[t]) / revenue_ex_grants[t]; "
                        "undefined where revenue ex-grants <= 0 (grants exceed revenue)",
            series=gm_ex, units="fraction", notes=note))


def _capex_in_opex(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    capex_comps = [it for it in ctx.items("opex_component") if "capex_expensed" in it.subtags]
    if not capex_comps:
        return
    rev = ctx.series("revenue_total")
    opex = ctx.series("opex_total")
    ebitda = ctx.series("ebitda_or_ebit")
    cap_sum: dict[str, Optional[float]] = {}
    for p in P:
        vals = [it.value_for(p) for it in capex_comps]
        vals = [v for v in vals if v is not None]
        cap_sum[p] = sum(vals) if vals else None
    cap_refs = [r for it in capex_comps for r in it.refs][:40]
    names = ", ".join(it.label.strip() for it in capex_comps[:3])
    if ebitda:
        out.metrics.append(mk(
            "ebitda_before_expensed_capex", "EBITDA before expensed capital equipment",
            applicability=f"capex-like rows inside opex: {names}",
            inputs=cap_refs + ctx.refs("ebitda_or_ebit"),
            computation="ebitda[t] + expensed_capex[t]",
            series=combine(P, lambda e, c: e + abs(c), ebitda, cap_sum), units=ctx.units_label(),
            notes="the model expenses capex inside opex; this restates EBITDA as if it were capitalized"))
    if opex:
        ex_capex = combine(P, lambda o, c: o - abs(c) if abs(o) >= abs(c) else o + abs(c), opex, cap_sum)
        out.metrics.append(mk(
            "opex_ex_capex_growth", "Expense growth excluding capital equipment",
            applicability=f"capex-like rows inside opex: {names}",
            inputs=cap_refs + ctx.refs("opex_total"),
            computation="growth(opex_total[t] - expensed_capex[t])",
            series=growth(ex_capex, P),
            units="x (period growth)",
            notes="the acquittal decomposition for capex-lumpiness expense anomalies"))
        if rev:
            out.metrics.append(mk(
                "opex_ex_capex_over_revenue", "Opex excluding expensed capex / revenue",
                applicability=f"capex-like rows inside opex: {names}",
                inputs=cap_refs + ctx.refs("opex_total", "revenue_total"),
                computation="(opex_total[t] - expensed_capex[t]) / revenue_total[t]",
                series=ratio(ex_capex, rev, P), units="fraction"))


_NATURE_KEYS = [
    ("grant", ("grant", "subsid", "subvent", "tax credit", "r&d credit", "sred")),
    ("recurring", ("royalt", "recurring", "subscription", "saas", "arr", "mrr", "licen",
                   "maintenance", "retainer", "hosting", "renewal", "support contract")),
    ("services", ("service", "consulting", "professional fee", "implementation", "training",
                  "integration", "installation service")),
    ("one-time", ("one-time", "one time", "onetime", "non-recurring", "setup", "nre",
                  "turnkey", "milestone", "upfront", "perpetual", "hardware sale", "equipment sale")),
]


def _revenue_nature(label: str) -> str:
    """Classify a revenue stream by nature so we can talk about MIX, not just
    growth. Heuristic on the label; 'product' is the default commercial bucket."""
    l = (label or "").lower()
    for nature, keys in _NATURE_KEYS:
        if any(k in l for k in keys):
            return nature
    return "product"


def _revenue_mix(ctx: Ctx, out: MetricsResult) -> None:
    """Revenue MIX by nature over time (recurring / one-time / services / grant /
    product). A shift in composition — e.g. recurring going 0% -> 78% — is usually
    the real story that top-line growth alone hides."""
    P = ctx.periods
    rev = ctx.series("revenue_total")
    if rev is None:
        return
    segs = [it for it in ctx.items("revenue_segment") if it.sheet == ctx.sheet]
    if len(segs) < 2:
        return
    segs, _ = resolve_hierarchy(segs, P)
    buckets: dict[str, list[LineItem]] = {}
    for it in segs:
        buckets.setdefault(_revenue_nature(it.instance or it.label), []).append(it)
    if len(buckets) < 2:                       # need an actual mix to decompose
        return
    emitted = 0
    for nature in ("recurring", "one-time", "services", "product", "grant"):
        items = buckets.get(nature)
        if not items:
            continue
        tot: dict[str, Optional[float]] = {}
        for p in P:
            vals = [it.value_for(p) for it in items if it.value_for(p) is not None]
            tot[p] = sum(vals) if vals else None
        share = ratio(tot, rev, P)
        if not any((share.get(p) or 0) >= 0.05 for p in P):   # immaterial — skip
            continue
        vals = [share[p] for p in P if share.get(p) is not None]
        shift = (max(vals) - min(vals)) if len(vals) >= 2 else 0.0
        names = ", ".join((it.instance or it.label).strip() for it in items[:3])
        note = f"share of revenue from {nature} streams"
        if shift >= 0.20:
            note += f" — composition shifts ~{shift*100:.0f}pp over the plan (the real story under the top line)"
        out.metrics.append(mk(
            f"revenue_mix_{nature.replace('-', '_')}", f"Revenue mix: {nature} share",
            applicability=f"{nature} streams: {names}",
            inputs=[r for it in items for r in it.refs][:40] + ctx.refs("revenue_total"),
            computation=f"sum({nature} segment rows)[t] / revenue_total[t]",
            series=share, units="fraction", notes=note))
        emitted += 1
    if emitted:
        out.notes.append("revenue decomposed by nature (recurring / one-time / services / grant / product)")


def _seg_share_series(it, rmap, P):
    return ratio({p: it.value_for(p) for p in P}, rmap, P)


def _seg_terminal(sh, P):
    return next((sh[p] for p in reversed(P) if sh.get(p) is not None), 0.0) or 0.0


def _is_real_revenue_share(it, sh, P) -> bool:
    """A 'Revenue share' row needs a REVENUE numerator with a plausible share.
    Drops all-zero rows and price/unit/market-sizing lines (Rule B)."""
    vals = [sh[p] for p in P if sh.get(p) is not None]
    if not vals or not any(abs(v) > 1e-9 for v in vals):
        return False                                  # all zero / null -> noise
    if sum(1 for v in vals if -0.2 <= v <= 1.2) < (len(vals) + 1) // 2:
        return False                                  # mostly out of [0,1] -> price/unit/market basis
    lab = f"{it.instance or ''} {it.label or ''}".lower()
    if any(k in lab for k in ("price", "/unit", "€/unit", "per unit", "units", " asp", "market")):
        return False                                  # explicit non-revenue line
    return True


def _segments(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    rev = ctx.series("revenue_total")
    if rev is None:
        return
    # candidates: each revenue_segment whose OWN sheet carries a revenue_total
    # (a share is only meaningful against the revenue it rolls into)
    cands = []
    for it in ctx.mapping.get_all("revenue_segment"):
        rt = ctx.mapping.get("revenue_total", it.sheet)
        if rt is None or rt.sheet != it.sheet:
            continue
        cands.append((it, {p: rt.value_for(p) for p in P}))
    if not cands:
        return

    # Emit segments from ONE sheet so every share uses the SAME revenue base and
    # they sum sensibly (mixing sheets' bases double-counts -> sums far past 100%).
    # Prefer the primary sheet (where total revenue is built); else the richest.
    by_sheet: dict[str, list] = {}
    for it, rmap in cands:
        sh = _seg_share_series(it, rmap, P)
        if not _is_real_revenue_share(it, sh, P):     # Rule B suppression
            continue
        by_sheet.setdefault(it.sheet, []).append((it, sh))
    if not by_sheet:
        return
    seg_sheet = ctx.sheet if ctx.sheet in by_sheet else max(by_sheet, key=lambda s: len(by_sheet[s]))
    # P0-3 denominator provenance: divide by (and cite) the SAME revenue row that
    # sits on the segments' sheet — the total shown in provenance, not some other.
    rev_base_it = ctx.mapping.get("revenue_total", seg_sheet)
    rev_base = {p: rev_base_it.value_for(p) for p in P} if rev_base_it else rev
    rev_base_refs = list(rev_base_it.refs) if rev_base_it else ctx.refs("revenue_total")

    # Rule A: ONE row per segment name on that sheet (best terminal share wins) —
    # never a same-named impostor pulling units/prices off another sheet.
    by_name: dict[str, tuple] = {}
    for it, sh in by_sheet[seg_sheet]:
        name = (it.instance or it.label.strip())[:40]
        score = _seg_terminal(sh, P)
        if name not in by_name or score > by_name[name][0]:
            by_name[name] = (score, it, sh)
    chosen = sorted(by_name.values(), key=lambda e: -e[0])         # terminal share desc

    # P0-2 component-to-parent tie-out: the components must not EXCEED the parent
    # shown as denominator. Shares summing well past 100% mean the component rows
    # sit on a different (phantom) base than the displayed revenue — 100% can be
    # satisfied on a phantom base, so the binding check is components<=displayed
    # total, not shares==100%. If they exceed it, do not emit shares; flag it.
    checked = over = 0
    for p in P:
        rv = rev_base.get(p)
        if rv is None or abs(rv) <= EPS:
            continue
        checked += 1
        if sum((sh.get(p) or 0) for (_s, _it, sh) in chosen) > 1.02:
            over += 1
    if checked and over > checked / 2:
        out.notes.append(
            f"CITATION segment tie-out: the mapped segments on {seg_sheet} sum to more than the revenue "
            f"shown as their denominator (components do not reconcile to the displayed total); revenue-"
            f"share rows suppressed until the source block is resolved.")
        return

    for (_score, it, sh) in chosen[:8]:
        name = (it.instance or it.label.strip())[:40]
        out.metrics.append(mk(
            f"segment_share::{name}", f"Revenue share: {name}",
            applicability="revenue segments mapped",
            inputs=[r for r in it.refs][:30] + rev_base_refs,
            computation=f"{name}[t] / revenue_total[t] (denominator: {seg_sheet} revenue)",
            series=sh, units="fraction"))

    # Rule A.3: explicit 'other / unattributed' remainder (same base) so the
    # segments sum to ~100% — a big residual flags a wrong/missing source.
    rem: dict[str, Optional[float]] = {}
    for p in P:
        rv = rev_base.get(p)
        if rv is None or abs(rv) <= EPS:
            rem[p] = None
            continue
        r = 1.0 - sum((sh.get(p) or 0) for (_s, _it, sh) in chosen)
        rem[p] = r if r > 0.01 else None
    if any(v is not None for v in rem.values()):
        out.metrics.append(mk(
            "segment_share::other", "Revenue share: other / unattributed",
            applicability="remainder so mapped segments sum to ~100% of revenue",
            inputs=ctx.refs("revenue_total"),
            computation="1 - sum(mapped segment shares)[t]",
            series=rem, units="fraction",
            notes="revenue not attributed to a mapped segment"))

    # terminal concentration — largest real segment's terminal share
    if chosen:
        top = max(chosen, key=lambda e: _seg_terminal(e[2], P))
        name = (top[1].instance or top[1].label.strip())[:40]
        share = _seg_terminal(top[2], P)
        if 0 < share <= 1.2:
            out.metrics.append(mk(
                "terminal_concentration", f"Largest terminal-period segment share ({name})",
                applicability="revenue segments mapped",
                inputs=[r for (_s, it, _sh) in chosen for r in it.refs][:40],
                computation="max over segments of segment[T] / revenue_total[T]",
                scalar=share, units="fraction", notes=f"segment: {name}"))

    units_items = ctx.items("units_sold")
    if units_items:
        total_units: dict[str, Optional[float]] = {}
        for p in P:
            vals = [it.value_for(p) for it in units_items]
            vals = [v for v in vals if v is not None]
            total_units[p] = sum(vals) if vals else None
        blended = combine(P, lambda r, u: r / u if u and abs(u) > EPS else None, rev, total_units)
        out.metrics.append(mk(
            "blended_asp", "Blended ASP (revenue / total units)",
            applicability="units_sold rows mapped",
            inputs=[r for it in units_items for r in it.refs][:40] + ctx.refs("revenue_total"),
            computation="revenue_total[t] / sum(units_sold rows)[t]",
            series=blended, units=ctx.units_label() + " per unit",
            notes="divergence from typed ASPs means mix shift"))


_SEG_STOP = re.compile(r"\b(revenue|sales|cogs|cost|gross|margin|net|total|the|of|and|&)\b", re.I)


def _seg_tokens(label: str) -> set[str]:
    base = _SEG_STOP.sub(" ", label.lower())
    return {t for t in re.split(r"[^a-z0-9]+", base) if len(t) > 2}


def _segment_detail(ctx: Ctx, out: MetricsResult) -> None:
    """Per-segment growth, COGS/revenue, and gross margin — the line-by-line
    decomposition an analyst runs to see which product line drives the story."""
    P = ctx.periods
    segs = [it for it in ctx.items("revenue_segment") if it.sheet == ctx.sheet]
    if not segs:
        return
    cogs_comps = [it for it in ctx.items("cogs_component") if it.sheet == ctx.sheet]

    for seg in segs[:8]:
        inst = (seg.instance or seg.label.strip())[:40]
        s = {p: seg.value_for(p) for p in P}
        out.metrics.append(mk(
            f"segment_growth::{inst}", f"Revenue growth: {inst}",
            applicability="revenue segments mapped", inputs=seg.refs[:30],
            computation=f"({inst}[t] - {inst}[t-1]) / |{inst}[t-1]|",
            series=growth(s, P), units="x (period growth)",
            notes="per-segment growth exposes which line actually drives the top-line story"))

        # match a COGS component to this segment by shared name token
        stoks = _seg_tokens(seg.label)
        match = None
        for c in cogs_comps:
            if stoks & _seg_tokens(c.label):
                match = c
                break
        if match is not None:
            cs = {p: match.value_for(p) for p in P}
            out.metrics.append(mk(
                f"segment_cogs_ratio::{inst}", f"COGS / revenue: {inst}",
                applicability="segment revenue and matching COGS component mapped",
                inputs=(seg.refs + match.refs)[:50],
                computation=f"|{match.label.strip()[:24]}|[t] / {inst}[t]",
                series=combine(P, lambda c, r: abs(c) / r if abs(r) > EPS else None, cs, s),
                units="fraction", notes=f"matched COGS row: {match.label.strip()!r}"))
            out.metrics.append(mk(
                f"segment_gross_margin::{inst}", f"Gross margin: {inst}",
                applicability="segment revenue and matching COGS component mapped",
                inputs=(seg.refs + match.refs)[:50],
                computation=f"({inst}[t] - |{match.label.strip()[:20]}|[t]) / {inst}[t]",
                series=combine(P, lambda r, c: (r - abs(c)) / r if abs(r) > EPS else None, s, cs),
                units="fraction", notes="per-line margin; divergence across lines flags mix risk"))


def _labor(ctx: Ctx, out: MetricsResult) -> None:
    """Payroll / revenue and payroll growth — does the org scale sub-linearly
    with revenue, or is headcount cost outrunning the top line?"""
    P = ctx.periods
    labor = [it for it in ctx.items("opex_component") if "labor" in it.subtags and it.sheet == ctx.sheet]
    rev = ctx.series("revenue_total")
    if not labor or rev is None:
        return
    # resolve the salary hierarchy first: never sum a 'Total salaries' subtotal
    # together with its component rows (that double-counts payroll).
    labor, parents = resolve_hierarchy(labor, P)
    payroll: dict[str, Optional[float]] = {}
    for p in P:
        vals = [it.value_for(p) for it in labor]
        vals = [v for v in vals if v is not None]
        payroll[p] = sum(vals) if vals else None
    refs = [r for it in labor for r in it.refs][:50]
    names = ", ".join(it.label.strip() for it in labor[:3])
    pre = prescale_periods(rev, P)
    note = "should fall over time as the company gains operating leverage"
    if parents:
        out.notes.append("hierarchy_resolved (payroll): treated as subtotal and dropped to avoid "
                         "double-counting: " + "; ".join(it.label.strip() for it in parents))
    out.metrics.append(mk(
        "payroll_over_revenue", "Payroll / revenue",
        applicability=f"labor opex rows mapped: {names}",
        inputs=refs + ctx.refs("revenue_total"),
        computation="sum(component salary rows)[t] / revenue_total[t]",
        series=mask(combine(P, lambda pay, r: abs(pay) / r if abs(r) > EPS else None, payroll, rev), pre),
        units="fraction", notes=note))
    out.metrics.append(mk(
        "payroll_growth", "Payroll growth",
        applicability=f"labor opex rows mapped: {names}", inputs=refs,
        computation="growth(sum(payroll rows))",
        series=growth({p: (abs(payroll[p]) if payroll[p] is not None else None) for p in P}, P),
        units="x (period growth)", notes="compare against revenue growth — labor should lag it"))


def _capacity_and_share(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    caps = ctx.items("capacity")
    rev = ctx.series("revenue_total")
    market = ctx.series("market_size")
    printed = ctx.series("market_share")
    _KNOWN_KINDS = ("operational", "contracted", "nameplate", "installed", "total")
    _UNIT_TOKEN = re.compile(r"\b(tpa|tpd|mw|gw|gwh|mwh|kg|tonnes?|tons?|units?|stacks?|wafers?)\b", re.I)

    def _tokens(label: str) -> set[str]:
        return {t.lower() for t in _UNIT_TOKEN.findall(label)}

    market_item = ctx.item("market_size")
    market_tokens = _tokens(market_item.label) if market_item else set()

    by_inst: dict[str, LineItem] = {}
    # stock rows ("Cumulative Capacity Contracted") beat flow rows ("New
    # Capacity Contracted") for the same instance — shares need stocks
    caps = sorted(caps, key=lambda it: bool(re.search(r"\b(added|new)\b", it.label, re.I)))
    for it in caps:
        inst = (it.instance or it.label.strip())[:40]
        by_inst.setdefault(inst, it)
    for inst, it in list(by_inst.items())[:5]:
        known = inst in _KNOWN_KINDS
        # share/revenue ratios on a flow row ("Capacity Added") or an
        # incompatible unit basis (stacks vs tpa) are nonsense — skip them
        if not known and not (_tokens(it.label) & market_tokens):
            continue
        if re.search(r"\badded\b|\bnew\b", it.label, re.I):
            continue
        s = {p: it.value_for(p) for p in P}
        if rev:
            out.metrics.append(mk(
                f"revenue_per_capacity::{inst}", f"Revenue per unit of {inst} capacity",
                applicability="capacity rows mapped",
                inputs=it.refs[:30] + ctx.refs("revenue_total"),
                computation=f"revenue_total[t] / capacity::{inst}[t]",
                series=combine(P, lambda r, c2: r / c2 if abs(c2) > EPS else None, rev, s),
                units=ctx.units_label() + " per capacity unit"))
        if market and (known or (_tokens(it.label) & market_tokens)):
            out.metrics.append(mk(
                f"market_share_recomputed::{inst}", f"Market share recomputed on {inst} basis",
                applicability="capacity and market_size mapped",
                inputs=it.refs[:30] + ctx.refs("market_size"),
                computation=f"capacity::{inst}[t] / market_size[t]",
                series=combine(P, lambda c2, m: c2 / m if abs(m) > EPS else None, s, market),
                units="fraction",
                notes="compare against the printed market-share row; basis mismatches flatter"))
    if len(by_inst) >= 2 and "operational" in by_inst and "contracted" in by_inst:
        op = {p: by_inst["operational"].value_for(p) for p in P}
        ct = {p: by_inst["contracted"].value_for(p) for p in P}
        out.metrics.append(mk(
            "operational_contracted_ratio", "Operational / contracted capacity",
            applicability="both capacity bases mapped",
            inputs=by_inst["operational"].refs[:20] + by_inst["contracted"].refs[:20],
            computation="capacity::operational[t] / capacity::contracted[t]",
            series=combine(P, lambda o, c2: o / c2 if abs(c2) > EPS else None, op, ct),
            units="fraction"))
    if printed is not None and market is not None and by_inst:
        pass  # printed-vs-recomputed comparison happens in findings with both series present


def _valuation(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    bm = ctx.benchmarks or {}
    finds = find_scalar_assumptions(ctx.wbd, ctx.structure,
                                    r"ebitda\s*multiple|multiple.*ebitda|x\s*ebitda", (0.5, 100))
    rev_mult = find_scalar_assumptions(ctx.wbd, ctx.structure,
                                       r"revenue\s*multiple|multiple.*revenue|x\s*revenue", (0.1, 100))
    dr = find_scalar_assumptions(ctx.wbd, ctx.structure,
                                 r"discount\s*rate|hurdle|wacc", (0.01, 0.99))
    exit_year = find_scalar_assumptions(ctx.wbd, ctx.structure,
                                        r"exit\s*year|valuation\s*year|year of exit", (1990, 2100))
    ebitda = ctx.series("ebitda_or_ebit")
    rev = ctx.series("revenue_total")

    def _exit_period(yr: Optional[float]) -> Optional[str]:
        if yr is not None:
            y = str(int(yr))
            for p in P:
                if p.startswith(y):
                    return p
        for p in reversed(P):
            if ebitda and ebitda.get(p) is not None:
                return p
        return None

    yr = exit_year[0][1] if exit_year else None
    exit_p = _exit_period(yr)
    if exit_p is None or not P:
        return

    # Discounting anchor: a typed "current year"/"valuation date" beats the
    # model's first historical period (real models start years in the past;
    # discounting from there overstates the horizon and understates nothing —
    # it just computes a different question than the model asks).
    current = find_scalar_assumptions(ctx.wbd, ctx.structure,
                                      r"current\s*year|valuation\s*(date|year)|model\s*start|base\s*year",
                                      (1990, 2100))
    base_year_ref: Optional[str] = None
    m0 = re.match(r"(\d{4})", P[0])
    base_year = int(m0.group(1)) if m0 else None
    if current:
        base_year_ref, by, _ = current[0]
        base_year = int(by)

    me = re.match(r"(\d{4})", exit_p)
    horizon = float(int(me.group(1)) - base_year) if (me and base_year is not None) else None

    for mults, metric_series, metric_name in ((finds, ebitda, "ebitda"), (rev_mult, rev, "revenue")):
        if not mults or metric_series is None:
            continue
        ref, mult, label = mults[0]
        mv = metric_series.get(exit_p)
        if mv is None or mv <= 0:
            continue
        exit_val = mult * mv
        m = mk(f"exit_value_rederived_{metric_name}",
               f"Exit value re-derived ({mult:g}x {metric_name} in {exit_p})",
               applicability=f"typed multiple found: {label!r}",
               inputs=[ref] + ctx.refs("ebitda_or_ebit" if metric_name == "ebitda" else "revenue_total"),
               computation=f"{mult:g} * {metric_name}[{exit_p}]",
               scalar=exit_val, units=ctx.units_label())
        out.metrics.append(m)
        if horizon and horizon > 0:
            # a model can carry several typed rates (DCF-at-exit vs to-exit VC
            # rate); re-derive PV under each rather than guessing which rules
            for dref, drate, dlabel in dr[:2]:
                slug = re.sub(r"[^a-z0-9]+", "_", dlabel.lower())[:32].strip("_")
                pv = exit_val / ((1.0 + drate) ** horizon)
                out.metrics.append(mk(
                    f"present_value_{metric_name}_at_{slug}",
                    f"PV of {metric_name}-multiple exit at {drate:.0%}/yr over {horizon:g}y ({dlabel!r})",
                    applicability=f"typed discount rate found: {dlabel!r}",
                    inputs=[ref, dref] + ([base_year_ref] if base_year_ref else []),
                    computation=f"exit_value / (1+{drate:g})^{horizon:g}, horizon anchored at {base_year}",
                    scalar=pv, units=ctx.units_label()))
        # external sensitivity
        alt_key = "ebitda_multiples" if metric_name == "ebitda" else "revenue_multiples"
        alts = (bm.get("valuation_sensitivity") or {}).get(alt_key) or []
        for alt in alts[:3]:
            out.metrics.append(mk(
                f"exit_value_at_{metric_name}_{alt:g}x".replace(".", "_"),
                f"External benchmark (not from workbook): exit at {alt:g}x {metric_name}",
                applicability="benchmarks.yaml sensitivity",
                inputs=ctx.refs("ebitda_or_ebit" if metric_name == "ebitda" else "revenue_total"),
                computation=f"{alt:g} * {metric_name}[{exit_p}] (multiple from benchmarks.yaml)",
                scalar=alt * mv, units=ctx.units_label(), external=True))


def _taxes(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    taxes = ctx.series("taxes")
    ni = ctx.series("net_income")
    if taxes is None:
        return
    # orient taxes: stored as expense (positive) or as outflow (negative)?
    tvals = [v for v in taxes.values() if v is not None]
    sign = -1.0 if tvals and sum(tvals) < 0 else 1.0
    # P2-1: prefer a row explicitly labelled pre-tax / EBT for the denominator;
    # only reconstruct pretax = net income + taxes when no such row exists. The
    # reconstruction is distorted by any credit / non-operating item that sits
    # BELOW the tax line (it breaks the net-income + taxes identity), so caveat it.
    pretax_series = ctx.series("pretax_income")
    have_ebt = pretax_series is not None and any(v is not None for v in pretax_series.values())
    if have_ebt:
        pretax = {p: pretax_series.get(p) for p in P}
        comp = "taxes[t] / pretax[t] (pretax = the model's labelled Earnings-Before-Tax row), capped to [-100%, 150%]"
        tax_inputs = ctx.refs("taxes", "pretax_income")
        denom_note = None
    elif ni is not None:
        pretax = combine(P, lambda n, t: n + t if sign > 0 else n - t, ni, taxes)
        comp = "taxes[t] / (net_income[t] + taxes[t]) where pretax > 0, capped to [-100%, 150%]"
        tax_inputs = ctx.refs("taxes", "net_income")
        denom_note = ("reconstructed pretax = net income + taxes; if the model books credits or "
                      "non-operating items below the tax line the reconstructed rate is distorted — "
                      "map a labelled Earnings-Before-Tax row for the exact rate")
    else:
        return
    # cap implausible rates (tiny pretax with a big tax line yields nonsense
    # effective rates of several hundred percent) so they don't pollute the series
    def _rate(t, pt):
        if pt is None or t is None or pt <= EPS:
            return None
        r = (sign * t) / pt
        return r if -1.0 <= r <= 1.5 else None
    eff = combine(P, _rate, taxes, pretax)
    out.metrics.append(mk(
        "effective_tax_rate", "Effective tax rate (profitable periods)",
        applicability="taxes and pretax mapped",
        inputs=tax_inputs, computation=comp, series=eff, units="fraction", notes=denom_note))
    cum_pretax = cumulative(pretax, P)
    nol = combine(P, lambda t, cpt: (sign * t) if (cpt is not None and cpt < 0 and t is not None and sign * t > EPS) else None,
                  taxes, cum_pretax)
    if any(v is not None for v in nol.values()):
        out.metrics.append(mk(
            "cash_taxes_despite_cumulative_losses", "Cash taxes paid while cumulatively loss-making",
            applicability="taxes and pretax mapped", inputs=tax_inputs,
            computation="taxes[t] where cumsum(pretax)[t] < 0 and taxes[t] > 0",
            series=nol, units=ctx.units_label(),
            notes="NOL carryforwards would normally shelter these periods"))
    refunds = combine(P, lambda t, pt: t if (pt is not None and pt < -EPS and t is not None and sign * t < -EPS) else None, taxes, pretax)
    if any(v is not None for v in refunds.values()):
        out.metrics.append(mk(
            "tax_refunds_in_loss_years", "Taxes booked as income in loss years",
            applicability="taxes and pretax mapped", inputs=tax_inputs,
            computation="taxes[t] where pretax[t] < 0 and taxes[t] credits income",
            series=refunds, units=ctx.units_label()))
    statutory = find_scalar_assumptions(ctx.wbd, ctx.structure, r"tax\s*rate", (0.0, 0.6))
    if statutory:
        ref, v, label = statutory[0]
        out.metrics.append(mk(
            "statutory_tax_rate_typed", f"Typed tax-rate assumption ({label!r})",
            applicability="typed tax rate found", inputs=[ref],
            computation="typed input", scalar=v, units="fraction"))


def _headcount(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    hc_item = ctx.item("headcount")
    hc = ctx.series("headcount")
    rev = ctx.series("revenue_total")
    scale = ctx.structure.units.get(ctx.sheet)
    unit_scale = scale.scale if scale else 1.0
    money = scale.label if scale else "model units"

    # The model's OWN headcount trajectory is the sanity-check an analyst reads
    # directly (e.g. 809 -> 1,461 FTEs); surface it rather than only a recompute.
    if hc:
        out.metrics.append(mk(
            "headcount_trajectory", "Headcount (the model's own FTE row)",
            applicability="headcount row mapped", inputs=ctx.refs("headcount"),
            computation="model's headcount row, as-is",
            series={p: hc.get(p) for p in P}, units="FTEs",
            notes="does the org size the model implies match the story?"))
    if hc and rev:
        # revenue is in `money` scale; report $ per FTE in whole currency
        out.metrics.append(mk(
            "revenue_per_employee", "Revenue per employee",
            applicability="headcount mapped",
            inputs=ctx.refs("headcount", "revenue_total"),
            computation=f"revenue_total[t] * {unit_scale:g} / headcount[t]",
            series=combine(P, lambda r, h: r * unit_scale / h if h and h > EPS else None, rev, hc),
            units="whole currency per FTE",
            notes=f"revenue scale: {money}"))

    # Implied FTEs from a typed cost-per-head — ONLY when the model has no
    # headcount row of its own (otherwise it just noisily recomputes a mapped
    # row, and the unit-scale mismatch made it wrong by ~1000x).
    if hc:
        return
    per_head = find_scalar_assumptions(
        ctx.wbd, ctx.structure,
        r"per\s+(corporate\s+)?(fte|head|employee)|cost\s+per\s+(fte|head|employee)|all.?in.*fte", (0.5, 1e9))
    labor = [it for it in ctx.items("opex_component") if "labor" in it.subtags]
    opex = ctx.series("opex_total")
    if per_head and (labor or opex):
        ref, v, label = per_head[0]
        monthly = "month" in label.lower() or "/mo" in label.lower()
        annual_cost = v * 12.0 if monthly else v
        # reconcile units: opex is in `money` scale, the typed cost-per-head is
        # whole currency — bring opex to whole currency before dividing
        base: dict[str, Optional[float]] = {}
        if labor:
            for p in P:
                vals = [it.value_for(p) for it in labor]
                vals = [x for x in vals if x is not None]
                base[p] = sum(vals) if vals else None
            base_refs = [r for it in labor for r in it.refs][:40]
            base_name = "labor opex rows"
        else:
            base = opex or {}
            base_refs = ctx.refs("opex_total")
            base_name = "opex_total (no labor rows mapped)"
        implied = combine(P, lambda b: abs(b) * unit_scale / annual_cost if annual_cost > EPS else None, base)
        out.metrics.append(mk(
            "implied_fte_from_cost_per_head", f"Implied FTEs from {label!r}",
            applicability="typed cost-per-head found and no headcount row mapped",
            inputs=[ref] + base_refs,
            computation=f"{base_name}[t] * {unit_scale:g} / {annual_cost:g} (annualized cost per head)",
            series=implied, units="FTEs",
            notes="sanity-check the implied organization size against the story"))


def _working_capital(ctx: Ctx, out: MetricsResult) -> None:
    P = ctx.periods
    present = [cid for cid in ("accounts_receivable", "accounts_payable", "inventory")
               if ctx.mapping.get(cid) is not None]
    out.metrics.append(mk(
        "working_capital_presence", "Working-capital rows present in the model",
        applicability="always computed",
        inputs=[r for cid in present for r in ctx.refs(cid)][:30],
        computation="count of AR/AP/inventory rows mapped",
        scalar=float(len(present)), units="count (0-3)",
        notes=("model carries " + ", ".join(present)) if present else
              "no AR/AP/inventory anywhere: cash-basis model with instant collection"))
    rev = ctx.series("revenue_total")

    # When AR/AP/inventory ARE present, compute the efficiency ratios the
    # analyst runs (Sample: AR/Revenue, AR/Beginning Cash).
    # AR/inventory ratios are only emitted when the source carries real data
    # (an all-zero / all-blank line is a conclusion-shaped row with nothing
    # behind it). The interpretive caption is conditional on the actual pattern.
    ar = ctx.series("accounts_receivable")
    ar_item = ctx.mapping.get("accounts_receivable")
    ar_switched_off = bool(ar) and not is_supported(ar, P) and pre_horizon_nonzero(ctx.wbd, ar_item) is not None
    if ar and rev and (is_supported(ar, P) or ar_switched_off):
        arr = ratio(ar, rev, P)
        if ar_switched_off:
            hv = pre_horizon_nonzero(ctx.wbd, ar_item)
            note = (f"receivables were ~{hv:,.0f} pre-forecast but are held at ZERO across the plan — "
                    "collection is assumed instant; working capital does not build with revenue")
        else:
            note = ("implied days sales outstanding — rising, tying up more cash in receivables"
                    if _rising(arr, P) else "implied days sales outstanding (broadly flat)")
        cov = coverage_note(ar, P)
        out.metrics.append(mk(
            "ar_over_revenue", "Accounts receivable / revenue",
            applicability="accounts_receivable and revenue mapped",
            inputs=ctx.refs("accounts_receivable", "revenue_total"),
            computation="accounts_receivable[t] / revenue_total[t]",
            series=arr, units="fraction",
            notes=note + ("; " + cov if cov else "")))
        beg = ctx.series("beginning_cash")
        if beg and is_supported(beg, P):
            out.metrics.append(mk(
                "ar_over_beginning_cash", "Accounts receivable / beginning cash",
                applicability="accounts_receivable and beginning_cash mapped",
                inputs=ctx.refs("accounts_receivable", "beginning_cash"),
                computation="accounts_receivable[t] / beginning_cash[t]",
                series=ratio(ar, beg, P), units="fraction",
                notes="receivables as a share of the cash on hand — liquidity strain if it climbs"))
    inv = ctx.series("inventory")
    if inv and rev and is_supported(inv, P):
        invr = ratio(inv, rev, P)
        note = ("inventory intensity is rising — can signal demand softening or overbuild"
                if _rising(invr, P) else "inventory intensity is flat/declining")
        cov = coverage_note(inv, P)
        out.metrics.append(mk(
            "inventory_over_revenue", "Inventory / revenue",
            applicability="inventory and revenue mapped",
            inputs=ctx.refs("inventory", "revenue_total"),
            computation="inventory[t] / revenue_total[t]",
            series=invr, units="fraction",
            notes=note + ("; " + cov if cov else "")))

    if rev and "accounts_receivable" not in present:
        days_list = (ctx.benchmarks or {}).get("receivables_scenario_days") or [30, 60]
        peak_rev = terminal_scale(rev, P)
        ppy = 12.0 / ctx.months_per_period
        for days in days_list[:2]:
            # annualized revenue basis: for sub-annual periods scale to a year
            annual_peak = peak_rev * ppy
            need = annual_peak * float(days) / 365.0
            out.metrics.append(mk(
                f"receivables_scenario_{int(days)}d", f"Scenario: cash tied up at {int(days)}-day receivables",
                applicability="model books revenue as instantly collected",
                inputs=ctx.refs("revenue_total"),
                computation=f"peak annualized revenue * {int(days)}/365 (tool scenario on model numbers)",
                scalar=need, units=ctx.units_label(),
                notes="tool-computed scenario, not a model value: extra peak cash need if customers paid this slowly"))


def _ebitda_ni_bridge(ctx: Ctx, out: MetricsResult) -> None:
    """P1-2: when both an EBITDA/operating-profit line and a net-income line exist,
    reconcile them. A large gap — or the wrong direction (net income ABOVE EBITDA) —
    means below-the-line items (interest, non-operating income, credits/grants)
    dominate the result, which is the single most important thing on a margins page
    and must not pass silently. Names the largest reconciling line for the worst
    period. Generic: keyed to the detected condition, no cell addresses."""
    eb = ctx.item("ebitda_or_ebit")
    ni = ctx.item("net_income")
    ebs = ctx.series("ebitda_or_ebit")
    nis = ctx.series("net_income")
    if not (eb and ni and ebs and nis):
        return
    P = ctx.periods
    rev = ctx.series("revenue_total")
    gap = combine(P, lambda n, e: (n - e) if (n is not None and e is not None) else None, nis, ebs)

    def material(p: str) -> bool:
        g, n, e = gap.get(p), nis.get(p), ebs.get(p)
        if g is None:
            return False
        if n is not None and e is not None and n > e + EPS:      # net income above EBITDA
            return True
        rv = rev.get(p) if rev else None
        return rv is not None and abs(rv) > EPS and abs(g) > 0.20 * abs(rv)

    flagged = [p for p in P if material(p)]
    if not flagged:
        return
    wrong_dir = any((nis.get(p) is not None and ebs.get(p) is not None and nis.get(p) > ebs.get(p) + EPS)
                    for p in flagged)
    driver_note = ""
    if eb.sheet == ni.sheet:                                      # scan the reconciling rows in between
        sd = ctx.wbd.sheets.get(eb.sheet)
        pworst = max(flagged, key=lambda p: abs(gap.get(p) or 0.0))
        col = next((c for (r, c), pp in zip(eb.coords, eb.periods) if pp == pworst), None)
        if sd is not None and col is not None:
            lo, hi = sorted((eb.index, ni.index))
            best_lab, best_val = None, 0.0
            for r in range(lo + 1, hi):
                rec = sd.cell(r, col)
                v = rec.value if rec else None
                if isinstance(v, (int, float)) and not isinstance(v, bool) and abs(v) > abs(best_val):
                    lab = next((sd.cell(r, cc).value.strip()[:40] for cc in range(1, col)
                                if sd.cell(r, cc) and isinstance(sd.cell(r, cc).value, str)
                                and sd.cell(r, cc).value.strip()), None)
                    if lab:
                        best_lab, best_val = lab, v
            if best_lab:
                rv = rev.get(pworst) if rev else None
                pct = f", {abs(best_val) / abs(rv) * 100:.0f}% of {pworst} revenue" if rv and abs(rv) > EPS else ""
                driver_note = f"; largest below-the-line item in {pworst}: {best_lab!r} = {best_val:.4g}{pct}"
    head = ("net income EXCEEDS EBITDA (wrong direction) — " if wrong_dir else "")
    out.metrics.append(mk(
        "ebitda_ni_bridge", "EBITDA-to-net-income gap",
        applicability="EBITDA/EBIT and net income mapped",
        inputs=ctx.refs("ebitda_or_ebit", "net_income"),
        computation="net_income[t] - ebitda_or_ebit[t]",
        series=gap, units=ctx.units_label(),
        notes=(head + "a large or wrong-direction gap means below-the-line items (interest, "
               "non-operating income, credits/grants) drive net income" + driver_note)))


BUILDERS = [_core_growth_and_margins, _capital_efficiency, _runway, _cumulative_capital,
            _grants, _capex_in_opex, _revenue_mix, _segments, _segment_detail, _labor,
            _capacity_and_share, _valuation, _taxes, _headcount, _working_capital,
            _ebitda_ni_bridge]


def compute_metrics(
    wbd: WorkbookData,
    structure: StructureResult,
    mapping: MappingResult,
    benchmarks: Optional[dict] = None,
) -> MetricsResult:
    out = MetricsResult()
    sheet = mapping.primary_sheet
    if sheet is None:
        out.notes.append("No mapped statement sheet; no metrics computable.")
        return out
    axis = structure.primary_axis(sheet)
    if axis is None:
        out.notes.append(f"Primary sheet {sheet!r} has no period axis; no metrics computable.")
        return out
    months = {Granularity.annual: 12.0, Granularity.quarterly: 3.0,
              Granularity.monthly: 1.0}.get(axis.granularity, 12.0)
    ctx = Ctx(wbd=wbd, structure=structure, mapping=mapping, benchmarks=benchmarks or {},
              sheet=sheet, periods=list(axis.periods), months_per_period=months)
    for builder in BUILDERS:
        try:
            builder(ctx, out)
        except Exception:
            log.exception("metric builder %s failed", builder.__name__)
            out.notes.append(f"metric builder {builder.__name__} failed; its metrics were skipped")
    return out

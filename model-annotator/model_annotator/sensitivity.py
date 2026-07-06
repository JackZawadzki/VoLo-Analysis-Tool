"""Sensitivity — built off the model's OWN input cells, presented the way the
DD financials tool presents it: a Shapley contribution bar (how much each input
drives the base→downside move) plus a two-way grid the analyst can drive
interactively (any two inputs at once, red→green).

Faithful and engine-free: each driver is a real model cell, each output is a
quantity the tool owns the closed form for, so the recompute is exact. To
"include every input that could affect the output", the income-statement output
is decomposed all the way to the model's line items — every mapped revenue
segment, COGS component, and opex component becomes its own driver, not just the
three totals.

Outputs:
  - linear_sum  : output = Σ coef·driver  (EBITDA = Σ segments − Σ COGS − Σ opex)
  - valuation_pv: PV = multiple × EBITDA_exit / (1+rate)^horizon
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from itertools import combinations
from math import factorial
from typing import Optional

from .ingest import is_number
from .mapping import LineItem, MappingResult
from .metrics import find_scalar_assumptions
from .schema import CellKind, SensitivityDriver, Tornado
from .structure import StructureResult

DEFAULT_PCT = 0.20
MAX_DRIVERS = 12            # keep the chart legible — show the inputs that move it most
MAX_RECOMPUTE_CONE = 50_000  # above this, per-input recompute is too slow → line-item flex


@dataclass
class _Spec:
    key: str
    label: str
    refs: list[str]
    base: float
    coef: float                 # +1 adds to output, −1 subtracts (linear_sum)
    unit: str = ""
    low_pct: float = -DEFAULT_PCT
    high_pct: float = DEFAULT_PCT
    low_val: Optional[float] = None
    high_val: Optional[float] = None
    # which direction is "adverse" (lowers the output): -1 means low value is adverse
    adverse: str = "low"        # "low" | "high"

    def low(self) -> float:
        return self.low_val if self.low_val is not None else self.base * (1 + self.low_pct)

    def high(self) -> float:
        return self.high_val if self.high_val is not None else self.base * (1 + self.high_pct)

    def adverse_val(self) -> float:
        return self.low() if self.adverse == "low" else self.high()

    def favorable_val(self) -> float:
        return self.high() if self.adverse == "low" else self.low()


def _terminal(it: Optional[LineItem]):
    if it is None:
        return None
    for p in reversed(it.periods):
        v = it.value_for(p)
        if v is not None:
            return p, float(v), it.ref_for(p)
    return None


def _terminal_for(it: LineItem, period: str):
    v = it.value_for(period)
    return (period, float(v), it.ref_for(period)) if v is not None else None


# ---------------------------------------------------------------------------
# Output evaluators
# ---------------------------------------------------------------------------

def _eval_linear(specs: list[_Spec], overrides: dict[str, float], offset: float = 0.0) -> float:
    return offset + sum(s.coef * overrides.get(s.key, s.base) for s in specs)


def _eval_pv(specs_by_key: dict[str, _Spec], overrides: dict[str, float], horizon: float) -> float:
    def g(k):
        return overrides.get(k, specs_by_key[k].base) if k in specs_by_key else 0.0
    rate = g("rate") if "rate" in specs_by_key else 0.0
    return g("multiple") * g("metric") / ((1 + rate) ** horizon)


# ---------------------------------------------------------------------------
# Shapley contribution of each driver to the base→downside move
# ---------------------------------------------------------------------------

def _shapley_linear(specs: list[_Spec]) -> dict[str, float]:
    # linear output has no interactions: contribution = coef·(adverse − base)
    return {s.key: s.coef * (s.adverse_val() - s.base) for s in specs}


def _shapley_exact(specs: list[_Spec], eval_fn) -> dict[str, float]:
    keys = [s.key for s in specs]
    base = {s.key: s.base for s in specs}
    adv = {s.key: s.adverse_val() for s in specs}
    cache: dict[frozenset, float] = {}

    def f(S: frozenset) -> float:
        if S in cache:
            return cache[S]
        ov = dict(base)
        for k in S:
            ov[k] = adv[k]
        val = eval_fn(ov)
        cache[S] = val
        return val

    n = len(keys)
    phi = {k: 0.0 for k in keys}
    for k in keys:
        rest = [x for x in keys if x != k]
        for r in range(len(rest) + 1):
            w = factorial(r) * factorial(n - r - 1) / factorial(n)
            for S in combinations(rest, r):
                phi[k] += w * (f(frozenset(S) | {k}) - f(frozenset(S)))
    return phi


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def _mk_driver(s: _Spec, output_low: float, output_high: float, shapley: float) -> SensitivityDriver:
    lo, hi = s.low(), s.high()
    base = s.base
    lp = (lo - base) / base if abs(base) > 1e-9 else s.low_pct
    hp = (hi - base) / base if abs(base) > 1e-9 else s.high_pct
    return SensitivityDriver(
        key=s.key, label=s.label, input_refs=s.refs, base=base, low=lo, high=hi,
        low_pct=round(lp, 4), high_pct=round(hp, 4),
        output_low=output_low, output_high=output_high, swing=abs(output_high - output_low),
        shapley=shapley, coef=s.coef, unit=s.unit)


def _components(ctx_items, sheet, period, coef, key_prefix, unit) -> list[_Spec]:
    specs = []
    for it in ctx_items:
        if it.sheet != sheet:
            continue
        t = _terminal_for(it, period)
        if t is None:
            continue
        _, v, ref = t
        inst = (it.instance or it.label.strip())[:32]
        specs.append(_Spec(key=f"{key_prefix}:{it.index}", label=f"{inst} ({period})",
                           refs=[ref], base=abs(v) if coef < 0 else v, coef=coef, unit=unit,
                           adverse="high" if coef < 0 else "low"))
    return specs


def _node_for_item(graph, item, period):
    """(graph node, value) for an item's terminal-period cell, or (None, None)."""
    t = _terminal_for(item, period)
    if not t:
        return None, None
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    from .graph import encode
    ref = t[2]
    sh = ref.split("!")[0].strip("'")
    if sh not in graph.sheet_index:
        return None, None
    col_l, row = coordinate_from_string(ref.split("!")[1])
    return encode(graph.sheet_index[sh], row, column_index_from_string(col_l)), t[1]


def _row_label(sd, row: int) -> str:
    for c in range(1, min(sd.max_col, 12) + 1):
        rec = sd.cell(row, c)
        if rec is not None and isinstance(rec.value, str) and rec.value.strip():
            return rec.value.strip()[:38]
    return f"row {row}"


def _sheet_bands(structure, sheet: str, max_col: int) -> list[tuple[int, int]]:
    """Horizontal table bands on a sheet, as (col_lo, col_hi) column ranges,
    derived from the detected period axes' spans. Sheets that stack two tables
    side-by-side (a blank-column gap between their period headers) yield TWO
    bands; ordinary sheets yield ONE band spanning the whole row (identical to
    the previous whole-row behaviour). This lets a driver's input cells be kept
    inside the table its label heads, instead of bleeding into a neighbouring
    table that merely shares the physical row."""
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    spans: list[tuple[int, int]] = []
    for ax in (structure.axes.get(sheet) or []):
        cols = []
        for ref in ax.header_cells:
            try:
                cl, _r = coordinate_from_string(ref.split("!")[-1])
                cols.append(column_index_from_string(cl))
            except Exception:
                continue
        if cols:
            spans.append((min(cols), max(cols)))
    if not spans:
        return [(1, max_col)]
    spans.sort()
    merged: list[list[int]] = [list(spans[0])]
    for lo, hi in spans[1:]:
        if lo <= merged[-1][1] + 1:                 # overlapping / adjacent -> one table
            merged[-1][1] = max(merged[-1][1], hi)
        else:
            merged.append([lo, hi])
    if len(merged) <= 1:                            # single table -> whole-row scan (unchanged)
        return [(1, max_col)]
    return [(lo, hi) for lo, hi in merged]


def _aggregate_groups(sd, blo: int, bhi: int, leaf_rows: list[int]) -> list[tuple[int, list[int]]]:
    """Roll leaf input rows up to the model's own total. Returns (total_row,
    member_rows) for each row whose per-band-period values equal the SUM of >=2 of
    the given leaf rows sitting just above it (blank rows tolerated) -- e.g. the
    per-segment 'new clients' rows summing into 'TOTAL New clients'. Sum-based, so
    it needs no naming convention; a row is never grouped into more than one total."""
    cols = list(range(blo, bhi + 1))

    def vec(r):
        out = []
        for c in cols:
            rec = sd.cell(r, c)
            v = rec.value if rec is not None else None
            out.append(v if isinstance(v, (int, float)) and not isinstance(v, bool) else None)
        return out
    leafset = sorted(set(leaf_rows))
    if len(leafset) < 2:
        return []
    groups: list[tuple[int, list[int]]] = []
    used: set[int] = set()
    for A in range(min(leafset) + 1, max(leafset) + 12):
        if A in leafset:                       # a total is not itself a leaf input
            continue
        av = vec(A)
        if not any(x not in (None, 0) for x in av):
            continue
        members = [r for r in leafset if 0 < A - r <= 12 and r not in used]
        if len(members) < 2:
            continue
        s = [0.0] * len(cols)
        for m in members:
            for i, x in enumerate(vec(m)):
                s[i] += (x or 0)
        present = [i for i in range(len(cols)) if av[i] is not None]
        if present and max(abs((av[i] or 0) - s[i]) for i in present) <= \
                0.02 * max(max((abs(av[i]) for i in present), default=1.0), 1.0):
            groups.append((A, members))
            used.update(members)
    return groups


def _band_label(sd, row: int, band_lo: int) -> str:
    """Label for a driver whose data sits in the band starting at band_lo: the
    nearest non-empty text cell to the LEFT of the band on this row, so a
    second-table driver is named from its OWN table's label (not the leftmost
    label, which belongs to the first table). Falls back to the row label."""
    if band_lo > 1:
        for c in range(band_lo - 1, 0, -1):
            rec = sd.cell(row, c)
            if rec is not None and isinstance(rec.value, str) and rec.value.strip():
                return rec.value.strip()[:38]
    return _row_label(sd, row)


def _section_label(sd, blo: int, bhi: int, row: int) -> Optional[str]:
    """The section header above a data row in band [blo,bhi]: the nearest row
    above with a text label (left of the band) but NO numeric data IN the band —
    i.e. a title row like 'Licensing (royalty per unit)' or 'Units x client'.
    Lets per-segment rows be grouped by the family their modeller put them under."""
    for rr in range(row - 1, max(0, row - 30), -1):
        lab = None
        for c in range(max(1, blo - 1), 0, -1):
            rec = sd.cell(rr, c)
            if rec is not None and isinstance(rec.value, str) and rec.value.strip():
                lab = rec.value.strip()
                break
        if not lab:
            continue
        has_num = False
        for c in range(blo, bhi + 1):
            rec = sd.cell(rr, c)
            if rec is not None and isinstance(rec.value, (int, float)) and not isinstance(rec.value, bool):
                has_num = True
                break
        if not has_num:                 # text + no numbers in the band -> a title row
            return lab
    return None


_YEAR_STR = None  # compiled lazily below


def _is_yearish(rec) -> bool:
    """A header-evidence cell: a year (1990-2100), a date, or a date-formatted
    number. Years/dates NEVER count as 'data in the band' when testing whether a
    row is a title/header row — a column-header row of 2026/2027/2028 IS the
    header, not data."""
    if rec is None:
        return False
    v = rec.value
    if hasattr(v, "year"):                       # datetime / date
        return True
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if 1990 <= v <= 2100 and abs(v - round(v)) < 1e-9:
            return True
        core = (rec.number_format or "").split(";")[0]
        import re as _r
        if _r.search(r"[ymd]", _r.sub(r"\[[^\]]*\]|\"[^\"]*\"", "", core), _r.I) and "%" not in core:
            return True
    if isinstance(v, str):
        s = v.strip()
        return len(s) == 4 and s.isdigit() and 1990 <= int(s) <= 2100
    return False


def _table_header(sd, blo, bhi, row, max_up=40):
    """The header row governing a data row, and everything derivable from it.

    Real input decks look like:  [title: 'Pricing']
                                 [banner: 'Base Case'   'Upside'   'Downside']
                                 [header: Customer  Unit  2026 2027 2028 | Customer ...]
                                 [data:   Anduril   KWh    45   45   45  | ...]
    The old nearest-left-text label grabbed unit vocab ('KWh'), banners ('Base
    Case') or entity headers ('Customer') — vague or plain wrong names. Instead:
    find the nearest row above with year/date cells over the band (the column-
    header row), type its label-side cells (entity / unit-of-measure columns),
    detect repeated header text as scenario-block starts (named by the banner row
    above), and take the MEASURE from the leftmost non-entity/non-unit header
    text — else from the nearest pure title row above. Returns a dict or None."""
    import re as _r
    from collections import Counter

    def _texts_on(rr, hi):
        out = []
        for c in range(1, hi + 1):
            rec = sd.cell(rr, c)
            if rec is not None and isinstance(rec.value, str) and rec.value.strip():
                out.append((c, rec.value.strip()))
        return out

    def _title_above(rr0):
        for rr in range(rr0 - 1, max(0, rr0 - 12), -1):
            txts = _texts_on(rr, bhi)
            if not txts:
                continue
            has_num = False
            for c in range(blo, bhi + 1):
                rec = sd.cell(rr, c)
                if rec is not None and isinstance(rec.value, (int, float)) \
                        and not isinstance(rec.value, bool) and not _is_yearish(rec):
                    has_num = True
                    break
            if not has_num and txts[0][0] < blo:
                cnt = Counter(t.lower() for _c, t in txts)
                if max(cnt.values()) == 1:           # banners repeat; titles don't
                    return txts[0][1]
        return None

    for rr in range(row - 1, max(0, row - max_up), -1):
        year_cols = [c for c in range(blo, bhi + 1) if _is_yearish(sd.cell(rr, c))]
        if not year_cols:
            continue
        texts = _texts_on(rr, bhi)
        cnt = Counter(t.lower() for _c, t in texts)
        entity_col = unit_col = None
        measure = None
        block_starts = sorted(c for c, t in texts if cnt[t.lower()] >= 2)
        for c, t in texts:
            if c >= blo and c not in year_cols and cnt[t.lower()] < 2:
                continue
            if _r.search(r"customer|client|segment|entity|product|name", t, _r.I) and entity_col is None:
                entity_col = c
            elif _r.match(r"units?\b", t, _r.I) and unit_col is None:
                unit_col = c
            elif measure is None and cnt[t.lower()] < 2 and c < blo:
                measure = t
        if measure is None:
            measure = _title_above(rr)
        blocks = []
        if len(block_starts) >= 2:
            for i, bs in enumerate(block_starts):
                be = block_starts[i + 1] - 1 if i + 1 < len(block_starts) else 10 ** 6
                name = None
                for rb in (rr - 1, rr - 2):
                    if rb < 1:
                        break
                    for c in range(bs, max(0, bs - 4), -1):
                        recb = sd.cell(rb, c)
                        if recb is not None and isinstance(recb.value, str) and recb.value.strip():
                            name = recb.value.strip()
                            break
                    if name:
                        break
                blocks.append(dict(lo=bs, hi=be, name=name or f"block {i + 1}"))
        years = {}
        for c in range(blo, bhi + 1):
            recy = sd.cell(rr, c)
            if recy is not None and _is_yearish(recy):
                v = recy.value
                try:
                    years[c] = str(v.year) if hasattr(v, "year") else str(int(float(str(v))))
                except Exception:
                    pass
        return dict(row=rr, measure=measure, entity_col=entity_col,
                    unit_col=unit_col, blocks=blocks, years=years)
    return None


def _clean_label(s: str) -> str:
    """Tidy a raw row/section label into a readable driver name: drop trailing
    unit notes (€/unit, /client), AVERAGE/TOTAL noise, and surrounding clutter."""
    import re as _r
    out = _r.sub(r"\s*\(?\s*€?\s*/\s*(unit|client|person|day|t|head|month|year)\b.*$", "", s, flags=_r.I)
    out = _r.sub(r"\s*€\s*/\s*\w+\s*$", "", out)
    out = _r.sub(r"\s+(?:AVERAGE|AVG)\s*$", "", out, flags=_r.I)   # trailing 'AVERAGE' noise only
    out = _r.sub(r"\s{2,}", " ", out).strip(" -—:·,")
    return (out or s.strip())[:40]


def _bracket_range(lo: float, hi: float, base: float) -> tuple[float, float]:
    """Guarantee a displayed sensitivity range BRACKETS its base and has non-zero
    width. The front-end slope is (outHigh-outLow)/(high-low); a range that clamped
    to one side of the base (a rate whose base sits outside its [0,1] box, a mixed
    aggregate) or collapsed to zero width would corrupt the live re-ranking. When
    the type-aware range violates this, fall back to a plain ±20% of the base,
    which always brackets it."""
    tol = 1e-9 * max(1.0, abs(base))
    if abs(hi - lo) < tol or not (min(lo, hi) - tol <= base <= max(lo, hi) + tol):
        return min(base * 0.8, base * 1.2), max(base * 0.8, base * 1.2)
    return (lo, hi) if lo <= hi else (hi, lo)


def _recompute_ebitda_tornado(wbd, structure, mapping, graph, U, periods) -> Optional[Tornado]:
    """Flex the model's REAL input cells and recompute EVERY mapped output at
    EVERY period exactly through the workbook's own formulas, so the UI can let
    the analyst pick which output/year to view. Returns None when the model can't
    be faithfully reproduced (the caller then uses the line-item flex)."""
    if graph is None:
        return None
    from .graph import encode
    from .recompute import RecomputeModel
    sheet = mapping.primary_sheet
    rev_item = mapping.get("revenue_total", sheet)
    eb_item = mapping.get("ebitda_or_ebit", sheet)
    if rev_item is None or eb_item is None:
        return None
    t = _terminal(rev_item)
    tp = t[0] if t else None
    if tp is None:
        return None
    eb_node, _ = _node_for_item(graph, eb_item, tp)
    rev_node, _ = _node_for_item(graph, rev_item, tp)
    if eb_node is None or rev_node is None:
        return None

    # Exact recompute costs ~O(model size) PER flexed input. On a very large model
    # (e.g. a 100k+-cell project-finance workbook) flexing dozens of inputs is
    # impractically slow, so fall back to the line-item flex there rather than
    # hang. Small/medium models recompute exactly.
    up = graph.upstream(eb_node, cap=400_000) | graph.upstream(rev_node, cap=400_000)
    if len(up) > MAX_RECOMPUTE_CONE:
        return None

    # outputs the analyst can view the change in (whichever are mapped), richest
    # first; the per-period cell of each is recomputed exactly.
    # ending_cash included: financing assumptions (raise sizes/timing) only move
    # CASH, and the user needs to see funding levers alongside revenue/cost ones
    want_full = [("ebitda", "EBITDA", eb_item), ("revenue", "Revenue", rev_item)]
    for okey, label, cid in [("gross_profit", "Gross profit", "gross_profit"),
                             ("net_income", "Net income", "net_income"),
                             ("cash", "Ending cash", "ending_cash")]:
        it = mapping.get(cid, sheet)
        if it is not None:
            want_full.append((okey, label, it))

    def _out_nodes(specs):
        nodes = {}
        for okey, _label, it in specs:
            for p in periods:
                nd, _v = _node_for_item(graph, it, p)
                if nd is not None:
                    nodes[(okey, p)] = nd
        return nodes

    model = None
    out_specs = None
    out_nodes = {}
    # try all outputs, then peel off the fragile below-EBITDA spines one at a
    # time (cash first, then net income — their cones often carry the workbook's
    # own error cells) so a failure there doesn't also cost gross profit
    ladders = [want_full,
               [s for s in want_full if s[0] != "cash"],
               [s for s in want_full if s[0] not in ("cash", "net_income")],
               want_full[:2]]
    seen_ladders = set()
    for specs in ladders:
        key = tuple(s[0] for s in specs)
        if key in seen_ladders:
            continue
        seen_ladders.add(key)
        out_nodes = _out_nodes(specs)
        check = list(set(out_nodes.values())) or [eb_node, rev_node]
        # per-output faithfulness: gate on EVERY requested output's terminal cell
        # (not just EBITDA+revenue), so a cube never shows an output the engine
        # can't reproduce — an unfaithful spine peels off via the ladder instead
        gates = [out_nodes.get((s[0], tp)) for s in specs]
        gates = [g for g in gates if g is not None] or [eb_node, rev_node]
        model = RecomputeModel.build(graph, check, gate_nodes=gates)
        if model is not None:
            out_specs = [s for s in specs if any((s[0], p) in out_nodes for p in periods)]
            break
    if model is None or not out_specs:
        return None
    all_nodes = list(set(out_nodes.values()))
    eb_base = model.values.get(eb_node)
    rev_base = model.values.get(rev_node)

    # ---- genuine-input discovery -------------------------------------------
    # A driver must be an ASSUMPTION a human typed (price, volume, cost rate,
    # tax, financing) that feeds a key output — never a unit-scale factor, enum
    # marker, year, or subtotal echo. Candidates span ALL columns (assumptions
    # commonly sit outside period-header bands: a lone value in column C), are
    # scored for genuineness, typed for context-appropriate ranges, and measured
    # by exact recompute — the measured swing does the ranking, not a fan-out
    # proxy. Generic: keyed on structure, labels and the dependency graph only.
    import re as _re
    from .input_discovery import (classify_type, ranged_vals, genuineness_score,
                                  is_enum_key, sheet_is_assumptionish)
    _COST = _re.compile(r"cost|personnel|payroll|opex|sg&a|r&d|salar|capex|overhead|wage", _re.I)

    def _clean_sheet(name):
        s = _re.sub(r"^(COSTS?-|M-)\s*", "", name).strip()
        return (s[:1].upper() + s[1:]) if s else name

    # one upstream cone per mapped output: financing levers only reach CASH and
    # tax levers only reach NET INCOME — an EBITDA∪revenue cone can't see them
    rank_nodes: dict[str, int] = {}
    for okey, _l, _it in out_specs:
        nd = out_nodes.get((okey, tp))
        if nd is not None:
            rank_nodes[okey] = nd
    cones = {okey: graph.upstream(nd, cap=400_000) for okey, nd in rank_nodes.items()}
    up_all = set().union(*cones.values()) if cones else up
    rev_cone = cones.get("revenue", set())
    eb_cone = cones.get("ebitda", set())
    conv_fill = getattr(structure.input_convention, "fill_key", None)

    sheet_leaves: dict[str, list[dict]] = {}
    n_cands = 0
    for sn, sd in wbd.sheets.items():
        cen = structure.censuses.get(sn)
        si = graph.sheet_index.get(sn)
        if cen is None or si is None:
            continue
        role = structure.roles.get(sn)
        role_v = role[0].value if role else None
        if role_v in ("documentation", "presentation", "scratch"):
            continue
        assumish = sheet_is_assumptionish(sn, role_v)
        # curve detection needs the FINEST axis on the sheet — the primary axis is
        # deliberately the annual one, but deployment ramps live on the monthly
        # (or quarterly) columns alongside it
        grans = {getattr(getattr(ax, "granularity", None), "value", "annual")
                 for ax in structure.axes.get(sn, [])} or {"annual"}
        gran = "monthly" if "monthly" in grans else ("quarterly" if "quarterly" in grans else "annual")
        shift_k = {"monthly": 6, "quarterly": 2}.get(gran, 1)   # columns ≈ 6 months
        for r in range(1, sd.max_row + 1):
            cells, zcells, fmt, fill_hit = [], [], "", False
            for c in range(1, sd.max_col + 1):
                if cen.kind(r, c) != CellKind.typed_input:
                    continue
                rec = sd.cell(r, c)
                if rec is None or not is_number(rec.value):
                    continue
                nd = encode(si, r, c)
                if nd not in up_all:
                    continue
                if rec.value == 0:
                    zcells.append((nd, c, 0.0))        # zero cells: shift targets only
                    continue
                cells.append((nd, c, float(rec.value)))
                fmt = fmt or (rec.number_format or "")
                fill_hit = fill_hit or (conv_fill is not None and rec.fill_key == conv_fill)
            if not cells:
                continue
            cells.sort(key=lambda x: x[1])
            v0 = cells[-1][2]
            blo, bhi = cells[0][1], cells[-1][1]

            # header recipe: measure / entity / unit-of-measure / scenario blocks
            hdr = _table_header(sd, blo, bhi, r)
            entity = unit_val = None
            block_name = None
            if hdr:
                if hdr.get("entity_col"):
                    rec_e = sd.cell(r, hdr["entity_col"])
                    if rec_e is not None and isinstance(rec_e.value, str) and rec_e.value.strip():
                        entity = rec_e.value.strip()[:34]
                if hdr.get("unit_col"):
                    rec_u = sd.cell(r, hdr["unit_col"])
                    if rec_u is not None and rec_u.value is not None:
                        unit_val = str(rec_u.value).strip()[:16]
                # scenario blocks: flex ONLY the live block — the inert Upside/
                # Downside columns don't feed the selected scenario, and summing
                # them corrupts the displayed base
                if hdr.get("blocks"):
                    def _blk_of(c):
                        for b in hdr["blocks"]:
                            if b["lo"] <= c <= b["hi"]:
                                return b
                        return None
                    per_blk: dict[int, list] = {}
                    for x in cells:
                        b = _blk_of(x[1])
                        per_blk.setdefault(id(b) if b else -1, []).append((b, x))
                    def _blk_score(entries):
                        b = entries[0][0]
                        pref = 2 if (b and _re.search(r"base", b["name"], _re.I)) else 0
                        live = sum(1 for _b, (nd, _c, _v) in entries if nd in up_all)
                        return (pref, live, len(entries))
                    best = max(per_blk.values(), key=_blk_score)
                    b0 = best[0][0]
                    if b0 is not None:
                        block_name = b0["name"][:20]
                        cells = sorted((x for _b, x in best), key=lambda x: x[1])
                        zcells = [z for z in zcells if b0["lo"] <= z[1] <= b0["hi"]]
                        v0 = cells[-1][2]
                        blo, bhi = cells[0][1], cells[-1][1]

            measure = (hdr or {}).get("measure")
            raw_label = _band_label(sd, r, blo)
            if measure and entity:
                label = f"{measure} — {entity}"
            elif measure:
                lbl2 = _clean_label(raw_label)
                label = measure if lbl2.lower() in (measure.lower(), "") else f"{measure} — {lbl2}"
            else:
                # no header table found: fall back to the section title for
                # context so a bare product/row name still reads on a tornado
                label = raw_label
                sec = _section_label(sd, blo, bhi, r)
                if sec and _clean_label(sec).lower() != _clean_label(raw_label).lower():
                    label = f"{_clean_label(raw_label)} ({_clean_label(sec)[:24]})"

            in_rev = any(nd in rev_cone for nd, _c, _v in cells)
            in_eb = any(nd in eb_cone for nd, _c, _v in cells)
            score = genuineness_score(label, fmt, v0, True, in_rev and in_eb,
                                      in_rev and not in_eb, fill_hit, assumish)
            if score <= 0:
                continue                               # scale factors, years, placeholders sink here
            if is_enum_key(graph, wbd, cells[-1][0], v0):
                continue                               # scenario switches / unit codes: not quantities

            # a normalized phasing/deployment curve (many small fractions summing
            # to ~1 per year) is a TIMING assumption: flex = start delay, not ±pp
            vals = [v for _n, _c, v in cells]
            curve_hint = (len(vals) >= 6 and all(0.0 <= v <= 1.0001 for v in vals)
                          and sum(vals) >= 0.8 and gran in ("monthly", "quarterly"))
            type_label = f"{measure or ''} {raw_label}".strip()
            itype = classify_type(type_label, fmt, v0, curve_hint=curve_hint)
            # a NAMED customer/contract row feeding revenue carries binary risk:
            # its downside runs to ZERO (lose the contract), not just -X% — the
            # per-customer child bar then shows true concentration exposure.
            # (The family aggregate keeps the symmetric dial via _agg_itype.)
            if entity and in_rev and getattr(itype, "mode", None) == "pct" and v0 > 0:
                from .input_discovery import InputType as _IT
                itype = _IT(itype.name + "@risk", "pct", -1.0, itype.high,
                            0.0, itype.hi_clamp, itype.integer)
            sheet_leaves.setdefault(sn, []).append(dict(
                sn=sn, r=r, band=blo, bhi=bhi, cells=cells, zcells=zcells, v0=v0,
                ref=graph.node_ref(cells[-1][0]), label=label, fmt=fmt,
                measure=measure, entity=entity, unit_val=unit_val,
                hdr_row=(hdr or {}).get("row"), block=block_name,
                shift_k=shift_k, score=score, itype=itype))
            n_cands += 1

    # safety cap: exact recompute per candidate is the cost driver; keep the
    # highest-genuineness rows if a model somehow yields thousands
    MAX_FLEX_ROWS = 400
    if n_cands > MAX_FLEX_ROWS:
        ranked = sorted((lf for lfs in sheet_leaves.values() for lf in lfs),
                        key=lambda lf: -lf["score"])[:MAX_FLEX_ROWS]
        keep_ids = {id(lf) for lf in ranked}
        for sn in list(sheet_leaves):
            sheet_leaves[sn] = [lf for lf in sheet_leaves[sn] if id(lf) in keep_ids]
            if not sheet_leaves[sn]:
                del sheet_leaves[sn]

    leaves = []
    aggregates = []
    child_of = {}                          # id(leaf) -> aggregate dict
    for sn, row_leaves in sheet_leaves.items():
        sd = wbd.sheets[sn]
        role = structure.roles.get(sn)
        role = role[0].value if role else ""
        if not row_leaves:
            continue
        is_cost = role == "cost_buildup" or _COST.search(sn) is not None
        if is_cost and len(row_leaves) >= 6:
            # a dense cost stack -> ONE lever (flex the whole stack), rows = children
            agg = dict(sn=sn, r=-1, band=0,
                       cells=[c for lf in row_leaves for c in lf["cells"]],
                       v0=sum(lf["v0"] for lf in row_leaves),
                       ref=f"{sn} · {len(row_leaves)} inputs",
                       label=f"{_clean_sheet(sn)} cost", children=row_leaves)
            aggregates.append(agg)
            for lf in row_leaves:
                child_of[id(lf)] = agg
        else:
            # ONE path for assumption sheets and schedules alike: rows sharing the
            # same TABLE HEADER (and scenario block) are one family — a pricing
            # deck, a volume deck, a phasing-curve block — grouped into a single
            # driver with per-entity children. Never grouped by the column band
            # (which splintered one curve family into per-band fragments) or by
            # nearest-left text (which named drivers 'Units'/'Base Case').
            leaves.extend(row_leaves)
            by_tbl: dict[tuple, list] = {}
            for lf in row_leaves:
                if lf.get("hdr_row") is not None and lf.get("measure"):
                    # SAME-TYPE constraint: a family is one lever only when its
                    # rows are the same KIND of assumption — a header that merely
                    # spans a block of unrelated rows (salary + tax + headcount)
                    # must not fuse them into one meaningless dial
                    tname = getattr(lf.get("itype"), "name", "") or ""
                    key = (lf["hdr_row"], lf.get("block") or "", tname.replace("@risk", ""))
                else:
                    key = (None, f"~row{lf['r']}", "")
                by_tbl.setdefault(key, []).append(lf)
            for key, lfs in by_tbl.items():
                if key[0] is None or len(lfs) < 2:
                    continue                        # single rows stand alone, already well-named
                lfs.sort(key=lambda l: l["r"])
                measure = lfs[0].get("measure") or _clean_label(lfs[0]["label"])

                def _entity_of(lf):
                    e = lf.get("entity")
                    if e:
                        return e
                    lbl = _clean_label(lf["label"])
                    # strip an embedded "measure — " prefix so entities read clean
                    if measure and lbl.lower().startswith(measure.lower()):
                        lbl = lbl[len(measure):].strip(" -—:·,") or lbl
                    return lbl
                ents = [_entity_of(lf)[:22] for lf in lfs]
                # mixed units of measure ($/KWh next to $/Unit) or wildly mixed
                # magnitudes: summing bases misleads — show the largest instead
                unitvals = {lf.get("unit_val") for lf in lfs if lf.get("unit_val")}
                mags = [abs(lf["v0"]) for lf in lfs if lf["v0"]]
                mixed = len(unitvals) > 1 or (mags and max(mags) / max(1e-12, min(mags)) > 100)
                v0_agg = (max(lfs, key=lambda l: abs(l["v0"]))["v0"] if mixed
                          else sum(lf["v0"] for lf in lfs))
                blocknote = f"; {lfs[0]['block']}" if lfs[0].get("block") else ""
                label = f"{measure} — {len(lfs)} rows: " + ", ".join(ents[:2]) \
                        + (f", +{len(lfs) - 2} more" if len(lfs) > 2 else "")
                agg = dict(sn=sn, r=-1, band=lfs[0]["band"],
                           cells=[c for lf in lfs for c in lf["cells"]],
                           zcells=[z for lf in lfs for z in lf.get("zcells", [])],
                           v0=v0_agg, mixed_units=mixed,
                           ref=f"{sn} · {measure[:24]}{blocknote}",
                           label=label[:70], measure=measure,
                           shift_k=lfs[0].get("shift_k", 1), children=lfs)
                for lf in lfs:                      # child = its entity/row name
                    lf["label"] = _entity_of(lf)[:38]
                aggregates.append(agg)
                for lf in lfs:
                    child_of[id(lf)] = agg

    # merge duplicate aggregates across replica blocks — but ONLY when they are
    # genuinely the same lever: same normalized measure AND same format family
    # AND comparable magnitude (a units-of-measure enum table must never merge
    # into a BOM deck again just because both were labeled 'Units')
    _bylabel: dict[tuple, dict] = {}
    _merged: list[dict] = []
    for agg in aggregates:
        base_mag = abs(agg["v0"]) or 1.0
        fmt0 = "%" if any("%" in (lf.get("fmt") or "") for lf in agg["children"]) else "#"
        k = (_re.sub(r"[^a-z0-9]+", "", (agg.get("measure") or agg["label"]).lower()), agg["sn"], fmt0)
        m = _bylabel.get(k)
        if m is not None and max(base_mag, abs(m["v0"]) or 1.0) / max(1e-12, min(base_mag, abs(m["v0"]) or 1.0)) < 1e3:
            m["cells"] = m["cells"] + agg["cells"]
            m["zcells"] = m.get("zcells", []) + agg.get("zcells", [])
            m["children"] = m["children"] + agg["children"]
            m["v0"] = m["v0"] + agg["v0"]
        else:
            _bylabel.setdefault(k, agg)
            _merged.append(agg)
    aggregates = _merged

    def _shift_overrides(rows_cells, k):
        """Delay a phasing curve by k columns: each row's value vector slides k
        columns later (deployment slips ~6 months); the vacated leading columns
        go to zero and the trailing typed zero-cells receive the displaced tail.
        The row's total is preserved wherever the model left room for the tail."""
        ov = {}
        for rcells in rows_cells:
            seq = sorted(rcells, key=lambda x: x[1])
            vals = [v for _n, _c, v in seq]
            shifted = [0.0] * min(k, len(vals)) + vals[:max(0, len(vals) - k)]
            for (nd, _c, _v), nv in zip(seq, shifted):
                ov[nd] = nv
        return ov

    def _flex(cells, itype, shift_rows=None, shift_k=1):
        # a phasing curve is flexed as a START DELAY (the analyst's 'what if
        # commercialization slips 6 months'), not by bending monthly fractions
        if itype is not None and getattr(itype, "mode", None) == "shift" and shift_rows:
            lo = model.outputs_with_overrides({}, all_nodes)          # on-time = base
            hi = model.outputs_with_overrides(_shift_overrides(shift_rows, shift_k), all_nodes)
            sw = max(abs((hi.get(nd) or 0) - (lo.get(nd) or 0)) for nd in rank_nodes.values())
            return lo, hi, sw
        # context-appropriate move per cell: rates shift in percentage POINTS,
        # level inputs scale by the type's ±%, all guard-clamped to validity
        ov_lo, ov_hi = {}, {}
        for (nd, _c, val) in cells:
            lo_v, hi_v = ranged_vals(itype, val)
            ov_lo[nd], ov_hi[nd] = lo_v, hi_v
        lo = model.outputs_with_overrides(ov_lo, all_nodes)
        hi = model.outputs_with_overrides(ov_hi, all_nodes)
        # rank by the OUTCOME the lever hits hardest across every mapped output —
        # volume levers move revenue, cost levers EBITDA, tax levers net income,
        # financing levers cash; a single-output swing would miss the last two
        sw = max(abs((hi.get(nd) or 0) - (lo.get(nd) or 0)) for nd in rank_nodes.values())
        return lo, hi, sw

    def _shift_rows_of(d):
        """Per-row ordered cell vectors (nonzero + typed zero tail cells) for a
        shift flex — the zeros are the landing zone for the displaced tail."""
        kids = d.get("children") or [d]
        return [sorted(list(lf["cells"]) + list(lf.get("zcells", [])), key=lambda x: x[1])
                for lf in kids]

    def _deresk(it):
        # the to-zero downside is a PER-CUSTOMER lens; the family aggregate flexes
        # as a symmetric dial (all our pct types are symmetric, so low = -high)
        if it is not None and it.name.endswith("@risk"):
            from .input_discovery import InputType as _IT
            return _IT(it.name[:-5], "pct", -it.high, it.high, it.lo_clamp, it.hi_clamp, it.integer)
        return it

    def _agg_itype(agg):
        # an aggregate flexes with its children's modal type (same-type grouping)
        kinds = [_deresk(lf["itype"]) for lf in agg.get("children", []) if lf.get("itype")]
        if not kinds:
            from .input_discovery import classify_type as _ct
            return _ct(agg.get("label", ""), "", agg.get("v0", 0.0))
        names = [k.name for k in kinds]
        return kinds[names.index(max(set(names), key=names.count))]

    gate = max(1.0, 1e-4 * abs(eb_base or 1))
    cand = []
    claimed: set[int] = set()      # GUARANTEE: each input cell is flexed by at most one driver
    # aggregates first (broader levers claim their cells); then standalone leaves
    for agg in aggregates:
        cells = [x for x in agg["cells"] if x[0] not in claimed]
        if not cells:
            continue
        agg["itype"] = _agg_itype(agg)
        shifting = getattr(agg["itype"], "mode", None) == "shift"
        lo, hi, sw = _flex(cells, agg["itype"],
                           shift_rows=_shift_rows_of(agg) if shifting else None,
                           shift_k=agg.get("shift_k", 1))
        if sw < gate:
            for lf in agg["children"]:
                child_of.pop(id(lf), None)
            continue
        cellset = {x[0] for x in cells}
        claimed |= cellset
        kids = []
        for lf in agg["children"]:          # children are this aggregate's own decomposition
            if not all(c[0] in cellset for c in lf["cells"]):
                continue
            k_shift = getattr(lf["itype"], "mode", None) == "shift"
            klo, khi, ksw = _flex(lf["cells"], lf["itype"],
                                  shift_rows=_shift_rows_of(lf) if k_shift else None,
                                  shift_k=lf.get("shift_k", 1))
            kids.append(dict(lf, lo=klo, hi=khi, swing=ksw))
        kids.sort(key=lambda d: -d["swing"])           # keep the most material rows as children
        cand.append(dict(agg, cells=cells, lo=lo, hi=hi, swing=sw, children=kids[:10]))
    for lf in leaves:                       # ungrouped leaves -> standalone drivers
        if id(lf) in child_of:
            continue
        cells = [x for x in lf["cells"] if x[0] not in claimed]
        if not cells:
            continue
        l_shift = getattr(lf["itype"], "mode", None) == "shift"
        lo, hi, sw = _flex(cells, lf["itype"],
                           shift_rows=_shift_rows_of(lf) if l_shift else None,
                           shift_k=lf.get("shift_k", 1))
        if sw < gate:
            continue
        claimed |= {x[0] for x in cells}
        cand.append(dict(lf, cells=cells, lo=lo, hi=hi, swing=sw, children=[]))
    if not cand:
        return None
    # guardrail: a swing orders of magnitude beyond the outputs means the flex
    # broke the model's branching (an enum/marker that slipped through) — a real
    # lever can't move EBITDA by 100x its own base. Tested on EVERY (output,
    # period) cell, not just the terminal ranking swing: an enum flex can look
    # sane in the terminal year yet put ±1e12 in an early-year cube cell.
    swing_cap = 100.0 * max(abs(eb_base or 1.0), abs(rev_base or 1.0))

    def _worst_cell_swing(d):
        return max((abs((d["hi"].get(nd) or 0) - (d["lo"].get(nd) or 0)) for nd in all_nodes),
                   default=0.0)

    cand = [d for d in cand if d["swing"] <= swing_cap and _worst_cell_swing(d) <= swing_cap]
    if not cand:
        return None
    cand.sort(key=lambda d: -d["swing"])
    cand = cand[:MAX_DRIVERS]
    # disambiguate duplicate labels (e.g. two 'Edge computing' rows) by cell
    lab_counts: dict[str, int] = {}
    for d in cand:
        lab_counts[d["label"]] = lab_counts.get(d["label"], 0) + 1
    for d in cand:
        if lab_counts[d["label"]] > 1:
            d["label"] = f"{d['label']} ({d['ref'].split('!')[-1]})"

    def _outmat(d):
        return {okey: {p: {"low": d["lo"].get(out_nodes[(okey, p)]),
                           "high": d["hi"].get(out_nodes[(okey, p)])}
                       for p in periods if (okey, p) in out_nodes}
                for okey, _l, _it in out_specs}

    # cube: every driver's effect on every output at every period (for the UI)
    cube = {
        "outputs": [{"key": okey, "label": label} for okey, label, _ in out_specs],
        "periods": list(periods),
        "defaultOutput": "ebitda",
        "defaultPeriod": tp,
        "unit": U,
        "base": {okey: {p: model.values.get(out_nodes[(okey, p)])
                        for p in periods if (okey, p) in out_nodes}
                 for okey, _l, _it in out_specs},
        "drivers": [],
    }
    drivers = []
    adv_ov: dict[int, float] = {}
    fav_ov: dict[int, float] = {}
    for i, d in enumerate(cand):
        # the enumerate index guarantees a UNIQUE key per driver: distinct
        # section-grouped drivers on one sheet can otherwise share (sheet, row,
        # band) -> a key collision that makes the JS maGet() read the wrong
        # driver's base, drifting the base eval and compounding the slopes into
        # nonsense (billions). Distinct keys keep every driver independent.
        dkey = f"in:{d['sn']}:{d['r']}:{d.get('band', 0)}:{i}"
        it_d = d.get("itype")
        d_shift = getattr(it_d, "mode", None) == "shift"
        base0 = d["v0"]
        if d_shift:
            # a timing lever: the editable input is the DELAY in months (0 = the
            # plan's own start), so base 0, range 0..6; outputs interpolate from
            # on-time (low) to slipped-6-months (high)
            months = 6.0
            d_val, d_lo, d_hi = 0.0, 0.0, months
            d_label = f"{d['label']} — start delay (months)"
        else:
            d_lo, d_hi = ranged_vals(it_d, base0) if it_d else (base0 * 0.8, base0 * 1.2)
            d_lo, d_hi = _bracket_range(d_lo, d_hi, base0)
            d_val, d_label = base0, d["label"]

        def _kid(ci, c):
            kt = c.get("itype")
            if getattr(kt, "mode", None) == "shift":
                kval, klo, khi = 0.0, 0.0, 6.0
            else:
                kval = c["v0"]
                klo, khi = ranged_vals(kt, kval) if kt else (kval * 0.8, kval * 1.2)
                klo, khi = _bracket_range(klo, khi, kval)      # same JS-slope invariant
            return {"key": f"{dkey}::{ci}", "label": c["label"], "refs": [c["ref"]],
                    "value": kval, "low": klo, "high": khi,
                    "rtype": kt.name if kt else None, "out": _outmat(c)}
        kids = [_kid(ci, c) for ci, c in enumerate(d.get("children", []))]
        # low/high = the type-appropriate analyst-editable defaults (a tax rate
        # moves ±5pp, a volume ±25%, a phasing curve 0-6 months of delay — never
        # one blanket ±20%); the JS reads them
        cube["drivers"].append({"key": dkey, "label": d_label, "refs": [d["ref"]],
                                "value": d_val, "low": d_lo, "high": d_hi,
                                "rtype": it_d.name if it_d else None,
                                "out": _outmat(d), "children": kids})
        eb_lo, eb_hi = d["lo"].get(eb_node), d["hi"].get(eb_node)
        rv_lo, rv_hi = d["lo"].get(rev_node), d["hi"].get(rev_node)
        base_v = d_val or 1.0
        drivers.append(SensitivityDriver(
            key=dkey, label=f"{d_label} ({tp})",
            input_refs=[d["ref"]], base=d_val, low=d_lo, high=d_hi,
            low_pct=round(d_lo / base_v - 1.0, 4), high_pct=round(d_hi / base_v - 1.0, 4),
            output_low=eb_lo, output_high=eb_hi, swing=d["swing"],
            out2_low=rv_lo, out2_high=rv_hi, unit=U))
        adverse_low = (eb_lo or 0) < (eb_hi or 0)
        if d_shift:
            # all-adverse = every curve slips; all-favorable = the plan holds
            adv_ov.update(_shift_overrides(_shift_rows_of(d), d.get("shift_k", 1)))
        else:
            for (nd, _c, val) in d["cells"]:
                c_lo, c_hi = ranged_vals(it_d, val) if it_d else (val * 0.8, val * 1.2)
                adv_ov[nd] = c_lo if adverse_low else c_hi
                fav_ov[nd] = c_hi if adverse_low else c_lo
    downside = model.outputs_with_overrides(adv_ov, [eb_node])[eb_node]
    upside = model.outputs_with_overrides(fav_ov, [eb_node])[eb_node]

    caveats = ["Each input is flexed through the model's actual formulas — an exact recompute, "
               "not a proportional estimate.",
               "Single-input moves are exact; all-adverse and the two-way grid combine drivers linearly."]
    # honesty over silence: when a mapped below-EBITDA output could not be
    # exactly recomputed (its spine carries the workbook's own error cells or
    # unsupported formulas), say which assumption families are therefore missing
    built = {s[0] for s in out_specs}
    blocked = []
    if mapping.get("net_income", sheet) is not None and "net_income" not in built:
        blocked.append("net income (tax / below-the-line assumptions)")
    if mapping.get("ending_cash", sheet) is not None and "cash" not in built:
        blocked.append("ending cash (financing / fundraising assumptions)")
    if blocked:
        caveats.append(
            "Not shown here: " + " and ".join(blocked) + " could not be exactly recomputed — "
            "that part of the workbook contains error cells or formulas the recompute engine "
            "can't evaluate, so flexing those assumptions would produce made-up numbers. "
            "They are still assumptions; review them via the flags and the raw sheet.")

    return Tornado(
        output_key="terminal_ebitda", output_label=f"Terminal EBITDA ({tp})",
        output_unit=U, output_base=eb_base, formula="recompute_linear",
        out2_base=rev_base, out2_label="Revenue", cube=cube,
        formula_note="every output recomputed through the model's own formulas as each input is flexed",
        drivers=drivers, downside=downside, upside=upside, caveats=caveats)


def build_sensitivities(wbd, structure: StructureResult, mapping: MappingResult,
                        benchmarks: dict, archetype: str = "default",
                        periods: Optional[list[str]] = None, graph=None) -> list[Tornado]:
    out: list[Tornado] = []
    sheet = mapping.primary_sheet
    if sheet is None:
        return out
    if periods is None:
        ax = structure.primary_axis(sheet)
        periods = list(ax.periods) if ax else []
    if not periods:
        return out
    units = structure.units.get(sheet)
    U = units.label if units else "model units"

    rev = mapping.get("revenue_total", sheet)
    tp = None
    if rev is not None:
        t = _terminal(rev)
        tp = t[0] if t else None

    # ---- Output 1: terminal EBITDA ----------------------------------------
    # Decompose revenue/COGS/opex into line items ONLY when those items
    # reconcile to the model's own total for that group; otherwise the mapped
    # "components" overlap (a subtotal like "TOTAL Salaries" plus its parts, or
    # an LLM-mapped bucket), which would double-count — so fall back to the one
    # total cell. This keeps every driver a real, non-overlapping model line.
    def _group(comp_items, total_item, coef, key_prefix, total_label):
        comps = _components(comp_items, sheet, tp, coef, key_prefix, U)
        tt = _terminal_for(total_item, tp) if total_item is not None else None
        if comps and tt is not None:
            comp_sum = sum(s.base for s in comps)            # bases are abs for costs
            total_abs = abs(tt[1])
            if total_abs > 1 and abs(comp_sum - total_abs) <= 0.02 * total_abs:
                return comps                                  # clean, non-overlapping decomposition
        if tt is not None:                                    # overlap or gap -> use the single total
            return [_Spec(key_prefix, f"{total_label} ({tp})", [tt[2]],
                          abs(tt[1]) if coef < 0 else tt[1], coef, U,
                          adverse="high" if coef < 0 else "low")]
        return comps

    # Preferred: flex the model's REAL input cells and recompute EBITDA exactly
    # through its own formulas. Falls back to the line-item closed form when the
    # workbook can't be faithfully recomputed (XLOOKUP/VBA/etc.).
    recomputed = None
    try:
        recomputed = _recompute_ebitda_tornado(wbd, structure, mapping, graph, U, periods)
    except Exception:
        recomputed = None
    if recomputed is not None:
        out.append(recomputed)

    if recomputed is None and rev is not None and tp is not None:
        specs: list[_Spec] = []
        specs += _group([it for it in mapping.get_all("revenue_segment") if it.sheet == sheet],
                        rev, +1.0, "rev", "Total revenue")
        specs += _group([it for it in mapping.get_all("cogs_component") if it.sheet == sheet],
                        mapping.get("cogs_total", sheet), -1.0, "cogs", "Total COGS")
        specs += _group([it for it in mapping.get_all("opex_component") if it.sheet == sheet],
                        mapping.get("opex_total", sheet), -1.0, "opex", "Total opex")

        if len(specs) >= 2:
            # Anchor the base to the model's OWN EBITDA/EBIT row when it exists,
            # so the output the analyst flexes equals the number the model
            # actually reports. The decomposed revenue−COGS−opex can differ (a
            # line not mapped, D&A handled differently, other income); the gap
            # becomes a constant offset, leaving every per-input delta exact.
            offset = 0.0
            anchored = False
            own_eb = mapping.get("ebitda_or_ebit", sheet)
            own_t = _terminal_for(own_eb, tp) if own_eb is not None else None
            if own_t is None and own_eb is not None:
                own_t = _terminal(own_eb)
            if own_t is not None:
                decomposed = _eval_linear(specs, {})
                offset = own_t[1] - decomposed
                anchored = abs(offset) > 0.01 * max(abs(own_t[1]), 1)

            base_out = _eval_linear(specs, {}, offset)
            downside = _eval_linear(specs, {s.key: s.adverse_val() for s in specs}, offset)
            upside = _eval_linear(specs, {s.key: s.favorable_val() for s in specs}, offset)
            shap = _shapley_linear(specs)            # contributions are differences; offset cancels
            drivers = []
            for s in specs:
                ol = _eval_linear(specs, {s.key: s.low()}, offset)
                oh = _eval_linear(specs, {s.key: s.high()}, offset)
                drivers.append(_mk_driver(s, ol, oh, shap[s.key]))
            drivers.sort(key=lambda d: -max(abs(d.shapley), d.swing))
            drivers = drivers[:MAX_DRIVERS]
            caveats = ["Every driver is one of the model's own line-item cells; "
                       "±20% default range, editable below."]
            if anchored:
                caveats.append("Base is anchored to the model's own EBITDA row; the part not "
                               "explained by the decomposed lines is held constant, so each "
                               "input's effect is exact while the base matches the model.")
            out.append(Tornado(
                output_key="terminal_ebitda", output_label=f"Terminal EBITDA ({tp})",
                output_unit=U, output_base=base_out, formula="linear_sum", offset=offset,
                formula_note="EBITDA = revenue − COGS − opex, decomposed to every mapped line item",
                drivers=drivers, downside=downside, upside=upside, caveats=caveats))

    # ---- Output 2: present value of exit ----------------------------------
    pv = _valuation_tornado(wbd, structure, mapping, benchmarks, archetype, periods, U)
    if pv is not None:
        out.append(pv)
    return out


def _valuation_tornado(wbd, structure, mapping, benchmarks, archetype, periods, U) -> Optional[Tornado]:
    sheet = mapping.primary_sheet
    ebitda_mult = find_scalar_assumptions(wbd, structure, r"ebitda\s*multiple|multiple.*ebitda|x\s*ebitda", (0.5, 100))
    dr = find_scalar_assumptions(wbd, structure, r"discount\s*rate|hurdle|wacc", (0.01, 0.99))
    exit_year = find_scalar_assumptions(wbd, structure, r"exit\s*year|valuation\s*year", (1990, 2100))
    ebitda = mapping.get("ebitda_or_ebit", sheet)
    if not ebitda_mult or ebitda is None or not periods:
        return None
    yr = str(int(exit_year[0][1])) if exit_year else None
    exit_p = next((p for p in periods if p.startswith(yr)), None) if yr else None
    if exit_p is None:
        for p in reversed(periods):
            if ebitda.value_for(p) is not None:
                exit_p = p
                break
    if exit_p is None or ebitda.value_for(exit_p) is None or ebitda.value_for(exit_p) <= 0:
        return None
    e_exit = float(ebitda.value_for(exit_p))
    mref, mval, mlabel = ebitda_mult[0]
    rate_ref = rate_val = None
    if dr:
        pref = [t for t in dr if re.search(r"current|to\s*exit|today|present", t[2], re.I)]
        rate_ref, rate_val, _ = (pref[0] if pref else dr[0])
    cur = find_scalar_assumptions(wbd, structure, r"current\s*year|valuation\s*(date|year)|base\s*year", (1990, 2100))
    base_year = int(cur[0][1]) if cur else (int(re.match(r"(\d{4})", periods[0]).group(1)) if re.match(r"(\d{4})", periods[0]) else None)
    me = re.match(r"(\d{4})", exit_p)
    horizon = float(int(me.group(1)) - base_year) if (me and base_year is not None) else 0.0

    rng = (benchmarks or {}).get("ev_ebitda_multiple", {}).get(archetype) or \
          (benchmarks or {}).get("ev_ebitda_multiple", {}).get("default") or {}

    specs: list[_Spec] = [
        _Spec("multiple", f"Exit EBITDA multiple ({mlabel})", [mref], mval, 0.0, "x", adverse="low",
              low_val=float(rng.get("low", mval * 0.8)) if rng else None,
              high_val=float(rng.get("high", mval * 1.2)) if rng else None),
        _Spec("metric", f"Exit-year EBITDA ({exit_p})", [ebitda.ref_for(exit_p)], e_exit, 0.0, U, adverse="low"),
    ]
    if rate_val is not None:
        specs.append(_Spec("rate", "Discount rate", [rate_ref], rate_val, 0.0, "fraction",
                           low_pct=-0.25, high_pct=0.25, adverse="high"))
    sbk = {s.key: s for s in specs}

    def ev(ov):
        return _eval_pv(sbk, ov, horizon)

    base_out = ev({})
    downside = ev({s.key: s.adverse_val() for s in specs})
    upside = ev({s.key: s.favorable_val() for s in specs})
    shap = _shapley_exact(specs, ev)
    drivers = []
    for s in specs:
        ol = ev({s.key: s.low()})
        oh = ev({s.key: s.high()})
        drivers.append(_mk_driver(s, ol, oh, shap[s.key]))
    drivers.sort(key=lambda d: -abs(d.shapley))
    return Tornado(
        output_key="valuation_pv", output_label=f"Present value of {exit_p} exit",
        output_unit=U, output_base=base_out, formula="valuation_pv",
        formula_note=f"PV = multiple × EBITDA[{exit_p}] / (1+discount)^{horizon:g}",
        horizon=horizon, drivers=drivers, downside=downside, upside=upside,
        caveats=[f"Discounts {horizon:g} years at the model's own rate; multiple range is the "
                 f"{archetype.replace('_', ' ')} benchmark band.",
                 "Shapley bars attribute the base→downside move across the inputs, including interactions."])

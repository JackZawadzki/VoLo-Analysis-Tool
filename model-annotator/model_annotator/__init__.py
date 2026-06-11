"""model-annotator — automated financial-model diligence.

Public API::

    from model_annotator import annotate
    report = annotate("model.xlsx")            # typed Report (pydantic)

The pipeline is a strict sequence of phases (ingest → structure → census →
tie-outs → metrics → findings → acquittals → severity/dedup → emit); each
phase writes intermediate JSON into the output directory so failures are
inspectable. Deterministic core; the LLM (optional) only classifies ambiguous
labels and polishes prose — it never produces a number.
"""
from __future__ import annotations

import csv
import datetime as _dt
import json
import logging
import re
from pathlib import Path
from typing import Optional

import yaml

from .schema import (  # noqa: F401  (public re-exports)
    Report,
    SCHEMA_VERSION,
    TOOL_VERSION,
    SheetMapEntry,
    TieOutStatus,
    WorkbookMap,
    format_ref,
    parse_ref,
)

__version__ = TOOL_VERSION
__all__ = ["annotate", "validate_citations", "Report", "__version__"]

log = logging.getLogger(__name__)

_A1 = re.compile(r"^[A-Z]{1,3}\d{1,7}$")


def _default_benchmarks_path() -> Path:
    return Path(__file__).parent / "benchmarks.yaml"


def annotate(
    path: str,
    out_dir: Optional[str] = None,
    *,
    no_llm: bool = False,
    benchmarks_path: Optional[str] = None,
    max_cells: int = 2_000_000,
    write_outputs: bool = True,
    json_only: bool = False,
    workbook_map: Optional[WorkbookMap] = None,
    llm_model: Optional[str] = None,
) -> Report:
    """Run the full diligence pipeline on one workbook.

    ``workbook_map`` may carry a previously computed map (from a prior run's
    ``workbook_map.json``); its row mappings are reused instead of re-scored,
    so a host engine can parse once and analyze many ways.
    """
    from .ingest import load_workbook_dual
    from .structure import build_structure
    from .graph import build_graph, build_assumptions_census
    from .mapping import build_mapping, mapping_from_workbook_map
    from .integrity import run_tie_outs
    from .metrics import compute_metrics
    from .findings import FCtx, build_findings
    from .llm import LLMClient
    from .narrate import polish_report
    from .report_md import render_markdown

    src = Path(path)
    out: Optional[Path] = None
    if write_outputs:
        stamp = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        out = Path(out_dir) if out_dir else Path("annotations") / f"{src.stem}-{stamp}"
        out.mkdir(parents=True, exist_ok=True)

    benchmarks = yaml.safe_load(Path(benchmarks_path or _default_benchmarks_path()).read_text()) or {}
    llm = LLMClient(enabled=not no_llm, model=llm_model)
    if llm.disabled_reason and not no_llm:
        log.info("LLM disabled: %s", llm.disabled_reason)

    # ---- Phase 0-1: ingest + structure + graph -----------------------------
    wbd = load_workbook_dual(str(src), max_cells=max_cells)
    structure = build_structure(wbd)
    graph = build_graph(wbd)

    # ---- mapping (heuristics + optional LLM label assist) ------------------
    if workbook_map is not None and workbook_map.row_mappings:
        mapping = mapping_from_workbook_map(wbd, structure, workbook_map)
    else:
        classifier = None
        if llm.available:
            def classifier(label: str, section: str, cands: list[str]):
                return llm.classify_label(label, section, cands)
        mapping = build_mapping(wbd, structure, llm_classify=classifier)

    stmt_sheets = [mapping.primary_sheet] if mapping.primary_sheet else []

    # ---- Phase 2: assumptions census ---------------------------------------
    census = build_assumptions_census(graph, stmt_sheets)

    # ---- Phase 3: integrity tie-outs (the trust gate) ----------------------
    integrity = run_tie_outs(wbd, mapping, graph)

    # ---- Phase 4: derived metrics ------------------------------------------
    metrics = compute_metrics(wbd, structure, mapping, benchmarks)

    # ---- Phases 5-7: findings / acquittals / severity-dedup ----------------
    fctx = FCtx(wbd=wbd, structure=structure, mapping=mapping, graph=graph,
                integrity=integrity, metrics=metrics, census=census, benchmarks=benchmarks)
    fres = build_findings(fctx)

    # ---- assemble the Report -----------------------------------------------
    wm = _build_workbook_map(wbd, structure, mapping, integrity)
    limitations = (list(wbd.limitations) + list(structure.limitations) +
                   list(graph.limitations) + list(mapping.notes) +
                   list(integrity.notes) + list(metrics.notes) + list(fres.notes))
    if llm.disabled_reason and not no_llm:
        limitations.append(f"LLM assist unavailable ({llm.disabled_reason}); heuristics-only run.")

    report = Report(
        source_file=str(src),
        sha256=wbd.sha256,
        analyzed_at=_dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds"),
        llm_used=llm.available,
        workbook_map=wm,
        assumptions_census=census,
        tie_outs=integrity.tie_outs,
        derived_metrics=metrics.metrics,
        findings=fres.findings,
        acquittals=fres.acquittals,
        clean_checks=fres.clean_checks,
        unmapped=mapping.unmapped,
        limitations=limitations,
    )

    # ---- Phase 8: polish (verified), validate citations, emit --------------
    if llm.available:
        polish_report(report, llm)

    violations = validate_citations(report, wbd)
    for v in violations[:20]:
        report.limitations.append(f"CITATION VALIDATION FAILURE: {v}")
    if violations:
        log.error("%d citation validation failure(s) — this is a tool bug", len(violations))

    if write_outputs and out is not None:
        _emit_outputs(report, mapping, metrics, out, json_only=json_only,
                      render_markdown=render_markdown)
    return report


# ---------------------------------------------------------------------------
# Workbook map assembly
# ---------------------------------------------------------------------------

def _build_workbook_map(wbd, structure, mapping, integrity) -> WorkbookMap:
    sheets = []
    for name in wbd.sheet_order:
        sd = wbd.sheets[name]
        cen = structure.censuses.get(name)
        role, conf = structure.roles.get(name, (None, 0.0))
        from .schema import SheetRole
        sheets.append(SheetMapEntry(
            name=name,
            role=role or SheetRole.unknown,
            role_confidence=conf,
            visibility=sd.visibility,
            n_rows=sd.max_row, n_cols=sd.max_col,
            n_formula_cells=cen.n_formula if cen else 0,
            n_typed_input_cells=cen.n_typed_input if cen else 0,
            n_label_cells=cen.n_label if cen else 0,
            n_error_cells=cen.n_error if cen else 0,
            period_axis=structure.primary_axis(name),
            units=structure.units.get(name),
            notes=(["truncated by cell cap"] if sd.truncated else []),
        ))
    primary = mapping.primary_sheet
    return WorkbookMap(
        sheets=sheets,
        period_axis=structure.primary_axis(primary) if primary else None,
        primary_statement_sheet=primary,
        units=dict(structure.units),
        row_mappings=mapping.to_row_mappings(),
        named_ranges={k: v for k, v in list(wbd.named_ranges.items())[:200]},
        has_macros=wbd.has_macros,
        hidden_sheets=[s.name for s in wbd.sheets.values() if s.visibility != "visible"],
        input_color_convention=structure.input_convention,
        trust_score=integrity.trust_score,
    )


# ---------------------------------------------------------------------------
# Citation validator (anti-hallucination rule 6)
# ---------------------------------------------------------------------------

def validate_citations(report: Report, wbd) -> list[str]:
    """Re-read the workbook and assert every cited cell exists and (for value
    citations) matches. Returns a list of violations — empty means clean."""
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

    problems: list[str] = []

    def check_ref(ref: str, value=None, where: str = "") -> None:
        if "!" not in ref or ref.startswith("<"):
            return
        try:
            sheet, cell = parse_ref(ref)
        except ValueError:
            problems.append(f"{where}: unparseable ref {ref!r}")
            return
        if not _A1.match(cell):
            return  # row-level pseudo refs ("row 12") are descriptive, not cell citations
        sd = wbd.sheets.get(sheet)
        if sd is None:
            problems.append(f"{where}: sheet {sheet!r} does not exist ({ref})")
            return
        cl, row = coordinate_from_string(cell)
        col = column_index_from_string(cl)
        if row > max(sd.max_row, 1) or col > max(sd.max_col, 1):
            problems.append(f"{where}: {ref} outside sheet bounds")
            return
        if value is None:
            return
        rec = sd.cell(row, col)
        actual = rec.value if rec else None
        if isinstance(value, float) and isinstance(actual, (int, float)) and not isinstance(actual, bool):
            if abs(float(actual) - value) > max(abs(value) * 1e-6, 1e-9):
                problems.append(f"{where}: {ref} cited {value!r} but workbook holds {actual!r}")
        elif isinstance(value, str) and isinstance(actual, str):
            if value[:40] != actual[:40]:
                problems.append(f"{where}: {ref} cited text differs from workbook")
        elif type(actual) is not type(value) and actual != value:
            # mixed types: compare loosely; None-vs-value is a real problem
            if actual is None:
                problems.append(f"{where}: {ref} cited {value!r} but the cell is empty")

    for f in report.findings:
        for e in f.evidence_values:
            check_ref(e.ref, e.value, where=f.id)
        for ref in f.evidence_cells:
            check_ref(ref, None, where=f.id)
    for t in report.tie_outs:
        for e in t.evidence:
            check_ref(e.ref, e.value, where=f"tie-out {t.id}")
    for m in report.derived_metrics:
        for ref in m.inputs:
            check_ref(ref, None, where=f"metric {m.metric_id}")
    for a in report.acquittals:
        for ref in a.resolution_cells:
            if ref:
                check_ref(ref, None, where=f"acquittal")
    for e in report.assumptions_census:
        check_ref(e.cell, None, where="census")
    return problems


# ---------------------------------------------------------------------------
# Output emission
# ---------------------------------------------------------------------------

def _emit_outputs(report: Report, mapping, metrics, out: Path, *, json_only: bool,
                  render_markdown) -> None:
    dump = dict(indent=2, default=str)

    (out / "report.json").write_text(report.model_dump_json(indent=2))
    (out / "workbook_map.json").write_text(report.workbook_map.model_dump_json(indent=2))

    # phase intermediates, for inspectability
    inter = {
        "phase2_assumptions_census.json": [e.model_dump() for e in report.assumptions_census],
        "phase3_tie_outs.json": [t.model_dump() for t in report.tie_outs],
        "phase4_derived_metrics.json": [m.model_dump() for m in report.derived_metrics],
        "phase5_findings.json": {
            "findings": [f.model_dump() for f in report.findings],
            "acquittals": [a.model_dump() for a in report.acquittals],
            "clean_checks": [c.model_dump() for c in report.clean_checks],
        },
    }
    for name, payload in inter.items():
        (out / name).write_text(json.dumps(payload, **dump))

    with (out / "derived_metrics.csv").open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["metric_id", "label", "sheet", "period", "value", "formula_inputs", "computation"])
        sheet = report.workbook_map.primary_statement_sheet or ""
        for m in report.derived_metrics:
            inputs = ";".join(m.inputs)
            if m.scalar is not None:
                w.writerow([m.metric_id, m.label, sheet, "(scalar)", m.scalar, inputs, m.computation])
            for period, value in m.series.items():
                w.writerow([m.metric_id, m.label, sheet, period, value, inputs, m.computation])

    if not json_only:
        (out / "report.md").write_text(render_markdown(report))
    log.info("outputs written to %s", out)

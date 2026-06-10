"""Formula dependency graph.

Parses formula *text* (never evaluates) into a precedents/dependents graph
over int-encoded cell nodes, then derives the structural facts later phases
need: error cells and whether they feed statements, orphaned computation rows,
circular references, external links, and downstream-impact ranks for typed
inputs.

Node encoding: ``(sheet_index << 44) | (row << 20) | col`` — keeps a
multi-million-edge graph affordable in CPython.
"""
from __future__ import annotations

import logging
import re
from bisect import bisect_left, bisect_right
from collections import deque
from dataclasses import dataclass, field
from typing import Iterable, Optional

from openpyxl.utils import column_index_from_string

from .ingest import SheetData, WorkbookData, a1, is_number
from .schema import format_ref

log = logging.getLogger(__name__)

ROW_BITS = 24
COL_BITS = 20
RANGE_EXPANSION_CAP = 20_000     # populated cells linked per single range token
AREA_NODE_MIN = 64               # areas with >= this many populated cells become shared range-nodes
MAX_EDGES = 30_000_000


def encode(sheet_idx: int, row: int, col: int) -> int:
    return (sheet_idx << (ROW_BITS + COL_BITS)) | (row << COL_BITS) | col


def decode(node: int) -> tuple[int, int, int]:
    col = node & ((1 << COL_BITS) - 1)
    row = (node >> COL_BITS) & ((1 << ROW_BITS) - 1)
    return node >> (ROW_BITS + COL_BITS), row, col


# ---------------------------------------------------------------------------
# Tokenizer
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class RefToken:
    sheet: Optional[str]          # None = same sheet
    r1: int
    c1: int
    r2: int
    c2: int
    external: bool = False
    whole_col: bool = False
    whole_row: bool = False


_RE_STRING = re.compile(r'"(?:[^"]|"")*"')
_SHEET_PART = r"(?:'(?:[^']|'')+'|[A-Za-z_\[\]][A-Za-z0-9_.\[\]]*)"
# multi-sheet refs like Sheet1:Sheet3!A1 are rare in startup models; the first
# sheet wins and the rest of the span is ignored (recorded by caller if needed)
_RE_REF = re.compile(
    rf"(?P<sheetpart>{_SHEET_PART}!)?"
    r"(?P<ref>"
    r"\$?[A-Za-z]{1,3}\$?\d{1,7}(?:\s*:\s*\$?[A-Za-z]{1,3}\$?\d{1,7})?"   # cell or area
    r"|\$?[A-Za-z]{1,3}\s*:\s*\$?[A-Za-z]{1,3}"                            # whole columns
    r"|\$?\d{1,7}\s*:\s*\$?\d{1,7}"                                        # whole rows
    r")"
)
_RE_NAME = re.compile(r"[A-Za-z_\\][A-Za-z0-9_.\\]*")
_RE_CELL_SHAPE = re.compile(r"^\$?[A-Za-z]{1,3}\$?\d{1,7}$")

_BOUNDARY_BEFORE = set("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_$.!:")


def _parse_cell(token: str) -> tuple[int, int]:
    token = token.replace("$", "")
    m = re.match(r"([A-Za-z]{1,3})(\d{1,7})$", token)
    return int(m.group(2)), column_index_from_string(m.group(1).upper())


def tokenize_refs(formula: str, sheet_names_lower: dict[str, str]) -> tuple[list[RefToken], list[str]]:
    """Extract cell/range references from formula text.

    Returns ``(tokens, name_candidates)`` where name_candidates are bare
    identifiers that may be named ranges (resolved by the caller).
    """
    body = _RE_STRING.sub('""', formula)
    tokens: list[RefToken] = []
    consumed: list[tuple[int, int]] = []

    for m in _RE_REF.finditer(body):
        start, end = m.span()
        if start > 0 and body[start - 1] in _BOUNDARY_BEFORE:
            continue
        nxt = body[end:end + 1]
        if nxt and (nxt in "(" or nxt.isalnum() or nxt in "_$"):
            # LOG10( … ) is a function, AB1CD is an identifier fragment
            continue
        sheetpart = m.group("sheetpart")
        ref = m.group("ref").replace(" ", "")
        sheet: Optional[str] = None
        external = False
        if sheetpart:
            raw_sheet = sheetpart[:-1]
            if raw_sheet.startswith("'") and raw_sheet.endswith("'"):
                raw_sheet = raw_sheet[1:-1].replace("''", "'")
            if "[" in raw_sheet:
                external = True
                sheet = raw_sheet
            else:
                resolved = sheet_names_lower.get(raw_sheet.lower())
                if resolved is None:
                    # reference to a sheet we don't have (deleted -> #REF!, or external)
                    external = True
                    sheet = raw_sheet
                else:
                    sheet = resolved
        else:
            # bare A1 without sheet: but guard against function names already done;
            # also guard plain column tokens matched as whole-col by requiring ':'
            pass

        if re.match(r"^\$?[A-Za-z]{1,3}:", ref) and not re.match(r"^\$?[A-Za-z]{1,3}\$?\d", ref):
            c1 = column_index_from_string(ref.split(":")[0].replace("$", "").upper())
            c2 = column_index_from_string(ref.split(":")[1].replace("$", "").upper())
            tokens.append(RefToken(sheet, 1, min(c1, c2), 1 << ROW_BITS, max(c1, c2),
                                   external=external, whole_col=True))
        elif re.match(r"^\$?\d", ref):
            r1 = int(ref.split(":")[0].replace("$", ""))
            r2 = int(ref.split(":")[1].replace("$", ""))
            tokens.append(RefToken(sheet, min(r1, r2), 1, max(r1, r2), 1 << COL_BITS,
                                   external=external, whole_row=True))
        elif ":" in ref:
            a, b = ref.split(":")
            ra, ca = _parse_cell(a)
            rb, cb = _parse_cell(b)
            tokens.append(RefToken(sheet, min(ra, rb), min(ca, cb), max(ra, rb), max(ca, cb), external=external))
        else:
            r, c = _parse_cell(ref)
            tokens.append(RefToken(sheet, r, c, r, c, external=external))
        consumed.append((start, end))

    # name candidates: identifiers outside consumed spans, not function calls
    names: list[str] = []
    for m in _RE_NAME.finditer(body):
        s, e = m.span()
        if any(cs <= s < ce for cs, ce in consumed):
            continue
        if body[e:e + 1] == "(":
            continue
        if s > 0 and body[s - 1] in "!'":
            continue
        tok = m.group(0)
        if _RE_CELL_SHAPE.match(tok) or tok.upper() in ("TRUE", "FALSE"):
            continue
        # bare column-letter shapes (AA, XFD) are stray range halves, not names
        if re.match(r"^[A-Za-z]{1,3}$", tok):
            continue
        names.append(tok)
    return tokens, names


# ---------------------------------------------------------------------------
# Graph
# ---------------------------------------------------------------------------

@dataclass
class FormulaGraph:
    wbd: WorkbookData
    sheet_index: dict[str, int]
    sheet_names: list[str]
    precedents: dict[int, list[int]] = field(default_factory=dict)
    dependents: dict[int, list[int]] = field(default_factory=dict)
    formula_nodes: set[int] = field(default_factory=set)
    external_ref_cells: list[tuple[str, str]] = field(default_factory=list)   # (cell ref, token text)
    unresolved_names: dict[str, int] = field(default_factory=dict)            # name -> count
    error_nodes: list[int] = field(default_factory=list)
    ref_error_formula_cells: list[str] = field(default_factory=list)          # formulas containing #REF!
    circular_groups: list[list[str]] = field(default_factory=list)            # qualified refs
    limitations: list[str] = field(default_factory=list)
    n_edges: int = 0

    # -- node helpers -------------------------------------------------------
    def node_ref(self, node: int) -> str:
        if node < 0:
            return f"<range#{-node}>"
        si, r, c = decode(node)
        return format_ref(self.sheet_names[si], a1(r, c))

    def node_of(self, sheet: str, row: int, col: int) -> int:
        return encode(self.sheet_index[sheet], row, col)

    # -- traversals ----------------------------------------------------------
    def downstream(self, start: int, cap: int = 200_000) -> set[int]:
        """All nodes reachable over dependent edges (bounded)."""
        seen: set[int] = set()
        dq = deque([start])
        while dq and len(seen) < cap:
            n = dq.popleft()
            for d in self.dependents.get(n, ()):  # noqa: B905
                if d not in seen:
                    seen.add(d)
                    dq.append(d)
        return seen

    def downstream_count_on_sheets(self, start: int, sheet_idxs: set[int], cap: int = 200_000) -> tuple[int, bool]:
        """(count of reachable cells on target sheets, truncated?)"""
        seen: set[int] = set()
        dq = deque([start])
        count = 0
        shift = ROW_BITS + COL_BITS
        while dq:
            if len(seen) >= cap:
                return count, True
            n = dq.popleft()
            for d in self.dependents.get(n, ()):
                if d not in seen:
                    seen.add(d)
                    if (d >> shift) in sheet_idxs:
                        count += 1
                    dq.append(d)
        return count, False

    def upstream(self, start: int, cap: int = 200_000) -> set[int]:
        seen: set[int] = set()
        dq = deque([start])
        while dq and len(seen) < cap:
            n = dq.popleft()
            for p in self.precedents.get(n, ()):
                if p not in seen:
                    seen.add(p)
                    dq.append(p)
        return seen

    def reaches_any(self, start: int, targets: set[int], cap: int = 200_000) -> bool:
        seen: set[int] = set()
        dq = deque([start])
        while dq and len(seen) < cap:
            n = dq.popleft()
            for d in self.dependents.get(n, ()):
                if d in targets:
                    return True
                if d not in seen:
                    seen.add(d)
                    dq.append(d)
        return False


def _row_cols_index(sd: SheetData) -> dict[int, list[int]]:
    idx: dict[int, list[int]] = {}
    for (r, c) in sd.cells:
        idx.setdefault(r, []).append(c)
    for cols in idx.values():
        cols.sort()
    return idx


def build_graph(wbd: WorkbookData) -> FormulaGraph:
    sheet_names = list(wbd.sheet_order)
    sheet_index = {n: i for i, n in enumerate(sheet_names)}
    sheet_lower = {n.lower(): n for n in sheet_names}
    g = FormulaGraph(wbd=wbd, sheet_index=sheet_index, sheet_names=sheet_names)

    # named-range resolution: name -> list of RefToken
    named: dict[str, list[RefToken]] = {}
    for name, target in wbd.named_ranges.items():
        toks: list[RefToken] = []
        for part in str(target).split(","):
            t, _ = tokenize_refs("=" + part.strip(), sheet_lower)
            toks.extend(t)
        if toks:
            named[name.lower()] = toks

    row_indexes = {name: _row_cols_index(sd) for name, sd in wbd.sheets.items()}

    # Large areas referenced by many formulas (spend curves, SUM rows) are
    # compressed into shared synthetic range-nodes (negative ids): the area's
    # members are linked once, and each referencing formula links to the node.
    # Traversals pass through them transparently.
    range_nodes: dict[tuple[int, int, int, int, int], int] = {}
    next_range_id = [-1]

    edges = 0
    for sname, sd in wbd.sheets.items():
        si = sheet_index[sname]
        for (r, c), rec in sd.cells.items():
            if rec.is_error:
                g.error_nodes.append(encode(si, r, c))
            if rec.formula is None:
                continue
            node = encode(si, r, c)
            g.formula_nodes.add(node)
            if "#REF!" in rec.formula:
                g.ref_error_formula_cells.append(format_ref(sname, a1(r, c)))
            toks, name_cands = tokenize_refs(rec.formula, sheet_lower)
            for nm in name_cands:
                if nm.lower() in named:
                    toks.extend(named[nm.lower()])
                elif nm.lower() in ("_xlfn",):
                    continue
                else:
                    # unknown identifier: only count it if it is plausibly a name
                    # (avoid noise from modern function prefixes etc.)
                    if "." not in nm and len(nm) > 1:
                        g.unresolved_names[nm] = g.unresolved_names.get(nm, 0) + 1
            precs: list[int] = []
            for tok in toks:
                if tok.external:
                    g.external_ref_cells.append((format_ref(sname, a1(r, c)), tok.sheet or "?"))
                    continue
                tsheet = tok.sheet or sname
                tsd = wbd.sheets.get(tsheet)
                if tsd is None:
                    continue
                tsi = sheet_index[tsheet]
                r1, c1 = tok.r1, tok.c1
                r2 = min(tok.r2, tsd.max_row)
                c2 = min(tok.c2, tsd.max_col)
                if r1 == r2 and c1 == c2:
                    precs.append(encode(tsi, r1, c1))
                    continue
                rkey = (tsi, r1, c1, r2, c2)
                if rkey in range_nodes:
                    precs.append(range_nodes[rkey])
                    continue
                # area: link only to populated cells (sparse sheets stay cheap)
                ridx = row_indexes[tsheet]
                members: list[int] = []
                truncated = False
                for rr in range(r1, r2 + 1):
                    cols = ridx.get(rr)
                    if not cols:
                        continue
                    lo = bisect_left(cols, c1)
                    hi = bisect_right(cols, c2)
                    for cc in cols[lo:hi]:
                        members.append(encode(tsi, rr, cc))
                        if len(members) >= RANGE_EXPANSION_CAP:
                            truncated = True
                            break
                    if truncated:
                        break
                if truncated:
                    g.limitations.append(
                        f"Range in {format_ref(sname, a1(r, c))} expanded to only the first "
                        f"{RANGE_EXPANSION_CAP:,} populated cells."
                    )
                if len(members) >= AREA_NODE_MIN:
                    rid = next_range_id[0]
                    next_range_id[0] -= 1
                    range_nodes[rkey] = rid
                    g.precedents[rid] = members
                    edges += len(members)
                    precs.append(rid)
                else:
                    precs.extend(members)
            if precs:
                # dedupe while preserving order
                seen: set[int] = set()
                uniq = [p for p in precs if not (p in seen or seen.add(p))]
                g.precedents[node] = uniq
                edges += len(uniq)
                if edges > MAX_EDGES:
                    g.limitations.append(
                        f"Formula graph truncated at {MAX_EDGES:,} edges; downstream analyses are partial."
                    )
                    g.n_edges = edges
                    _build_dependents(g)
                    return g
    g.n_edges = edges
    _build_dependents(g)
    _find_cycles(g)
    return g


def _build_dependents(g: FormulaGraph) -> None:
    deps: dict[int, list[int]] = {}
    for node, precs in g.precedents.items():
        for p in precs:
            deps.setdefault(p, []).append(node)
    g.dependents = deps


def _find_cycles(g: FormulaGraph, max_report: int = 10) -> None:
    """Iterative Tarjan SCC over formula nodes; reports non-trivial SCCs and self-loops."""
    index: dict[int, int] = {}
    low: dict[int, int] = {}
    on_stack: set[int] = set()
    stack: list[int] = []
    counter = 0
    sccs: list[list[int]] = []

    for root in g.precedents:
        if root in index:
            continue
        work = [(root, iter(g.precedents.get(root, ())))]
        index[root] = low[root] = counter
        counter += 1
        stack.append(root)
        on_stack.add(root)
        while work:
            node, it = work[-1]
            advanced = False
            for nxt in it:
                if nxt not in g.precedents:
                    continue  # only formula->formula edges can cycle
                if nxt not in index:
                    index[nxt] = low[nxt] = counter
                    counter += 1
                    stack.append(nxt)
                    on_stack.add(nxt)
                    work.append((nxt, iter(g.precedents.get(nxt, ()))))
                    advanced = True
                    break
                elif nxt in on_stack:
                    low[node] = min(low[node], index[nxt])
            if advanced:
                continue
            work.pop()
            if work:
                parent = work[-1][0]
                low[parent] = min(low[parent], low[node])
            if low[node] == index[node]:
                comp: list[int] = []
                while True:
                    w = stack.pop()
                    on_stack.discard(w)
                    comp.append(w)
                    if w == node:
                        break
                if len(comp) > 1 or node in g.precedents.get(node, ()):
                    sccs.append(comp)

    for comp in sccs[:max_report]:
        g.circular_groups.append([g.node_ref(n) for n in comp[:12]])
    if len(sccs) > max_report:
        g.limitations.append(f"{len(sccs)} circular-reference groups found; only {max_report} reported in detail.")


# ---------------------------------------------------------------------------
# Derived structural facts
# ---------------------------------------------------------------------------

def orphan_formula_rows(g: FormulaGraph, min_cells: int = 3) -> list[tuple[str, int, int]]:
    """Rows of formula cells nothing depends on: ``(sheet, row, n_cells)``.

    A computed block that feeds nothing is either presentation or a wired-to-
    nothing schedule (a real model expensed capex while a full depreciation
    schedule sat orphaned). Role-based filtering happens in findings.
    """
    from collections import Counter
    per_row: Counter[tuple[int, int]] = Counter()
    for node in g.formula_nodes:
        if g.dependents.get(node):
            continue
        si, r, _ = decode(node)
        per_row[(si, r)] += 1
    out = [(g.sheet_names[si], r, n) for (si, r), n in per_row.items() if n >= min_cells]
    out.sort(key=lambda t: (t[0], t[1]))
    return out


def errors_feeding_sheets(g: FormulaGraph, target_sheets: Iterable[str]) -> list[tuple[str, bool]]:
    """For each error cell: (qualified ref, does it feed any target sheet?)."""
    idxs = {g.sheet_index[s] for s in target_sheets if s in g.sheet_index}
    out: list[tuple[str, bool]] = []
    for node in g.error_nodes:
        cnt, _ = g.downstream_count_on_sheets(node, idxs, cap=50_000)
        out.append((g.node_ref(node), cnt > 0))
    return out


@dataclass
class InputImpact:
    sheet: str
    row: int
    col: int
    ref: str
    value: object
    out_degree: int
    statement_cells: int
    truncated: bool


def rank_inputs_by_impact(
    g: FormulaGraph,
    statement_sheets: Iterable[str],
    top_n: int = 400,
    bfs_cap: int = 150_000,
) -> list[InputImpact]:
    """Typed numeric inputs ranked by downstream statement impact.

    Exact bounded BFS for the ``top_n`` inputs by direct out-degree; anything
    beyond that cannot plausibly outrank them. Truncations are flagged, never
    silent.
    """
    idxs = {g.sheet_index[s] for s in statement_sheets if s in g.sheet_index}
    candidates: list[tuple[int, int]] = []   # (out_degree, node)
    for node, deps in g.dependents.items():
        if node in g.formula_nodes:
            continue
        si, r, c = decode(node)
        rec = g.wbd.sheets[g.sheet_names[si]].cell(r, c)
        if rec is None or rec.formula is not None or not is_number(rec.value):
            continue
        candidates.append((len(deps), node))
    candidates.sort(key=lambda t: -t[0])

    out: list[InputImpact] = []
    for deg, node in candidates[:top_n]:
        count, trunc = g.downstream_count_on_sheets(node, idxs, cap=bfs_cap)
        si, r, c = decode(node)
        sname = g.sheet_names[si]
        out.append(InputImpact(
            sheet=sname, row=r, col=c, ref=g.node_ref(node),
            value=g.wbd.sheets[sname].cell(r, c).value,
            out_degree=deg, statement_cells=count, truncated=trunc,
        ))
    out.sort(key=lambda x: (-x.statement_cells, -x.out_degree))
    return out

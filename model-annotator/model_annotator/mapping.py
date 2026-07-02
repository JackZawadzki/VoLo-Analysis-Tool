"""Canonical line-item mapping.

Rows (or columns, for transposed models) are extracted as LineItems and scored
against the canonical ontology with transparent regex/keyword heuristics.
Uncertain labels can optionally be classified by the LLM (labels only — the
LLM never sees or produces numbers). Mappings that fail their mandatory
arithmetic verification are demoted, never silently kept.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from typing import Callable, Optional

from .ingest import SheetData, WorkbookData, a1, is_number
from .schema import (
    Orientation,
    PeriodAxis,
    RowMapping,
    SheetRole,
    UnmappedRow,
    format_ref,
    parse_ref,
)
from .structure import StructureResult

log = logging.getLogger(__name__)

MAP_THRESHOLD = 0.55          # minimum confidence to accept a mapping
LLM_BAND = (0.30, 0.55)       # heuristic-score band where the LLM may assist


# ---------------------------------------------------------------------------
# Line-item extraction
# ---------------------------------------------------------------------------

@dataclass
class LineItem:
    sheet: str
    index: int                       # row (periods-in-columns) or column (transposed)
    label: str
    depth: int                       # leading-whitespace indentation level
    section: str                     # nearest preceding section header (normalized)
    periods: list[str]
    coords: list[tuple[int, int]]    # (row, col) per period
    values: list[Optional[float]]
    percentish: bool                 # values/format look like percentages
    n_numeric: int
    n_formula: int
    unit_scale_override: Optional[float] = None   # "($M)" in the label etc.

    # filled by mapping
    canonical_id: Optional[str] = None
    instance: Optional[str] = None
    subtags: list[str] = field(default_factory=list)
    confidence: float = 0.0
    evidence: str = ""
    demoted: bool = False

    @property
    def refs(self) -> list[str]:
        return [format_ref(self.sheet, a1(r, c)) for r, c in self.coords]

    def ref_for(self, period: str) -> Optional[str]:
        try:
            i = self.periods.index(period)
        except ValueError:
            return None
        r, c = self.coords[i]
        return format_ref(self.sheet, a1(r, c))

    def value_for(self, period: str) -> Optional[float]:
        try:
            return self.values[self.periods.index(period)]
        except ValueError:
            return None


_RE_ROW_UNIT = re.compile(r"\(\s*\$?\s*(m|mm|millions?)\s*\)", re.I)
_RE_ROW_UNIT_K = re.compile(r"\(\s*\$?\s*(k|000s?|thousands?)\s*\)", re.I)
# chart-feeder recap blocks ("For Summary Chart" + sign-flipped echoes of the
# statements) must never win mapping over the real statement rows
_RE_CHART_MARKER = re.compile(r"\bfor\b.*\bchart\b|chart data|^for summary\b", re.I)


def _coerce_value(v: object) -> Optional[float]:
    if is_number(v):
        return float(v)
    return None  # strings ("N/A"), dates, errors, blanks all become None


def extract_line_items(sd: SheetData, axis: PeriodAxis) -> list[LineItem]:
    """All candidate line items on a sheet for a given period axis."""
    header_coords: list[tuple[int, int]] = []
    for ref in axis.header_cells:
        _, cell = parse_ref(ref)
        m = re.match(r"([A-Z]+)(\d+)", cell)
        col_n = 0
        for ch in m.group(1):
            col_n = col_n * 26 + (ord(ch) - 64)
        header_coords.append((int(m.group(2)), col_n))

    items: list[LineItem] = []
    section = ""

    # the line index sitting ON the period-header axis is the header itself; its
    # "values" are the period labels (years), never a data row — exclude it so a
    # row like "Revenue Stream" headering the year row isn't mapped as a segment
    if axis.orientation == Orientation.periods_in_columns:
        period_cols = [c for _, c in header_coords]
        first_period_col = min(period_cols)
        line_indices = range(1, sd.max_row + 1)
        header_idx = {r for r, _ in header_coords}
    else:
        period_rows = [r for r, _ in header_coords]
        first_period_row = min(period_rows)
        line_indices = range(1, sd.max_col + 1)
        header_idx = {c for _, c in header_coords}

    for idx in line_indices:
        if idx in header_idx:
            continue
        if axis.orientation == Orientation.periods_in_columns:
            coords = [(idx, c) for c in period_cols]
            label_cells = [(idx, c) for c in range(1, first_period_col)]
        else:
            coords = [(r, idx) for r in period_rows]
            label_cells = [(r, idx) for r in range(1, first_period_row)]

        # Choose the most-specific label. Multi-column label blocks put the line
        # name in the left column and a generic category to its right
        # ("6010 PR & Marketing" in A, "Marketing" in C) — keep the leftmost
        # descriptive cell (has letters, not prose), not the rightmost generic.
        candidates = []
        for (lr, lc) in label_cells:
            rec = sd.cell(lr, lc)
            if rec is not None and isinstance(rec.value, str) and rec.value.strip():
                candidates.append(rec.value)
        if not candidates:
            continue
        raw_label = next((c for c in candidates if re.search(r"[A-Za-z]", c) and len(c) <= 80),
                         candidates[0])
        depth = len(raw_label) - len(raw_label.lstrip())
        label = raw_label.strip()

        if _RE_CHART_MARKER.search(label):
            section = _norm(label)   # poisons the rows below it even though this row has numbers
            continue

        recs = [sd.cell(r, c) for r, c in coords]
        values = [_coerce_value(rec.value) if rec else None for rec in recs]
        n_numeric = sum(1 for v in values if v is not None)
        n_formula = sum(1 for rec in recs if rec is not None and rec.is_formula)

        if n_numeric == 0:
            # no numbers across the axis: a section header
            if len(label) < 60:
                section = _norm(label)
            continue

        pct_fmt = sum(1 for rec in recs if rec is not None and "%" in (rec.number_format or ""))
        in_band = sum(1 for v in values if v is not None and -5.0 <= v <= 5.0)
        percentish = pct_fmt >= max(1, n_numeric // 2) or (
            "%" in label or "margin" in label.lower() or "rate" in label.lower()
        ) and in_band == n_numeric

        unit_override = None
        if _RE_ROW_UNIT.search(label):
            unit_override = 1_000_000.0
        elif _RE_ROW_UNIT_K.search(label):
            unit_override = 1_000.0

        items.append(LineItem(
            sheet=sd.name, index=idx, label=label, depth=depth, section=section,
            periods=list(axis.periods), coords=coords, values=values,
            percentish=percentish, n_numeric=n_numeric, n_formula=n_formula,
            unit_scale_override=unit_override,
        ))
    return items


def _norm(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^a-z0-9&%/$]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# ---------------------------------------------------------------------------
# Ontology scoring
# ---------------------------------------------------------------------------

# Each rule: (canonical_id, exact-ish patterns, keyword patterns, vetoes, base boosts)
# Exact patterns score 3, keywords 1.5 each (capped), vetoes kill the rule.

_DERIVED_VETO = re.compile(
    r"growth|% of|/ ?revenue|/ ?ebitda|per (unit|share|employee|fte|tpa|mwh|kg|machine)"
    r"|cumulative|cushion|\bratio\b|\bconversion\b|months of|runway|\bcheck\b|tie ?out"
    r"|ex ?grants?|share on",
)


@dataclass(frozen=True)
class OntologyRule:
    canonical_id: str
    exact: tuple[str, ...] = ()
    keywords: tuple[str, ...] = ()
    veto: tuple[str, ...] = ()
    percent_expected: Optional[bool] = None
    section_hint: tuple[str, ...] = ()       # bonus when the section matches
    section_veto: tuple[str, ...] = ()
    weak_keywords: bool = False              # generic keywords ("revenue") score below threshold alone
    prefix_suffix: tuple[str, ...] = ()      # "x revenue" / "cogs x" patterns; strong segment signal
    allow_cumulative: bool = False           # "Cumulative Capacity ..." is a real model row, not an analyst row


RULES: list[OntologyRule] = [
    OntologyRule("revenue_total",
                 exact=("total revenue", "total revenues", "revenue", "revenues", "total net revenue",
                        "total sales", "net revenue", "net sales"),
                 keywords=("total revenue",),
                 veto=("cogs", "cost", "%", "growth", "ex grant", "per ", "deferred"),
                 percent_expected=False),
    OntologyRule("grant_or_subsidy_revenue",
                 exact=("grant revenue", "grants", "grant income", "subsidy revenue", "government grants",
                        "government subsidies", "subsidies", "grants & subsidies"),
                 # keyword matching is whole-word, so list the inflections that
                 # actually appear ("grants", "subsidy", "subsidies") — a bare
                 # "grant" stem would miss "...including grants".
                 keywords=("grant", "grants", "subsidy", "subsidies",
                           "government funding", "government grant"),
                 # bare "/" removed: it blocked legitimate lines like "Other
                 # revenues/expenses including grants"; ratio slashes (/revenue,
                 # /ebitda) are still caught by _DERIVED_VETO.
                 veto=("%", "growth", "change in", "received"),
                 percent_expected=False),
    OntologyRule("revenue_segment",
                 keywords=("revenue", "sales"),
                 weak_keywords=True,
                 prefix_suffix=(" revenue", "revenue "),
                 veto=("total", "cogs", "cost", "%", "growth", "ex grant", "per ", "grant", "deferred",
                       "cumulative", "share", "subsid"),
                 percent_expected=False),
    OntologyRule("cogs_total",
                 exact=("cogs", "total cogs", "cost of goods sold", "cost of sales", "cost of revenue",
                        "total cost of goods sold", "total cost of revenue"),
                 keywords=("cost of goods", "cost of revenue", "cost of sales"),
                 veto=("%", "growth", "/", "per "),
                 percent_expected=False),
    OntologyRule("cogs_component",
                 keywords=("material", "freight", "warranty", "direct labor", "manufacturing cost",
                           "stack cost", "unit cost"),
                 prefix_suffix=("cogs ", " cogs"),
                 veto=("total", "%", "growth", "/", "change in"),
                 section_hint=("cogs", "cost of goods", "cost of revenue", "cost of sales"),
                 section_veto=("expense", "development", "opex", "operating", "balance", "liabilit",
                               "assets", "cash flow", "activities"),
                 percent_expected=False),
    OntologyRule("gross_profit",
                 exact=("gross profit", "total gross profit", "gross margin $", "gross profit loss"),
                 keywords=("gross profit",),
                 veto=("%", "growth", "ex grant"),
                 percent_expected=False),
    OntologyRule("gross_margin_pct",
                 exact=("gross margin", "gross margin %", "total gross profit margin", "gross profit margin",
                        "gross profit %"),
                 keywords=("gross margin",),
                 veto=("ex grant", "growth", "$"),
                 percent_expected=True),
    OntologyRule("opex_total",
                 exact=("total opex", "total operating expenses", "operating expenses", "opex",
                        "total expenses", "total operating costs", "total sg&a",
                        "total cash operating expenses", "cash operating expenses"),
                 keywords=("total operating", "operating expenses"),
                 veto=("%", "growth", "/", "cogs"),
                 percent_expected=False),
    OntologyRule("opex_component",
                 keywords=("r&d", "research", "development", "sales & marketing", "marketing",
                           "g&a", "general & admin", "general and admin", "payroll", "salaries",
                           "compensation", "rent", "facilities", "lease", "travel", "legal",
                           "professional", "insurance", "software", "overhead", "support",
                           "capital equipment", "equipment purchase", "capex", "capital expenditure",
                           "lab supplies", "supplies", "materials", "utilities", "labor",
                           "advertising", "bank charges", "recruitment", "phantom share",
                           "stock comp", "share-based comp", "contractors"),
                 veto=("total", "%", "growth", "cogs", "cost of goods", "depreciation", "change in"),
                 section_veto=("cost of goods", "cogs", "cash flow", "investing", "financing",
                               "activities", "balance", "liabilit", "assets", "working capital"),
                 percent_expected=False),
    OntologyRule("ebitda_or_ebit",
                 exact=("ebitda", "ebit", "operating income", "operating profit", "adjusted ebitda",
                        "ebitda before expensed capex", "operating income loss"),
                 keywords=("ebitda",),
                 veto=("%", "margin", "growth", "/", "interest", "taxes"),
                 percent_expected=False),
    OntologyRule("depreciation",
                 exact=("depreciation", "depreciation & amortization", "d&a", "depreciation and amortization"),
                 keywords=("depreciation", "amortization"),
                 veto=("%", "schedule", "accumulated"),
                 percent_expected=False),
    OntologyRule("interest_net",
                 exact=("interest", "net interest", "interest expense", "interest income", "interest net",
                        "total net interest income", "net interest income", "net interest expense"),
                 keywords=("interest",),
                 veto=("%", "/", "rate", "ebitda"),
                 percent_expected=False),
    OntologyRule("taxes",
                 exact=("taxes", "tax", "income tax", "income taxes", "tax expense", "taxes paid"),
                 keywords=("income tax",),
                 veto=("%", "/", "rate", "credit", "deferred", "payroll"),
                 percent_expected=False),
    OntologyRule("net_income",
                 exact=("net income", "net profit", "net earnings", "net income loss", "net loss",
                        "net profit after tax", "profit after tax"),
                 keywords=("net income",),
                 veto=("%", "margin", "growth"),
                 percent_expected=False),
    OntologyRule("cf_operating",
                 exact=("cash from operations", "operating cash flow", "net cash from operating activities",
                        "cash flow from operations", "cash flows from operating activities",
                        "net cash provided by operating activities", "total operating activities",
                        "net cf from operating activities", "cf from operating activities"),
                 keywords=("operating activities", "cash from operations"),
                 veto=("%", "investing", "financing"),
                 section_hint=("cash flow", "statement of cash"),
                 percent_expected=False),
    OntologyRule("capex",
                 exact=("capex", "capital expenditures", "capital expenditure", "purchases of property and equipment",
                        "purchase of pp&e", "equipment purchases", "manufacturing capex", "facility capex"),
                 keywords=("capital expenditure", "purchases of property"),
                 veto=("%", "/", "cumulative", "depreciation"),
                 section_hint=("investing", "cash flow"),
                 percent_expected=False),
    OntologyRule("cf_investing",
                 exact=("cash from investing", "net cash used in investing activities",
                        "cash flows from investing activities", "total investing activities",
                        "investing activities"),
                 keywords=("investing activities",),
                 veto=("%",),
                 percent_expected=False),
    OntologyRule("debt_proceeds",
                 exact=("debt proceeds", "proceeds from debt", "new debt", "debt raised", "change in debt",
                        "loan proceeds", "borrowings", "equipment financing", "debt draw", "venture debt"),
                 keywords=("debt", "loan", "borrowing"),
                 veto=("%", "repayment", "service", "interest", "balance", "/"),
                 section_hint=("financing",),
                 percent_expected=False),
    OntologyRule("equity_proceeds",
                 exact=("equity proceeds", "equity raise", "proceeds from equity", "equity raised",
                        "new equity", "equity financing", "capital raise", "equity investment",
                        "series a", "series b", "series c", "preferred stock issuance"),
                 keywords=("equity raise", "equity proceeds", "stock issuance"),
                 veto=("%", "cumulative", "balance", "/"),
                 section_hint=("financing",),
                 percent_expected=False),
    OntologyRule("cf_financing",
                 exact=("cash from financing", "net cash from financing activities",
                        "cash flows from financing activities", "total financing activities",
                        "financing activities"),
                 keywords=("financing activities",),
                 veto=("%",),
                 percent_expected=False),
    OntologyRule("net_change_cash",
                 exact=("net change in cash", "change in cash", "net cash flow", "net increase in cash",
                        "net increase decrease in cash", "total cash flow", "net change"),
                 keywords=("net change in cash", "change in cash"),
                 veto=("%", "cumulative"),
                 percent_expected=False),
    OntologyRule("beginning_cash",
                 exact=("beginning cash", "cash beginning", "cash at beginning of period", "starting cash",
                        "beginning cash balance", "cash beginning of year", "bop cash"),
                 keywords=("beginning cash", "beginning of period"),
                 veto=("%",),
                 percent_expected=False),
    OntologyRule("ending_cash",
                 exact=("ending cash", "cash ending", "cash at end of period", "ending cash balance",
                        "cash balance", "cash & equivalents", "cash and cash equivalents", "eop cash",
                        "cash end of year", "total cash"),
                 keywords=("ending cash", "end of period"),
                 veto=("%", "months", "runway", "beginning"),
                 percent_expected=False),
    OntologyRule("balance_check",
                 exact=("check", "balance check", "balance sheet check", "tie out", "cash check",
                        "negative cash check"),
                 keywords=("check",),
                 veto=("%",),
                 percent_expected=None),
    OntologyRule("headcount",
                 exact=("headcount", "total headcount", "ftes", "total ftes", "employees",
                        "total employees", "total staff"),
                 keywords=("headcount", "fte"),
                 veto=("%", "cost", "per ", "revenue"),
                 percent_expected=False),
    OntologyRule("units_sold",
                 keywords=("units sold", "units shipped", "systems sold", "tools sold", "deployments",
                           "installs", "units delivered", "plants delivered", "projects completed",
                           "number of units", "unit sales"),
                 veto=("%", "price", "cost", "revenue", "capacity"),
                 percent_expected=False),
    OntologyRule("asp",
                 keywords=("asp", "average selling price", "selling price", "price per unit",
                           "price per system", "unit price"),
                 veto=("%", "cost", "growth"),
                 percent_expected=False),
    OntologyRule("capacity",
                 keywords=("capacity", "nameplate", "contracted tpa", "operational tpa", "tpa",
                           "production volume", "throughput"),
                 veto=("%", "per ", "utilization", "share", "revenue", "market"),
                 percent_expected=False,
                 allow_cumulative=True),
    OntologyRule("market_size",
                 exact=("market size", "tam", "sam", "som", "total addressable market", "total market"),
                 keywords=("market size", "addressable market", "market demand"),
                 veto=("%", "share"),
                 percent_expected=False),
    OntologyRule("market_share",
                 exact=("market share", "share of market", "penetration", "market penetration"),
                 keywords=("market share", "penetration"),
                 veto=("growth",),
                 percent_expected=True),
    OntologyRule("valuation_multiple",
                 keywords=("multiple", "exit multiple", "ebitda multiple", "revenue multiple", "x ebitda",
                           "x revenue"),
                 veto=("%",),
                 percent_expected=False),
    OntologyRule("discount_rate",
                 exact=("discount rate", "wacc", "hurdle rate"),
                 keywords=("discount rate",),
                 percent_expected=True),
    OntologyRule("investment_in",
                 exact=("investment", "our investment", "check size", "investment amount"),
                 keywords=("investment",),
                 veto=("%", "cumulative", "return", "activities", "income"),
                 percent_expected=False),
    OntologyRule("accounts_receivable",
                 exact=("accounts receivable", "a/r", "ar", "receivables", "trade receivables"),
                 keywords=("receivable",),
                 veto=("%", "days", "/ revenue", "/ beginning", "/ ebitda"),
                 percent_expected=False),
    OntologyRule("accounts_payable",
                 exact=("accounts payable", "a/p", "ap", "payables", "trade payables"),
                 keywords=("payable",),
                 veto=("%", "days", "/ revenue", "/ beginning", "/ ebitda"),
                 percent_expected=False),
    OntologyRule("inventory",
                 exact=("inventory", "inventories"),
                 keywords=("inventory",),
                 veto=("%", "days", "turns"),
                 percent_expected=False),
]

_CAPEX_LIKE = re.compile(r"capital equipment|equipment purchase|capex|capital expenditure")
_OPEX_SUBTAG_PATTERNS: list[tuple[str, re.Pattern]] = [
    ("capex_expensed", _CAPEX_LIKE),
    ("rnd", re.compile(r"r&d|research|development|engineering")),
    ("sm", re.compile(r"sales|marketing|s&m|business development")),
    ("ga", re.compile(r"g&a|general|admin|legal|professional|insurance|accounting")),
    ("labor", re.compile(r"payroll|salar|compensation|labor|wages|benefits")),
    ("rent", re.compile(r"rent|facilit|lease|office")),
    ("support", re.compile(r"support|service|maintenance|customer success")),
]


def score_label(item: LineItem) -> list[tuple[str, float, str]]:
    """All (canonical_id, score 0..1, evidence) candidates for a line item, best first."""
    norm = _norm(item.label)
    if not norm:
        return []
    if "chart" in item.section:
        return []
    derived_hit = _DERIVED_VETO.search(norm)
    derived_hit_ex_cum = _DERIVED_VETO.search(norm.replace("cumulative", ""))
    out: list[tuple[str, float, str]] = []
    for rule in RULES:
        if any(v in norm for v in rule.veto):
            continue
        if rule.section_veto and any(sv in item.section for sv in rule.section_veto):
            continue
        # derived/analyst rows (growth, ratios, cumulative) never map, except checks
        hit = derived_hit_ex_cum if rule.allow_cumulative else derived_hit
        if hit and rule.canonical_id != "balance_check":
            continue
        score = 0.0
        ev = ""
        if norm in rule.exact:
            score = 3.0
            ev = f"label {item.label!r} exactly matches {rule.canonical_id}"
        elif rule.prefix_suffix and any(
                norm.endswith(ps) if ps.startswith(" ") else norm.startswith(ps)
                for ps in rule.prefix_suffix):
            score = 2.2
            ev = f"label {item.label!r} matches {rule.canonical_id} segment pattern"
        else:
            hits = [kw for kw in rule.keywords
                    if re.search(rf"(?<![a-z0-9]){re.escape(kw)}(?![a-z0-9])", norm)]
            if hits:
                base = 1.2 if rule.weak_keywords else 1.8
                score = min(2.4, base + 0.4 * (len(hits) - 1))
                ev = f"label {item.label!r} contains {hits!r}"
        if score == 0.0:
            continue
        if rule.section_hint and any(sh in item.section for sh in rule.section_hint):
            score += 0.5
            ev += f"; section {item.section!r}"
        if rule.percent_expected is True and not item.percentish:
            score -= 1.0
        if rule.percent_expected is False and item.percentish:
            score -= 1.2
        if score <= 0:
            continue
        out.append((rule.canonical_id, min(0.95, score / 3.0), ev))
    out.sort(key=lambda t: -t[1])
    return out


def _instance_from_label(canonical_id: str, label: str) -> Optional[str]:
    lab = label.strip()
    low = lab.lower()
    if canonical_id == "revenue_segment":
        inst = re.sub(r"\b(revenue|sales)\b", "", low, flags=re.I).strip(" -–—:_")
        return inst or None
    if canonical_id in ("cogs_component", "opex_component", "units_sold", "asp", "cost_tier"):
        return lab
    if canonical_id == "capacity":
        for kind in ("operational", "contracted", "nameplate", "installed", "total"):
            if kind in low:
                return kind
        return lab
    return None


# ---------------------------------------------------------------------------
# Mapping result
# ---------------------------------------------------------------------------

@dataclass
class MappingResult:
    items: list[LineItem] = field(default_factory=list)              # all extracted items, mapped or not
    primary_sheet: Optional[str] = None
    unmapped: list[UnmappedRow] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    alternates: dict[str, list[str]] = field(default_factory=dict)   # canonical_id -> losing candidate refs

    def mapped(self) -> list[LineItem]:
        return [it for it in self.items if it.canonical_id and not it.demoted]

    def get(self, canonical_id: str, sheet: Optional[str] = None) -> Optional[LineItem]:
        """Best non-demoted item for a scalar canonical id (primary sheet preferred)."""
        sheet = sheet or self.primary_sheet
        cands = [it for it in self.items if it.canonical_id == canonical_id and not it.demoted]
        if not cands:
            return None
        cands.sort(key=lambda it: (it.sheet != sheet, -it.confidence))
        return cands[0]

    def get_all(self, canonical_id: str, sheet: Optional[str] = None) -> list[LineItem]:
        out = [it for it in self.items if it.canonical_id == canonical_id and not it.demoted]
        if sheet:
            primary = [it for it in out if it.sheet == sheet]
            return primary or out
        return out

    def to_row_mappings(self) -> list[RowMapping]:
        out = []
        for it in self.items:
            if not it.canonical_id:
                continue
            out.append(RowMapping(
                canonical_id=it.canonical_id, instance=it.instance, subtags=it.subtags,
                sheet=it.sheet, row=it.index, label_text=it.label,
                confidence=round(it.confidence, 2), evidence=it.evidence, demoted=it.demoted,
                cells={p: format_ref(it.sheet, a1(r, c))
                       for p, (r, c) in zip(it.periods, it.coords)},
            ))
        return out


LabelClassifier = Callable[[str, str, list[str]], Optional[tuple[str, float]]]
# (label, section, ontology_ids) -> (canonical_id, confidence) | None


def _match_period(cellval, period: str) -> bool:
    """True if a header cell denotes the given canonical period label (year-level)."""
    if cellval is None:
        return False
    try:
        from datetime import datetime, date
        if isinstance(cellval, (datetime, date)):
            return str(cellval.year) == period
    except Exception:
        pass
    s = str(cellval).strip()
    if s == period:
        return True
    try:
        if str(int(float(s))) == period:
            return True
    except Exception:
        pass
    return bool(period) and period in s


def _reconcile_multiblock_sheets(res: "MappingResult", structure: StructureResult, wbd) -> None:
    """A source sheet can carry more than one period-column block for the SAME
    years (prior vs current presentation, base vs downside, budget vs actual).
    Extraction reads the leftmost block by default, which silently corrupts every
    metric built on that sheet. Detect repeated period runs in the header row, and
    re-point the sheet's rows to whichever block reconciles (column totals within
    2%) to the PRIMARY output sheet's revenue. Generic: keyed to the detected
    condition, no cell addresses. If none reconciles, flag ambiguity and keep the
    leftmost. No-ops on single-block sheets."""
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    prim = res.primary_sheet
    if not prim:
        return
    prim_rev = next((it for it in res.items if it.sheet == prim
                     and it.canonical_id == "revenue_total" and not it.demoted), None)
    if prim_rev is None:
        return
    ref = {p: prim_rev.value_for(p) for p in prim_rev.periods}
    ref_periods = [p for p in prim_rev.periods if ref.get(p) is not None]
    if not ref_periods:
        return

    for sheet in {it.sheet for it in res.items}:
        if sheet == prim:
            continue
        axes = [ax for ax in structure.axes.get(sheet, [])
                if ax.orientation == Orientation.periods_in_columns]
        if not axes:
            continue
        periods = list(axes[0].periods)
        if len(periods) < 2:
            continue
        sd = wbd.sheets.get(sheet)
        if sd is None:
            continue
        try:
            hdr_row = coordinate_from_string(axes[0].header_cells[0].split("!")[-1])[1]
        except Exception:
            continue
        # scan the header row for EVERY column matching each period, then form
        # blocks from the k-th occurrence of each period (block 0 leftmost)
        occ = {p: [c for c in range(1, sd.max_col + 1)
                   if _match_period(sd.cell(hdr_row, c).value if sd.cell(hdr_row, c) else None, p)]
               for p in periods}
        nblocks = min((len(occ[p]) for p in periods), default=0)
        if nblocks < 2:
            continue                                     # single block -> nothing to disambiguate
        blocks = [{p: occ[p][k] for p in periods} for k in range(nblocks)]

        # the sheet's own parent/total row to reconcile: its revenue_total row,
        # else the sum of its segment rows
        rev_it = next((it for it in res.items if it.sheet == sheet
                       and it.canonical_id == "revenue_total"), None)
        seg_rows = [it.index for it in res.items if it.sheet == sheet
                    and it.canonical_id == "revenue_segment"]

        def block_total(cols):
            out = {}
            for p in ref_periods:
                c = cols.get(p)
                if c is None:
                    out[p] = None
                    continue
                if rev_it is not None:
                    rec = sd.cell(rev_it.index, c)
                    out[p] = rec.value if rec else None
                elif seg_rows:
                    s, ok = 0.0, False
                    for rr in seg_rows:
                        rec = sd.cell(rr, c)
                        v = rec.value if rec else None
                        if isinstance(v, (int, float)) and not isinstance(v, bool):
                            s += v; ok = True
                    out[p] = s if ok else None
                else:
                    out[p] = None
            return out

        def reconciles(vals):
            n = ok = 0
            for p in ref_periods:
                r, v = ref.get(p), vals.get(p)
                if r is None or not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                n += 1
                if abs(r) < 1e-9:
                    ok += (abs(v) < 1e-6)
                elif abs(v - r) <= 0.02 * abs(r):
                    ok += 1
            return n > 0 and ok == n

        chosen = next((b for b in blocks if reconciles(block_total(b))), None)
        leftmost = blocks[0]
        if chosen is None:
            res.notes.append(
                f"CITATION source ambiguity: {sheet} carries multiple presentation blocks and none "
                f"reconciles to {prim} revenue; kept the leftmost block — verify which is current.")
            continue
        if chosen == leftmost:
            continue                                     # leftmost already the right block
        # re-point every row on this sheet to the reconciling block's columns
        moved = 0
        for it in res.items:
            if it.sheet != sheet:
                continue
            if any(p not in chosen for p in it.periods):
                continue
            it.coords = [(it.index, chosen[p]) for p in it.periods]
            it.values = [_coerce_value(sd.cell(r, c).value if sd.cell(r, c) else None)
                         for (r, c) in it.coords]
            it.evidence = (it.evidence or "") + "; re-pointed to the presentation block reconciling to primary"
            moved += 1
        if moved:
            res.notes.append(
                f"{sheet}: multiple presentation blocks for {periods[0]}–{periods[-1]}; selected the "
                f"block whose totals reconcile to {prim} revenue and re-pointed {moved} rows to it.")


def _infer_segments_by_sum(res: "MappingResult", scalar_best: dict) -> None:
    """Mark the unmapped numeric rows just above a sheet's revenue_total whose
    per-period values SUM to that total as its revenue segments.

    This recovers a revenue breakdown *generically* — independent of how the
    segment rows are labelled (product names like "Consumer electronics" carry no
    revenue keyword for score_label to catch) — using the one property that is
    always true of a breakdown: the parts add up to the whole. Replaces what the
    LLM was doing for segment discovery with a deterministic, repeatable rule.
    """
    for (sheet, cid), rev in list(scalar_best.items()):
        if cid != "revenue_total":
            continue
        P = list(rev.periods)
        rev_vals = {p: rev.value_for(p) for p in P}
        cands = []
        for it in res.items:
            if it.sheet != sheet or it.canonical_id or it.demoted:
                continue
            if not (rev.index - 15 <= it.index < rev.index):     # window just above the total
                continue
            if it.percentish or _DERIVED_VETO.search(_norm(it.label)):
                continue
            if not any(v not in (None, 0) for v in it.values):    # has real numbers
                continue
            cands.append(it)
        cands.sort(key=lambda it: it.index)
        if len(cands) < 2:                                        # a breakdown has >=2 parts
            continue

        def _sums(block) -> bool:
            for p in P:
                tot = rev_vals.get(p)
                if tot is None:
                    continue
                s = sum((it.value_for(p) or 0.0) for it in block)
                if abs(tot) < 1.0:
                    if abs(s) > 1.0:
                        return False
                elif abs(s - tot) > 0.02 * abs(tot):              # within 2% every period
                    return False
            return True

        # the breakdown is the largest contiguous tail (ending just above the
        # total) whose parts add to the whole — trims stray non-segment rows / a
        # repeated subtotal sitting on top of the real segments
        block = None
        for start in range(len(cands) - 1):
            tail = cands[start:]
            if _sums(tail):
                block = tail
                break
        if not block:
            continue
        for it in block:
            it.canonical_id = "revenue_segment"
            it.confidence = 0.7
            it.instance = it.label.strip()[:40]
            it.evidence = f"sums with siblings to {rev.label!r} (revenue breakdown, {sheet})"


def build_mapping(
    wbd: WorkbookData,
    structure: StructureResult,
    llm_classify: Optional[LabelClassifier] = None,
) -> MappingResult:
    res = MappingResult()

    # Extract line items from every sheet that has a period axis.
    for name, sd in wbd.sheets.items():
        axis = structure.primary_axis(name)
        if axis is None:
            continue
        role, _ = structure.roles.get(name, (SheetRole.unknown, 0.0))
        if role in (SheetRole.documentation, SheetRole.presentation, SheetRole.scratch):
            continue
        res.items.extend(extract_line_items(sd, axis))

    # Score and assign.
    scalar_best: dict[tuple[str, str], LineItem] = {}   # (sheet, canonical_id) -> winner
    section_seg_candidates: list[LineItem] = []
    for it in res.items:
        cands = score_label(it)
        if not cands:
            # section-based segment inference candidate: numeric rows sitting
            # under a "Revenue" section header can be revenue segments even
            # when their labels carry no revenue keyword ("NRE — allocation").
            # Resolved AFTER scoring, with magnitude guards against valuation/
            # factor rows that merely live under revenue-ish headers.
            norm = _norm(it.label)
            if (re.match(r"^revenue\b", it.section) and not _DERIVED_VETO.search(norm)
                    and not it.percentish
                    and not any(v in it.section for v in ("valuation", "multiple", "exit", "factor", "adjust"))
                    and not re.match(r"^(gp|gm|ebitda|ebit)\b", norm)
                    and not any(v in norm for v in ("total", "cumulative", "grant", "subsid",
                                                    "valuation", "exit", "npv", "irr", "multiple",
                                                    "gross profit", "gross margin", "gp ", "gm ",
                                                    "cogs", "cost", "ebitda", "ebit", "net income",
                                                    "opex", "expense"))):
                section_seg_candidates.append(it)
                continue
        if not cands and llm_classify is not None:
            llm_out = llm_classify(it.label, it.section, [])
            if llm_out:
                cid, conf = llm_out
                it.canonical_id, it.confidence = cid, min(conf, 0.7)
                it.evidence = "LLM label classification (heuristics scored 0)"
            continue
        if not cands:
            continue
        cid, conf, ev = cands[0]
        if conf < MAP_THRESHOLD and llm_classify is not None and conf >= LLM_BAND[0]:
            llm_out = llm_classify(it.label, it.section, [c[0] for c in cands[:4]])
            if llm_out:
                cid2, conf2 = llm_out
                if cid2 in {c[0] for c in cands[:4]} or conf2 >= 0.7:
                    cid, conf, ev = cid2, min(conf2, 0.7), ev + "; LLM-assisted"
        if conf < MAP_THRESHOLD:
            continue
        it.canonical_id, it.confidence, it.evidence = cid, conf, ev
        it.instance = _instance_from_label(cid, it.label)
        if cid == "opex_component":
            it.subtags = [tag for tag, rx in _OPEX_SUBTAG_PATTERNS if rx.search(_norm(it.label))]

    # Scalar ids: one winner per sheet; losers become alternates/unmapped.
    # Grant/subsidy rows are exempt: models legitimately carry several
    # (grants + subsidies + credits) and metrics sum them.
    # Ties resolve toward the sheet's main statement block (distance to the
    # median index of confidently-mapped rows) — recap/presentation echoes of
    # the statements sit far from the main block and must not win.
    from .schema import CANONICAL_IDS
    median_idx: dict[str, float] = {}
    for sheet in {it.sheet for it in res.items}:
        idxs = sorted(it.index for it in res.items
                      if it.sheet == sheet and it.canonical_id and it.confidence >= 0.9)
        if idxs:
            median_idx[sheet] = idxs[len(idxs) // 2]

    def _beats_in_sheet(a: LineItem, b: LineItem) -> bool:
        if abs(a.confidence - b.confidence) > 0.05:
            return a.confidence > b.confidence
        if a.n_formula != b.n_formula:
            return a.n_formula > b.n_formula
        med = median_idx.get(a.sheet)
        if med is not None:
            return abs(a.index - med) < abs(b.index - med)
        return a.index > b.index

    for it in res.items:
        if it.canonical_id == "grant_or_subsidy_revenue":
            it.instance = it.label
            continue
        if it.canonical_id not in CANONICAL_IDS:
            continue
        key = (it.sheet, it.canonical_id)
        cur = scalar_best.get(key)
        if cur is None:
            scalar_best[key] = it
        else:
            better, worse = (it, cur) if _beats_in_sheet(it, cur) else (cur, it)
            scalar_best[key] = better
            worse_ref = format_ref(worse.sheet, f"row {worse.index}")
            res.alternates.setdefault(worse.canonical_id or "", []).append(
                f"{worse_ref} ({worse.label!r}, conf {worse.confidence:.2f})")
            worse.canonical_id = None
            worse.confidence = 0.0

    # Resolve section-based segment candidates now that revenue_total winners
    # are known: same-sheet revenue must exist and the candidate's magnitude
    # must sit inside a sane share band of it.
    for it in section_seg_candidates:
        rev = scalar_best.get((it.sheet, "revenue_total"))
        if rev is None:
            continue
        seg_max = max((abs(v) for v in it.values if v is not None), default=0.0)
        rev_max = max((abs(v) for v in rev.values if v is not None), default=0.0)
        if rev_max <= 0 or not (0.001 * rev_max <= seg_max <= 1.5 * rev_max):
            continue
        it.canonical_id = "revenue_segment"
        it.confidence = 0.6
        it.instance = it.label.strip()[:40]
        it.evidence = f"numeric row under section {it.section!r} within revenue magnitude band"

    # Structural segment inference: rows that ADD UP to a sheet's revenue_total
    # are its segment breakdown (works when labels carry no revenue keyword).
    _infer_segments_by_sum(res, scalar_best)

    # De-duplicate multi-instance rows that are echoes of each other: the same
    # canonical id with the same instance label AND an identical value series is
    # the same line repeated (e.g. a revenue segment shown in the income
    # statement and again in a per-segment summary). Keep the one with the most
    # formula cells (the live one); demote the rest so they aren't double-counted.
    from collections import defaultdict
    groups: dict[tuple, list[LineItem]] = defaultdict(list)
    for it in res.items:
        if it.canonical_id in ("revenue_segment", "cogs_component", "opex_component") and not it.demoted:
            sig = tuple(round(v, 4) if v is not None else None for v in it.values)
            inst = (it.instance or it.label.strip()).lower()
            groups[(it.sheet, it.canonical_id, inst, sig)].append(it)
    for dupes in groups.values():
        if len(dupes) > 1:
            keep = max(dupes, key=lambda it: (it.n_formula, -it.index))
            for it in dupes:
                if it is not keep:
                    it.demoted = True
                    it.evidence += "; demoted: duplicate echo of the same line"

    # Grant/subsidy echoes: two grant lines with a byte-identical value series are
    # the same grant mirrored across statements (the P&L grant line and the cash-
    # flow "(+) increase in grants"). Grants are intentionally multi-instance, so
    # match on the SERIES alone (the echo is deliberately relabelled) and keep one
    # so the grant total isn't double-counted.
    grant_groups: dict[tuple, list[LineItem]] = defaultdict(list)
    for it in res.items:
        if it.canonical_id == "grant_or_subsidy_revenue" and not it.demoted:
            if not any(v not in (None, 0) for v in it.values):
                continue
            sig = tuple(round(v, 4) if v is not None else None for v in it.values)
            grant_groups[sig].append(it)
    for dupes in grant_groups.values():
        if len(dupes) > 1:
            keep = max(dupes, key=lambda it: (it.n_formula, -it.index))
            for it in dupes:
                if it is not keep:
                    it.demoted = True
                    it.evidence += "; demoted: duplicate grant echo (same series)"

    # Unmapped bucket: numeric rows on statement-role sheets that didn't map.
    for it in res.items:
        if it.canonical_id is None and it.n_numeric >= 2:
            role, _ = structure.roles.get(it.sheet, (SheetRole.unknown, 0.0))
            if role == SheetRole.statements and not _DERIVED_VETO.search(_norm(it.label)):
                res.unmapped.append(UnmappedRow(
                    sheet=it.sheet, row=it.index, label=it.label[:120],
                    reason="no ontology rule scored above threshold"))

    res.primary_sheet = _pick_primary_sheet(res, structure)
    # P0-1: a sheet may carry several period-column blocks for the same years
    # (prior vs current); re-point each to the block reconciling to primary revenue
    _reconcile_multiblock_sheets(res, structure, wbd)
    _verify_and_demote(res)

    # surface any row whose own unit marker ("Revenue ($M)") conflicts with its
    # sheet scale — otherwise that row's magnitude is silently off by the ratio
    for it in res.items:
        if it.canonical_id and it.unit_scale_override is not None:
            su = structure.units.get(it.sheet)
            sheet_scale = su.scale if su else 1.0
            if abs(it.unit_scale_override - sheet_scale) > 1e-6:
                res.notes.append(
                    f"Row {it.sheet}!{it.index} {it.label.strip()!r} carries its own unit marker "
                    f"(scale {it.unit_scale_override:g}) different from the sheet scale "
                    f"({sheet_scale:g}); its absolute magnitude may be off by "
                    f"{it.unit_scale_override / sheet_scale:g}x — verify before relying on it.")
    return res


_CORE_IDS = ("revenue_total", "cogs_total", "gross_profit", "opex_total", "ebitda_or_ebit",
             "net_income", "ending_cash", "net_change_cash", "cf_operating")


def _pick_primary_sheet(res: MappingResult, structure: StructureResult) -> Optional[str]:
    scores: dict[str, float] = {}
    for it in res.items:
        if it.canonical_id in _CORE_IDS:
            role, _ = structure.roles.get(it.sheet, (SheetRole.unknown, 0.0))
            w = 1.0 + (0.5 if role == SheetRole.statements else 0.0)
            scores[it.sheet] = scores.get(it.sheet, 0.0) + w
    if not scores:
        return None
    return max(scores.items(), key=lambda kv: kv[1])[0]


# ---------------------------------------------------------------------------
# Mandatory mapping verifications
# ---------------------------------------------------------------------------

def _series(it: Optional[LineItem]) -> dict[str, float]:
    if it is None:
        return {}
    return {p: v for p, v in zip(it.periods, it.values) if v is not None}


def _verify_and_demote(res: MappingResult) -> None:
    """(a) gross_profit ≈ revenue − cogs; (b) ending_cash roll-forward.

    Failure demotes the weakest mapping in the identity rather than silently
    keeping a wrong row. Integrity tie-outs later re-run these as formal
    tie-outs with evidence; this pass is purely about mapping quality.
    """
    sheets = {it.sheet for it in res.items if it.canonical_id}
    for sheet in sheets:
        rev = res.get("revenue_total", sheet)
        cogs = res.get("cogs_total", sheet)
        gp = res.get("gross_profit", sheet)
        if rev and cogs and gp and rev.sheet == cogs.sheet == gp.sheet == sheet:
            r, c, g = _series(rev), _series(cogs), _series(gp)
            common = [p for p in g if p in r and p in c]
            if len(common) >= 2:
                bad = 0
                for p in common:
                    expect = r[p] - c[p]
                    tol = max(abs(r[p]), abs(c[p]), 1.0) * 0.02
                    if abs(g[p] - expect) > tol:
                        bad += 1
                if bad > len(common) / 2:
                    weakest = min((gp, cogs, rev), key=lambda it: it.confidence)
                    weakest.demoted = True
                    weakest.evidence += (
                        f"; DEMOTED: gross_profit != revenue - cogs in {bad}/{len(common)} periods "
                        f"on sheet {sheet!r}")
                    res.notes.append(
                        f"Mapping verification failed on {sheet!r}: gross profit identity off in "
                        f"{bad}/{len(common)} periods; demoted {weakest.label!r}.")

        end = res.get("ending_cash", sheet)
        chg = res.get("net_change_cash", sheet)
        if end and chg and end.sheet == chg.sheet == sheet:
            e, n = _series(end), _series(chg)
            common = [p for p in end.periods if p in e and p in n]
            if len(common) >= 3:
                bad = 0
                checked = 0
                for prev, cur in zip(common, common[1:]):
                    expect = e[prev] + n[cur]
                    tol = max(abs(e[prev]), abs(e[cur]), 1.0) * 0.02
                    checked += 1
                    if abs(e[cur] - expect) > tol:
                        bad += 1
                if checked and bad == checked:
                    # every period fails -> almost certainly the wrong row mapped
                    weakest = min((end, chg), key=lambda it: it.confidence)
                    weakest.demoted = True
                    weakest.evidence += "; DEMOTED: cash roll-forward fails in every period"
                    res.notes.append(
                        f"Mapping verification failed on {sheet!r}: ending_cash roll-forward off in "
                        f"all {checked} transitions; demoted {weakest.label!r}.")


# ---------------------------------------------------------------------------
# Reuse path: rebuild a MappingResult from a previously computed WorkbookMap
# ---------------------------------------------------------------------------

def mapping_from_workbook_map(wbd: WorkbookData, structure: StructureResult,
                              wm) -> MappingResult:
    """Rehydrate mappings from a prior run's workbook_map.json so a host
    engine can parse once and analyze many ways. Values are re-read from the
    workbook; only the row→ontology assignment is reused."""
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

    res = MappingResult()
    res.primary_sheet = wm.primary_statement_sheet
    for rm in wm.row_mappings:
        if rm.demoted:
            continue
        sd = wbd.sheets.get(rm.sheet)
        if sd is None:
            res.notes.append(f"precomputed mapping references missing sheet {rm.sheet!r}; skipped")
            continue
        periods: list[str] = []
        coords: list[tuple[int, int]] = []
        values: list[Optional[float]] = []
        for period, ref in rm.cells.items():
            try:
                _, cell = ref.rsplit("!", 1)
                cl, row = coordinate_from_string(cell)
                col = column_index_from_string(cl)
            except Exception:
                continue
            periods.append(period)
            coords.append((row, col))
            rec = sd.cell(row, col)
            values.append(_coerce_value(rec.value) if rec else None)
        item = LineItem(
            sheet=rm.sheet, index=rm.row, label=rm.label_text, depth=0, section="",
            periods=periods, coords=coords, values=values,
            percentish=rm.canonical_id.endswith("_pct"),
            n_numeric=sum(1 for v in values if v is not None),
            n_formula=0,
        )
        item.canonical_id = rm.canonical_id
        item.instance = rm.instance
        item.subtags = list(rm.subtags)
        item.confidence = rm.confidence
        item.evidence = (rm.evidence or "") + " [reused from precomputed workbook_map]"
        res.items.append(item)
    _verify_and_demote(res)
    return res

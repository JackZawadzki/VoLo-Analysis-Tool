"""Annotation-table builder — the analyst worksheet view.

For each derived metric, reconstruct exactly what an analyst lays out by hand:
the company's own source rows IN FULL (so the number can be cross-checked),
then the NEW derived row IN FULL directly beneath, with the values that jump
out highlighted and a hover comment explaining why.

Source rows are recovered from each metric's `inputs` (cell citations) via the
mapping, so nothing is re-computed — the same cells the metric consumed are the
cells shown.
"""
from __future__ import annotations

import re
from typing import Optional

from .findings import FindingsResult
from .highlight import compute_highlights
from .ingest import WorkbookData
from .mapping import LineItem, MappingResult
from .metrics import MetricsResult
from .schema import (
    AnnotatedCell,
    AnnotatedRow,
    AnnotationTable,
    CellKind,
    DerivedMetric,
    Finding,
    Severity,
    format_ref,
)
from .structure import StructureResult

# metric_id -> (family key, human family title). Four broad groups that share
# the same source rows; order here drives report order. Longest matching key wins
# (see _family), so specific ids land in the right bucket.
_GROWTH = "Growth & revenue quality"
_MARGIN = "Margins & profitability"
_CASH = "Cash, burn & capital"
_VAL = "Valuation, returns & team"

FAMILIES: list[tuple[str, str]] = [
    # 1) top line: how fast it grows and how real the revenue is
    ("growth", _GROWTH),
    ("cagr", _GROWTH),
    ("revenue_mix", _GROWTH),
    ("segment", _GROWTH),
    ("terminal_concentration", _GROWTH),
    ("blended_asp", _GROWTH),
    ("grant", _GROWTH),
    ("revenue_ex_grants", _GROWTH),
    ("capacity", _GROWTH),
    ("market_share", _GROWTH),
    ("operational", _GROWTH),
    ("revenue_per_capacity", _GROWTH),
    # 2) profitability and unit economics
    ("margin", _MARGIN),
    ("opex_over_revenue", _MARGIN),
    ("opex_ex_capex", _MARGIN),
    ("ebitda_before", _MARGIN),
    ("rule_of_40", _MARGIN),
    ("operating_leverage", _MARGIN),
    ("tax", _MARGIN),
    ("effective_tax", _MARGIN),
    # 3) cash, burn, capital intensity
    ("runway", _CASH),
    ("minimum_cash", _CASH),
    ("avg_monthly_cash", _CASH),
    ("cumulative", _CASH),
    ("cushion", _CASH),
    ("peak", _CASH),
    ("capex", _CASH),
    ("working_capital", _CASH),
    ("receivables", _CASH),
    ("ar_", _CASH),
    ("inventory", _CASH),
    ("cash_conversion", _CASH),
    # 4) value created and the team behind it
    ("exit_value", _VAL),
    ("present_value", _VAL),
    ("revenue_per_employee", _VAL),
    ("implied_fte", _VAL),
    ("payroll", _VAL),
    ("llm_directed", _GROWTH),
]


def _family(metric_id: str) -> tuple[str, str]:
    # most specific (longest) matching key wins, so e.g. "revenue_ex_grants_growth"
    # maps to Grant dependence (via "revenue_ex_grants") rather than Growth ("growth")
    best: Optional[tuple[str, str]] = None
    for key, title in FAMILIES:
        if metric_id.startswith(key) or key in metric_id:
            if best is None or len(key) > len(best[0]):
                best = (key, title)
    return best or ("other", "Other")


# Headline calcs that always tell the trajectory story (good or bad) and are
# worth showing even with no flag firing. Everything else must earn its place
# by carrying a flag (a highlight or a finding) — so the report stays selective:
# every block reveals something, nothing is filler.
_HEADLINE = {
    "revenue_growth", "revenue_cagr", "gross_margin", "ebitda_margin",
    "opex_over_revenue", "grant_share_of_revenue", "revenue_ex_grants_growth",
    "runway_forward_months", "burn_multiple", "cash_conversion",
    "terminal_concentration", "cushion_ratio", "rule_of_40",
    "operational_contracted_ratio",
    # 2.3 — where diligence lives: mix, survival, conversion, concentration
    "revenue_mix_recurring", "revenue_mix_one_time", "revenue_mix_grant",
    "minimum_cash", "cash_conversion_cumulative",
}


def _is_headline(metric_id: str) -> bool:
    return metric_id in _HEADLINE or metric_id.startswith("revenue_mix_")


def _worth_a_table(t: AnnotationTable) -> bool:
    # A calc earns a place only if it reveals something: a finding, a genuinely
    # notable standout (high/critical — not a routine medium "largest jump"), an
    # LLM-directed company-specific calc, or a headline trajectory metric.
    strong_hl = t.derived_row is not None and any(
        c.highlighted and c.severity is not None and c.severity.value in ("critical", "high")
        for c in t.derived_row.cells)
    return bool(t.related_finding_ids or strong_hl or t.llm_directed
                or _is_headline(t.metric_id)
                or t.metric_id.startswith("market_share_recomputed"))


def _is_percentish(m: DerivedMetric) -> bool:
    if m.units in ("fraction", "x (period growth)"):
        return True
    return bool(re.search(r"margin|share|_growth|effective_tax|ratio|opex_over|conversion", m.metric_id))


def _row_index(mapping: MappingResult) -> tuple[dict[str, LineItem], dict[str, tuple[str, int]]]:
    """(row-key -> LineItem, cell-ref -> row-key)."""
    by_key: dict[str, LineItem] = {}
    ref_to_key: dict[str, tuple[str, int]] = {}
    for it in mapping.items:
        key = f"{it.sheet}#{it.index}"
        by_key[key] = it
        for ref in it.refs:
            ref_to_key[ref] = key
    return by_key, ref_to_key


def _kind_lookup(structure: StructureResult):
    def is_input(sheet: str, row: int, col: int) -> bool:
        cen = structure.censuses.get(sheet)
        return cen is not None and cen.kind(row, col) == CellKind.typed_input
    return is_input


def _source_row(it: LineItem, periods: list[str], is_input, transposed: bool = False) -> AnnotatedRow:
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
    own = {p: (it.coords[i], it.values[i]) for i, p in enumerate(it.periods)}
    cells: list[AnnotatedCell] = []
    for p in periods:
        if p in own:
            (r, c), v = own[p]
            cells.append(AnnotatedCell(period=p, value=v, ref=format_ref(it.sheet, _a1(r, c)),
                                       is_input=is_input(it.sheet, r, c)))
        else:
            cells.append(AnnotatedCell(period=p, value=None, ref=None))
    # precise, navigable citation: sheet + Excel row (e.g. "Lithios_TopCo · row 31").
    # For transposed (periods-in-rows) models, it.index is a COLUMN, so cite the
    # column span instead of a row span.
    if transposed:
        from openpyxl.utils import get_column_letter
        col = get_column_letter(it.index)
        coord = format_ref(it.sheet, f"{col}:{col}")
    else:
        coord = format_ref(it.sheet, f"{it.index}:{it.index}")
    return AnnotatedRow(label=it.label.strip(), sheet=it.sheet, row_index=it.index,
                        coordinate=coord, kind="source",
                        canonical_id=it.canonical_id, units=None,
                        is_percent=it.percentish, cells=cells)


def _a1(row: int, col: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(col)}{row}"


def build_annotation_tables(
    wbd: WorkbookData,
    structure: StructureResult,
    mapping: MappingResult,
    metrics: MetricsResult,
    findings: FindingsResult,
) -> list[AnnotationTable]:
    sheet = mapping.primary_sheet
    if sheet is None:
        return []
    axis = structure.primary_axis(sheet)
    periods = list(axis.periods) if axis else []
    if not periods:
        return []

    by_key, ref_to_key = _row_index(mapping)
    is_input = _kind_lookup(structure)

    # Structural / tool-coverage findings are about the WORKBOOK as a whole, not
    # any one calculation — they must never become a calc's flag (they belong in
    # Model-level flags). A calc's flag should be about that calc's own behaviour.
    _MODEL_LEVEL_ONLY = {"unit_scale_inconsistency", "data_quality", "arithmetic_integrity"}

    # finding lookup: metric_id-or-row -> finding ids
    fk_by_metric: dict[str, list[str]] = {}
    fk_by_ref: dict[str, list[str]] = {}
    for f in findings.findings:
        if f.category in _MODEL_LEVEL_ONLY:
            continue
        if f.quantified_impact and f.quantified_impact.metric:
            base = f.quantified_impact.metric.split("[")[0]
            fk_by_metric.setdefault(base, []).append(f.id)
        for ref in f.evidence_cells:
            fk_by_ref.setdefault(ref, []).append(f.id)

    tables: list[AnnotationTable] = []
    for m in metrics.metrics:
        # resolve distinct source rows from the metric's input cells
        seen: list[str] = []
        for ref in m.inputs:
            key = ref_to_key.get(ref)
            if key and key not in seen:
                seen.append(key)
        source_items = [by_key[k] for k in seen]
        if not source_items and m.scalar is None and not m.series:
            continue

        from .schema import Orientation
        transposed = axis.orientation == Orientation.periods_in_rows
        src_rows = [_source_row(it, periods, is_input, transposed) for it in source_items[:6]]

        # derived row
        is_pct = _is_percentish(m)
        d_cells: list[AnnotatedCell] = []
        if m.series:
            highs = {h.period: h for h in compute_highlights(m.metric_id, m.series, periods)}
            for p in periods:
                v = m.series.get(p)
                h = highs.get(p)
                d_cells.append(AnnotatedCell(
                    period=p, value=v, ref=None,
                    highlighted=h is not None and v is not None,
                    comment=h.comment if h else None,
                    severity=h.severity if h else None))
        elif m.scalar is not None:
            d_cells.append(AnnotatedCell(period=periods[-1], value=m.scalar, ref=None,
                                         highlighted=False))
        derived = AnnotatedRow(label=m.label, kind="derived", canonical_id=m.metric_id,
                               units=m.units, is_percent=is_pct, cells=d_cells)
        n_hl = sum(1 for c in d_cells if c.highlighted)

        # related findings
        fids: list[str] = list(fk_by_metric.get(m.metric_id, []))
        for it in source_items:
            for ref in it.refs:
                fids.extend(fk_by_ref.get(ref, []))
        fids = sorted(set(fids))

        fam_key, fam_title = _family(m.metric_id)
        tables.append(AnnotationTable(
            id=f"T_{m.metric_id}".replace("::", "__").replace(" ", "_")[:80],
            title=m.label,
            metric_id=m.metric_id,
            family=fam_title,
            rationale=m.notes or m.applicability,
            computation=m.computation,
            periods=periods,
            source_rows=src_rows,
            derived_row=derived,
            related_finding_ids=fids,
            llm_directed=m.metric_id.startswith("llm_directed"),
            n_highlights=n_hl,
        ))

    tables = [t for t in tables if _worth_a_table(t)]

    # order: family order, then highlighted/finding-linked first within family
    fam_order = {title: i for i, (_, title) in enumerate(FAMILIES)}
    fam_order["Other"] = len(FAMILIES)
    tables.sort(key=lambda t: (fam_order.get(t.family, 99),
                               0 if (t.n_highlights or t.related_finding_ids) else 1,
                               -t.n_highlights, t.metric_id))
    return tables

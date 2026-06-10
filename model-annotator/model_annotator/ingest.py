"""Workbook ingestion: dual openpyxl load (cached values + formula text).

Produces the internal :class:`WorkbookData` cell model that every later phase
consumes. No formula evaluation happens anywhere in this tool — we read the
cached values Excel last saved, and the formula *text* for graph analysis.
"""
from __future__ import annotations

import hashlib
import logging
import zipfile
from dataclasses import dataclass, field
from typing import Any, Iterator, Optional

import openpyxl
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

DEFAULT_MAX_CELLS = 2_000_000

ERROR_TOKENS = frozenset({
    "#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!", "#SPILL!", "#CALC!",
})


class ModelParseError(Exception):
    """Raised when the workbook cannot be parsed at all (CLI exit code 2)."""


def is_number(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def is_error_value(v: Any) -> bool:
    return isinstance(v, str) and v.strip().upper() in ERROR_TOKENS


@dataclass
class CellRecord:
    value: Any = None                 # cached value from the data_only pass
    formula: Optional[str] = None     # formula text ("=...") if this is a formula cell
    number_format: str = "General"
    fill_key: Optional[str] = None    # "rgb:FFD9D9D9" | "theme:0/tint:-0.05" | "idx:22"
    font_rgb: Optional[str] = None
    comment: Optional[str] = None
    is_merged: bool = False
    merged_anchor: Optional[tuple[int, int]] = None  # (row, col) of the range's top-left

    @property
    def is_formula(self) -> bool:
        return self.formula is not None

    @property
    def is_error(self) -> bool:
        return is_error_value(self.value)


@dataclass
class SheetData:
    name: str
    visibility: str = "visible"       # visible | hidden | veryHidden
    max_row: int = 0
    max_col: int = 0
    cells: dict[tuple[int, int], CellRecord] = field(default_factory=dict)
    hidden_rows: set[int] = field(default_factory=set)
    hidden_cols: set[int] = field(default_factory=set)
    merged_ranges: list[str] = field(default_factory=list)
    truncated: bool = False

    def cell(self, row: int, col: int) -> Optional[CellRecord]:
        return self.cells.get((row, col))

    def value(self, row: int, col: int) -> Any:
        rec = self.cells.get((row, col))
        return rec.value if rec else None

    def iter_cells(self) -> Iterator[tuple[int, int, CellRecord]]:
        for (r, c), rec in self.cells.items():
            yield r, c, rec


@dataclass
class WorkbookData:
    path: str
    sha256: str
    sheets: dict[str, SheetData] = field(default_factory=dict)
    sheet_order: list[str] = field(default_factory=list)
    named_ranges: dict[str, str] = field(default_factory=dict)
    has_macros: bool = False
    n_cells: int = 0
    n_formula_cells: int = 0
    n_formula_cells_without_cached_value: int = 0
    limitations: list[str] = field(default_factory=list)

    @property
    def missing_cached_values(self) -> bool:
        if self.n_formula_cells == 0:
            return False
        frac = self.n_formula_cells_without_cached_value / self.n_formula_cells
        return self.n_formula_cells_without_cached_value > 50 and frac > 0.2

    def sheet(self, name: str) -> SheetData:
        return self.sheets[name]

    def get_cell(self, sheet: str, row: int, col: int) -> Optional[CellRecord]:
        sd = self.sheets.get(sheet)
        return sd.cell(row, col) if sd else None

    def get_value_by_a1(self, sheet: str, a1: str) -> Any:
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        col_letter, row = coordinate_from_string(a1)
        rec = self.get_cell(sheet, row, column_index_from_string(col_letter))
        return rec.value if rec else None


# ---------------------------------------------------------------------------
# Style normalization
# ---------------------------------------------------------------------------

def _fill_key(cell) -> Optional[str]:
    f = cell.fill
    if f is None or f.fill_type in (None, "none"):
        return None
    c = f.start_color
    if c is None:
        return None
    if c.type == "rgb" and c.rgb and c.rgb != "00000000":
        return f"rgb:{c.rgb}"
    if c.type == "theme" and c.theme is not None:
        return f"theme:{c.theme}/tint:{round(c.tint or 0.0, 3)}"
    if c.type == "indexed" and c.indexed not in (None, 64, 65):
        return f"idx:{c.indexed}"
    return None


def _font_rgb(cell) -> Optional[str]:
    fo = cell.font
    if fo and fo.color and fo.color.type == "rgb" and fo.color.rgb and fo.color.rgb != "FF000000":
        return fo.color.rgb
    return None


def _formula_text(raw: Any) -> Optional[str]:
    """Extract formula text from a data_only=False cell value, else None."""
    if isinstance(raw, str) and raw.startswith("="):
        return raw
    # openpyxl represents array formulas as ArrayFormula objects
    cls = type(raw).__name__
    if cls == "ArrayFormula":
        text = getattr(raw, "text", None)
        if isinstance(text, str):
            return text if text.startswith("=") else "=" + text
    if cls == "DataTableFormula":
        return "=TABLE()"
    return None


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def _open_both(path: str):
    try:
        wb_formulas = openpyxl.load_workbook(path, data_only=False, keep_vba=False, rich_text=False)
        wb_values = openpyxl.load_workbook(path, data_only=True, keep_vba=False, rich_text=False)
    except zipfile.BadZipFile as exc:
        raise ModelParseError(
            f"Cannot open {path!r}: not a readable .xlsx/.xlsm zip archive "
            "(password-protected/encrypted workbooks are not supported)."
        ) from exc
    except Exception as exc:  # openpyxl raises a variety of exceptions on corrupt files
        raise ModelParseError(f"Cannot open {path!r}: {exc}") from exc
    return wb_formulas, wb_values


def _sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def load_workbook_dual(path: str, max_cells: int = DEFAULT_MAX_CELLS) -> WorkbookData:
    """Load a workbook into the aligned cell model.

    Never modifies the input file. Enforces ``max_cells`` as a safety cap:
    sheets beyond the cap are truncated/skipped with a limitation recorded.
    """
    wb_f, wb_v = _open_both(path)

    data = WorkbookData(path=path, sha256=_sha256(path))
    data.has_macros = str(path).lower().endswith(".xlsm") or getattr(wb_f, "vba_archive", None) is not None
    if data.has_macros:
        data.limitations.append(
            "Workbook contains macros (.xlsm); macros were NOT executed. "
            "Any macro-driven values are whatever was last saved."
        )

    # Named ranges (workbook-level)
    try:
        for name, defn in wb_f.defined_names.items():
            data.named_ranges[name] = str(defn.attr_text or "")
    except Exception:  # malformed defined names should never kill the parse
        log.warning("could not read defined names", exc_info=True)

    budget = max_cells
    for ws_f in wb_f.worksheets:
        name = ws_f.title
        try:
            ws_v = wb_v[name]
        except KeyError:
            ws_v = None
        sd = SheetData(
            name=name,
            visibility=ws_f.sheet_state,
            max_row=ws_f.max_row or 0,
            max_col=ws_f.max_column or 0,
        )
        data.sheets[name] = sd
        data.sheet_order.append(name)

        if budget <= 0:
            sd.truncated = True
            data.limitations.append(f"Sheet '{name}' skipped entirely: workbook exceeds the {max_cells:,}-cell safety cap.")
            continue

        # merged cells: value lives at the top-left anchor
        anchors: dict[tuple[int, int], tuple[int, int]] = {}
        for rng in ws_f.merged_cells.ranges:
            sd.merged_ranges.append(str(rng))
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    anchors[(r, c)] = (rng.min_row, rng.min_col)

        for rd, dim in ws_f.row_dimensions.items():
            if dim.hidden:
                sd.hidden_rows.add(rd)
        for cd, dim in ws_f.column_dimensions.items():
            if dim.hidden:
                try:
                    from openpyxl.utils import column_index_from_string
                    sd.hidden_cols.add(column_index_from_string(cd))
                except ValueError:
                    pass

        n_read = 0
        for row_f in ws_f.iter_rows():
            if budget <= 0:
                sd.truncated = True
                data.limitations.append(
                    f"Sheet '{name}' truncated at the {max_cells:,}-cell safety cap; rows beyond row "
                    f"{row_f[0].row if row_f else '?'} were not read."
                )
                break
            for cell_f in row_f:
                raw = cell_f.value
                comment = cell_f.comment.text if cell_f.comment else None
                fill = _fill_key(cell_f)
                if raw is None and comment is None and fill is None:
                    continue
                budget -= 1
                n_read += 1
                formula = _formula_text(raw)
                if formula is not None:
                    cached = ws_v.cell(row=cell_f.row, column=cell_f.column).value if ws_v is not None else None
                    data.n_formula_cells += 1
                    if cached is None:
                        data.n_formula_cells_without_cached_value += 1
                else:
                    cached = raw
                rec = CellRecord(
                    value=cached,
                    formula=formula,
                    number_format=cell_f.number_format or "General",
                    fill_key=fill,
                    font_rgb=_font_rgb(cell_f),
                    comment=comment,
                )
                key = (cell_f.row, cell_f.column)
                if key in anchors:
                    rec.is_merged = True
                    rec.merged_anchor = anchors[key]
                sd.cells[key] = rec
        data.n_cells += n_read

    wb_f.close()
    wb_v.close()

    if data.missing_cached_values:
        data.limitations.append(
            f"{data.n_formula_cells_without_cached_value:,} of {data.n_formula_cells:,} formula cells have no "
            "cached value (file likely saved without calculation). Numeric analysis of those cells is impossible "
            "without opening and re-saving the model in Excel; downstream results are degraded."
        )

    hidden = [s.name for s in data.sheets.values() if s.visibility != "visible"]
    if hidden:
        log.info("hidden sheets included in analysis: %s", hidden)

    return data


def a1(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"

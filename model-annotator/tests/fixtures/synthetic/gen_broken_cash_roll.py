"""Synthetic fixture: broken cash roll-forward.

A small annual model whose statements are internally consistent everywhere
EXCEPT one mid-horizon Ending Cash value: 2027 ending cash is hardcoded
11,700 above what Beginning Cash + Net Change in Cash produces. The 2028+
years roll forward from the broken printed value (as a real model would if
someone fat-fingered one ending-cash cell), so exactly one transition of the
cash roll-forward tie-out fails — far above integrity.py's 0.5%-of-scale
tolerance (series max ~61.6k -> tol ~308).

Beginning-cash continuity still holds (beginning[t] = printed ending[t-1]),
the gross-profit identity holds, and every cell is a literal value so no
data-quality gate fires.
"""
from __future__ import annotations

import openpyxl

NAME = "broken_cash_roll"

YEARS = [2024, 2025, 2026, 2027, 2028, 2029, 2030]

REVENUE = [100, 600, 2_400, 6_500, 14_000, 27_000, 48_000]
COGS = [40, 240, 960, 2_600, 5_600, 10_800, 19_200]
OPEX = [1_800, 2_600, 3_800, 5_200, 6_800, 8_600, 10_500]

BROKEN_YEAR = 2027
BREAK_AMOUNT = 11_700.0
INITIAL_CASH = 30_000.0


def _series() -> dict[str, list[float]]:
    gross = [r - c for r, c in zip(REVENUE, COGS)]
    ebitda = [g - o for g, o in zip(gross, OPEX)]
    net_change = list(ebitda)  # all-cash model: no capex/financing rows

    beginning: list[float] = []
    ending: list[float] = []
    prev_end = INITIAL_CASH
    for year, chg in zip(YEARS, net_change):
        beginning.append(prev_end)
        end = prev_end + chg
        if year == BROKEN_YEAR:
            end += BREAK_AMOUNT  # the single broken printed value
        ending.append(end)
        prev_end = end  # later years roll from the printed (broken) number

    return {
        "Total Revenue": [float(v) for v in REVENUE],
        "Total COGS": [float(v) for v in COGS],
        "Gross Profit": [float(v) for v in gross],
        "Total Operating Expenses": [float(v) for v in OPEX],
        "EBITDA": [float(v) for v in ebitda],
        "Net Change in Cash": [float(v) for v in net_change],
        "Beginning Cash": beginning,
        "Ending Cash": ending,
    }


def build(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"

    for j, year in enumerate(YEARS):
        ws.cell(row=1, column=2 + j, value=year)

    for i, (label, values) in enumerate(_series().items()):
        row = 2 + i
        ws.cell(row=row, column=1, value=label)
        for j, v in enumerate(values):
            ws.cell(row=row, column=2 + j, value=v)

    wb.save(path)

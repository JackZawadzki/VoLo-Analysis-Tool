"""Fixture: one_way_margin.

One-way margin expansion: on an "Assumptions" sheet, the typed cost-side row
"Unit cost learning factor" declines ~-29% over the horizon while the typed
price-side row "ASP" is held exactly flat. Every learning effect accrues to
the company, none to customers.

Targets _rule_one_way_margin in model_annotator/findings.py:
  - scans ctx.mapping.items on sheets whose role is assumptions /
    cost_buildup / schedule (sheet named "Assumptions" -> assumptions via the
    name hint), keeping rows where n_formula <= n_numeric / 2 (all literals
    here) with >= 4 non-None values;
  - cost side: label matches cost|learning|decline|reduction|efficiency and
    (last-first)/|first| < -0.10  -> "Unit cost learning factor" 1.00 -> 0.71;
  - price side: label matches price|asp (and not the cost regex) with
    (max-min)/|max| < 0.02        -> "ASP" 4500 flat every year;
  -> one_way_margin_expansion, severity medium.

The P&L sheet carries mappable literal statements (revenue / COGS / gross
profit / opex components+total / EBITDA / cash roll) whose identities hold
exactly, so no tie-out fires and exit code stays 0. Zero formula cells.
"""
from __future__ import annotations

import openpyxl

NAME = "one_way_margin"

YEARS = [2024, 2025, 2026, 2027, 2028, 2029, 2030]

# --- Assumptions ($/unit and unitless factors), all typed literals ---------
UNITS_SOLD = [90, 210, 420, 760, 1250, 1900, 2700]
ASP = [4500] * 7                                            # exactly flat
LEARNING = [1.00, 0.93, 0.87, 0.81, 0.77, 0.74, 0.71]       # -29% over horizon
BASE_UNIT_COST = 2.7                                        # $K per unit in 2024

# --- P&L ($000s), consistent with the assumptions ---------------------------
REVENUE = [u * 4.5 for u in UNITS_SOLD]                     # ASP $4,500 -> $K
COGS = [round(u * BASE_UNIT_COST * f, 1) for u, f in zip(UNITS_SOLD, LEARNING)]
PAYROLL = [600, 820, 1080, 1380, 1720, 2100, 2520]
RENT = [120, 130, 140, 150, 165, 180, 200]
SALES_MARKETING = [200, 320, 480, 680, 900, 1150, 1400]


def build(path: str) -> None:
    wb = openpyxl.Workbook()

    # ---- Sheet 1: P&L (statements, all literals) ----------------------------
    pl = wb.active
    pl.title = "P&L"
    pl["A1"] = "P&L ($000s)"
    for j, yr in enumerate(YEARS):
        pl.cell(row=1, column=2 + j, value=yr)

    gross_profit = [r - c for r, c in zip(REVENUE, COGS)]
    opex_total = [p + r + s for p, r, s in zip(PAYROLL, RENT, SALES_MARKETING)]
    ebitda = [g - o for g, o in zip(gross_profit, opex_total)]
    net_change = list(ebitda)
    beginning = [8000.0]
    for ch in net_change[:-1]:
        beginning.append(beginning[-1] + ch)
    ending = [b + ch for b, ch in zip(beginning, net_change)]

    pl_rows = [
        (3, "Total Revenue", REVENUE),
        (4, "Total COGS", COGS),
        (5, "Gross Profit", gross_profit),
        # row 6 intentionally blank
        (7, "Payroll", PAYROLL),
        (8, "Rent", RENT),
        (9, "Sales & Marketing", SALES_MARKETING),
        (10, "Total Operating Expenses", opex_total),
        (11, "EBITDA", ebitda),
        # row 12 intentionally blank
        (13, "Net Change in Cash", net_change),
        (14, "Beginning Cash", beginning),
        (15, "Ending Cash", ending),
    ]
    for r, label, series in pl_rows:
        pl.cell(row=r, column=1, value=label)
        for j, v in enumerate(series):
            pl.cell(row=r, column=2 + j, value=v)

    # ---- Sheet 2: Assumptions (typed drivers, same period axis) -------------
    asm = wb.create_sheet("Assumptions")
    asm["A1"] = "Assumptions"
    for j, yr in enumerate(YEARS):
        asm.cell(row=1, column=2 + j, value=yr)

    asm_rows = [
        (3, "Units Sold", UNITS_SOLD),
        (4, "ASP", ASP),
        (5, "Unit cost learning factor", LEARNING),
    ]
    for r, label, series in asm_rows:
        asm.cell(row=r, column=1, value=label)
        for j, v in enumerate(series):
            asm.cell(row=r, column=2 + j, value=v)

    wb.save(path)

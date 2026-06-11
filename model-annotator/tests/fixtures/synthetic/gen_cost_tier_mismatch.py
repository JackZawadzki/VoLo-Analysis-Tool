"""Cost-tier mismatch: the cost sheet defines two cost levels for the same
product ("Widget cost - premium tier" = 120, "Widget cost - standard tier"
= 70) and prices the product at the premium tier (ASP 250), but the P&L's
Total COGS chain references the cheap standard-tier row. Gross margin is
overstated by construction.

Workbook mechanics (what the tool actually reads):
- "Cost Model" sheet name + "tier"/"cost model" labels -> SheetRole.cost_buildup.
- "Total COGS" maps via its 6 literal cells (n_numeric >= 2); only the LAST
  TWO period cells are formulas, whose *text* references 'Cost Model'!$B$4
  (the cheap row) — exactly the refs[-2:] cells _rule_cost_tier_mismatch
  walks upstream. 2 valueless formula cells stay far below the 50-cell
  missing-cached-values gate.
- All other rows are literals and internally consistent (GP = Rev - COGS on
  cached periods, Opex = Payroll + Rent), so no arithmetic tie-out fires and
  the exit code stays 0 (no cash rows -> cash roll not applicable).
"""
from openpyxl import Workbook

NAME = "cost_tier_mismatch"

YEARS = list(range(2024, 2032))  # B..I

REVENUE = [500, 1200, 2600, 5200, 9800, 16800, 26500, 38500]      # $K, ASP 250
UNITS = [2000, 4800, 10400, 20800, 39200, 67200, 106000, 154000]  # rev = units*250/1000
COGS_LIT = [140, 336, 728, 1456, 2744, 4704]                      # units * $70 / 1000, first 6 only
GROSS = [360, 864, 1872, 3744, 7056, 12096, 19080, 27720]         # rev - cogs (cheap tier throughout)
PAYROLL = [400, 700, 1100, 1700, 2500, 3600, 5000, 6800]
RENT = [100, 150, 220, 330, 480, 700, 1000, 1400]
OPEX = [p + r for p, r in zip(PAYROLL, RENT)]
EBITDA = [g - o for g, o in zip(GROSS, OPEX)]


def build(path: str) -> None:
    wb = Workbook()

    pl = wb.active
    pl.title = "P&L"
    pl["A1"] = "P&L"
    for i, y in enumerate(YEARS):
        pl.cell(row=1, column=2 + i, value=y)

    def row(r: int, label: str, vals) -> None:
        pl.cell(row=r, column=1, value=label)
        for i, v in enumerate(vals):
            pl.cell(row=r, column=2 + i, value=v)

    row(2, "Total Revenue", REVENUE)
    row(3, "Units Sold", UNITS)
    # Total COGS: literals 2024-2029, formula text only for 2030/2031 (H4, I4)
    # — the chain into the CHEAP standard-tier cost row.
    row(4, "Total COGS", COGS_LIT)
    pl["H4"] = "='Cost Model'!$B$4*H3/1000"
    pl["I4"] = "='Cost Model'!$B$4*I3/1000"
    row(5, "Gross Profit", GROSS)
    row(6, "Payroll", PAYROLL)
    row(7, "Rent", RENT)
    row(8, "Total Operating Expenses", OPEX)
    row(9, "EBITDA", EBITDA)

    cm = wb.create_sheet("Cost Model")
    cm["A1"] = "Cost Model — unit cost build-up"
    cm["A3"] = "Widget cost - premium tier"
    cm["B3"] = 120
    cm["A4"] = "Widget cost - standard tier"
    cm["B4"] = 70
    cm["A6"] = "ASP (premium tier)"
    cm["B6"] = 250

    wb.save(path)

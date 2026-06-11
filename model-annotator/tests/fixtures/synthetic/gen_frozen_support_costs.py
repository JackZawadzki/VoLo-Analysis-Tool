"""Fixture: frozen_support_costs.

Revenue grows >10x across 7 years while the opex components
"Customer Support" and "Rent" are held EXACTLY constant every year and
"Payroll" grows. "Total Operating Expenses" is the literal sum of the three
components each year, and the gross-profit identity holds exactly, so no
integrity tie-out fires. Every number is a literal (openpyxl writes no
cached formula values).

Targets _rule_frozen_costs in model_annotator/findings.py:
  - revenue_total mapped, >=5 periods with rev > EPS, last/first >= 10x
  - opex_component items on the primary sheet with (max-min)/max < 0.02
    across >=5 valid periods -> frozen -> high-severity finding citing the
    frozen rows at the first and last valid periods.
"""
from __future__ import annotations

import openpyxl

NAME = "frozen_support_costs"

YEARS = [2024, 2025, 2026, 2027, 2028, 2029, 2030]

# $K-scale startup: ~31x revenue growth over the horizon
REVENUE = [400, 900, 1800, 3400, 5800, 8900, 12500]
COGS = [160, 340, 650, 1150, 1850, 2700, 3600]

SUPPORT = [300] * 7          # frozen (row 6)
RENT = [240] * 7             # frozen (row 7)
PAYROLL = [500, 700, 950, 1250, 1600, 2000, 2450]  # grows (row 8)


def build(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"

    # Row 1: annual period headers starting at column B
    for j, yr in enumerate(YEARS):
        ws.cell(row=1, column=2 + j, value=yr)

    gross_profit = [r - c for r, c in zip(REVENUE, COGS)]
    opex_total = [s + r + p for s, r, p in zip(SUPPORT, RENT, PAYROLL)]
    ebitda = [g - o for g, o in zip(gross_profit, opex_total)]

    rows = [
        (2, "Total Revenue", REVENUE),
        (3, "Total COGS", COGS),
        (4, "Gross Profit", gross_profit),
        # row 5 intentionally blank
        (6, "Customer Support", SUPPORT),
        (7, "Rent", RENT),
        (8, "Payroll", PAYROLL),
        (9, "Total Operating Expenses", opex_total),
        (10, "EBITDA", ebitda),
    ]
    for r, label, series in rows:
        ws.cell(row=r, column=1, value=label)
        for j, v in enumerate(series):
            ws.cell(row=r, column=2 + j, value=v)

    wb.save(path)

"""Forward-runway cliff with a typed rescue raise.

Ending cash entering 2026 covers ~0.2 months of 2027's opex (forward
runway < 3 months), and a typed (non-formula) "Equity Proceeds" raise of
8,000 lands in 2027 to restore cash. The cash roll (Beginning / Net
Change / Ending) is all-literal and exact, gross profit and the opex
component sum tie out, and there are zero formula cells — so every
integrity tie-out passes and trust stays 1.0.
"""
from openpyxl import Workbook

NAME = "runway_cliff_rescue"

YEARS = [2024, 2025, 2026, 2027, 2028, 2029, 2030]

# ($K) Revenue growth decelerates 75% -> 25% (no hockey-stick run, no
# whiplash, end/start < 10x). EBITDA = GP - opex; Net Change = EBITDA +
# Equity Proceeds; cash roll is exact.
ROWS = [
    ("Total Revenue",            [800, 1400, 2240, 3360, 4700, 6100, 7600]),
    ("Total COGS",               [240, 420, 672, 1008, 1410, 1830, 2280]),
    ("Gross Profit",             [560, 980, 1568, 2352, 3290, 4270, 5320]),
    ("Payroll",                  [1500, 1900, 2400, 3000, 3600, 4200, 4800]),
    ("Rent",                     [200, 250, 300, 360, 420, 480, 540]),
    ("Total Operating Expenses", [1700, 2150, 2700, 3360, 4020, 4680, 5340]),
    ("EBITDA",                   [-1140, -1170, -1132, -1008, -730, -410, -20]),
    ("Equity Proceeds",          [3000, 0, 0, 8000, 0, 0, 0]),
    ("Net Change in Cash",       [1860, -1170, -1132, 6992, -730, -410, -20]),
    ("Beginning Cash",           [500, 2360, 1190, 58, 7050, 6320, 5910]),
    # fwd runway 2026: 58 / (3360/12) = 0.21 months -> cliff
    ("Ending Cash",              [2360, 1190, 58, 7050, 6320, 5910, 5890]),
]


def build(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    for j, y in enumerate(YEARS):
        ws.cell(row=1, column=2 + j, value=y)
    for i, (label, vals) in enumerate(ROWS):
        r = 2 + i
        ws.cell(row=r, column=1, value=label)
        for j, v in enumerate(vals):
            ws.cell(row=r, column=2 + j, value=v)
    wb.save(path)

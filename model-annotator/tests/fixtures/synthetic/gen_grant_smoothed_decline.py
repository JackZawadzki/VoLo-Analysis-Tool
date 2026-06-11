"""Synthetic fixture: grant revenue smooths a commercial decline.

Product Revenue declines ~-10% in 2027 while Government Grants step up so
Total Revenue still grows. Grant share of revenue exceeds 30% in four years
(peak ~64% in 2027). All cells are literals; Total Revenue = Product Revenue
+ Government Grants, Gross Profit = Total Revenue - Total COGS exactly.
"""
from openpyxl import Workbook

NAME = "grant_smoothed_decline"

YEARS = [2024, 2025, 2026, 2027, 2028, 2029, 2030]

PRODUCT_REV = [800, 1200, 1700, 1530, 2100, 2900, 3800]   # -10.0% in 2027
GRANTS = [250, 600, 1300, 2700, 2600, 2300, 1900]          # steps up into 2027
TOTAL_REV = [p + g for p, g in zip(PRODUCT_REV, GRANTS)]   # grows every year
COGS = [360, 540, 765, 690, 945, 1305, 1710]               # ~45% of product rev
GROSS_PROFIT = [t - c for t, c in zip(TOTAL_REV, COGS)]


def build(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    ws["A1"] = "P&L ($000s)"
    for j, year in enumerate(YEARS):
        ws.cell(row=1, column=2 + j, value=year)

    rows = [
        ("Product Revenue", PRODUCT_REV),
        ("Government Grants", GRANTS),
        ("Total Revenue", TOTAL_REV),
        ("Total COGS", COGS),
        ("Gross Profit", GROSS_PROFIT),
    ]
    for i, (label, series) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=label)
        for j, v in enumerate(series):
            ws.cell(row=r, column=2 + j, value=v)

    wb.save(path)

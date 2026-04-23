"""Tests for the deterministic Excel reader.

Uses a synthetic workbook generated at test time — no dependency on any real
financial model, no company-specific assumptions.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from banker.excel_tools import ExcelWorkbook, address_range, row_range


@pytest.fixture
def synthetic_workbook(tmp_path: Path) -> Path:
    """Build a toy two-sheet workbook for testing."""
    wb = Workbook()

    # Sheet 1: P&L-shaped
    s1 = wb.active
    s1.title = "Sheet One"
    s1["A1"] = "Label"
    s1["B1"] = 2026
    s1["C1"] = 2027
    s1["D1"] = 2028
    s1["A2"] = "Revenue"
    s1["B2"] = 100
    s1["C2"] = 200
    s1["D2"] = 400
    s1["A3"] = "Costs"
    s1["B3"] = 50
    s1["C3"] = 80
    s1["D3"] = 140
    s1["A4"] = "Profit"
    s1["B4"] = "=B2-B3"
    s1["C4"] = "=C2-C3"
    s1["D4"] = "=D2-D3"

    # Sheet 2: assumptions-shaped
    s2 = wb.create_sheet("Assumptions Sheet")
    s2["A1"] = "Unit Price"
    s2["B1"] = 25.5
    s2["A2"] = "Churn"
    s2["B2"] = 0.03
    s2["A3"] = "Notes"
    s2["B3"] = "Applies to commercial tier only"

    path = tmp_path / "synthetic.xlsx"
    wb.save(path)
    return path


def test_list_sheets(synthetic_workbook):
    wb = ExcelWorkbook(synthetic_workbook)
    sheets = wb.list_sheets()
    names = [s["name"] for s in sheets]
    assert "Sheet One" in names
    assert "Assumptions Sheet" in names
    sheet_one = next(s for s in sheets if s["name"] == "Sheet One")
    assert sheet_one["max_row"] >= 4
    assert sheet_one["max_col"] >= 4


def test_preview_sheet(synthetic_workbook):
    wb = ExcelWorkbook(synthetic_workbook)
    preview = wb.preview_sheet("Sheet One", rows=5)
    assert preview["rows_returned"] >= 4
    # Grid is [row][col]; row 0 col 0 is A1
    a1 = preview["grid"][0][0]
    assert a1["address"] == "A1"
    assert a1["value"] == "Label"
    # Row 1 col 1 is B2 (value 100)
    b2 = preview["grid"][1][1]
    assert b2["address"] == "B2"
    assert b2["value"] == 100


def test_read_range_row(synthetic_workbook):
    wb = ExcelWorkbook(synthetic_workbook)
    result = wb.read_range("Sheet One", "B2:D2")
    assert result["shape"] == "1d"
    values = [c["value"] for c in result["cells"]]
    assert values == [100, 200, 400]


def test_read_range_rectangle(synthetic_workbook):
    wb = ExcelWorkbook(synthetic_workbook)
    result = wb.read_range("Sheet One", "A1:D3")
    assert result["shape"] == "2d"
    assert result["rows"] == 3
    assert result["cols"] == 4


def test_read_cell_formula(synthetic_workbook):
    wb = ExcelWorkbook(synthetic_workbook)
    # B4 is =B2-B3. openpyxl won't evaluate it (no Excel calc engine),
    # but the formula should be surfaced.
    cell = wb.read_cell("Sheet One", "B4")
    assert cell["address"] == "B4"
    assert cell["formula"] == "=B2-B3"


def test_find_label(synthetic_workbook):
    wb = ExcelWorkbook(synthetic_workbook)
    hits = wb.find_label(r"churn", case_insensitive=True)
    assert any(h["sheet"] == "Assumptions Sheet" and h["text"] == "Churn" for h in hits)


def test_find_label_invalid_regex(synthetic_workbook):
    wb = ExcelWorkbook(synthetic_workbook)
    hits = wb.find_label(r"([")  # broken regex
    assert hits and "error" in hits[0]


def test_missing_sheet_raises(synthetic_workbook):
    wb = ExcelWorkbook(synthetic_workbook)
    with pytest.raises(ValueError):
        wb.preview_sheet("NonexistentSheet")


def test_address_helpers():
    assert address_range("A1", "Z99") == "A1:Z99"
    assert row_range("C12", "I") == "C12:I12"
